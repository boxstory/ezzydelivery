"""
Import Tools

AI tools for automating order imports from configured sources
(OneDrive, Google Sheets, Shopify, WooCommerce) and raw text parsing.
All tools are staff-only.
"""

import logging
import uuid
from decimal import Decimal
from typing import Any, Dict, List, Optional

from django.db.models import Q
from django.utils import timezone

from ai_agent.tools.base import BaseTool, ToolError, register_tool

logger = logging.getLogger(__name__)


def _resolve_business(business_id=None, business_name=None):
    """Resolve a business by ID or name search."""
    from business.models import Business

    if business_id:
        try:
            return Business.objects.get(business_id=business_id)
        except Business.DoesNotExist:
            raise ToolError(f'Business {business_id} not found', 'NOT_FOUND')

    if business_name:
        matches = Business.objects.filter(
            business_name__icontains=business_name,
            business_status='active',
        )
        if matches.count() == 1:
            return matches.first()
        if matches.count() > 1:
            names = [f"{b.business_id}: {b.business_name}" for b in matches[:5]]
            raise ToolError(
                f'Multiple businesses match "{business_name}": {", ".join(names)}. '
                'Please specify business_id.',
                'AMBIGUOUS'
            )
        raise ToolError(f'No active business matching "{business_name}"', 'NOT_FOUND')

    raise ToolError('Provide business_id or business_name', 'MISSING_PARAM')


@register_tool
class ListImportSourcesTool(BaseTool):
    """List all configured import sources for a business."""

    name = 'list_import_sources'
    allowed_roles = ['staff']
    description = '''List all configured import sources for a business.

    Shows:
    - OneDrive sources (label, link, active status, last import info)
    - API integrations (Shopify, WooCommerce, Google Sheets)

    Use this before triggering an import to see what sources are available.
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'business_id': {
                'type': 'integer',
                'description': 'Business ID'
            },
            'business_name': {
                'type': 'string',
                'description': 'Business name (partial match, use if ID unknown)'
            },
        },
    }

    def execute(
        self,
        business_id: Optional[int] = None,
        business_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        from orders.models import OneDriveSource
        from business.models import BusinessApiSettings

        business = _resolve_business(business_id, business_name)

        # OneDrive sources
        od_sources = OneDriveSource.objects.filter(business=business).order_by('-created_at')
        onedrive_list = []
        for s in od_sources:
            onedrive_list.append({
                'id': s.id,
                'label': s.label,
                'is_active': s.is_active,
                'last_import_at': s.last_import_at.isoformat() if s.last_import_at else None,
                'last_import_count': s.last_import_count or 0,
                'last_sheet_name': s.last_sheet_name or '',
                'has_saved_mapping': bool(s.last_column_mapping),
            })

        # API integrations
        api_settings = BusinessApiSettings.objects.filter(
            business=business, is_active=True
        )
        api_list = []
        for api in api_settings:
            api_list.append({
                'id': api.id,
                'type': api.api_type,
                'label': api.get_api_type_display() if hasattr(api, 'get_api_type_display') else api.api_type,
                'site_url': api.site_api_url or '',
                'has_credentials': bool(api.api_key or api.api_access_token),
                'google_sheet_url': getattr(api, 'google_sheet_url', '') or '',
            })

        return {
            'business_id': business.business_id,
            'business_name': business.business_name,
            'onedrive_sources': onedrive_list,
            'api_integrations': api_list,
            'total_sources': len(onedrive_list) + len(api_list),
        }


@register_tool
class ImportFromOneDriveTool(BaseTool):
    """Trigger an automated import from a configured OneDrive source."""

    name = 'import_from_onedrive'
    allowed_roles = ['staff']
    description = '''Import new orders from a configured OneDrive Excel source.

    Automatically:
    1. Downloads the Excel file from the OneDrive share link
    2. Uses the saved column mapping (or auto-detects columns)
    3. Skips already-imported rows and duplicates
    4. Creates orders with status "to_review" (temp orders)
    5. Logs the import in ImportLog

    Requires: OneDrive source ID (use list_import_sources first).
    Optional: sheet_name, max_rows limit.
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'source_id': {
                'type': 'integer',
                'description': 'OneDrive source ID (from list_import_sources)'
            },
            'sheet_name': {
                'type': 'string',
                'description': 'Sheet name to import from (uses last used or first sheet if not specified)'
            },
            'max_rows': {
                'type': 'integer',
                'description': 'Maximum rows to import (default 200, max 500)',
                'default': 200
            },
        },
        'required': ['source_id']
    }

    def execute(
        self,
        source_id: int,
        sheet_name: Optional[str] = None,
        max_rows: int = 200,
    ) -> Dict[str, Any]:
        from orders.models import OneDriveSource, Order, ImportLog, OrderItem
        from core.utils import (
            contains_arabic, translate_to_english,
            convert_arabic_numerals, format_whatsapp_number
        )

        try:
            source = OneDriveSource.objects.select_related('business').get(id=source_id)
        except OneDriveSource.DoesNotExist:
            raise ToolError(f'OneDrive source {source_id} not found', 'NOT_FOUND')

        if not source.is_active:
            raise ToolError(f'Source "{source.label}" is inactive', 'INACTIVE_SOURCE')

        business = source.business
        max_rows = min(max_rows, 500)

        # Download the file
        file_bytes, err = self._download_file(source)
        if err:
            raise ToolError(err, 'DOWNLOAD_ERROR')

        # Parse workbook
        wb = self._load_workbook(file_bytes)

        # Select sheet
        target_sheet = sheet_name or source.last_sheet_name
        if target_sheet and target_sheet in wb.sheetnames:
            ws = wb[target_sheet]
        else:
            ws = wb.active
            target_sheet = ws.title

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            raise ToolError('Sheet has no data rows', 'EMPTY_SHEET')

        headers = [str(c).strip().lower() if c else '' for c in rows[0]]
        data_rows = rows[1:]

        # Use saved mapping or auto-detect
        column_mapping = source.last_column_mapping or {}
        if not column_mapping:
            column_mapping = self._auto_detect_mapping(headers)

        # Build field_indices: {db_field: col_idx}
        field_indices = {}
        for col_idx_str, db_field in column_mapping.items():
            if db_field:
                field_indices[db_field] = int(col_idx_str)

        required = ['customer_name', 'customer_phone']
        missing = [f for f in required if f not in field_indices]
        if missing:
            raise ToolError(
                f'Cannot map required fields: {", ".join(missing)}. '
                'Configure column mapping in OneDrive Sources page first.',
                'MAPPING_ERROR'
            )

        # Determine which rows are already imported
        already_imported = set()
        if source.last_imported_rows:
            for ir in source.last_imported_rows:
                if isinstance(ir, dict) and 'row' in ir:
                    already_imported.add(ir['row'])

        def get_val(row_data, field):
            idx = field_indices.get(field)
            if idx is not None and idx < len(row_data):
                val = row_data[idx]
                return str(val).strip() if val is not None else ''
            return ''

        def safe_int(val):
            try:
                return int(float(val)) if val else 0
            except (ValueError, TypeError):
                return 0

        def safe_int_or_none(val):
            try:
                if not val or (isinstance(val, str) and not val.strip()):
                    return None
                return int(float(val))
            except (ValueError, TypeError):
                return None

        # Create ImportLog
        import_log = ImportLog.objects.create(
            business=business,
            source='onedrive',
            status='processing',
            total_rows=0,
            onedrive_source=source,
            source_meta={
                'source_label': source.label,
                'sheet_name': target_sheet,
                'auto_import': True,
            },
            column_mapping=column_mapping,
        )

        saved = 0
        skipped = 0
        errors = []
        created_orders = []
        imported_rows_data = []

        for row_num, row_data in enumerate(data_rows, start=2):
            if row_num in already_imported:
                skipped += 1
                continue

            if not any(row_data):
                continue

            if saved >= max_rows:
                break

            customer_name = get_val(row_data, 'customer_name')
            customer_phone = get_val(row_data, 'customer_phone')

            if not customer_name or not customer_phone:
                skipped += 1
                continue

            client_order_code = get_val(row_data, 'client_order_code')
            if not client_order_code:
                client_order_code = f"OD-{uuid.uuid4().hex[:8].upper()}"

            if Order.objects.filter(client_order_code=client_order_code).exists():
                skipped += 1
                continue

            try:
                customer_phone = convert_arabic_numerals(customer_phone)
                raw_whatsapp = get_val(row_data, 'customer_whatsapp')
                customer_whatsapp = format_whatsapp_number(raw_whatsapp) if raw_whatsapp else format_whatsapp_number(customer_phone)
                customer_address = get_val(row_data, 'customer_address')

                if contains_arabic(customer_name):
                    customer_name = translate_to_english(customer_name)
                if customer_address and contains_arabic(customer_address):
                    customer_address = translate_to_english(customer_address)

                seller_notes = get_val(row_data, 'seller_notes')
                internal_notes = get_val(row_data, 'internal_notes')
                notes_parts = []
                if seller_notes:
                    notes_parts.append(f"Seller: {seller_notes}")
                if internal_notes:
                    notes_parts.append(f"Ezzy: {internal_notes}")
                combined_notes = ' | '.join(notes_parts) if notes_parts else ''

                order = Order(
                    business=business,
                    client_order_code=client_order_code,
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    customer_whatsapp=customer_whatsapp,
                    customer_address=customer_address or '',
                    dl_zone=safe_int_or_none(get_val(row_data, 'dl_zone')),
                    dl_street=safe_int_or_none(get_val(row_data, 'dl_street')),
                    dl_building=safe_int_or_none(get_val(row_data, 'dl_building')),
                    cod_amount=safe_int(get_val(row_data, 'cod_amount')),
                    dl_amount=safe_int(get_val(row_data, 'dl_amount') if 'dl_amount' in field_indices else '0'),
                    order_notes=combined_notes[:100] if combined_notes else '',
                    deadline_date=get_val(row_data, 'deadline_date'),
                    order_status='to_review',
                    verification_status='pending',
                    is_transferred=False,
                    original_order_data={
                        'source': 'ai_onedrive_import',
                        'onedrive_source_id': source.id,
                        'sheet_name': target_sheet,
                        'row_number': row_num,
                    },
                )

                # Package description
                package_desc_val = get_val(row_data, 'package_desc')
                if package_desc_val:
                    order.package_description = package_desc_val[:255]
                    order.package_qty = safe_int(get_val(row_data, 'package_qty')) or 1
                else:
                    desc_parts = []
                    total_qty = 0
                    for i in range(1, 6):
                        pn = get_val(row_data, f'product_{i}')
                        if pn:
                            pc = safe_int(get_val(row_data, f'count_{i}')) or 1
                            desc_parts.append(f"{pn} x{pc}")
                            total_qty += pc
                    if desc_parts:
                        order.package_description = ', '.join(desc_parts)[:255]
                        order.package_qty = total_qty

                order.save()

                # Match products to catalog
                from product.models import Product as ProductModel
                product_names = []
                if package_desc_val:
                    product_names.append((package_desc_val, safe_int(get_val(row_data, 'package_qty')) or 1))
                for i in range(1, 6):
                    pn = get_val(row_data, f'product_{i}')
                    if pn:
                        pc = safe_int(get_val(row_data, f'count_{i}')) or 1
                        product_names.append((pn, pc))
                for pname, pqty in product_names:
                    matched = ProductModel.objects.filter(
                        business=business, item_name__iexact=pname.strip()
                    ).first()
                    if not matched:
                        matched = ProductModel.objects.filter(
                            business=business, item_sku__iexact=pname.strip()
                        ).first()
                    if not matched:
                        matched = ProductModel.objects.filter(
                            business=business, barcode__iexact=pname.strip()
                        ).first()
                    if matched:
                        OrderItem.objects.create(
                            order=order, product=matched, quantity=pqty,
                            unit_price=matched.item_price or 0,
                            notes=matched.item_name
                        )

                imported_rows_data.append({'row': row_num, 'order': order.order_number})
                created_orders.append(order)
                saved += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        # Update ImportLog
        import_log.total_rows = saved + skipped + len(errors)
        if created_orders:
            import_log.orders.add(*created_orders)
        import_log.orders_created = saved
        import_log.orders_skipped = skipped
        import_log.orders_failed = len(errors)
        import_log.errors = errors[:50]
        import_log.mark_completed()

        # Update source tracking
        existing_imported = source.last_imported_rows or []
        existing_imported.extend(imported_rows_data)
        source.last_import_at = timezone.now()
        source.last_import_count = saved
        source.last_sheet_name = target_sheet
        source.last_imported_rows = existing_imported
        if not source.last_column_mapping:
            source.last_column_mapping = column_mapping
        source.save()

        return {
            'source_label': source.label,
            'business_name': business.business_name,
            'sheet_name': target_sheet,
            'orders_created': saved,
            'orders_skipped': skipped,
            'errors_count': len(errors),
            'errors': errors[:10],
            'import_log_id': import_log.id,
        }

    def _download_file(self, source):
        """Download Excel file from OneDrive source."""
        import requests as http_requests

        link = source.share_link.strip()
        try:
            resp1 = http_requests.get(link, timeout=30, allow_redirects=False)
            if resp1.status_code in (301, 302, 307, 308):
                redirect_url = resp1.headers.get('Location', '')
                if redirect_url:
                    sep = '&' if '?' in redirect_url else '?'
                    download_url = redirect_url + sep + 'download=1'
                    resp = http_requests.get(download_url, timeout=60, allow_redirects=True)
                    content_type = resp.headers.get('Content-Type', '')
                    if resp.status_code == 200 and 'text/html' not in content_type:
                        return resp.content, None

            download_url = source.get_download_url()
            resp = http_requests.get(download_url, timeout=30, allow_redirects=True)
            content_type = resp.headers.get('Content-Type', '')
            if resp.status_code == 200 and 'text/html' not in content_type:
                return resp.content, None

            return None, 'Could not download. Ensure the file is shared with "Anyone with the link".'
        except http_requests.exceptions.Timeout:
            return None, 'Download timed out.'
        except Exception as e:
            return None, f'Download error: {str(e)}'

    def _load_workbook(self, file_bytes):
        """Load Excel workbook with fallback for corrupted stylesheets."""
        import openpyxl
        from io import BytesIO
        try:
            return openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            if 'stylesheet' in str(e).lower():
                from openpyxl.reader import excel as openpyxl_excel
                orig = openpyxl_excel.apply_stylesheet
                openpyxl_excel.apply_stylesheet = lambda *a, **kw: None
                try:
                    return openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
                finally:
                    openpyxl_excel.apply_stylesheet = orig
            raise ToolError(f'Cannot read Excel file: {str(e)}', 'PARSE_ERROR')

    def _auto_detect_mapping(self, headers):
        """Auto-detect column mapping from header names."""
        mapping_rules = {
            'customer_name': ['name', 'customer name', 'customer', 'recipient', 'receiver'],
            'customer_phone': ['phone', 'mobile', 'customer phone', 'tel', 'contact'],
            'customer_address': ['address', 'delivery address', 'customer address', 'location'],
            'client_order_code': ['order', 'order number', 'order id', 'order #', 'order code', 'ref'],
            'cod_amount': ['price', 'cod', 'amount', 'cod amount', 'total', 'value', 'cash'],
            'dl_zone': ['zone', 'zone number', 'delivery zone', 'area'],
            'package_desc': ['product', 'product name', 'item', 'description', 'package', 'goods'],
            'package_qty': ['qty', 'quantity', 'count', 'units', 'pcs'],
            'seller_notes': ['notes', 'seller notes', 'remarks', 'comment'],
            'deadline_date': ['deadline', 'delivery date', 'day', 'preferred'],
        }

        mapping = {}
        used_fields = set()
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            h = header.lower().strip()
            for field, keywords in mapping_rules.items():
                if field in used_fields:
                    continue
                if h in keywords or any(kw in h for kw in keywords):
                    mapping[str(col_idx)] = field
                    used_fields.add(field)
                    break

        return mapping


@register_tool
class ImportFromApiTool(BaseTool):
    """Import new orders from Shopify, WooCommerce, or Google Sheets."""

    name = 'import_from_api'
    allowed_roles = ['staff']
    description = '''Import new orders from a configured API source (Shopify, WooCommerce, or Google Sheets).

    Automatically:
    1. Connects to the platform using saved API credentials
    2. Fetches recent orders/rows
    3. Skips already-imported orders (dedup by platform_id)
    4. Creates orders with status "to_review" (temp orders)
    5. Logs the import in ImportLog

    Requires: api_settings_id (use list_import_sources first).
    Optional: max_orders limit (default 50).
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'api_settings_id': {
                'type': 'integer',
                'description': 'API settings ID (from list_import_sources)'
            },
            'max_orders': {
                'type': 'integer',
                'description': 'Maximum orders to import (default 50, max 200)',
                'default': 50
            },
        },
        'required': ['api_settings_id']
    }

    def execute(
        self,
        api_settings_id: int,
        max_orders: int = 50,
    ) -> Dict[str, Any]:
        from business.models import BusinessApiSettings
        from orders.models import Order, ImportLog

        try:
            api = BusinessApiSettings.objects.select_related('business').get(
                id=api_settings_id, is_active=True
            )
        except BusinessApiSettings.DoesNotExist:
            raise ToolError(f'API settings {api_settings_id} not found or inactive', 'NOT_FOUND')

        business = api.business
        max_orders = min(max_orders, 200)

        if api.api_type == 'shopify':
            return self._import_shopify(api, business, max_orders)
        elif api.api_type == 'woocommerce':
            return self._import_woocommerce(api, business, max_orders)
        elif api.api_type == 'google_sheet':
            return self._import_google_sheet(api, business, max_orders)
        else:
            raise ToolError(f'Unsupported API type: {api.api_type}', 'UNSUPPORTED')

    def _import_shopify(self, api, business, max_orders):
        import shopify
        from orders.models import Order, ImportLog

        shop_name = (api.site_api_url or '').replace('https://', '').replace('http://', '').replace('.myshopify.com', '').strip()
        session = shopify.Session(shop_name, api.api_version or '2023-10', api.api_access_token)
        shopify.ShopifyResource.activate_session(session)

        import_log = ImportLog.objects.create(
            business=business, source='shopify', status='processing',
            source_meta={'api_type': 'shopify', 'api_settings_id': api.id, 'site_url': api.site_api_url or '', 'auto_import': True},
        )

        try:
            orders = shopify.Order.find(status='any', limit=max_orders)
            created = 0
            skipped = 0
            errors = []
            created_orders = []

            for o in orders:
                pid = str(o.id)
                if Order.objects.filter(
                    business=business,
                    original_order_data__platform_id=pid,
                    original_order_data__source='shopify'
                ).exists():
                    skipped += 1
                    continue

                try:
                    shipping = getattr(o, 'shipping_address', None)
                    billing = getattr(o, 'billing_address', None)
                    addr = shipping or billing
                    addr_str = ' '.join(filter(None, [
                        getattr(addr, 'address1', '') or '',
                        getattr(addr, 'city', '') or '',
                    ])) if addr else ''
                    phone = getattr(addr, 'phone', '') or ''
                    customer = getattr(o, 'customer', None)
                    customer_name = f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip() if customer else ''

                    order = Order.objects.create(
                        business=business,
                        customer_name=customer_name or o.name,
                        customer_phone=phone,
                        customer_address=addr_str,
                        cod_amount=o.total_price,
                        order_status='to_review',
                        is_transferred=False,
                        original_order_data={
                            'source': 'shopify', 'platform_id': pid,
                            'name': o.name, 'total_price': str(o.total_price),
                            'auto_import': True,
                        },
                    )
                    created_orders.append(order)
                    created += 1
                except Exception as e:
                    errors.append(f"Shopify #{pid}: {str(e)}")

            shopify.ShopifyResource.clear_session()

            import_log.total_rows = created + skipped + len(errors)
            if created_orders:
                import_log.orders.add(*created_orders)
            import_log.orders_created = created
            import_log.orders_skipped = skipped
            import_log.orders_failed = len(errors)
            import_log.errors = errors[:50]
            import_log.mark_completed()

            return {
                'source': 'shopify',
                'business_name': business.business_name,
                'orders_created': created,
                'orders_skipped': skipped,
                'errors_count': len(errors),
                'errors': errors[:10],
                'import_log_id': import_log.id,
            }
        except Exception as e:
            import_log.mark_failed(str(e))
            raise ToolError(f'Shopify import failed: {str(e)}', 'IMPORT_ERROR')

    def _import_woocommerce(self, api, business, max_orders):
        from woocommerce import API as WooAPI
        from orders.models import Order, ImportLog

        wcapi = WooAPI(
            url=api.site_api_url or '',
            consumer_key=api.api_key or '',
            consumer_secret=api.api_secret or '',
            version='wc/v3', timeout=15,
        )

        import_log = ImportLog.objects.create(
            business=business, source='woocommerce', status='processing',
            source_meta={'api_type': 'woocommerce', 'api_settings_id': api.id, 'site_url': api.site_api_url or '', 'auto_import': True},
        )

        try:
            r = wcapi.get('orders', params={'per_page': max_orders, 'status': 'processing'})
            if r.status_code != 200:
                import_log.mark_failed(f'WooCommerce API error: HTTP {r.status_code}')
                raise ToolError(f'WooCommerce API error: HTTP {r.status_code}', 'API_ERROR')

            woo_orders = r.json()
            created = 0
            skipped = 0
            errors = []
            created_orders = []

            for o in woo_orders:
                pid = str(o.get('id'))
                if Order.objects.filter(
                    business=business,
                    original_order_data__platform_id=pid,
                    original_order_data__source='woocommerce'
                ).exists():
                    skipped += 1
                    continue

                try:
                    billing = o.get('billing', {})
                    shipping = o.get('shipping', {})
                    addr_parts = [
                        shipping.get('address_1') or billing.get('address_1', ''),
                        shipping.get('city') or billing.get('city', ''),
                    ]
                    addr_str = ', '.join(filter(None, addr_parts))
                    customer_name = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()

                    order = Order.objects.create(
                        business=business,
                        customer_name=customer_name or f"#{o.get('number')}",
                        customer_phone=billing.get('phone', ''),
                        customer_address=addr_str,
                        cod_amount=o.get('total') or 0,
                        order_status='to_review',
                        is_transferred=False,
                        original_order_data={
                            'source': 'woocommerce', 'platform_id': pid,
                            'number': o.get('number'), 'total': str(o.get('total')),
                            'auto_import': True,
                        },
                    )
                    created_orders.append(order)
                    created += 1
                except Exception as e:
                    errors.append(f"WooCommerce #{pid}: {str(e)}")

            import_log.total_rows = created + skipped + len(errors)
            if created_orders:
                import_log.orders.add(*created_orders)
            import_log.orders_created = created
            import_log.orders_skipped = skipped
            import_log.orders_failed = len(errors)
            import_log.errors = errors[:50]
            import_log.mark_completed()

            return {
                'source': 'woocommerce',
                'business_name': business.business_name,
                'orders_created': created,
                'orders_skipped': skipped,
                'errors_count': len(errors),
                'errors': errors[:10],
                'import_log_id': import_log.id,
            }
        except ToolError:
            raise
        except Exception as e:
            import_log.mark_failed(str(e))
            raise ToolError(f'WooCommerce import failed: {str(e)}', 'IMPORT_ERROR')

    def _import_google_sheet(self, api, business, max_orders):
        import re as _re
        import gspread
        import json as _json
        from pathlib import Path as _Path
        from google.oauth2.credentials import Credentials as _GCreds
        from google.auth.transport.requests import Request as _GReq
        from django.conf import settings as django_settings
        from orders.models import Order, ImportLog, OrderItem

        sheet_url = api.google_sheet_url or api.site_api_url or ''
        token_path = _Path(django_settings.BASE_DIR) / getattr(
            django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
        )
        if not token_path.exists():
            raise ToolError('Google Sheets not authorized. Token file missing.', 'AUTH_ERROR')

        _td = _json.loads(token_path.read_text())
        creds = _GCreds(
            token=_td.get('access_token'),
            refresh_token=_td.get('refresh_token'),
            token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
            client_id=_td.get('client_id'),
            client_secret=_td.get('client_secret'),
            scopes=_td.get('scope', '').split(),
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(_GReq())
            _td['access_token'] = creds.token
            token_path.write_text(_json.dumps(_td, indent=2))

        gc = gspread.authorize(creds)
        match = _re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
        if not match:
            raise ToolError('Invalid Google Sheet URL in API settings.', 'CONFIG_ERROR')

        sheet_id = match.group(1)
        gid_match = _re.search(r'gid=(\d+)', sheet_url)
        gid = int(gid_match.group(1)) if gid_match else 0

        spreadsheet = gc.open_by_key(sheet_id)
        worksheet = None
        for ws in spreadsheet.worksheets():
            if ws.id == gid:
                worksheet = ws
                break
        if worksheet is None:
            worksheet = spreadsheet.sheet1

        all_values = worksheet.get_all_values()
        if len(all_values) <= 1:
            raise ToolError('Sheet has no data rows', 'EMPTY_SHEET')

        headers = all_values[0]
        data_rows = all_values[1:]

        # Use saved column mapping
        saved_mapping = api.column_mapping or {}

        def col_idx(field_name, *alt_headers):
            """Find column index by saved mapping or header name match."""
            if saved_mapping.get(field_name) is not None:
                try:
                    return int(saved_mapping[field_name])
                except (ValueError, TypeError):
                    pass
            search_terms = [field_name.lower()] + [a.lower() for a in alt_headers]
            for i, h in enumerate(headers):
                if h.lower().strip() in search_terms:
                    return i
            return None

        idx_name = col_idx('customer_name', 'name', 'customer', 'recipient')
        idx_phone = col_idx('customer_phone', 'phone', 'mobile', 'tel')
        idx_address = col_idx('customer_address', 'address', 'delivery address')
        idx_cod = col_idx('cod_amount', 'price', 'cod', 'amount', 'total')
        idx_order = col_idx('client_order_code', 'order', 'order number', 'order id')
        idx_zone = col_idx('dl_zone', 'zone', 'zone number')

        if idx_name is None or idx_phone is None:
            raise ToolError(
                'Cannot find customer_name and customer_phone columns. '
                'Configure column mapping in API settings.',
                'MAPPING_ERROR'
            )

        import_log = ImportLog.objects.create(
            business=business, source='google_sheet', status='processing',
            total_rows=len(data_rows),
            source_meta={
                'api_type': 'google_sheet', 'api_settings_id': api.id,
                'sheet_url': sheet_url, 'auto_import': True,
            },
        )

        created = 0
        skipped = 0
        errors = []
        created_orders = []

        for row_idx, row in enumerate(data_rows[:max_orders]):
            def cell(idx):
                if idx is not None and idx < len(row):
                    return str(row[idx]).strip() if row[idx] else ''
                return ''

            name = cell(idx_name)
            phone = cell(idx_phone)
            if not name or not phone:
                skipped += 1
                continue

            order_code = cell(idx_order)
            if not order_code:
                order_code = f"GS-{uuid.uuid4().hex[:8].upper()}"

            if Order.objects.filter(client_order_code=order_code, business=business).exists():
                skipped += 1
                continue

            try:
                from core.utils import contains_arabic, translate_to_english, convert_arabic_numerals, format_whatsapp_number
                phone = convert_arabic_numerals(phone)
                address = cell(idx_address)
                if contains_arabic(name):
                    name = translate_to_english(name)
                if address and contains_arabic(address):
                    address = translate_to_english(address)

                cod = 0
                if idx_cod is not None:
                    try:
                        cod = int(float(cell(idx_cod) or '0'))
                    except (ValueError, TypeError):
                        cod = 0

                order = Order.objects.create(
                    business=business,
                    client_order_code=order_code,
                    customer_name=name,
                    customer_phone=phone,
                    customer_whatsapp=format_whatsapp_number(phone),
                    customer_address=address,
                    dl_zone=int(float(cell(idx_zone))) if idx_zone and cell(idx_zone) else None,
                    cod_amount=cod,
                    order_status='to_review',
                    verification_status='pending',
                    is_transferred=False,
                    original_order_data={
                        'source': 'google_sheet',
                        'sheet_row': row_idx + 2,
                        'auto_import': True,
                    },
                )
                created_orders.append(order)
                created += 1
            except Exception as e:
                errors.append(f"Row {row_idx + 2}: {str(e)}")

        import_log.total_rows = created + skipped + len(errors)
        if created_orders:
            import_log.orders.add(*created_orders)
        import_log.orders_created = created
        import_log.orders_skipped = skipped
        import_log.orders_failed = len(errors)
        import_log.errors = errors[:50]
        import_log.mark_completed()

        return {
            'source': 'google_sheet',
            'business_name': business.business_name,
            'orders_created': created,
            'orders_skipped': skipped,
            'errors_count': len(errors),
            'errors': errors[:10],
            'import_log_id': import_log.id,
        }


@register_tool
class ParseTextToOrdersTool(BaseTool):
    """Parse raw text (WhatsApp, pasted data) into temp orders."""

    name = 'parse_text_to_orders'
    allowed_roles = ['staff']
    description = '''Parse raw text (e.g. WhatsApp messages, pasted order lists) into temp orders.

    Accepts unstructured text containing order information and extracts:
    - Customer name
    - Phone number
    - Address
    - COD amount
    - Notes

    Creates orders with status "to_review" (temp orders list).
    Each order block should be separated by a blank line or "---".

    Example input:
    "Ahmed 55123456
    Zone 10 Building 5
    COD 150 QAR

    Sara 33987654
    Al Wakrah near souq
    COD 200"
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'text': {
                'type': 'string',
                'description': 'Raw text containing order information'
            },
            'business_id': {
                'type': 'integer',
                'description': 'Business ID to create orders for'
            },
            'business_name': {
                'type': 'string',
                'description': 'Business name (partial match, use if ID unknown)'
            },
        },
        'required': ['text']
    }

    def execute(
        self,
        text: str,
        business_id: Optional[int] = None,
        business_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        import re
        from orders.models import Order, ImportLog
        from core.utils import (
            contains_arabic, translate_to_english,
            convert_arabic_numerals, format_whatsapp_number
        )

        business = _resolve_business(business_id, business_name)

        # Split text into order blocks
        blocks = re.split(r'\n\s*\n|---+', text.strip())
        blocks = [b.strip() for b in blocks if b.strip()]

        if not blocks:
            raise ToolError('No order data found in text', 'EMPTY_INPUT')

        import_log = ImportLog.objects.create(
            business=business,
            source='csv_upload',  # closest match for raw text
            status='processing',
            total_rows=len(blocks),
            raw_data=[{'text': b} for b in blocks[:100]],
            source_meta={'method': 'ai_text_parse', 'auto_import': True},
        )

        # Phone patterns for Qatar
        phone_pattern = re.compile(r'(?:\+?974[\s-]?)?([3-7]\d{7})')
        cod_pattern = re.compile(r'(?:cod|price|amount|total|قيمة|مبلغ)[\s:]*(\d+(?:\.\d+)?)', re.IGNORECASE)
        zone_pattern = re.compile(r'(?:zone|منطقة)[\s:#]*(\d+)', re.IGNORECASE)

        created = 0
        skipped = 0
        errors = []
        created_orders = []
        parsed_results = []

        for idx, block in enumerate(blocks):
            lines = [l.strip() for l in block.split('\n') if l.strip()]
            if not lines:
                skipped += 1
                continue

            # Extract phone
            phone_match = phone_pattern.search(block)
            if not phone_match:
                skipped += 1
                errors.append(f"Block {idx + 1}: No phone number found")
                continue

            raw_phone = phone_match.group(0)
            phone = convert_arabic_numerals(raw_phone)

            # Extract COD
            cod_match = cod_pattern.search(block)
            cod_amount = int(float(cod_match.group(1))) if cod_match else 0

            # Extract zone
            zone_match = zone_pattern.search(block)
            zone_num = int(zone_match.group(1)) if zone_match else None

            # First line (or line without phone/cod) is likely the name
            name = ''
            address_parts = []
            for line in lines:
                line_clean = line.strip()
                # Skip lines that are just phone/cod/zone
                if phone_pattern.fullmatch(line_clean.replace('+', '').replace('-', '').replace(' ', '')):
                    continue
                if cod_pattern.match(line_clean):
                    continue
                if zone_pattern.match(line_clean):
                    continue

                if not name:
                    # First non-data line is the name
                    name = line_clean
                else:
                    address_parts.append(line_clean)

            if not name:
                name = f"Customer {phone[-4:]}"

            address = ' '.join(address_parts)

            # Translate Arabic if needed
            if contains_arabic(name):
                name = translate_to_english(name)
            if address and contains_arabic(address):
                address = translate_to_english(address)

            try:
                order = Order.objects.create(
                    business=business,
                    client_order_code=f"TX-{uuid.uuid4().hex[:8].upper()}",
                    customer_name=name,
                    customer_phone=phone,
                    customer_whatsapp=format_whatsapp_number(phone),
                    customer_address=address,
                    dl_zone=zone_num,
                    cod_amount=cod_amount,
                    order_status='to_review',
                    verification_status='pending',
                    is_transferred=False,
                    original_order_data={
                        'source': 'ai_text_parse',
                        'raw_text': block[:500],
                        'auto_import': True,
                    },
                )
                created_orders.append(order)
                created += 1
                parsed_results.append({
                    'order_number': order.order_number,
                    'name': name,
                    'phone': phone,
                    'address': address[:80],
                    'cod': cod_amount,
                    'zone': zone_num,
                })
            except Exception as e:
                errors.append(f"Block {idx + 1}: {str(e)}")

        import_log.total_rows = len(blocks)
        if created_orders:
            import_log.orders.add(*created_orders)
        import_log.orders_created = created
        import_log.orders_skipped = skipped
        import_log.orders_failed = len(errors)
        import_log.errors = errors[:50]
        import_log.mark_completed()

        return {
            'business_name': business.business_name,
            'orders_created': created,
            'orders_skipped': skipped,
            'errors_count': len(errors),
            'errors': errors[:10],
            'parsed_orders': parsed_results[:20],
            'import_log_id': import_log.id,
        }


@register_tool
class GetImportHistoryTool(BaseTool):
    """Get import history for a business."""

    name = 'get_import_history'
    allowed_roles = ['staff']
    description = '''Get recent import history for a business.

    Shows past imports with:
    - Source type, status, timestamps
    - Orders created/skipped/failed counts
    - Error details if any

    Useful for checking if a recent import succeeded or troubleshooting failures.
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'business_id': {
                'type': 'integer',
                'description': 'Business ID'
            },
            'business_name': {
                'type': 'string',
                'description': 'Business name (partial match)'
            },
            'source': {
                'type': 'string',
                'description': 'Filter by source: csv_upload, onedrive, shopify, woocommerce, google_sheet'
            },
            'limit': {
                'type': 'integer',
                'description': 'Max results (default 10)',
                'default': 10
            },
        },
    }

    def execute(
        self,
        business_id: Optional[int] = None,
        business_name: Optional[str] = None,
        source: Optional[str] = None,
        limit: int = 10,
    ) -> Dict[str, Any]:
        from orders.models import ImportLog

        business = _resolve_business(business_id, business_name)
        limit = min(limit, 50)

        qs = ImportLog.objects.filter(business=business).order_by('-started_at')
        if source:
            qs = qs.filter(source=source)

        logs = qs[:limit]
        results = []
        for log in logs:
            results.append({
                'id': log.id,
                'source': log.get_source_display(),
                'status': log.status,
                'started_at': log.started_at.isoformat(),
                'completed_at': log.completed_at.isoformat() if log.completed_at else None,
                'total_rows': log.total_rows,
                'orders_created': log.orders_created,
                'orders_skipped': log.orders_skipped,
                'orders_failed': log.orders_failed,
                'errors': log.errors[:5] if log.errors else [],
                'source_meta': {k: v for k, v in (log.source_meta or {}).items() if k not in ('platform_ids', 'row_numbers')},
            })

        return {
            'business_name': business.business_name,
            'total_imports': qs.count(),
            'imports': results,
        }


@register_tool
class GetTempOrdersTool(BaseTool):
    """Get temp orders (to_review) with full details."""

    name = 'get_temp_orders'
    allowed_roles = ['staff']
    description = '''Get temp orders (status "to_review") awaiting staff review.

    Returns orders that are in the temp/review queue with full details:
    - Customer name, phone, address
    - Business name
    - COD amount, zone, package info
    - Verification status, import source
    - Created date

    Filter by business, date range, or search by customer name/phone.
    '''

    parameters_schema = {
        'type': 'object',
        'properties': {
            'business_id': {
                'type': 'integer',
                'description': 'Filter by business ID'
            },
            'business_name': {
                'type': 'string',
                'description': 'Filter by business name (partial match)'
            },
            'search': {
                'type': 'string',
                'description': 'Search by customer name or phone'
            },
            'date_from': {
                'type': 'string',
                'description': 'Start date (YYYY-MM-DD)'
            },
            'date_to': {
                'type': 'string',
                'description': 'End date (YYYY-MM-DD)'
            },
            'verification_status': {
                'type': 'string',
                'description': 'Filter by verification: pending, address_verified, verified, rejected'
            },
            'limit': {
                'type': 'integer',
                'description': 'Max results (default 50, max 200)',
                'default': 50
            },
        },
    }

    def execute(
        self,
        business_id: Optional[int] = None,
        business_name: Optional[str] = None,
        search: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        verification_status: Optional[str] = None,
        limit: int = 50,
    ) -> Dict[str, Any]:
        from orders.models import Order
        from datetime import datetime

        qs = Order.objects.filter(
            order_status='to_review',
        ).select_related('business').prefetch_related('order_items').order_by('-created_at')

        # Business filter
        if business_id:
            qs = qs.filter(business_id=business_id)
        elif business_name:
            qs = qs.filter(business__business_name__icontains=business_name)

        # Search
        if search:
            qs = qs.filter(
                Q(customer_name__icontains=search) |
                Q(customer_phone__icontains=search) |
                Q(order_number__icontains=search) |
                Q(client_order_code__icontains=search)
            )

        # Date filters
        if date_from:
            try:
                qs = qs.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass
        if date_to:
            try:
                qs = qs.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d'))
            except ValueError:
                pass

        # Verification status
        if verification_status:
            qs = qs.filter(verification_status=verification_status)

        total_count = qs.count()
        limit = min(limit, 200)
        orders = qs[:limit]

        results = []
        for o in orders:
            source = ''
            if o.original_order_data:
                source = o.original_order_data.get('source', '')

            items = []
            for item in o.order_items.all():
                items.append({
                    'name': item.display_name,
                    'quantity': item.quantity,
                    'price': float(item.unit_price) if item.unit_price else 0,
                })

            results.append({
                'order_number': o.order_number,
                'client_order_code': o.client_order_code or '',
                'business': {
                    'id': o.business_id,
                    'name': o.business.business_name if o.business else '',
                },
                'customer': {
                    'name': o.customer_name,
                    'phone': o.customer_phone,
                    'address': o.customer_address,
                },
                'delivery': {
                    'zone': o.dl_zone,
                    'street': o.dl_street,
                    'building': o.dl_building,
                },
                'cod_amount': float(o.cod_amount) if o.cod_amount else 0,
                'package': {
                    'description': o.package_description or '',
                    'qty': o.package_qty or 0,
                },
                'items': items,
                'verification_status': o.verification_status,
                'is_transferred': o.is_transferred,
                'import_source': source,
                'notes': o.order_notes or '',
                'created_at': o.created_at.isoformat() if o.created_at else None,
            })

        # Summary stats
        from django.db.models import Count, Sum
        stats = Order.objects.filter(order_status='to_review').aggregate(
            total=Count('id'),
            total_cod=Sum('cod_amount'),
        )
        biz_counts = Order.objects.filter(
            order_status='to_review'
        ).values('business__business_name').annotate(
            count=Count('id')
        ).order_by('-count')[:10]

        return {
            'total_temp_orders': total_count,
            'showing': len(results),
            'orders': results,
            'summary': {
                'total_all_businesses': stats['total'] or 0,
                'total_cod': float(stats['total_cod'] or 0),
                'by_business': [
                    {'name': b['business__business_name'], 'count': b['count']}
                    for b in biz_counts
                ],
            },
        }
