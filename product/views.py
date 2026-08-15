"""
Product Views Module
====================

This module handles product catalog management for businesses.

Views:
    Product List:
        - product_all_list: Display all products (table view)
        - product_all_list_card: Display all products (card view)

    Product CRUD:
        - product_single_add: Add new product
        - product_single_update: Edit existing product
        - product_single_delete: Delete product

    Inventory:
        - product_inventory: Display inventory management page

    Categories:
        - product_categories: Display product categories

Security:
    All views (except product_categories) require authentication.
    IDOR protection ensures users can only access their business's products.

Query Optimization:
    Views use select_related() to prevent N+1 queries on ForeignKey fields
    (color, unit, business, product_category).

Related:
    - product.models: Product, ProductCategory, ColorVariant, UnitVariant
    - product.forms: AddItemsForm
    - business.models: Business
"""

import logging
import json
import csv
import io
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.urls import reverse


from product import models as product_models
from business import models as business_models
from warehouse import models as warehouse_models
from product import forms as product_forms
from core.context_processors import get_cached_business
from product.image_import import attach_product_image
from product.catalog_import import find_existing_product, update_product_fields
from core.pagination import paginate
from django.core.paginator import Paginator
from django.db.models import Sum, F, Q
from core.exports import safe_csv_writer

# Local aliases for commonly used models
Product = product_models.Product
ProductCategory = product_models.ProductCategory
ColorVariant = product_models.ColorVariant
UnitVariant = product_models.UnitVariant
Business = business_models.Business

logger = logging.getLogger('product')


# =============================================================================
# PAGINATION HELPERS (shared brand pagination component)
# =============================================================================

PER_PAGE_CHOICES = [10, 25, 50, 100]


def get_per_page(request, default=50):
    """
    Validated `per_page` for the shared pagination component.
    Returns a STRING — the component compares it as a string when marking the
    selected <option>, so an int would leave the selector blank.
    """
    raw = request.GET.get('per_page')
    if raw:
        try:
            value = int(raw)
            if value in PER_PAGE_CHOICES:
                return str(value)
        except (ValueError, TypeError):
            pass
    return str(default)


def get_filter_params(request):
    """
    Urlencoded query string of every GET param except `page` / `per_page`
    (the pagination component supplies those itself). Keeping the whole
    QueryDict means no filter can silently be dropped on page 2.
    """
    params = request.GET.copy()
    params.pop('page', None)
    params.pop('per_page', None)
    return params.urlencode()


# =============================================================================
# PRODUCT LIST VIEWS
# =============================================================================


# -----------------------------------------------------------------------------
# product_all_list / product_all_list_card: Redirect to table view.
# Grid and card views have been removed; table is the only product list view.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_all_list(request):
    return redirect('product:product_all_list_table')


@login_required(login_url='account_login')
def product_all_list_card(request):
    return redirect('product:product_all_list_table')


# -----------------------------------------------------------------------------
# product_single_add: Add new product to business catalog.
# Sets business FK automatically from logged-in user, or from ?business_id=
# when staff add on a seller's behalf from the seller detail page.
# Template: product/product_single_add.html
# Form: AddItemsForm
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_single_add(request):
    """Add one product manually.

    Two contexts (mirrors product_csv_import):
      - Client dashboard: adds into the logged-in user's own business.
      - Staff dashboard (seller section): staff pass ?business_id=<pk> to add
        into that seller's catalog, and land back on the seller detail page.
    """
    biz_id = request.POST.get('business_id') or request.GET.get('business_id')
    staff_mode = bool(biz_id) and is_staff_user(request)

    if staff_mode:
        business = get_object_or_404(Business, pk=biz_id)
    else:
        business = get_cached_business(request)
        if not business:
            logger.warning(f"User {request.user.id} has no associated business")
            messages.error(request, "No business associated with your account")
            return redirect('business:business_dashboard')

    logger.info(f"User {request.user.id} accessing product add page for business {business.business_id}")

    if request.method == 'POST':
        form = product_forms.AddItemsForm(request.POST, request.FILES)
        if form.is_valid():
            logger.info(f"Valid product form submitted by user {request.user.id}")
            product = form.save(commit=False)
            product.business = business
            product.save()
            logger.info(f"Product {product.id} created for business {business.business_id}")
            messages.success(request, 'Product added successfully!')
            if staff_mode:
                return redirect(
                    reverse('workforce:seller_detail', args=[business.business_id]) + '#products'
                )
            return redirect('product:product_all_list_table')
        else:
            logger.warning(f"Invalid product form submitted by user {request.user.id}: {form.errors}")
            messages.error(request, 'Error adding product. Please check the form.')
    else:
        form = product_forms.AddItemsForm()

    data = {
        'form': form,
        'business': business,
        'staff_mode': staff_mode,
        'base_template': 'wf_dashboard_base.html' if staff_mode else 'business_dashboard_base.html',
    }
    return render(request, 'product/product_single_add.html', data)


# -----------------------------------------------------------------------------
# product_single_delete: Delete product from business catalog.
# SECURITY: Verifies product belongs to user's business before deleting.
# Template: product/product_single_delete.html
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(["POST"])
def product_single_delete(request, product_id):
    """Delete product. POST only, redirects back to product table."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    try:
        product = product_models.Product.objects.get(id=product_id, business=business)
        logger.info(f"User {request.user.id} deleting product {product_id} from business {business.business_id}")
        product.delete()
        messages.success(request, 'Product deleted successfully!')
    except product_models.Product.DoesNotExist:
        logger.warning(f"User {request.user.id} attempted to delete non-existent or unauthorized product {product_id}")
        messages.error(request, "Product not found or unauthorized")

    return redirect('product:product_all_list_table')


# -----------------------------------------------------------------------------
# product_single_update: Update existing product details.
# SECURITY: Verifies product belongs to user's business.
# OPTIMIZATION: Uses select_related to fetch related data.
# Template: product/product_single_update.html
# Form: AddItemsForm (with instance)
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_single_update(request, product_id):
    business = get_cached_business(request)
    if not business:
        logger.warning(f"User {request.user.id} has no associated business")
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    try:
        # OPTIMIZATION: Use select_related to fetch related data in single query
        # SECURITY FIX: Verify product belongs to user's business
        product = product_models.Product.objects.select_related(
            'color', 'unit', 'business', 'product_category'
        ).get(id=product_id, business=business)

        logger.info(f"User {request.user.id} updating product {product_id}")
    except product_models.Product.DoesNotExist:
        logger.warning(f"User {request.user.id} attempted to update non-existent or unauthorized product {product_id}")
        messages.error(request, "Product not found or unauthorized")
        return redirect('product:product_all_list_table')

    if request.method == 'POST':
        form = product_forms.AddItemsForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            logger.info(f"Product {product_id} updated successfully by user {request.user.id}")
            messages.success(request, 'Your product details have been updated!')
            return redirect('product:product_all_list_table')
        else:
            logger.warning(f"Invalid update form for product {product_id}: {form.errors}")
            messages.error(request, 'Error updating product. Please check the form.')
    else:
        form = product_forms.AddItemsForm(instance=product)

    data = {
        'business': business,
        'form': form,
        'product': product,
    }
    return render(request, 'product/product_single_update.html', data)


# -----------------------------------------------------------------------------
# product_inventory: Display product inventory management page.
# Shows all business products with their stock quantities.
# Supports search by name/SKU and low-stock filtering.
# Template: product/product_inventory.html
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_inventory(request):
    business = get_cached_business(request)
    if not business:
        logger.warning(f"User {request.user.id} has no associated business")
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')
    logger.info(f"User {request.user.id} accessing inventory for business {business.business_id}")

    search_q = request.GET.get('q', '').strip()
    show_low = request.GET.get('low_stock', '') == '1'

    qs = product_models.Product.objects.filter(
        business=business
    ).select_related(
        'color', 'unit', 'product_category'
    ).annotate(
        warehouse_stock=Sum('stock_levels__quantity_on_hand')
    ).order_by('-created_at')

    if search_q:
        qs = qs.filter(
            Q(item_name__icontains=search_q) |
            Q(brand_name__icontains=search_q) |
            Q(item_sku__icontains=search_q) |
            Q(barcode__icontains=search_q)
        )

    if show_low:
        # Low stock: no warehouse stock or <= 5
        qs = qs.filter(Q(warehouse_stock__isnull=True) | Q(warehouse_stock__lte=5))

    total_products = qs.count()
    total_in_stock = qs.filter(warehouse_stock__gt=0).count()
    low_stock_count = product_models.Product.objects.filter(
        business=business
    ).annotate(warehouse_stock=Sum('stock_levels__quantity_on_hand')).filter(
        Q(warehouse_stock__isnull=True) | Q(warehouse_stock__lte=5)
    ).count()

    per_page = get_per_page(request, default=25)
    paginator = Paginator(qs, int(per_page))
    page_obj = paginator.get_page(request.GET.get('page'))

    data = {
        'business': business,
        'page_obj': page_obj,
        'search_q': search_q,
        'show_low': show_low,
        'total_products': total_products,
        'total_in_stock': total_in_stock,
        'low_stock_count': low_stock_count,
        # Pagination component: carries `q` + `low_stock` (and anything else)
        'per_page': per_page,
        'filter_params': get_filter_params(request),
    }
    return render(request, 'product/product_inventory.html', data)


# -----------------------------------------------------------------------------
# inventory_qty_update: AJAX endpoint to update a product's inventory quantity.
# POST: {qty: int}  → updates or creates ProductInventory record.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(['POST'])
def inventory_qty_update(request, product_id):
    business = get_cached_business(request)
    if not business:
        return JsonResponse({'success': False, 'error': 'No business'}, status=403)

    product = get_object_or_404(product_models.Product, id=product_id, business=business)

    try:
        body = json.loads(request.body)
        qty = int(body.get('qty', 0))
    except (ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'success': False, 'error': 'Invalid quantity'}, status=400)

    inv, _ = product_models.ProductInventory.objects.update_or_create(
        item_sku=product,
        defaults={'item_quantity': qty}
    )
    return JsonResponse({'success': True, 'qty': inv.item_quantity})



# -----------------------------------------------------------------------------
# product_categories: Display product categories page.
# PUBLIC: Does not require authentication.
# Template: product/product_categories.html
# -----------------------------------------------------------------------------
def product_categories(request):
    data = {

    }
    return render(request, 'product/product_categories.html', data)


# -----------------------------------------------------------------------------
# test_images: Diagnostic page to test image rendering
# PUBLIC: For troubleshooting only
# Template: product/test_images.html
# -----------------------------------------------------------------------------
def test_images(request):
    business = get_cached_business(request)
    if not business:
        products = []
    else:
        products = product_models.Product.objects.filter(
            business=business
        ).select_related('business')[:10]

    data = {
        'products': products
    }
    return render(request, 'product/test_images.html', data)


# =============================================================================
# PRODUCT TABLE VIEW WITH INLINE EDITING
# =============================================================================


# -----------------------------------------------------------------------------
# product_all_list_table: Display all products in editable table format.
# Template: product/product_all_list_table.html
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_all_list_table(request):
    """Display all products in an editable table format."""
    business = get_cached_business(request)
    if not business:
        logger.warning(f"User {request.user.id} has no associated business")
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')
    logger.info(f"User {request.user.id} accessing product table for business {business.business_id}")

    products = product_models.Product.objects.filter(
        business=business
    ).select_related(
        'color',
        'unit',
        'business',
        'product_category',
    ).order_by('-created_at')

    # Get dropdown options for inline editing
    categories = product_models.ProductCategory.objects.all()
    colors = product_models.ColorVariant.objects.all()
    units = product_models.UnitVariant.objects.all()

    # Paginate — this table renders an editable row per product with four
    # dropdowns each, so a full catalogue was the heaviest page in the app.
    products_page, products_total = paginate(request, products)

    data = {
        'products': products_page,
        'products_page': products_page,
        'products_total': products_total,
        'business': business,
        'categories': categories,
        'colors': colors,
        'units': units,
    }
    return render(request, 'product/product_all_list_table.html', data)


# -----------------------------------------------------------------------------
# product_inline_update: API endpoint for inline cell editing.
# Accepts JSON payload with field and value to update.
# SECURITY: Verifies product belongs to user's business.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(["POST"])
def product_inline_update(request, product_id):
    """
    API endpoint for inline product updates.

    Expects JSON payload:
    {
        "field": "item_name|item_sku|item_price|...",
        "value": "new value"
    }

    Returns JSON:
    {
        "success": true/false,
        "message": "...",
        "value": "formatted value for display"
    }
    """
    business = get_cached_business(request)
    if not business:
        return JsonResponse({
            'success': False,
            'message': 'No business associated with your account'
        }, status=403)

    try:
        product = product_models.Product.objects.get(id=product_id, business=business)
    except product_models.Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found or unauthorized'
        }, status=404)

    try:
        data = json.loads(request.body)
        field = data.get('field')
        value = data.get('value')
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)

    # Define allowed fields for inline editing
    allowed_fields = [
        'item_name', 'brand_name', 'item_sku', 'barcode',
        'item_price', 'size', 'item_discription',
        'color', 'unit', 'product_category'
    ]

    if field not in allowed_fields:
        return JsonResponse({
            'success': False,
            'message': f'Field "{field}" is not editable'
        }, status=400)

    try:
        # Handle FK fields
        if field == 'color':
            if value:
                product.color = product_models.ColorVariant.objects.get(id=value)
            else:
                product.color = None
            display_value = product.color.color_variant if product.color else '-'
        elif field == 'unit':
            if value:
                product.unit = product_models.UnitVariant.objects.get(id=value)
            else:
                product.unit = None
            display_value = product.unit.unit_variant if product.unit else '-'
        elif field == 'product_category':
            if value:
                product.product_category = product_models.ProductCategory.objects.get(id=value)
            else:
                product.product_category = None
            display_value = product.product_category.category_name if product.product_category else '-'
        elif field == 'item_price':
            product.item_price = int(value) if value else 0
            display_value = f'QAR {product.item_price}'
        else:
            setattr(product, field, value)
            display_value = value or '-'

        product.save()
        logger.info(f"Product {product_id} field '{field}' updated to '{value}' by user {request.user.id}")

        return JsonResponse({
            'success': True,
            'message': 'Product updated successfully',
            'value': display_value
        })

    except (product_models.ColorVariant.DoesNotExist,
            product_models.UnitVariant.DoesNotExist,
            product_models.ProductCategory.DoesNotExist) as e:
        return JsonResponse({
            'success': False,
            'message': 'Invalid selection'
        }, status=400)
    except Exception as e:
        logger.error(f"Error updating product {product_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Error updating product'
        }, status=500)


# =============================================================================
# API PRODUCT IMPORT WIZARD (Shopify / WooCommerce)
# Steps: 1) Fetch from API  2) Field Mapping  3) Verify  4) Confirm Import
# =============================================================================


def _fetch_api_products(business):
    """Shared helper: fetch products from Shopify/WooCommerce for a business.
    Returns (api_products_list, error_string, platform_string).
    """
    import signal

    api_settings = business_models.BusinessApiSettings.objects.filter(business=business)
    first_api = api_settings.first() if api_settings else None

    api_products = []
    api_products_error = ''
    api_products_platform = ''

    if not first_api or first_api.api_type not in ('shopify', 'woocommerce'):
        return [], 'No Shopify or WooCommerce API configured for your business.', ''

    api_products_platform = first_api.api_type

    class ApiTimeout(Exception):
        pass

    def timeout_handler(signum, frame):
        raise ApiTimeout("API request timed out after 60 seconds")

    old_handler = signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(60)

    try:
        if first_api.api_type == 'shopify':
            import shopify
            shop_name = (first_api.site_api_url or '').replace('https://', '').replace('http://', '').replace('.myshopify.com', '').strip().rstrip('/')
            session = shopify.Session(shop_name, first_api.api_version or '2023-10', first_api.api_access_token)
            shopify.ShopifyResource.activate_session(session)
            try:
                page_products = shopify.Product.find(limit=250)
                while page_products:
                    for p in page_products:
                        product_title = p.title or ''
                        product_image = ''
                        if hasattr(p, 'image') and p.image:
                            product_image = getattr(p.image, 'src', '') or ''
                        elif hasattr(p, 'images') and p.images:
                            product_image = getattr(p.images[0], 'src', '') or ''
                        vendor = getattr(p, 'vendor', '') or ''
                        product_type = getattr(p, 'product_type', '') or ''

                        variants = getattr(p, 'variants', []) or []
                        if variants:
                            for v in variants:
                                variant_image = ''
                                if getattr(v, 'image_id', None) and hasattr(p, 'images'):
                                    for img in (p.images or []):
                                        if getattr(img, 'id', None) == v.image_id:
                                            variant_image = getattr(img, 'src', '') or ''
                                            break
                                api_products.append({
                                    'platform_id': str(p.id),
                                    'variant_id': str(v.id),
                                    'title': product_title,
                                    'variant_title': getattr(v, 'title', '') or '',
                                    'sku': getattr(v, 'sku', '') or '',
                                    'barcode': getattr(v, 'barcode', '') or '',
                                    'price': getattr(v, 'price', '') or '',
                                    'inventory_qty': getattr(v, 'inventory_quantity', 0) or 0,
                                    'option1': getattr(v, 'option1', '') or '',
                                    'option2': getattr(v, 'option2', '') or '',
                                    'option3': getattr(v, 'option3', '') or '',
                                    'image': variant_image or product_image,
                                    'vendor': vendor,
                                    'product_type': product_type,
                                })
                        else:
                            api_products.append({
                                'platform_id': str(p.id), 'variant_id': '',
                                'title': product_title, 'variant_title': '',
                                'sku': '', 'barcode': '', 'price': '',
                                'inventory_qty': 0, 'option1': '', 'option2': '', 'option3': '',
                                'image': product_image, 'vendor': vendor, 'product_type': product_type,
                            })
                    # Fetch next page if available
                    if page_products.has_next_page():
                        page_products = page_products.next_page()
                    else:
                        break
            finally:
                shopify.ShopifyResource.clear_session()

        elif first_api.api_type == 'woocommerce':
            from woocommerce import API as WooAPI
            wcapi = WooAPI(
                url=first_api.site_api_url or '',
                consumer_key=first_api.api_key or '',
                consumer_secret=first_api.api_secret or '',
                version='wc/v3', timeout=10,
            )
            page = 1
            while True:
                r = wcapi.get('products', params={'per_page': 100, 'page': page, 'orderby': 'date', 'order': 'desc'})
                if r.status_code != 200:
                    api_products_error = f'WooCommerce API error {r.status_code}'
                    break
                products_page = r.json()
                if not products_page:
                    break
                for p in products_page:
                    product_title = p.get('name', '')
                    product_image = ''
                    images = p.get('images', [])
                    if images:
                        product_image = images[0].get('src', '')
                    product_type = p.get('type', '')

                    variations = p.get('variations', [])
                    if p.get('type') == 'variable' and variations:
                        for var_id in variations:
                            vr = wcapi.get(f'products/{p["id"]}/variations/{var_id}')
                            if vr.status_code == 200:
                                v = vr.json()
                                var_attrs = v.get('attributes', [])
                                option1 = var_attrs[0].get('option', '') if len(var_attrs) > 0 else ''
                                option2 = var_attrs[1].get('option', '') if len(var_attrs) > 1 else ''
                                option3 = var_attrs[2].get('option', '') if len(var_attrs) > 2 else ''
                                var_image = v.get('image', {}).get('src', '') if v.get('image') else ''
                                api_products.append({
                                    'platform_id': str(p['id']), 'variant_id': str(v['id']),
                                    'title': product_title,
                                    'variant_title': ' / '.join(a.get('option', '') for a in var_attrs),
                                    'sku': v.get('sku', ''), 'barcode': '',
                                    'price': v.get('price', ''),
                                    'inventory_qty': v.get('stock_quantity', 0) or 0,
                                    'option1': option1, 'option2': option2, 'option3': option3,
                                    'image': var_image or product_image, 'vendor': '', 'product_type': product_type,
                                })
                    else:
                        api_products.append({
                            'platform_id': str(p['id']), 'variant_id': '',
                            'title': product_title, 'variant_title': '',
                            'sku': p.get('sku', ''), 'barcode': '',
                            'price': p.get('price', ''),
                            'inventory_qty': p.get('stock_quantity', 0) or 0,
                            'option1': '', 'option2': '', 'option3': '',
                            'image': product_image, 'vendor': '', 'product_type': product_type,
                        })
                # Check if there are more pages
                total_pages = int(r.headers.get('X-WP-TotalPages', 1))
                if page >= total_pages:
                    break
                page += 1
    except ApiTimeout:
        api_products_error = 'API request timed out. Please try again.'
    except Exception as e:
        api_products_error = str(e)
    finally:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, old_handler)

    return api_products, api_products_error, api_products_platform


# -----------------------------------------------------------------------------
# product_api_wizard: Full-page wizard for API product import.
# Renders the wizard page; API fetch happens via HTMX on button click.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_api_wizard(request):
    """Render the API product import wizard page."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('product:product_all_list_table')

    api_settings = business_models.BusinessApiSettings.objects.filter(business=business)
    first_api = api_settings.first() if api_settings else None
    api_platform = first_api.api_type if first_api and first_api.api_type in ('shopify', 'woocommerce') else ''

    if not api_platform:
        messages.error(request, "No Shopify or WooCommerce API configured for your business.")
        return redirect('product:product_all_list_table')

    return render(request, 'product/product_api_wizard.html', {
        'business': business,
        'api_platform': api_platform,
    })


# -----------------------------------------------------------------------------
# product_api_fetch: HTMX endpoint to fetch API products and return JSON.
# Called by the wizard page step 1.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_api_fetch(request):
    """HTMX endpoint: fetch API products and return as JSON for the wizard."""
    business = get_cached_business(request)
    if not business:
        return JsonResponse({'success': False, 'error': 'No business found'}, status=400)

    api_products, error, platform = _fetch_api_products(business)

    if error:
        return JsonResponse({'success': False, 'error': error, 'platform': platform})

    return JsonResponse({
        'success': True,
        'platform': platform,
        'products': api_products,
        'count': len(api_products),
    })


# -----------------------------------------------------------------------------
# product_api_import: Import selected & mapped products into the local database.
# Accepts JSON payload with list of products and their mapped field values.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(["POST"])
def product_api_import(request):
    """Import selected products from API into the local product catalog using custom field mapping."""
    business = get_cached_business(request)
    if not business:
        return JsonResponse({'success': False, 'message': 'No business found'}, status=400)

    try:
        data = json.loads(request.body)
        products_to_import = data.get('products', [])
        # Update mode: refresh products that already exist instead of skipping them.
        update_existing = bool(data.get('update_existing'))
    except (json.JSONDecodeError, KeyError):
        return JsonResponse({'success': False, 'message': 'Invalid request data'}, status=400)

    if not products_to_import:
        return JsonResponse({'success': False, 'message': 'No products selected'}, status=400)

    imported_count = 0
    updated_count = 0
    skipped_count = 0
    skipped_no_sku = 0
    skipped_duplicate = 0
    skipped_unchanged = 0
    skipped_failed = 0
    generated_sku_count = 0
    image_count = 0

    # SKUs already in the catalogue — checked in memory so a batch can't create
    # duplicates within itself (the per-row DB query could not see them).
    existing_skus = set(
        s for s in Product.objects
        .filter(business=business)
        .exclude(item_sku='')
        .values_list('item_sku', flat=True)
        if s
    )

    for p in products_to_import:
        item_name = (p.get('item_name') or '').strip()[:100]
        sku = (p.get('item_sku') or '').strip()[:100]

        if not item_name:
            skipped_count += 1
            continue

        variant_id = (p.get('variant_id') or '').strip()
        platform_id = (p.get('platform_id') or '').strip()

        # Update mode: refresh the row this product already maps to. Matching also
        # covers the EZ-{platform id} form, so a SKU added in the store later lands
        # on the existing product instead of creating a second copy.
        if update_existing:
            existing = find_existing_product(Product, business, sku, variant_id, platform_id)
            if existing:
                try:
                    if update_product_fields(existing, p, new_sku=sku):
                        updated_count += 1
                    else:
                        skipped_count += 1
                        skipped_unchanged += 1
                except Exception:
                    logger.exception(
                        "Product API update failed for business %s (sku=%r)",
                        business.pk, sku,
                    )
                    skipped_count += 1
                    skipped_failed += 1
                    continue
                # Fill in a photo the product never had; existing photos are kept.
                if not existing.product_image:
                    try:
                        if attach_product_image(existing, p.get('image_url')):
                            image_count += 1
                    except Exception:
                        logger.exception("Product image update failed for %s", existing.pk)
                continue

        # Auto-generate a SKU when the merchant didn't set one — same rule as the
        # staff-side importer (workforce.seller_api_products_import). Prefer the
        # platform's variant id (stable across re-imports), then the product id,
        # then a per-business counter. Products with no SKU cannot be saved at all
        # for fulfillment-enabled businesses (Product.save raises ValueError).
        if not sku:
            ext_id = variant_id or platform_id
            if ext_id:
                sku = f'EZ-{ext_id}'[:100]
            else:
                n = len([s for s in existing_skus
                         if s.startswith(f'EZ-{business.business_id}-')]) + 1
                sku = f'EZ-{business.business_id}-{n:05d}'[:100]
            generated_sku_count += 1

        # Skip if a product with the same SKU already exists for this business
        if sku in existing_skus:
            skipped_count += 1
            skipped_duplicate += 1
            continue

        # Parse price - item_price is PositiveIntegerField
        try:
            price_val = int(float(p.get('item_price') or 0))
        except (ValueError, TypeError):
            price_val = 0

        try:
            new_product = Product.objects.create(
                business=business,
                item_name=item_name,
                brand_name=(p.get('brand_name') or '')[:100],
                item_sku=sku,
                barcode=(p.get('barcode') or '')[:100],
                item_price=price_val,
                size=(p.get('size') or '')[:100],
                item_discription=(p.get('item_discription') or '')[:100],
            )
        except Exception:
            logger.exception(
                "Product API import failed for business %s (sku=%r, name=%r)",
                business.pk, sku, item_name,
            )
            skipped_count += 1
            skipped_failed += 1
            continue

        existing_skus.add(sku)
        imported_count += 1

        # Pull the store's photo into local storage. Kept outside the create
        # block on purpose: the product is already saved, so a dead CDN must
        # never turn a successful import into a "failed" row.
        try:
            if attach_product_image(new_product, p.get('image_url')):
                image_count += 1
        except Exception:
            logger.exception("Product image import failed for product %s", new_product.pk)

    reasons = []
    if skipped_duplicate:
        reasons.append(f'{skipped_duplicate} already imported (same SKU)')
    if skipped_unchanged:
        reasons.append(f'{skipped_unchanged} already up to date')
    if skipped_failed:
        reasons.append(f'{skipped_failed} failed to save')
    other = skipped_count - skipped_duplicate - skipped_unchanged - skipped_failed
    if other > 0:
        reasons.append(f'{other} missing name')

    message = f'Imported {imported_count} product{"s" if imported_count != 1 else ""}.'
    if updated_count:
        message += f' Updated {updated_count} existing product{"s" if updated_count != 1 else ""}.'
    if skipped_count:
        message += f' Skipped {skipped_count} ({", ".join(reasons)}).'
    if image_count:
        message += f' {image_count} photo{"s" if image_count != 1 else ""} saved.'
    if generated_sku_count:
        message += (f' {generated_sku_count} product'
                    f'{"s" if generated_sku_count != 1 else ""} had no SKU in your store — '
                    f'an EZ- SKU was generated automatically.')

    logger.info(
        "API product import for business %s: imported=%s updated=%s skipped=%s "
        "(dup=%s unchanged=%s failed=%s) generated_sku=%s images=%s update_mode=%s",
        business.business_id, imported_count, updated_count, skipped_count,
        skipped_duplicate, skipped_unchanged, skipped_failed,
        generated_sku_count, image_count, update_existing,
    )

    return JsonResponse({
        'success': True,
        'message': message,
        'imported': imported_count,
        'updated': updated_count,
        'skipped': skipped_count,
        'skipped_unchanged': skipped_unchanged,
        'skipped_no_sku': skipped_no_sku,
        'skipped_duplicate': skipped_duplicate,
        'skipped_failed': skipped_failed,
        'generated_sku': generated_sku_count,
        'images_saved': image_count,
    })


# =============================================================================
# INVENTORY MANAGEMENT (Moved from warehouse app)
# =============================================================================


def get_user_business(request):
    """Helper to get business for current user - uses cached version"""
    return get_cached_business(request)


def is_staff_user(request):
    """Check if user is a staff member (can access all warehouses)"""
    _prof = getattr(request.user, 'profile', None)
    return request.user.is_staff or getattr(_prof, 'is_staff', False) or getattr(_prof, 'is_superadmin', False)


def get_business_filter(request):
    """
    Get business filter for queries.
    Staff users see all data, regular users see only their business data.
    Returns (business_obj_or_none, is_staff_bool)
    """
    if is_staff_user(request):
        return None, True
    business = get_user_business(request)
    return business, False


@login_required(login_url='account_login')
def inventory_list(request):
    """List all inventory items with detailed product information"""
    try:
        business, is_staff = get_business_filter(request)
    except Exception as e:
        logger.error(f"Error in get_business_filter: {e}", exc_info=True)
        messages.error(request, "An error occurred. Please try again.")
        return redirect('workforce:wf_dashboard')

    if not is_staff and not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    # Filters
    warehouse_id = request.GET.get('warehouse')
    search = request.GET.get('search', '')
    low_stock_only = request.GET.get('low_stock') == '1'
    category_id = request.GET.get('category')

    try:
        if is_staff:
            stock_levels = warehouse_models.StockLevel.objects.all()
            warehouses = warehouse_models.Warehouse.objects.filter(is_active=True)
        else:
            # Get warehouses linked to this business
            linked_warehouse_ids = warehouse_models.SellerWarehouseLink.objects.filter(
                business=business, is_active=True
            ).values_list('warehouse_id', flat=True)
            stock_levels = warehouse_models.StockLevel.objects.filter(warehouse_id__in=linked_warehouse_ids)
            warehouses = warehouse_models.Warehouse.objects.filter(id__in=linked_warehouse_ids, is_active=True)

        # Select related to avoid N+1 queries for product details
        stock_levels = stock_levels.select_related(
            'product',
            'product__business',
            'product__product_category',
            'product__color',
            'product__unit',
            'warehouse',
            'location'
        ).order_by('product__item_name')
    except Exception as e:
        logger.error(f"Error setting up initial queries in inventory_list: {e}", exc_info=True)
        messages.error(request, "An error occurred while loading inventory data.")
        return redirect('workforce:wf_dashboard')

    if warehouse_id:
        stock_levels = stock_levels.filter(warehouse_id=warehouse_id)

    if search:
        stock_levels = stock_levels.filter(
            Q(product__item_name__icontains=search) |
            Q(product__item_sku__icontains=search) |
            Q(product__brand_name__icontains=search) |
            Q(product__product_id__icontains=search) |
            Q(product__barcode__icontains=search)
        )

    if category_id:
        stock_levels = stock_levels.filter(product__product_category_id=category_id)

    if low_stock_only:
        stock_levels = stock_levels.filter(
            quantity_on_hand__lte=F('reorder_point') + F('quantity_reserved')
        )

    # Get categories for filter (from products in stock)
    try:
        category_ids = stock_levels.values_list('product__product_category_id', flat=True).distinct()
        categories = product_models.ProductCategory.objects.filter(
            id__in=[cat_id for cat_id in category_ids if cat_id is not None]
        ).order_by('category_name')
    except Exception as e:
        logger.error(f"Error fetching categories: {e}")
        categories = product_models.ProductCategory.objects.none()

    # Pagination
    per_page = get_per_page(request, default=25)
    paginator = Paginator(stock_levels, int(per_page))
    page = request.GET.get('page', 1)
    items = paginator.get_page(page)

    # Calculate total inventory value
    try:
        total_value = sum(
            (stock.quantity_on_hand or 0) * (stock.product.item_price if stock.product and stock.product.item_price else 0)
            for stock in stock_levels
        )
    except Exception as e:
        logger.error(f"Error calculating total inventory value: {e}")
        total_value = 0

    context = {
        'stock_levels': items,
        'warehouses': warehouses,
        'categories': categories,
        'search': search,
        'selected_warehouse': warehouse_id,
        'selected_category': category_id,
        'low_stock_only': low_stock_only,
        'total_value': total_value,
        'is_staff': is_staff,
        # Pagination component: carries search / warehouse / category / low_stock
        'per_page': per_page,
        'filter_params': get_filter_params(request),
    }
    return render(request, 'product/inventory_list.html', context)


@login_required(login_url='account_login')
def stock_card(request, product_id):
    """Stock card - detailed view for a single product"""
    business, is_staff = get_business_filter(request)

    if is_staff:
        product = get_object_or_404(product_models.Product, pk=product_id)
        stock_levels = warehouse_models.StockLevel.objects.filter(product=product)
        transactions = warehouse_models.InventoryTransaction.objects.filter(product=product)
    else:
        if not business:
            messages.error(request, "No business associated with your account")
            return redirect('business:business_dashboard')
        product = get_object_or_404(product_models.Product, pk=product_id, business=business)
        # Get warehouses linked to this business
        linked_warehouse_ids = warehouse_models.SellerWarehouseLink.objects.filter(
            business=business, is_active=True
        ).values_list('warehouse_id', flat=True)
        stock_levels = warehouse_models.StockLevel.objects.filter(
            product=product, warehouse_id__in=linked_warehouse_ids
        )
        transactions = warehouse_models.InventoryTransaction.objects.filter(
            product=product, warehouse_id__in=linked_warehouse_ids
        )

    stock_levels = stock_levels.select_related('warehouse', 'location')
    transactions = transactions.select_related('warehouse', 'location', 'created_by').order_by('-created_at')[:50]

    context = {
        'product': product,
        'stock_levels': stock_levels,
        'transactions': transactions,
        'is_staff': is_staff,
    }
    return render(request, 'product/stock_card.html', context)


@login_required(login_url='account_login')
def transaction_list(request):
    """List all inventory transactions"""
    business, is_staff = get_business_filter(request)

    if not is_staff and not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    if is_staff:
        transactions = warehouse_models.InventoryTransaction.objects.all()
        warehouses = warehouse_models.Warehouse.objects.filter(is_active=True)
    else:
        # Get warehouses linked to this business
        linked_warehouse_ids = warehouse_models.SellerWarehouseLink.objects.filter(
            business=business, is_active=True
        ).values_list('warehouse_id', flat=True)
        transactions = warehouse_models.InventoryTransaction.objects.filter(warehouse_id__in=linked_warehouse_ids)
        warehouses = warehouse_models.Warehouse.objects.filter(id__in=linked_warehouse_ids, is_active=True)

    transactions = transactions.select_related('product', 'warehouse', 'location', 'created_by').order_by('-created_at')

    # Filters
    transaction_type = request.GET.get('type')
    warehouse_id = request.GET.get('warehouse')

    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    if warehouse_id:
        transactions = transactions.filter(warehouse_id=warehouse_id)

    # Pagination
    per_page = get_per_page(request, default=50)
    paginator = Paginator(transactions, int(per_page))
    page = request.GET.get('page', 1)
    items = paginator.get_page(page)

    # Get transaction type choices from warehouse models
    try:
        transaction_types = warehouse_models.TRANSACTION_TYPE_CHOICES
    except AttributeError:
        transaction_types = []

    context = {
        'transactions': items,
        'warehouses': warehouses,
        'transaction_types': transaction_types,
        'selected_type': transaction_type,
        'selected_warehouse': warehouse_id,
        'is_staff': is_staff,
        # Pagination component: carries type / warehouse
        'per_page': per_page,
        'filter_params': get_filter_params(request),
    }
    return render(request, 'product/transaction_list.html', context)


# =============================================================================
# PRODUCT COMBO / BUNDLE VIEWS
# =============================================================================

ProductCombo = product_models.ProductCombo
ProductComboItem = product_models.ProductComboItem


@login_required(login_url='account_login')
def combo_list(request):
    """List all combos/bundles for the current business."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    combos = ProductCombo.objects.filter(
        business=business
    ).prefetch_related('items__product').order_by('-created_at')

    context = {
        'combos': combos,
    }
    return render(request, 'product/combo_list.html', context)


@login_required(login_url='account_login')
def combo_create(request):
    """Create a new product combo/bundle."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    products = Product.objects.filter(business=business).order_by('item_name')

    if request.method == 'POST':
        combo_name = request.POST.get('combo_name', '').strip()
        combo_sku = request.POST.get('combo_sku', '').strip()
        description = request.POST.get('description', '').strip()
        combo_price = request.POST.get('combo_price', '').strip()

        if not combo_name:
            messages.error(request, "Combo name is required.")
            return render(request, 'product/combo_form.html', {
                'products': products, 'mode': 'create',
            })

        combo = ProductCombo.objects.create(
            business=business,
            combo_name=combo_name,
            combo_sku=combo_sku,
            description=description,
            combo_price=combo_price if combo_price else None,
        )

        # Process combo items from POST data
        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')

        items_added = 0
        for pid, qty in zip(product_ids, quantities):
            if pid and qty:
                try:
                    product = Product.objects.get(id=int(pid), business=business)
                    ProductComboItem.objects.create(
                        combo=combo,
                        product=product,
                        quantity=max(1, int(qty)),
                    )
                    items_added += 1
                except (Product.DoesNotExist, ValueError):
                    continue

        if items_added == 0:
            combo.delete()
            messages.error(request, "Please add at least one product to the combo.")
            return render(request, 'product/combo_form.html', {
                'products': products, 'mode': 'create',
            })

        messages.success(request, f'Combo "{combo.combo_name}" created with {items_added} product(s).')
        return redirect('product:combo_list')

    context = {
        'products': products,
        'mode': 'create',
    }
    return render(request, 'product/combo_form.html', context)


@login_required(login_url='account_login')
def combo_update(request, combo_id):
    """Edit an existing product combo/bundle."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    combo = get_object_or_404(ProductCombo, id=combo_id, business=business)
    products = Product.objects.filter(business=business).order_by('item_name')

    if request.method == 'POST':
        combo_name = request.POST.get('combo_name', '').strip()
        combo_sku = request.POST.get('combo_sku', '').strip()
        description = request.POST.get('description', '').strip()
        combo_price = request.POST.get('combo_price', '').strip()

        if not combo_name:
            messages.error(request, "Combo name is required.")
            return render(request, 'product/combo_form.html', {
                'combo': combo, 'products': products, 'mode': 'update',
            })

        combo.combo_name = combo_name
        combo.combo_sku = combo_sku
        combo.description = description
        combo.combo_price = combo_price if combo_price else None
        combo.is_active = request.POST.get('is_active') == 'on'
        combo.save()

        # Replace combo items
        combo.items.all().delete()

        product_ids = request.POST.getlist('product_id')
        quantities = request.POST.getlist('quantity')

        items_added = 0
        for pid, qty in zip(product_ids, quantities):
            if pid and qty:
                try:
                    product = Product.objects.get(id=int(pid), business=business)
                    ProductComboItem.objects.create(
                        combo=combo,
                        product=product,
                        quantity=max(1, int(qty)),
                    )
                    items_added += 1
                except (Product.DoesNotExist, ValueError):
                    continue

        messages.success(request, f'Combo "{combo.combo_name}" updated.')
        return redirect('product:combo_list')

    existing_items = combo.items.select_related('product').all()

    context = {
        'combo': combo,
        'products': products,
        'existing_items': existing_items,
        'mode': 'update',
    }
    return render(request, 'product/combo_form.html', context)


@login_required(login_url='account_login')
@require_http_methods(["POST"])
def combo_delete(request, combo_id):
    """Delete a product combo/bundle (POST only)."""
    business = get_cached_business(request)
    if not business:
        messages.error(request, "No business associated with your account")
        return redirect('business:business_dashboard')

    combo = get_object_or_404(ProductCombo, id=combo_id, business=business)
    name = combo.combo_name
    combo.delete()
    messages.success(request, f'Combo "{name}" deleted.')
    return redirect('product:combo_list')


# =============================================================================
# STAFF BULK TABLE EDIT (seller detail → Products tab)
# =============================================================================

# Fields a staff member may edit straight in the seller's products table.
# Scalars carry their model max_length; the three FKs are resolved by pk.
BULK_EDIT_TEXT_FIELDS = {
    'item_name': 100,
    'brand_name': 100,
    'item_sku': 100,
    'barcode': 100,
    'size': 100,
    'item_discription': 100,
}
BULK_EDIT_FK_FIELDS = {
    'color': (ColorVariant, 'color_variant'),
    'unit': (UnitVariant, 'unit_variant'),
    'product_category': (ProductCategory, 'category_name'),
}
# Fields that must not be blanked out.
BULK_EDIT_REQUIRED = {'item_name'}
# Columns Product allows to be empty string vs NULL.
BULK_EDIT_NULLABLE = {'barcode', 'size', 'item_discription'}


# -----------------------------------------------------------------------------
# product_staff_bulk_update: Save many inline cell edits for one seller at once.
# POST JSON {business_id, changes: [{id, field, value}, ...]} -> per-row results.
# SECURITY: staff-only, and every product is re-checked against business_id so a
# crafted payload cannot touch another seller's catalog.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(["POST"])
def product_staff_bulk_update(request):
    """Apply a batch of inline edits to one seller's products."""
    if not is_staff_user(request):
        return JsonResponse({'success': False, 'message': 'Staff access required.'}, status=403)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON.'}, status=400)

    business_id = payload.get('business_id')
    changes = payload.get('changes') or []
    if not business_id:
        return JsonResponse({'success': False, 'message': 'Missing business_id.'}, status=400)
    if not isinstance(changes, list) or not changes:
        return JsonResponse({'success': False, 'message': 'No changes to save.'}, status=400)

    business = get_object_or_404(Business, pk=business_id)

    # Group the flat change list by product so each row saves once.
    by_product = {}
    for change in changes:
        try:
            pid = int(change.get('id'))
        except (TypeError, ValueError):
            continue
        field = change.get('field')
        if field not in BULK_EDIT_TEXT_FIELDS and field not in BULK_EDIT_FK_FIELDS and field != 'item_price':
            return JsonResponse(
                {'success': False, 'message': f'Field "{field}" is not editable.'}, status=400
            )
        by_product.setdefault(pid, {})[field] = change.get('value')

    products = {
        p.id: p for p in Product.objects.filter(id__in=by_product.keys(), business=business)
    }

    updated = 0
    errors = []

    for pid, fields in by_product.items():
        product = products.get(pid)
        if not product:
            errors.append(f'Product #{pid} does not belong to this seller — skipped.')
            continue

        label = product.item_name or f'#{pid}'
        try:
            for field, raw in fields.items():
                value = (raw or '').strip() if isinstance(raw, str) else raw

                if field == 'item_price':
                    try:
                        price = int(float(value)) if value not in (None, '') else 0
                    except (TypeError, ValueError):
                        raise ValueError(f"'{value}' is not a valid price")
                    product.item_price = max(price, 0)

                elif field in BULK_EDIT_FK_FIELDS:
                    model, _label_field = BULK_EDIT_FK_FIELDS[field]
                    if value in (None, '', '0'):
                        setattr(product, field, None)
                    else:
                        try:
                            setattr(product, field, model.objects.get(pk=value))
                        except (model.DoesNotExist, ValueError, TypeError):
                            raise ValueError(f'invalid {field} selection')

                else:
                    if field in BULK_EDIT_REQUIRED and not value:
                        raise ValueError(f'{field} cannot be empty')
                    trimmed = (value or '')[:BULK_EDIT_TEXT_FIELDS[field]]
                    setattr(product, field, trimmed or (None if field in BULK_EDIT_NULLABLE else ''))

            product.save()
            updated += 1
        except ValueError as exc:
            # Product.save() raises ValueError for the fulfillment SKU rule too.
            errors.append(f'{label}: {exc}')
        except Exception as exc:
            logger.exception(f'Bulk edit failed for product {pid}: {exc}')
            errors.append(f'{label}: could not be saved.')

    logger.info(
        f'Staff {request.user.id} bulk-edited {updated} product(s) '
        f'for business {business.business_id} ({len(errors)} failed)'
    )

    return JsonResponse({
        'success': not errors,
        'updated': updated,
        'errors': errors,
        'message': (
            f"Saved {updated} product{'s' if updated != 1 else ''}."
            if updated else 'Nothing was saved.'
        ),
    })


# =============================================================================
# CSV PRODUCT IMPORT
# =============================================================================

# Columns accepted in the upload. Only `item_name` is required; the rest are
# optional. Header matching is case-insensitive and ignores surrounding spaces.
CSV_IMPORT_COLUMNS = [
    'item_name', 'brand_name', 'item_sku', 'barcode', 'item_price',
    'size', 'category', 'color', 'unit', 'item_discription', 'client_names',
]


# -----------------------------------------------------------------------------
# product_staff_delete: Staff-only delete of a product from a seller's catalog.
# Unlike product_single_delete (scoped to the logged-in user's own business),
# staff have no business of their own, so this deletes by product id after a
# staff check and redirects back to that seller's detail page.
# POST only. Template: none (redirect).
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
@require_http_methods(["POST"])
def product_staff_delete(request, product_id):
    """Delete a product on behalf of staff, returning to the seller page."""
    if not is_staff_user(request):
        messages.error(request, "You don't have permission to delete this product.")
        return redirect('business:business_dashboard')

    product = get_object_or_404(Product, id=product_id)
    biz = product.business
    name = product.item_name
    logger.info(f"Staff {request.user.id} deleting product {product_id} ('{name}') from business {biz.business_id if biz else 'none'}")
    product.delete()
    messages.success(request, f'Deleted "{name}".')

    if biz:
        return redirect(reverse('workforce:seller_detail', args=[biz.business_id]) + '#products')
    return redirect('product:product_all_list_table')


# -----------------------------------------------------------------------------
# _read_import_rows: Turn an uploaded catalog file into (headers, row dicts).
# Handles .xlsx/.xlsm via openpyxl and .csv/.txt with any of the four common
# separators — Excel writes ';' instead of ',' on Arabic / Gulf locales, and
# such a file used to import as one squashed column.
# Raises ValueError with a user-facing message when the file can't be read.
# -----------------------------------------------------------------------------
def _read_import_rows(upload, filename):
    """Return (fieldnames, [row dicts]) from an uploaded CSV or Excel file."""
    if filename.endswith(('.xlsx', '.xlsm')):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(upload.read()), read_only=True, data_only=True)
        except Exception:
            raise ValueError("Could not read that Excel file. Please re-save it as .xlsx or .csv and retry.")

        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return [], []

        def as_text(value):
            if value is None:
                return ''
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        fieldnames = [as_text(h) for h in header_row]
        data_rows = []
        for raw_row in rows:
            values = [as_text(v) for v in raw_row]
            if not any(values):
                continue
            data_rows.append({
                fieldnames[i]: (values[i] if i < len(values) else '')
                for i in range(len(fieldnames))
            })
        wb.close()
        return fieldnames, data_rows

    # Decode with utf-8-sig to transparently strip Excel's BOM.
    payload = upload.read()
    try:
        raw = payload.decode('utf-8-sig')
    except UnicodeDecodeError:
        try:
            raw = payload.decode('latin-1')
        except Exception:
            raise ValueError("Could not read the file. Please save it as UTF-8 CSV and retry.")

    # Normalise line endings first. "Save as CSV (Macintosh)" writes bare \r,
    # which io.StringIO does not treat as a line break — the csv module then
    # sees the \r inside a field and raises _csv.Error instead of parsing.
    raw = raw.replace('\r\n', '\n').replace('\r', '\n')

    # Excel writes a "sep=;" hint line above the header — drop it and honour it.
    delimiter = None
    lines = raw.split('\n')
    if lines and lines[0].lower().startswith('sep=') and len(lines[0]) >= 5:
        delimiter = lines[0][4]
        raw = '\n'.join(lines[1:])

    if delimiter is None:
        header_line = next((ln for ln in raw.split('\n') if ln.strip()), '')
        # Pick whichever separator actually splits the header row.
        delimiter = max([',', ';', '\t', '|'], key=header_line.count)
        if header_line.count(delimiter) == 0:
            delimiter = ','

    try:
        reader = csv.DictReader(io.StringIO(raw, newline=''), delimiter=delimiter)
        return (reader.fieldnames or []), list(reader)
    except csv.Error as exc:
        logger.warning(f"CSV parse failed for '{filename}': {exc}")
        raise ValueError(
            "Could not parse that CSV — the rows or quotes look malformed. "
            "Open it in Excel and save it as .xlsx (or as normal CSV UTF-8), then retry."
        )


CSV_SAMPLE_ROW = [
    'Oud Royal 50ml', 'Ezzy Perfumes', 'OUD-50-001', '6291000000017',
    '120', '50ml', 'Perfumes', 'Gold', 'piece',
    'Premium oud fragrance', 'royal oud, special oud',
]


# -----------------------------------------------------------------------------
# product_csv_sample: Download a ready-to-fill import template with the correct
# headers and one example row.
# Default is .xlsx — a plain CSV opens as a single squashed column in Excel
# whenever the machine's list separator is not a comma (common on Arabic /
# Gulf locales, where it is ';'). ?format=csv still serves the raw CSV.
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_csv_sample(request):
    """Return a sample product-import file (headers + one example row)."""
    fmt = (request.GET.get('format') or 'xlsx').lower()

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_import_sample.csv"'
        # utf-8-sig BOM so Excel opens Arabic / special characters correctly
        response.write('﻿')
        writer = safe_csv_writer(response)
        writer.writerow(CSV_IMPORT_COLUMNS)
        writer.writerow(CSV_SAMPLE_ROW)
        return response

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(CSV_IMPORT_COLUMNS)
    ws.append(CSV_SAMPLE_ROW)

    head_fill = PatternFill('solid', fgColor='1B2A4A')
    for idx, column in enumerate(CSV_IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = head_fill
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(column) + 4)
    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_import_sample.xlsx"'
    wb.save(response)
    return response


# -----------------------------------------------------------------------------
# product_csv_import: Upload a CSV and bulk-create products for the business.
# GET  -> render the upload page (with sample-file link and column guide).
# POST -> parse the CSV, create products, report imported/skipped/errors.
# Duplicate SKUs (per business) and rows without item_name are skipped.
# Template: product/product_csv_import.html
# -----------------------------------------------------------------------------
@login_required(login_url='account_login')
def product_csv_import(request):
    """Bulk-import products from an uploaded CSV file.

    Two contexts:
      - Client dashboard: imports into the logged-in user's own business.
      - Staff dashboard (seller section): staff pass ?business_id=<pk> to import
        into that seller's catalog. Redirects back to the seller detail page.
    """
    # Staff may target a specific seller's business via business_id.
    biz_id = request.POST.get('business_id') or request.GET.get('business_id')
    staff_mode = bool(biz_id) and is_staff_user(request)

    if staff_mode:
        business = get_object_or_404(Business, pk=biz_id)
    else:
        business = get_cached_business(request)
        if not business:
            messages.error(request, "No business associated with your account")
            return redirect('business:business_dashboard')

    context = {
        'business': business,
        'csv_columns': CSV_IMPORT_COLUMNS,
        'staff_mode': staff_mode,
        'base_template': 'wf_dashboard_base.html' if staff_mode else 'business_dashboard_base.html',
    }

    if request.method != 'POST':
        return render(request, 'product/product_csv_import.html', context)

    upload = request.FILES.get('csv_file')
    if not upload:
        messages.error(request, "Please choose a file to upload.")
        return render(request, 'product/product_csv_import.html', context)

    name = upload.name.lower()
    if name.endswith('.xls'):
        messages.error(
            request,
            "Old .xls files are not supported. Open it in Excel and re-save as .xlsx or .csv."
        )
        return render(request, 'product/product_csv_import.html', context)

    if not name.endswith(('.csv', '.txt', '.xlsx', '.xlsm')):
        messages.error(request, "Unsupported file type. Please upload a .csv or .xlsx file.")
        return render(request, 'product/product_csv_import.html', context)

    try:
        fieldnames, data_rows = _read_import_rows(upload, name)
    except ValueError as exc:
        messages.error(request, str(exc))
        return render(request, 'product/product_csv_import.html', context)
    except Exception as exc:
        # A malformed upload must never take the page down with a 500.
        logger.exception(f"Product import failed to read '{upload.name}': {exc}")
        messages.error(
            request,
            "Could not read that file. Please re-save it as .xlsx or CSV UTF-8 and try again."
        )
        return render(request, 'product/product_csv_import.html', context)

    if not fieldnames:
        messages.error(request, "The file appears to be empty.")
        return render(request, 'product/product_csv_import.html', context)

    # Normalise headers -> lowercase/stripped so mapping is forgiving.
    header_map = {(h or '').strip().lower(): h for h in fieldnames}
    if 'item_name' not in header_map:
        messages.error(
            request,
            "The file must include an 'item_name' column — found: "
            f"{', '.join(h for h in fieldnames if h) or '(no headers)'}. "
            "Download the template for the correct format."
        )
        return render(request, 'product/product_csv_import.html', context)

    def cell(row, key):
        source = header_map.get(key)
        return (row.get(source) or '').strip() if source else ''

    # Pre-load existing SKUs for this business to skip duplicates cheaply.
    existing_skus = set(
        Product.objects.filter(business=business)
        .exclude(item_sku='')
        .values_list('item_sku', flat=True)
    )

    # Fulfillment sellers reject SKU-less products at Product.save() — catch it
    # per row so the page names the offending rows instead of a raw ValueError.
    requires_sku = bool(getattr(business, 'fulfillment_service_enabled', False))

    imported = 0
    skipped = 0
    errors = []
    seen_skus = set()

    # Row 1 is the header, so data rows are numbered from 2 (matches Excel).
    for idx, row in enumerate(data_rows, start=2):
        item_name = cell(row, 'item_name')[:100]
        if not item_name:
            skipped += 1
            continue

        sku = cell(row, 'item_sku')[:100]
        if sku and (sku in existing_skus or sku in seen_skus):
            skipped += 1
            errors.append(f"Row {idx}: skipped — SKU '{sku}' already exists.")
            continue

        if requires_sku and not sku:
            skipped += 1
            errors.append(
                f"Row {idx}: skipped — '{item_name}' has no item_sku, "
                "and this seller has fulfillment enabled (SKU is mandatory)."
            )
            continue

        price_raw = cell(row, 'item_price')
        try:
            price_val = int(float(price_raw)) if price_raw else 0
            if price_val < 0:
                price_val = 0
        except (ValueError, TypeError):
            price_val = 0

        # Optional related lookups — matched by name, never auto-created blindly
        # except category which is safe to get-or-create per business need.
        category = None
        cat_name = cell(row, 'category')
        if cat_name:
            category = ProductCategory.objects.filter(category_name__iexact=cat_name).first()
            if not category:
                category = ProductCategory.objects.create(
                    category_name=cat_name[:100], discription=cat_name[:100]
                )

        color = None
        color_name = cell(row, 'color')
        if color_name:
            color = ColorVariant.objects.filter(color_variant__iexact=color_name).first()

        unit = None
        unit_name = cell(row, 'unit')
        if unit_name:
            unit = UnitVariant.objects.filter(unit_variant__iexact=unit_name).first()

        try:
            Product.objects.create(
                business=business,
                item_name=item_name,
                brand_name=cell(row, 'brand_name')[:100],
                item_sku=sku,
                barcode=cell(row, 'barcode')[:100] or None,
                item_price=price_val,
                size=cell(row, 'size')[:100] or None,
                item_discription=cell(row, 'item_discription')[:100] or None,
                client_names=cell(row, 'client_names') or None,
                product_category=category,
                color=color,
                unit=unit,
            )
        except Exception as exc:
            skipped += 1
            errors.append(f"Row {idx}: {exc}")
            logger.warning(f"CSV import row {idx} failed for business {business.business_id}: {exc}")
            continue

        if sku:
            seen_skus.add(sku)
        imported += 1

    logger.info(
        f"User {request.user.id} CSV-imported {imported} products "
        f"({skipped} skipped) for business {business.business_id}"
    )

    if imported:
        messages.success(request, f"Imported {imported} product{'s' if imported != 1 else ''}.")
    if skipped:
        reason = "see the details below" if errors else "no item_name"
        messages.warning(request, f"Skipped {skipped} row{'s' if skipped != 1 else ''} — {reason}.")
    if not imported and not skipped:
        messages.info(request, "No data rows found in the file.")

    # Show up to 15 row-level errors back on the page for troubleshooting.
    context['import_errors'] = errors[:15]
    context['imported'] = imported
    context['skipped'] = skipped
    if imported and not errors:
        if staff_mode:
            return redirect(
                reverse('workforce:seller_detail', args=[business.business_id]) + '#products'
            )
        return redirect('product:product_all_list_table')
    return render(request, 'product/product_csv_import.html', context)