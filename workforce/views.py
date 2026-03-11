"""
Workforce Views Module
======================

This module handles internal staff/admin operations for managing
orders, deliveries, drivers, businesses, and user verification.

View Categories:
    Dashboard:
        - wf_dashboard: Main workforce dashboard with pending counts

    Order Management:
        - all_orders: List/filter all orders (paginated)
        - order_detail_modal: AJAX order details

    Delivery Management:
        - all_deliveries: List all delivery tasks
        - delivery_detail: View delivery details
        - publish_task_to_fleets: Publish single task to fleet drivers

    Driver Management:
        - all_drivers: List all drivers with status
        - driver_detail: Driver profile and documents
        - driver_documents_verification: Verify driver documents
        - fleet_cod_in_hand: View driver COD balances
        - fleet_drivers_earnings: View driver earnings
        - fleet_transactions: View financial transactions

    Business Management:
        - all_stores: List all businesses
        - store_detail: Business profile details
        - business_license_detail: Verify business documents

    User Verification:
        - user_verification: List pending verifications
        - user_verification_detail: View user details
        - approve_user_verification: Approve user
        - reject_user_verification: Reject user

    Reports:
        - staff_reports: Staff activity reports
        - staff_contacts: Contact management

Integrations:
    - ShipDay API: For DMS order publishing

Security:
    All views require authentication and typically staff permissions.
    IDOR protection on all detail views.

Related:
    - workforce.models: (empty - uses core.Profile)
    - All app models: orders, delivery, fleet, business
"""

from django.http import Http404, HttpResponse, JsonResponse
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
import csv
from django.contrib.auth.decorators import login_required
from core.decorators import staff_required
from django.views.decorators.http import require_http_methods, require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
import json
import logging
from django.core.paginator import (
    Paginator,
    EmptyPage,
    PageNotAnInteger,
)
from django.urls import reverse
from django.db.models import Count, Q

from business import models as business_models
from core import models as core_models
from orders import models as orders_models
from delivery import models as delivery_models
from fleet import models as fleet_models
from product import models as product_models
from webpages import models as webpages_models

from business import forms as business_forms
from core.context_processors import get_cached_profile, get_cached_business

# Local aliases for commonly used models
Business = business_models.Business
BusinessProfile = business_models.BusinessProfile
Profile = core_models.Profile
Order = orders_models.Order
DeliveryTask = delivery_models.DeliveryTask
Driver = fleet_models.Driver
DriverDocument = fleet_models.DriverDocument


logger = logging.getLogger(__name__)

# Import shared utilities from core
from core.utils import (
    contains_arabic,
    translate_to_english,
    convert_arabic_numerals,
    format_whatsapp_number,
)


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


def get_per_page(request, default=10):
    """
    Get the validated per_page value from request.
    Valid per_page values: 10, 25, 50, 100 (defaults to provided default if invalid)
    Returns the per_page value as a string for template use.
    """
    per_page = request.GET.get('per_page')
    if per_page:
        try:
            per_page = int(per_page)
            if per_page not in [10, 25, 50, 100]:
                per_page = default
        except (ValueError, TypeError):
            per_page = default
    else:
        per_page = default
    return str(per_page)


def paginate_queryset(request, queryset, items_per_page=25):
    """
    A helper function to handle pagination for a given queryset.
    Reads 'per_page' from request.GET to allow users to change page size.
    Valid per_page values: 10, 25, 50, 100 (defaults to items_per_page if invalid)
    """
    # Read per_page from request, with validation
    per_page = request.GET.get('per_page')
    if per_page:
        try:
            per_page = int(per_page)
            # Only allow valid page sizes to prevent abuse
            if per_page not in [10, 25, 50, 100]:
                per_page = items_per_page
        except (ValueError, TypeError):
            per_page = items_per_page
    else:
        per_page = items_per_page

    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    return page_obj




@login_required(login_url='/accounts/login/')
@staff_required
def wf_dashboard(request):
    from django.utils import timezone
    from django.db.models import Sum, Count, Q
    from datetime import timedelta

    # Use cached profile to avoid duplicate queries
    profile = get_cached_profile(request)
    if not profile:
        logger.warning(f"User {request.user.id} has no profile. Redirecting to profile creation.")
        return redirect('core:profile_view')

    today = timezone.localdate()

    # Order Statistics - single aggregate query instead of 5 separate counts
    from django.db.models import Case, When, IntegerField
    order_stats = Order.objects.aggregate(
        total=Count('id'),
        today=Count(Case(When(order_date=today, then=1), output_field=IntegerField())),
        not_published=Count(Case(When(
            task_created=False,
            then=Case(When(
                ~Q(order_status__in=['cancelled', 'delivered', 'fulfilled']),
                then=1
            ), output_field=IntegerField())
        ), output_field=IntegerField())),
        published=Count(Case(When(task_created=True, then=1), output_field=IntegerField())),
        loc_reconfirm=Count(Case(When(
            verification_status__in=['pending', 'needs_review'], then=1
        ), output_field=IntegerField())),
    )
    total_orders = order_stats['total']
    orders_today = order_stats['today']
    not_published = order_stats['not_published']
    published_orders = order_stats['published']
    loc_reconfirm = order_stats['loc_reconfirm']

    # Follow up count - tasks that need attention
    follow_up_count = DeliveryTask.objects.filter(
        dl_task_status__in=['failed', 'rescheduled', 'customer_unavailable']
    ).count()

    # User Verification pending
    pending_verifications = core_models.Profile.objects.filter(
        verification_status='pending'
    ).count()

    # Driver and Seller counts - 2 aggregate queries instead of 4 separate counts
    driver_stats = Driver.objects.aggregate(
        active=Count(Case(When(driver_status='approved', then=1), output_field=IntegerField())),
        pending=Count(Case(When(driver_status='pending', then=1), output_field=IntegerField())),
    )
    active_drivers = driver_stats['active']
    pending_drivers = driver_stats['pending']
    # COD in hand: sum of collected but unsettled COD from delivery tasks
    cod_in_hand = DeliveryTask.objects.filter(
        cod_collected=True, cod_settled=False
    ).aggregate(total=Sum('cod_collected_amount'))['total'] or 0

    seller_stats = Business.objects.aggregate(
        active=Count(Case(When(business_status='Approved', then=1), output_field=IntegerField())),
        pending=Count(Case(When(business_status='Pending on Review', then=1), output_field=IntegerField())),
    )
    active_sellers = seller_stats['active']
    pending_sellers = seller_stats['pending']

    # Recent orders (last 10 updated)
    orders = Order.objects.select_related('business').order_by('-updated_at')[:10]

    # Orders trend data for the last 7 days - based on delivered_at date
    from django.db.models.functions import TruncDate
    week_ago = today - timedelta(days=6)
    order_counts_by_date = dict(
        Order.objects.filter(delivered_at__date__gte=week_ago, delivered_at__date__lte=today)
        .annotate(delivered_date=TruncDate('delivered_at'))
        .values('delivered_date')
        .annotate(count=Count('id'))
        .values_list('delivered_date', 'count')
    )
    orders_trend = []
    max_orders = 1  # Prevent division by zero
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        count = order_counts_by_date.get(date, 0)
        if count > max_orders:
            max_orders = count
        orders_trend.append({
            'date': date.strftime('%a'),
            'count': count
        })

    data = {
        'profile': profile,
        'today': today,
        # Order stats
        'total_orders': total_orders,
        'orders_today': orders_today,
        'not_published': not_published,
        'published_orders': published_orders,
        'loc_reconfirm': loc_reconfirm,
        'follow_up_count': follow_up_count,
        # Verification
        'pending_verifications': pending_verifications,
        # Driver/Seller stats
        'active_drivers': active_drivers,
        'pending_drivers': pending_drivers,
        'active_sellers': active_sellers,
        'pending_sellers': pending_sellers,
        'cod_in_hand': cod_in_hand,
        # Recent data
        'orders': orders,
        'orders_trend': orders_trend,
        'max_orders': max_orders,
    }
    return render(request, 'workforce/wf_base_dashboard.html', data)


# Orders section  ------------------------------------------------------------------------------------------------------
@login_required(login_url='/accounts/login/')
@staff_required
def all_orders(request):
    from django.db.models import Count

    # Start with all orders, prefetch related data to avoid N+1 queries
    orders = orders_models.Order.objects.select_related(
        'business', 'pickup_location'
    ).prefetch_related('order_comments', 'delivery_task', 'order_items')

    # Apply filters based on GET parameters
    dl_code = request.GET.get('dlCode', '').strip()
    search = request.GET.get('search', '').strip()
    c_status = request.GET.get('cStatus', '').strip()
    dl_task_status = request.GET.get('dlTaskStatus', '').strip()
    business_id = request.GET.get('business', '').strip()
    date_from = request.GET.get('dateFrom', '').strip()
    date_to = request.GET.get('dateTo', '').strip()
    date_preset = request.GET.get('datePreset', '').strip()

    # Filter by Business ID
    if business_id:
        orders = orders.filter(business_id=business_id)

    # Filter by DL Code (delivery task code)
    if dl_code:
        orders = orders.filter(delivery_task__dl_task_number__icontains=dl_code)

    # Filter by Customer Name or Mobile (combined search)
    if search:
        orders = orders.filter(
            Q(customer_name__icontains=search) | Q(customer_phone__icontains=search)
        )

    # Filter by Order Status
    if c_status:
        orders = orders.filter(order_status=c_status)

    # Filter by Delivery Task Status
    if dl_task_status:
        orders = orders.filter(delivery_task__dl_task_status=dl_task_status)

    # Filter by date range
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    # Annotate with comment count (for now, all comments are counted as unread)
    orders = orders.annotate(unread_comments_count=Count('order_comments'))

    # Order by created date
    orders = orders.order_by('-created_at')

    # Paginate results
    orders = paginate_queryset(request, orders)

    # Get all businesses for filter dropdown
    all_businesses = business_models.Business.objects.all().order_by('business_name')

    # Build filter_params string for pagination to preserve filters
    filter_params_list = []
    if dl_code:
        filter_params_list.append(f'dlCode={dl_code}')
    if search:
        filter_params_list.append(f'search={search}')
    if c_status:
        filter_params_list.append(f'cStatus={c_status}')
    if dl_task_status:
        filter_params_list.append(f'dlTaskStatus={dl_task_status}')
    if business_id:
        filter_params_list.append(f'business={business_id}')

    filter_params = '&'.join(filter_params_list)

    data = {
        'orders': orders,
        'all_businesses': all_businesses,
        'filters': {
            'dlCode': dl_code,
            'search': search,
            'cStatus': c_status,
            'dlTaskStatus': dl_task_status,
            'business': business_id,
            'dateFrom': date_from,
            'dateTo': date_to,
            'datePreset': date_preset,
        },
        'filter_params': filter_params,
        'per_page': request.GET.get('per_page', '25'),
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def fulfilled_clients_orders(request):
    """Orders from businesses with fulfillment service enabled"""
    from django.db.models import Count

    # Filter orders from businesses with fulfillment service enabled
    orders = orders_models.Order.objects.select_related(
        'business', 'pickup_location'
    ).prefetch_related('order_comments', 'delivery_task', 'order_items').filter(
        business__fulfillment_service_enabled=True
    )

    # Apply filters based on GET parameters
    dl_code = request.GET.get('dlCode', '').strip()
    search = request.GET.get('search', '').strip()
    c_status = request.GET.get('cStatus', '').strip()
    dl_task_status = request.GET.get('dlTaskStatus', '').strip()
    business_id = request.GET.get('business', '').strip()

    if business_id:
        orders = orders.filter(business_id=business_id)
    if dl_code:
        orders = orders.filter(delivery_task__dl_task_number__icontains=dl_code)
    if search:
        orders = orders.filter(
            Q(customer_name__icontains=search) | Q(customer_phone__icontains=search)
        )
    if c_status:
        orders = orders.filter(order_status=c_status)
    if dl_task_status:
        orders = orders.filter(delivery_task__dl_task_status=dl_task_status)

    # Annotate with comment count
    orders = orders.annotate(unread_comments_count=Count('order_comments'))
    orders = orders.order_by('-created_at')
    orders = paginate_queryset(request, orders)

    # Get all fulfillment-enabled businesses for filter
    all_businesses = business_models.Business.objects.filter(
        fulfillment_service_enabled=True
    ).order_by('business_name')

    # Build filter_params string for pagination
    filter_params_list = []
    if dl_code:
        filter_params_list.append(f'dlCode={dl_code}')
    if search:
        filter_params_list.append(f'search={search}')
    if c_status:
        filter_params_list.append(f'cStatus={c_status}')
    if dl_task_status:
        filter_params_list.append(f'dlTaskStatus={dl_task_status}')
    if business_id:
        filter_params_list.append(f'business={business_id}')

    filter_params = '&'.join(filter_params_list)

    data = {
        'orders': orders,
        'all_businesses': all_businesses,
        'filters': {
            'dlCode': dl_code,
            'search': search,
            'cStatus': c_status,
            'dlTaskStatus': dl_task_status,
            'business': business_id,
        },
        'filter_params': filter_params,
        'per_page': request.GET.get('per_page', '25'),
        'page_title': 'Fulfilled Clients Orders',
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def non_fulfilled_clients_orders(request):
    """Orders from businesses without fulfillment service"""
    from django.db.models import Count

    # Filter orders from businesses without active fulfillment service
    orders = orders_models.Order.objects.select_related(
        'business', 'pickup_location'
    ).prefetch_related('order_comments', 'delivery_task', 'order_items').exclude(
        business__fulfillment_service_status='active'
    )

    # Apply filters based on GET parameters
    dl_code = request.GET.get('dlCode', '').strip()
    search = request.GET.get('search', '').strip()
    c_status = request.GET.get('cStatus', '').strip()
    dl_task_status = request.GET.get('dlTaskStatus', '').strip()
    business_id = request.GET.get('business', '').strip()

    if business_id:
        orders = orders.filter(business_id=business_id)
    if dl_code:
        orders = orders.filter(delivery_task__dl_task_number__icontains=dl_code)
    if search:
        orders = orders.filter(
            Q(customer_name__icontains=search) | Q(customer_phone__icontains=search)
        )
    if c_status:
        orders = orders.filter(order_status=c_status)
    if dl_task_status:
        orders = orders.filter(delivery_task__dl_task_status=dl_task_status)

    # Annotate with comment count
    orders = orders.annotate(unread_comments_count=Count('order_comments'))
    orders = orders.order_by('-created_at')
    orders = paginate_queryset(request, orders)

    # Get all non-fulfillment businesses for filter
    all_businesses = business_models.Business.objects.exclude(
        fulfillment_service_status='active'
    ).order_by('business_name')

    # Build filter_params string for pagination
    filter_params_list = []
    if dl_code:
        filter_params_list.append(f'dlCode={dl_code}')
    if search:
        filter_params_list.append(f'search={search}')
    if c_status:
        filter_params_list.append(f'cStatus={c_status}')
    if dl_task_status:
        filter_params_list.append(f'dlTaskStatus={dl_task_status}')
    if business_id:
        filter_params_list.append(f'business={business_id}')

    filter_params = '&'.join(filter_params_list)

    data = {
        'orders': orders,
        'all_businesses': all_businesses,
        'filters': {
            'dlCode': dl_code,
            'search': search,
            'cStatus': c_status,
            'dlTaskStatus': dl_task_status,
            'business': business_id,
        },
        'filter_params': filter_params,
        'per_page': request.GET.get('per_page', '25'),
        'page_title': 'Non-Fulfilled Clients Orders',
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def export_orders_csv(request):
    """
    Export orders to CSV file with applied filters
    """
    from django.utils import timezone

    # Start with all orders
    orders = orders_models.Order.objects.select_related(
        'business'
    ).prefetch_related('delivery_task')

    # Apply same filters as all_orders view
    dl_code = request.GET.get('dlCode', '').strip()
    search = request.GET.get('search', '').strip()
    c_status = request.GET.get('cStatus', '').strip()
    dl_task_status = request.GET.get('dlTaskStatus', '').strip()
    business_id = request.GET.get('business', '').strip()
    date_from = request.GET.get('dateFrom', '').strip()
    date_to = request.GET.get('dateTo', '').strip()

    if business_id:
        orders = orders.filter(business_id=business_id)
    if dl_code:
        orders = orders.filter(delivery_task__dl_task_number__icontains=dl_code)
    if search:
        orders = orders.filter(
            Q(customer_name__icontains=search) | Q(customer_phone__icontains=search)
        )
    if c_status:
        orders = orders.filter(order_status=c_status)
    if dl_task_status:
        orders = orders.filter(delivery_task__dl_task_status=dl_task_status)
    if date_from:
        orders = orders.filter(order_date__gte=date_from)
    if date_to:
        orders = orders.filter(order_date__lte=date_to)

    orders = orders.order_by('-created_at')

    # Create the HttpResponse object with CSV content type
    response = HttpResponse(content_type='text/csv')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="orders_export_{timestamp}.csv"'

    writer = csv.writer(response)
    # Write header row
    writer.writerow([
        'Order Number',
        'Client Order Code',
        'Business',
        'Customer Name',
        'Customer Phone',
        'Customer Address',
        'Order Status',
        'Task Status',
        'COD Amount',
        'Order Date',
        'Created At',
    ])

    # Write data rows
    for order in orders[:5000]:  # Limit to 5000 rows for performance
        writer.writerow([
            order.order_number,
            order.client_order_code,
            order.business.business_name if order.business else '',
            order.customer_name,
            order.customer_phone,
            order.customer_address,
            order.order_status,
            order.task_status,
            order.cod_amount,
            order.order_date,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else '',
        ])

    return response


@login_required(login_url='/accounts/login/')
@staff_required
def export_drivers_csv(request):
    """
    Export drivers to CSV file
    """
    from django.utils import timezone

    drivers = fleet_models.Driver.objects.select_related(
        'user'
    ).prefetch_related('driver_vehicle')

    # Apply filters
    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')

    if search:
        from django.db.models import Q
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(driver_code__icontains=search) |
            Q(driver_phone__icontains=search)
        )
    if status_filter:
        drivers = drivers.filter(driver_status=status_filter)

    drivers = drivers.order_by('-driver_id')

    response = HttpResponse(content_type='text/csv')
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="drivers_export_{timestamp}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Driver ID',
        'Driver Code',
        'Name',
        'Phone',
        'WhatsApp',
        'Email',
        'Status',
        'Rating',
        'Wallet Balance',
        'COD In Hand',
        'Date Joined',
    ])

    for driver in drivers[:2000]:
        writer.writerow([
            driver.driver_id,
            driver.driver_code,
            f"{driver.user.first_name} {driver.user.last_name}",
            driver.driver_phone,
            driver.driver_whatsapp,
            driver.user.email,
            driver.driver_status,
            driver.driver_rating,
            driver.wallet_balance,
            getattr(driver, 'cod_in_hand', 0),
            driver.user.date_joined.strftime('%Y-%m-%d') if driver.user.date_joined else '',
        ])

    return response


@login_required(login_url='/accounts/login/')
@staff_required
def orders_by_seller(request):
    """
    View to display orders grouped by seller/business
    """
    from django.db.models import Count, Q

    search = request.GET.get('search', '').strip()
    date_from = request.GET.get('dateFrom', '').strip()
    date_to = request.GET.get('dateTo', '').strip()
    date_preset = request.GET.get('datePreset', '').strip()

    # Build date filter for order counts
    date_q = Q()
    if date_from:
        date_q &= Q(order__created_at__date__gte=date_from)
    if date_to:
        date_q &= Q(order__created_at__date__lte=date_to)

    # All active sellers for bubbles
    from django.db.models import Count as _Count
    all_sellers = business_models.Business.objects.filter(
        business_status='active'
    ).annotate(order_count=_Count('order')
    ).values('business_id', 'business_name', 'business_code', 'order_count').order_by('business_name')

    # All businesses (show zero-order businesses too)
    businesses = business_models.Business.objects.annotate(
        total_orders=Count('order', filter=date_q),
        pending_orders=Count('order', filter=date_q & Q(order__order_status='pending')),
        processing_orders=Count('order', filter=date_q & Q(order__order_status='processing')),
        completed_orders=Count('order', filter=date_q & Q(order__order_status='delivered')),
        failed_orders=Count('order', filter=date_q & Q(order__order_status='cancelled')),
    )

    if search:
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search) |
            Q(business_code__icontains=search)
        )

    businesses = businesses.order_by('-total_orders', 'business_name')

    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Orders by Seller',
        'page_obj': page_obj,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'date_preset': date_preset,
        'all_sellers': all_sellers,
    }

    return render(request, 'workforce/wf_orders_by_seller.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def orders_to_publish(request):
    orders = orders_models.Order.objects.select_related(
        'business'
    ).prefetch_related('order_comments', 'delivery_task').filter(task_created=False).order_by('-created_at')
    orders = paginate_queryset(request, orders)

    data = {
        'orders': orders,
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def orders_published(request):
    orders = orders_models.Order.objects.select_related(
        'business', 'pickup_location'
    ).prefetch_related('order_comments', 'delivery_task').filter(task_created=True).order_by('-created_at')
    orders = paginate_queryset(request, orders)

    data = {
        'orders': orders,
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)

@login_required(login_url='/accounts/login/')
@staff_required
def submit_to_task(request, order_id):
    """Submit order to delivery task - now uses verification workflow"""
    from django.utils import timezone
    from orders.signals import _create_delivery_task_from_order
    
    order = get_object_or_404(
        orders_models.Order.objects.select_related('business'),
        id=order_id
    )

    # Use the automated function that handles DMS push
    delivery_task = _create_delivery_task_from_order(order)
    
    if delivery_task:
        from django.contrib import messages
        messages.success(request, f'Delivery task created and pushed to DMS: {delivery_task.dl_task_number}')
    else:
        from django.contrib import messages
        messages.error(request, 'Failed to create delivery task')
    
    return redirect(reverse('workforce:wf_orders_all'))


@login_required(login_url='/accounts/login/')
@staff_required
def verify_order_address(request, order_id):
    """Verify order address - workforce view"""
    from django.utils import timezone
    from orders.models import AddressVerification, OrderVerificationLog

    order = get_object_or_404(
        orders_models.Order.objects.select_related('business'),
        id=order_id
    )
    
    if request.method == 'POST':
        verified_address = request.POST.get('verified_address', order.customer_address)
        verification_result = request.POST.get('verification_result', 'valid')
        latitude = request.POST.get('latitude')
        longitude = request.POST.get('longitude')
        zone_number = request.POST.get('zone_number')
        street_number = request.POST.get('street_number')
        building_number = request.POST.get('building_number')
        notes = request.POST.get('notes', '')
        
        # Create or update address verification
        address_verification, created = AddressVerification.objects.get_or_create(
            order=order,
            defaults={
                'original_address': order.customer_address,
                'verified_address': verified_address,
                'verification_result': verification_result,
                'verified_by': request.user,
                'verified_at': timezone.now(),
                'latitude': latitude,
                'longitude': longitude,
                'zone_number': zone_number,
                'street_number': street_number,
                'building_number': building_number,
                'notes': notes,
            }
        )
        
        if not created:
            address_verification.verified_address = verified_address
            address_verification.verification_result = verification_result
            address_verification.verified_by = request.user
            address_verification.verified_at = timezone.now()
            if latitude:
                address_verification.latitude = latitude
            if longitude:
                address_verification.longitude = longitude
            if zone_number:
                address_verification.zone_number = zone_number
            if street_number:
                address_verification.street_number = street_number
            if building_number:
                address_verification.building_number = building_number
            address_verification.notes = notes
            address_verification.save()
        
        # Update order
        order.address_verified = True
        order.address_verified_by = request.user
        order.address_verified_at = timezone.now()
        
        if verification_result == 'valid':
            order.verification_status = 'address_verified'
        elif verification_result == 'needs_update':
            order.verification_status = 'address_needs_update'
            if verified_address:
                order.customer_address = verified_address
            if zone_number:
                order.dl_zone = zone_number
            if street_number:
                order.dl_street = street_number
            if building_number:
                order.dl_building = building_number
        
        order.save()
        
        # Log verification
        OrderVerificationLog.objects.create(
            order=order,
            verified_by=request.user,
            action='address_verified',
            notes=notes,
            new_status=order.verification_status
        )
        
        from django.contrib import messages
        messages.success(request, 'Address verified successfully')
        return redirect(reverse('workforce:wf_orders_all'))
    
    # GET request - show verification form
    address_verification = AddressVerification.objects.filter(order=order).first()
    context = {
        'order': order,
        'address_verification': address_verification,
    }
    return render(request, 'workforce/parts/verify_address.html', context)



@login_required(login_url='/accounts/login/')
@staff_required
def orders_pending_verification(request):
    """List orders pending verification"""
    verification_status = request.GET.get('verification_status', 'pending')

    orders = orders_models.Order.objects.select_related(
        'business'
    ).prefetch_related('order_comments', 'delivery_task').filter(verification_status=verification_status).order_by('-created_at')
    orders = paginate_queryset(request, orders)
    
    data = {
        'orders': orders,
        'verification_status': verification_status,
    }
    return render(request, 'workforce/parts/lists/orders_list_view.html', data)


    

    





# Staff Order Creation section  ------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
def add_order(request):
    """
    Staff view to add an order for a client/seller.
    Seller is selected at the top of the form.
    """
    import uuid
    from django.contrib import messages
    from core.utils import (
        contains_arabic, translate_to_english,
        convert_arabic_numerals, format_whatsapp_number
    )

    # Get all active businesses for seller dropdown
    businesses = business_models.Business.objects.filter(
        business_status='active'
    ).select_related('business_profile').order_by('business_name')

    # Get selected business if provided
    selected_business_id = request.GET.get('business', '')
    selected_business = None
    pickup_locations = []
    warehouse_products = []

    if selected_business_id:
        try:
            selected_business = business_models.Business.objects.get(business_id=selected_business_id)
            # Show fulfillment stores first when fulfillment service is enabled
            pickup_locations = business_models.PickupLocation.objects.filter(
                business=selected_business
            ).order_by('-is_fulfilment_center', 'pickup_location_title')

            # If fulfillment service is enabled, fetch warehouse inventory products
            if selected_business.fulfillment_service_enabled:
                from warehouse import models as warehouse_models
                from django.db.models import Sum

                # Get warehouse links for this business
                warehouse_links = warehouse_models.SellerWarehouseLink.objects.filter(
                    business=selected_business,
                    is_active=True
                ).select_related('warehouse', 'default_location')

                # Get products from delivered orders (order items) for this business
                if warehouse_links.exists():
                    warehouse_ids = warehouse_links.values_list('warehouse_id', flat=True)

                    # Get products from OrderItems where order status is delivered
                    # and the product exists in warehouse stock
                    delivered_order_products = orders_models.OrderItem.objects.filter(
                        order__business=selected_business,
                        order__order_status='delivered',
                        product__isnull=False,
                    ).values('product_id').annotate(
                        total_delivered=Sum('quantity')
                    ).values_list('product_id', flat=True).distinct()

                    # Get stock levels for these products from linked warehouses
                    warehouse_products = warehouse_models.StockLevel.objects.filter(
                        warehouse_id__in=warehouse_ids,
                        product_id__in=delivered_order_products,
                        quantity_on_hand__gt=0,
                    ).select_related(
                        'product', 'warehouse', 'location'
                    ).order_by('product__item_name')[:100]  # Limit to 100 products for performance
        except business_models.Business.DoesNotExist:
            pass

    if request.method == 'POST':
        try:
            # Get form data
            business_id = request.POST.get('business')
            business = business_models.Business.objects.get(business_id=business_id)

            # Generate client order code if not provided
            client_order_code = request.POST.get('client_order_code', '').strip()
            if not client_order_code:
                client_order_code = f"WF-{uuid.uuid4().hex[:8].upper()}"

            # Get and process customer data
            customer_name = request.POST.get('customer_name', '').strip()
            customer_address = request.POST.get('customer_address', '').strip()
            customer_phone = request.POST.get('customer_phone', '').strip()

            # Store originals before translation
            customer_name_original = customer_name
            customer_address_original = customer_address

            # Translate Arabic text to English
            customer_name_en = translate_to_english(customer_name)
            customer_address_en = translate_to_english(customer_address)

            if contains_arabic(customer_name):
                customer_name = customer_name_en
            if contains_arabic(customer_address):
                customer_address = customer_address_en

            # Convert Arabic numerals and format phone numbers
            customer_phone = convert_arabic_numerals(customer_phone)
            raw_whatsapp = request.POST.get('customer_whatsapp', '').strip()
            if raw_whatsapp:
                customer_whatsapp = format_whatsapp_number(raw_whatsapp)
            else:
                customer_whatsapp = format_whatsapp_number(customer_phone)

            # Helper to safely parse integers
            def safe_int(val):
                try:
                    return int(float(val)) if val else 0
                except (ValueError, TypeError):
                    return 0

            # Build combined notes
            notes_parts = []
            order_notes = request.POST.get('order_notes', '').strip()
            seller_notes = request.POST.get('seller_notes', '').strip()
            if order_notes:
                notes_parts.append(f"Customer: {order_notes}")
            if seller_notes:
                notes_parts.append(f"Seller: {seller_notes}")
            combined_notes = ' | '.join(notes_parts) if notes_parts else order_notes

            # Prepare original_order_data with extra fields
            original_data = {
                'source': 'staff_dashboard',
                'created_by': request.user.username,
                'extra_fields': {
                    'customer_name_original': customer_name_original,
                    'customer_address_original': customer_address_original,
                    'customer_name_en': customer_name_en,
                    'customer_address_en': customer_address_en,
                    'name_was_translated': customer_name_original != customer_name_en,
                    'address_was_translated': customer_address_original != customer_address_en,
                    'dl_landmark': request.POST.get('dl_landmark', '').strip(),
                    'location_link': request.POST.get('location_link', '').strip(),
                    'product_name': request.POST.get('product_name', '').strip(),
                    'quantity': request.POST.get('quantity', '1'),
                    'seller_notes': seller_notes,
                    'internal_notes': request.POST.get('internal_notes', '').strip(),
                }
            }

            # Parse scheduled date and time if provided
            scheduled_delivery = request.POST.get('scheduled_delivery') == 'on'
            scheduled_date = None
            scheduled_time = None
            if scheduled_delivery:
                from datetime import datetime
                if request.POST.get('scheduled_date'):
                    try:
                        scheduled_date = datetime.strptime(request.POST.get('scheduled_date'), '%Y-%m-%d').date()
                    except ValueError:
                        scheduled_date = None
                if request.POST.get('scheduled_time'):
                    try:
                        scheduled_time = datetime.strptime(request.POST.get('scheduled_time'), '%H:%M').time()
                    except ValueError:
                        scheduled_time = None

            # Create order
            order = orders_models.Order(
                business=business,
                client_order_code=client_order_code,
                customer_name=customer_name,
                customer_phone=customer_phone,
                customer_whatsapp=customer_whatsapp,
                customer_address=customer_address,
                dl_zone=safe_int(request.POST.get('dl_zone')),
                dl_street=safe_int(request.POST.get('dl_street')),
                dl_building=safe_int(request.POST.get('dl_building')),
                latitude=request.POST.get('latitude') or None,
                longitude=request.POST.get('longitude') or None,
                coords_accuracy=request.POST.get('coords_accuracy') or None,
                cod_amount=safe_int(request.POST.get('cod_amount')),
                dl_amount=safe_int(request.POST.get('dl_amount')),
                order_type=request.POST.get('order_type', 'normal_delivery'),
                scheduled_delivery=scheduled_delivery,
                scheduled_date=scheduled_date,
                scheduled_time=scheduled_time,
                order_notes=combined_notes[:100] if combined_notes else '',
                deadline_date=request.POST.get('deadline_date', '').strip(),
                order_status='to_review',
                verification_status='pending',
                original_order_data=original_data,
            )

            # Set pickup location if provided
            pickup_location_id = request.POST.get('pickup_location')
            if pickup_location_id:
                try:
                    order.pickup_location = business_models.PickupLocation.objects.get(
                        id=pickup_location_id, business=business
                    )
                except business_models.PickupLocation.DoesNotExist:
                    pass

            order.save()

            # Create OrderItems from inline product rows
            inline_product_ids = request.POST.getlist('inline_product_id[]')
            inline_quantities = request.POST.getlist('inline_quantity[]')
            inline_prices = request.POST.getlist('inline_unit_price[]')
            inline_notes = request.POST.getlist('inline_notes[]')
            items_created = 0
            for i, product_id in enumerate(inline_product_ids):
                product_id = product_id.strip()
                qty = safe_int(inline_quantities[i]) if i < len(inline_quantities) else 1
                qty = qty or 1
                try:
                    price_raw = inline_prices[i] if i < len(inline_prices) else ''
                    price = float(price_raw) if price_raw.strip() else None
                except (ValueError, TypeError):
                    price = None
                note = inline_notes[i] if i < len(inline_notes) else ''
                if product_id:
                    try:
                        from product.models import Product
                        product_obj = Product.objects.get(pk=product_id)
                        orders_models.OrderItem.objects.create(
                            order=order, product=product_obj, quantity=qty,
                            unit_price=price, notes=note
                        )
                        items_created += 1
                    except Exception:
                        pass
                elif note:
                    # No product selected — save as notes-only item
                    orders_models.OrderItem.objects.create(
                        order=order, quantity=qty, unit_price=price, notes=note
                    )
                    items_created += 1

            # Fallback: old plain product_name → save as order description
            if items_created == 0:
                product_name = request.POST.get('product_name', '').strip()
                if product_name:
                    order.package_description = product_name[:255]
                    order.total_quantity = safe_int(request.POST.get('quantity', '1')) or 1
                    order.save(update_fields=['package_description', 'total_quantity'])

            messages.success(request, f'Order {order.order_number} created successfully.')

            # Return JSON response for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Order {order.order_number} created successfully.',
                    'order_id': order.id,
                    'order_number': order.order_number
                })

            return redirect(reverse('workforce:wf_orders_all'))

        except business_models.Business.DoesNotExist:
            error_msg = 'Please select a valid seller.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            messages.error(request, error_msg)

        except Exception as e:
            logger.exception('Error creating order: %s', str(e))
            error_msg = 'An error occurred while creating the order. Please try again.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            messages.error(request, error_msg)

    from datetime import date
    context = {
        'businesses': businesses,
        'selected_business': selected_business,
        'selected_business_id': selected_business_id,
        'pickup_locations': pickup_locations,
        'warehouse_products': warehouse_products,
        'has_fulfillment': selected_business.fulfillment_service_enabled if selected_business else False,
        'today': date.today().isoformat(),
    }
    return render(request, 'workforce/orders_add.html', context)


# Bulk import views removed - now using shared views from orders app
# See orders/views.py: bulk_import_orders, bulk_import_preview, bulk_import_save


@login_required(login_url='/accounts/login/')
@staff_required
def bulk_transfer_api_orders(request):
    """
    Bulk-mark selected API orders as transferred into main workflow.
    POST: order_ids (list), sets is_transferred=True, transferred_at, transferred_by.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    from django.utils import timezone
    order_ids = request.POST.getlist('order_ids[]')
    if not order_ids:
        return JsonResponse({'success': False, 'error': 'No orders selected'}, status=400)

    try:
        updated = orders_models.Order.objects.filter(
            id__in=order_ids,
            is_transferred=False
        ).update(
            is_transferred=True,
            transferred_at=timezone.now(),
            transferred_by=request.user,
            order_status='to_review',
        )
        return JsonResponse({'success': True, 'updated': updated})
    except Exception as e:
        logger.exception('bulk_transfer_api_orders error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def preview_api_import(request):
    """
    POST: Return preview of mapped data for selected Google Sheet rows.
    Used in the import wizard before actual import.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    import json as _json, re as _re
    business_id = request.POST.get('business_id', '').strip()
    platform_ids = request.POST.getlist('platform_ids[]')

    if not business_id or not platform_ids:
        return JsonResponse({'success': False, 'error': 'Missing business_id or platform_ids'}, status=400)

    try:
        business = business_models.Business.objects.get(business_id=business_id)
        api = business.business_settings_api.filter(is_verify_api=True).first()
        if not api:
            return JsonResponse({'success': False, 'error': 'No approved API config'}, status=400)
    except business_models.Business.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Business not found'}, status=404)

    if api.api_type != 'google_sheet':
        # For Shopify/WooCommerce, just return the platform_ids as-is (preview not needed)
        return JsonResponse({'success': True, 'preview': [], 'api_type': api.api_type})

    try:
        import gspread
        from google.oauth2.credentials import Credentials as _GCreds
        from google.auth.transport.requests import Request as _GReq
        from django.conf import settings as django_settings
        from pathlib import Path as _Path

        sheet_url = api.google_sheet_url or api.site_api_url or ''
        token_path = _Path(django_settings.BASE_DIR) / getattr(
            django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
        )
        if not token_path.exists():
            return JsonResponse({'success': False, 'error': 'Google Sheets not authorized.'}, status=400)

        _td = _json.loads(token_path.read_text())
        creds = _GCreds(
            token=_td.get('access_token'),
            refresh_token=_td.get('refresh_token'),
            token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
            client_id=_td.get('client_id'),
            client_secret=_td.get('client_secret'),
            scopes=_td.get('scope', '').split(),
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(_GReq())
            _td['access_token'] = creds.token
            token_path.write_text(_json.dumps(_td, indent=2))

        gc = gspread.authorize(creds)
        match = _re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
        if not match:
            return JsonResponse({'success': False, 'error': 'Invalid Google Sheet URL.'}, status=400)
        sheet_id = match.group(1)
        gid_match = _re.search(r'gid=(\d+)', sheet_url)
        gid = int(gid_match.group(1)) if gid_match else 0

        spreadsheet = gc.open_by_key(sheet_id)
        worksheet = None
        for ws in spreadsheet.worksheets():
            if ws.id == gid:
                worksheet = ws
                break
        if worksheet is None:
            worksheet = spreadsheet.sheet1

        all_values = worksheet.get_all_values()
        if len(all_values) <= 1:
            for ws in spreadsheet.worksheets():
                if ws.id == (worksheet.id if worksheet else -1):
                    continue
                candidate = ws.get_all_values()
                if len(candidate) > 1:
                    worksheet = ws
                    all_values = candidate
                    break

        headers = all_values[0] if all_values else []
        data_rows = all_values[1:] if all_values else []

        saved_mapping = api.column_mapping or {}

        def col_idx_by_header(header_name):
            for j, h in enumerate(headers):
                if h.strip() == header_name.strip():
                    return j
            return None

        def col_idx_auto(*keys):
            for k in keys:
                for j, h in enumerate(headers):
                    if h.strip().lower() == k.lower():
                        return j
            return None

        def get_idx(field_name, *auto_keys):
            if field_name in saved_mapping:
                return col_idx_by_header(saved_mapping[field_name])
            return col_idx_auto(*auto_keys)

        idx_name = get_idx('customer_name', 'name', 'customer name', 'customer', 'client name')
        idx_phone = get_idx('customer_phone', 'phone', 'mobile', 'phone number', 'contact')
        idx_whatsapp = get_idx('customer_whatsapp', 'whatsapp', 'phone 2', 'alternate phone')
        idx_email = get_idx('customer_email', 'email', 'e-mail', 'customer email')
        idx_address = get_idx('customer_address', 'address', 'delivery address', 'location')
        idx_landmark = get_idx('dl_landmark', 'city', 'landmark', 'customer city', 'area')
        idx_building = get_idx('dl_building', 'building', 'villa', 'building no')
        idx_street = get_idx('dl_street', 'street', 'street no', 'street number')
        idx_zone = get_idx('dl_zone', 'zone', 'zone no', 'zone number')
        idx_cod = get_idx('cod_amount', 'cod', 'amount', 'cod amount', 'price', 'total')
        idx_dl_amount = get_idx('dl_amount', 'delivery fee', 'shipping fee', 'delivery amount')
        idx_order = get_idx('client_order_code', 'order', 'order number', 'order #', 'order id')
        idx_product = get_idx('product_name', 'product', 'product name', 'item', 'items')
        idx_sku = get_idx('sku', 'sku', 'product sku', 'item sku', 'article')
        idx_qty = get_idx('quantity', 'qty', 'quantity', 'count')
        idx_deadline = get_idx('deadline_date', 'deadline', 'delivery date', 'preferred', 'day')
        idx_notes = get_idx('seller_notes', 'notes', 'seller notes', 'note', 'remarks', 'comment')
        idx_internal = get_idx('internal_notes', 'internal notes', 'staff notes', 'notes by ezzy')

        def cell(row, idx):
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else ''
            return ''

        # Build all field indices with labels and mapped sheet column
        all_fields = [
            ('row_number', 'Row', None),
            ('client_order_code', 'Order ID', idx_order),
            ('customer_name', 'Customer Name', idx_name),
            ('customer_phone', 'Phone 1', idx_phone),
            ('customer_whatsapp', 'Phone 2 / WhatsApp', idx_whatsapp),
            ('customer_email', 'Email', idx_email),
            ('customer_address', 'Customer Address', idx_address),
            ('dl_landmark', 'City / Landmark', idx_landmark),
            ('dl_building', 'Villa / Building No', idx_building),
            ('dl_street', 'Street No', idx_street),
            ('dl_zone', 'Zone No', idx_zone),
            ('cod_amount', 'Price / COD Amount', idx_cod),
            ('dl_amount', 'Delivery Fee', idx_dl_amount),
            ('product_name', 'Product Name', idx_product),
            ('sku', 'SKU', idx_sku),
            ('quantity', 'Quantity', idx_qty),
            ('deadline_date', 'Day & Time', idx_deadline),
            ('seller_notes', 'Seller Notes', idx_notes),
            ('internal_notes', 'Internal Notes', idx_internal),
        ]

        # Only include fields that have a mapped column (non-None idx)
        active_fields = [('row_number', 'Row', '')]  # always show row
        for fname, flabel, fidx in all_fields:
            if fname == 'row_number':
                continue
            if fidx is not None:
                # Get the sheet column header name for display
                sheet_col = headers[fidx] if fidx < len(headers) else ''
                active_fields.append((fname, flabel, sheet_col))

        preview_rows = []
        already_imported = []
        for pid in platform_ids:
            parts = pid.split('_', 2)
            if len(parts) < 2:
                continue
            row_idx = int(parts[1])
            sheet_row = row_idx + 2

            # Check if already imported
            is_imported = orders_models.Order.objects.filter(
                business=business,
            ).filter(
                Q(original_order_data__platform_id=pid) |
                Q(original_order_data__row_number=sheet_row)
            ).exists()
            if is_imported:
                already_imported.append(pid)
                continue

            if row_idx >= len(data_rows):
                continue

            row = data_rows[row_idx]
            # Build row_data dynamically from active_fields
            row_data = {'platform_id': pid}
            field_idx_map = {
                'row_number': None,
                'client_order_code': idx_order,
                'customer_name': idx_name,
                'customer_phone': idx_phone,
                'customer_whatsapp': idx_whatsapp,
                'customer_email': idx_email,
                'customer_address': idx_address,
                'dl_landmark': idx_landmark,
                'dl_building': idx_building,
                'dl_street': idx_street,
                'dl_zone': idx_zone,
                'cod_amount': idx_cod,
                'dl_amount': idx_dl_amount,
                'product_name': idx_product,
                'sku': idx_sku,
                'quantity': idx_qty,
                'deadline_date': idx_deadline,
                'seller_notes': idx_notes,
                'internal_notes': idx_internal,
            }
            for fname, flabel, scol in active_fields:
                if fname == 'row_number':
                    row_data['row_number'] = sheet_row
                elif fname == 'customer_name':
                    row_data['customer_name'] = cell(row, field_idx_map.get(fname)) or f'Row {sheet_row}'
                elif fname == 'client_order_code':
                    row_data['client_order_code'] = cell(row, field_idx_map.get(fname)) or str(sheet_row)
                else:
                    row_data[fname] = cell(row, field_idx_map.get(fname))
            preview_rows.append(row_data)

        return JsonResponse({
            'success': True,
            'api_type': 'google_sheet',
            'preview': preview_rows,
            'fields': [{'name': f[0], 'label': f[1], 'sheet_col': f[2]} for f in active_fields],
            'already_imported': already_imported,
            'has_mapping': bool(saved_mapping),
            'mapping_fields': len(saved_mapping),
        })
    except Exception as e:
        logger.exception('preview_api_import error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def import_api_orders(request):
    """
    POST: Fetch selected platform orders by platform_id from Shopify/WooCommerce
    and create Order records in the DB (is_transferred=False, ready for review).
    POST params: business_id, platform_ids[] (list of platform order IDs), source
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    import json
    from django.utils import timezone as dj_tz

    business_id = request.POST.get('business_id', '').strip()
    platform_ids = request.POST.getlist('platform_ids[]')
    source = request.POST.get('source', '').strip()

    if not business_id or not platform_ids:
        return JsonResponse({'success': False, 'error': 'Missing business_id or platform_ids'}, status=400)

    try:
        business = business_models.Business.objects.get(business_id=business_id)
        api = business.business_settings_api.filter(is_verify_api=True).first()
        if not api:
            return JsonResponse({'success': False, 'error': 'No approved API config for this business'}, status=400)
    except business_models.Business.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Business not found'}, status=404)

    created = 0
    skipped = 0
    errors = []

    try:
        if api.api_type == 'shopify':
            import shopify
            shop_name = (api.site_api_url or '').replace('https://', '').replace('http://', '').replace('.myshopify.com', '').strip()
            session = shopify.Session(shop_name, api.api_version or '2023-10', api.api_access_token)
            shopify.ShopifyResource.activate_session(session)

            for pid in platform_ids:
                try:
                    # Skip if already imported
                    if orders_models.Order.objects.filter(
                        business=business,
                        original_order_data__platform_id=pid,
                        original_order_data__source='shopify'
                    ).exists():
                        skipped += 1
                        continue

                    o = shopify.Order.find(pid)
                    shipping = getattr(o, 'shipping_address', None)
                    billing = getattr(o, 'billing_address', None)
                    addr = shipping or billing
                    addr_str = ' '.join(filter(None, [
                        getattr(addr, 'address1', '') or '',
                        getattr(addr, 'city', '') or '',
                    ])) if addr else ''
                    phone = getattr(addr, 'phone', '') or ''
                    customer = getattr(o, 'customer', None)
                    customer_name = f"{getattr(customer, 'first_name', '')} {getattr(customer, 'last_name', '')}".strip() if customer else ''
                    original_data = {
                        'source': 'shopify',
                        'platform_id': str(o.id),
                        'name': o.name,
                        'financial_status': o.financial_status,
                        'fulfillment_status': o.fulfillment_status,
                        'total_price': str(o.total_price),
                        'currency': o.currency,
                        'created_at': str(o.created_at),
                    }

                    orders_models.Order.objects.create(
                        business=business,
                        customer_name=customer_name or o.name,
                        customer_phone=phone,
                        customer_address=addr_str,
                        cod_amount=o.total_price,
                        order_status='to_review',
                        is_transferred=False,
                        original_order_data=original_data,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"{pid}: {e}")

            shopify.ShopifyResource.clear_session()

        elif api.api_type == 'woocommerce':
            from woocommerce import API as WooAPI
            wcapi = WooAPI(
                url=api.site_api_url or '',
                consumer_key=api.api_key or '',
                consumer_secret=api.api_secret or '',
                version='wc/v3',
                timeout=15,
            )
            for pid in platform_ids:
                try:
                    if orders_models.Order.objects.filter(
                        business=business,
                        original_order_data__platform_id=pid,
                        original_order_data__source='woocommerce'
                    ).exists():
                        skipped += 1
                        continue

                    r = wcapi.get(f'orders/{pid}')
                    if r.status_code != 200:
                        errors.append(f"#{pid}: HTTP {r.status_code}")
                        continue
                    o = r.json()
                    billing = o.get('billing', {})
                    shipping = o.get('shipping', {})
                    addr_parts = [
                        shipping.get('address_1') or billing.get('address_1', ''),
                        shipping.get('city') or billing.get('city', ''),
                    ]
                    addr_str = ', '.join(filter(None, addr_parts))
                    customer_name = f"{billing.get('first_name','')} {billing.get('last_name','')}".strip()

                    original_data = {
                        'source': 'woocommerce',
                        'platform_id': str(o.get('id')),
                        'number': o.get('number'),
                        'status': o.get('status'),
                        'total': str(o.get('total')),
                        'currency': o.get('currency'),
                        'date_created': o.get('date_created'),
                    }

                    orders_models.Order.objects.create(
                        business=business,
                        customer_name=customer_name or f"#{o.get('number')}",
                        customer_phone=billing.get('phone', ''),
                        customer_address=addr_str,
                        cod_amount=o.get('total') or 0,
                        order_status='to_review',
                        is_transferred=False,
                        original_order_data=original_data,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"{pid}: {e}")
        elif api.api_type == 'google_sheet':
            import re as _re
            import gspread
            from google.oauth2.credentials import Credentials as _GCreds
            from google.auth.transport.requests import Request as _GReq
            from django.conf import settings as django_settings
            from pathlib import Path as _Path
            import json as _json

            sheet_url = api.google_sheet_url or api.site_api_url or ''
            token_path = _Path(django_settings.BASE_DIR) / getattr(
                django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
            )
            if not token_path.exists():
                return JsonResponse({'success': False, 'error': 'Google Sheets not authorized.'}, status=400)

            _td = _json.loads(token_path.read_text())
            creds = _GCreds(
                token=_td.get('access_token'),
                refresh_token=_td.get('refresh_token'),
                token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=_td.get('client_id'),
                client_secret=_td.get('client_secret'),
                scopes=_td.get('scope', '').split(),
            )
            if creds.expired and creds.refresh_token:
                creds.refresh(_GReq())
                _td['access_token'] = creds.token
                token_path.write_text(_json.dumps(_td, indent=2))

            gc = gspread.authorize(creds)
            match = _re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
            if not match:
                return JsonResponse({'success': False, 'error': 'Invalid Google Sheet URL.'}, status=400)
            sheet_id = match.group(1)
            gid_match = _re.search(r'gid=(\d+)', sheet_url)
            gid = int(gid_match.group(1)) if gid_match else 0

            spreadsheet = gc.open_by_key(sheet_id)
            worksheet = None
            for ws in spreadsheet.worksheets():
                if ws.id == gid:
                    worksheet = ws
                    break
            if worksheet is None:
                worksheet = spreadsheet.sheet1

            all_values = worksheet.get_all_values()
            if len(all_values) <= 1:
                for ws in spreadsheet.worksheets():
                    if ws.id == (worksheet.id if worksheet else -1):
                        continue
                    candidate = ws.get_all_values()
                    if len(candidate) > 1:
                        worksheet = ws
                        all_values = candidate
                        break

            headers = all_values[0] if all_values else []
            data_rows = all_values[1:] if all_values else []

            # Use saved column mapping if available
            saved_mapping = api.column_mapping or {}

            def col_idx_by_header(header_name):
                for j, h in enumerate(headers):
                    if h.strip() == header_name.strip():
                        return j
                return None

            def col_idx_auto(*keys):
                for k in keys:
                    for j, h in enumerate(headers):
                        if h.strip().lower() == k.lower():
                            return j
                return None

            def get_idx(field_name, *auto_keys):
                if field_name in saved_mapping:
                    return col_idx_by_header(saved_mapping[field_name])
                return col_idx_auto(*auto_keys)

            idx_name = get_idx('customer_name', 'name', 'customer name', 'customer', 'client name', 'client', 'first name')
            idx_phone = get_idx('customer_phone', 'phone', 'mobile', 'phone number', 'contact')
            idx_whatsapp = get_idx('customer_whatsapp', 'whatsapp', 'phone 2', 'alternate phone')
            idx_email = get_idx('customer_email', 'email', 'e-mail', 'customer email')
            idx_address = get_idx('customer_address', 'address', 'delivery address', 'location', 'area', 'city')
            idx_landmark = get_idx('dl_landmark', 'city', 'landmark', 'customer city', 'area')
            idx_building = get_idx('dl_building', 'building', 'villa', 'building no', 'villa no')
            idx_street = get_idx('dl_street', 'street', 'street no', 'street number')
            idx_zone = get_idx('dl_zone', 'zone', 'zone no', 'zone number')
            idx_location_link = get_idx('location_link', 'location link', 'google map', 'map link', 'gps')
            idx_lat = get_idx('dl_latitude', 'latitude', 'lat')
            idx_lng = get_idx('dl_longitude', 'longitude', 'lng', 'long')
            idx_cod = get_idx('cod_amount', 'cod', 'amount', 'cod amount', 'price', 'total')
            idx_dl_amount = get_idx('dl_amount', 'delivery fee', 'delivery amount', 'shipping')
            idx_order = get_idx('client_order_code', 'order', 'order number', 'order #', 'order id', 'ref', 'reference')
            idx_date = get_idx('order_date', 'date', 'order date', 'created')
            idx_deadline = get_idx('deadline_date', 'deadline', 'day', 'time', 'preferred', 'delivery date')
            idx_product = get_idx('product_name', 'product', 'product name', 'item', 'items', 'description')
            idx_sku = get_idx('sku', 'sku', 'product sku', 'item sku', 'article')
            idx_product_url = get_idx('product_url', 'product url', 'url', 'product link')
            idx_qty = get_idx('quantity', 'qty', 'quantity', 'count', 'pcs')
            idx_seller_notes = get_idx('seller_notes', 'notes', 'seller notes', 'note', 'remarks', 'comment')
            idx_internal_notes = get_idx('internal_notes', 'notes by ezzy', 'internal notes', 'staff notes')
            # Additional product columns
            idx_products = {}
            for i in range(1, 6):
                idx_products[f'product_{i}'] = get_idx(f'product_{i}', f'product {i}', f'product:{i}', f'item {i}')
                idx_products[f'count_{i}'] = get_idx(f'count_{i}', f'count {i}', f'count:{i}', f'qty {i}')

            def cell(row, idx):
                if idx is not None and idx < len(row):
                    v = row[idx]
                    return str(v).strip() if v is not None else ''
                return ''

            def safe_int(val):
                try:
                    return int(float(val)) if val else 0
                except (ValueError, TypeError):
                    return 0

            def safe_int_or_none(val):
                try:
                    if val is None or val == '' or (isinstance(val, str) and not val.strip()):
                        return None
                    return int(float(val))
                except (ValueError, TypeError):
                    return None

            def safe_decimal(val):
                try:
                    from decimal import Decimal
                    return Decimal(str(val)) if val else None
                except (ValueError, TypeError):
                    return None

            # Parse row indices from platform_ids (format: gs_{row_idx}_{order_num})
            for pid in platform_ids:
                try:
                    parts = pid.split('_', 2)  # gs, row_idx, order_num
                    if len(parts) < 2:
                        errors.append(f"{pid}: invalid format")
                        continue
                    row_idx = int(parts[1])
                    sheet_row = row_idx + 2  # +2 for header + 1-based

                    # Check if already imported
                    if orders_models.Order.objects.filter(
                        business=business,
                    ).filter(
                        Q(original_order_data__platform_id=pid) |
                        Q(original_order_data__row_number=sheet_row)
                    ).exists():
                        skipped += 1
                        continue

                    if row_idx >= len(data_rows):
                        errors.append(f"Row {sheet_row}: out of range")
                        continue

                    row = data_rows[row_idx]
                    customer = cell(row, idx_name) or f"Row {sheet_row}"
                    phone = cell(row, idx_phone)
                    whatsapp_raw = cell(row, idx_whatsapp)
                    address = cell(row, idx_address)
                    cod_val = cell(row, idx_cod)
                    order_num = cell(row, idx_order) or str(sheet_row)
                    product_desc = cell(row, idx_product)

                    # Format WhatsApp
                    if whatsapp_raw:
                        customer_whatsapp = format_whatsapp_number(whatsapp_raw)
                    elif phone:
                        customer_whatsapp = format_whatsapp_number(phone)
                    else:
                        customer_whatsapp = ''

                    # Parse COD amount
                    try:
                        cod_amount = float(_re.sub(r'[^\d.]', '', cod_val)) if cod_val else 0
                    except (ValueError, TypeError):
                        cod_amount = 0

                    # Parse delivery amount
                    dl_amount_val = cell(row, idx_dl_amount)
                    try:
                        dl_amount = float(_re.sub(r'[^\d.]', '', dl_amount_val)) if dl_amount_val else 0
                    except (ValueError, TypeError):
                        dl_amount = 0

                    # Build notes
                    notes_parts = []
                    seller_notes = cell(row, idx_seller_notes)
                    internal_notes = cell(row, idx_internal_notes)
                    if seller_notes:
                        notes_parts.append(f"Seller: {seller_notes}")
                    if internal_notes:
                        notes_parts.append(f"Staff: {internal_notes}")
                    combined_notes = ' | '.join(notes_parts)

                    original_data = {
                        'source': 'google_sheet',
                        'platform_id': pid,
                        'row_number': sheet_row,
                        'order_number': order_num,
                        'import_data': {h: cell(row, i) for i, h in enumerate(headers)},
                        'extra_fields': {
                            'customer_email': cell(row, idx_email),
                            'dl_landmark': cell(row, idx_landmark),
                            'location_link': cell(row, idx_location_link),
                            'dl_latitude': cell(row, idx_lat),
                            'dl_longitude': cell(row, idx_lng),
                            'product_name': product_desc,
                            'sku': cell(row, idx_sku),
                            'product_url': cell(row, idx_product_url),
                            'quantity': cell(row, idx_qty),
                            'seller_notes': seller_notes,
                            'internal_notes': internal_notes,
                        },
                    }

                    order = orders_models.Order.objects.create(
                        business=business,
                        client_order_code=order_num if order_num != str(sheet_row) else '',
                        customer_name=customer,
                        customer_phone=phone,
                        customer_whatsapp=customer_whatsapp,
                        customer_address=address,
                        dl_zone=safe_int_or_none(cell(row, idx_zone)),
                        dl_street=safe_int_or_none(cell(row, idx_street)),
                        dl_building=safe_int_or_none(cell(row, idx_building)),
                        cod_amount=cod_amount,
                        dl_amount=dl_amount,
                        package_description=product_desc[:255] if product_desc else '',
                        total_quantity=safe_int(cell(row, idx_qty)) or 1,
                        deadline_date=cell(row, idx_deadline),
                        order_notes=combined_notes[:100] if combined_notes else '',
                        order_status='to_review',
                        verification_status='pending',
                        is_transferred=False,
                        original_order_data=original_data,
                    )

                    # Build package_description & total_quantity from all product fields
                    desc_parts = []
                    total_qty = 0
                    sku_val = cell(row, idx_sku)
                    main_qty = safe_int(cell(row, idx_qty)) or 1
                    if product_desc:
                        label = product_desc
                        if sku_val:
                            label += f" (SKU:{sku_val})"
                        desc_parts.append(f"{label} x{main_qty}")
                        total_qty += main_qty

                    for i in range(1, 6):
                        prod_name = cell(row, idx_products.get(f'product_{i}'))
                        prod_count = safe_int(cell(row, idx_products.get(f'count_{i}'))) or 1
                        if prod_name:
                            desc_parts.append(f"{prod_name} x{prod_count}")
                            total_qty += prod_count

                    if desc_parts:
                        order.package_description = ', '.join(desc_parts)[:255]
                        order.total_quantity = total_qty
                        order.save(update_fields=['package_description', 'total_quantity'])

                    # Create AddressVerification if lat/long provided
                    lat = safe_decimal(cell(row, idx_lat))
                    lng = safe_decimal(cell(row, idx_lng))
                    if lat and lng:
                        orders_models.AddressVerification.objects.create(
                            order=order,
                            original_address=address,
                            latitude=lat,
                            longitude=lng,
                            zone_number=safe_int(cell(row, idx_zone)) or None,
                            street_number=safe_int(cell(row, idx_street)) or None,
                            building_number=safe_int(cell(row, idx_building)) or None,
                            notes=f"Landmark: {cell(row, idx_landmark)} | Link: {cell(row, idx_location_link)}",
                            verification_result='pending'
                        )

                    created += 1
                except Exception as e:
                    errors.append(f"{pid}: {e}")

        else:
            return JsonResponse({'success': False, 'error': f'Import not supported for {api.api_type}'}, status=400)

    except Exception as e:
        logger.exception('import_api_orders error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({
        'success': True,
        'created': created,
        'skipped': skipped,
        'errors': errors,
    })


@login_required(login_url='/accounts/login/')
@staff_required
def wf_sheet_headers(request):
    """GET: Return the column headers from a business's Google Sheet config."""
    business_id = request.GET.get('business_id', '').strip()
    if not business_id:
        return JsonResponse({'success': False, 'error': 'Missing business_id'}, status=400)

    try:
        api = business_models.BusinessApiSettings.objects.get(
            business__business_id=business_id, api_type='google_sheet', is_verify_api=True
        )
    except business_models.BusinessApiSettings.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No approved Google Sheet config'}, status=404)

    import re as _re, json as _json, gspread
    from google.oauth2.credentials import Credentials as _GCreds
    from google.auth.transport.requests import Request as _GReq
    from django.conf import settings as django_settings
    from pathlib import Path as _Path

    sheet_url = api.google_sheet_url or api.site_api_url or ''
    token_path = _Path(django_settings.BASE_DIR) / getattr(
        django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
    )
    if not token_path.exists():
        return JsonResponse({'success': False, 'error': 'Google Sheets not authorized.'}, status=400)

    try:
        _td = _json.loads(token_path.read_text())
        creds = _GCreds(
            token=_td.get('access_token'),
            refresh_token=_td.get('refresh_token'),
            token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
            client_id=_td.get('client_id'),
            client_secret=_td.get('client_secret'),
            scopes=_td.get('scope', '').split(),
        )
        if creds.expired and creds.refresh_token:
            creds.refresh(_GReq())
            _td['access_token'] = creds.token
            token_path.write_text(_json.dumps(_td, indent=2))

        gc = gspread.authorize(creds)
        match = _re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
        if not match:
            return JsonResponse({'success': False, 'error': 'Invalid Sheet URL'}, status=400)

        spreadsheet = gc.open_by_key(match.group(1))
        gid_match = _re.search(r'gid=(\d+)', sheet_url)
        gid = int(gid_match.group(1)) if gid_match else 0

        worksheet = None
        for ws in spreadsheet.worksheets():
            if ws.id == gid:
                worksheet = ws
                break
        if worksheet is None:
            worksheet = spreadsheet.sheet1

        all_values = worksheet.get_all_values()
        if len(all_values) <= 1:
            for ws in spreadsheet.worksheets():
                if ws.id == (worksheet.id if worksheet else -1):
                    continue
                candidate = ws.get_all_values()
                if len(candidate) > 1:
                    all_values = candidate
                    break

        headers = all_values[0] if all_values else []
        # Also return sample data (first data row) for preview
        sample = all_values[1] if len(all_values) > 1 else []

        return JsonResponse({
            'success': True,
            'headers': headers,
            'sample': sample,
            'saved_mapping': api.column_mapping or {},
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def wf_save_column_mapping(request):
    """POST: Save column mapping for a business's Google Sheet config."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    import json
    business_id = request.POST.get('business_id', '').strip()
    mapping_json = request.POST.get('mapping', '{}')

    if not business_id:
        return JsonResponse({'success': False, 'error': 'Missing business_id'}, status=400)

    try:
        api = business_models.BusinessApiSettings.objects.get(
            business__business_id=business_id, api_type='google_sheet', is_verify_api=True
        )
    except business_models.BusinessApiSettings.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'No approved Google Sheet config'}, status=404)

    try:
        mapping = json.loads(mapping_json)
    except (json.JSONDecodeError, TypeError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)

    api.column_mapping = mapping
    api.save(update_fields=['column_mapping'])
    return JsonResponse({'success': True, 'mapping': mapping})


@login_required(login_url='/accounts/login/')
@staff_required
def wf_api_orders(request):
    """
    List orders from approved API businesses.
    When a business is selected, fetch live orders from their Shopify/WooCommerce API.
    """
    # Businesses with at least one approved API config
    api_businesses = business_models.Business.objects.filter(
        business_settings_api__is_verify_api=True
    ).prefetch_related('business_settings_api').distinct().order_by('business_name')

    selected_business_id = request.GET.get('business', '')
    status_filter = request.GET.get('status', '')
    selected_business = None
    selected_api = None
    live_orders = []
    live_error = None
    live_total = None

    if selected_business_id:
        try:
            selected_business = business_models.Business.objects.prefetch_related(
                'business_settings_api'
            ).get(business_id=selected_business_id)
            # Get the approved API config for this business
            selected_api = selected_business.business_settings_api.filter(
                is_verify_api=True
            ).first()
        except business_models.Business.DoesNotExist:
            pass

    # Fetch live orders from the platform API when a business with approved config is selected
    if selected_api:
        try:
            if selected_api.api_type == 'shopify':
                import shopify
                shop_name = (selected_api.site_api_url or '').replace('https://', '').replace('http://', '').replace('.myshopify.com', '').strip()
                session = shopify.Session(shop_name, selected_api.api_version or '2023-10', selected_api.api_access_token)
                shopify.ShopifyResource.activate_session(session)
                orders = shopify.Order.find(limit=50, status='any')
                for o in orders:
                    ship = getattr(o, 'shipping_address', None)
                    live_orders.append({
                        'platform_id': str(o.id),
                        'name': o.name,
                        'customer': getattr(o, 'contact_email', '') or (getattr(o.customer, 'email', '') if hasattr(o, 'customer') and o.customer else ''),
                        'phone': getattr(ship, 'phone', '') if ship else '',
                        'address': getattr(ship, 'address1', '') if ship else '',
                        'country': getattr(ship, 'country', '') if ship else '',
                        'cod': o.total_price,
                        'currency': o.currency,
                        'status': o.financial_status,
                        'fulfillment': o.fulfillment_status or 'unfulfilled',
                        'date': o.created_at,
                        'source': 'shopify',
                    })
                try:
                    live_total = shopify.Order.count(status='any')
                except Exception:
                    live_total = len(live_orders)
                shopify.ShopifyResource.clear_session()

            elif selected_api.api_type == 'woocommerce':
                from woocommerce import API as WooAPI
                wcapi = WooAPI(
                    url=selected_api.site_api_url or '',
                    consumer_key=selected_api.api_key or '',
                    consumer_secret=selected_api.api_secret or '',
                    version='wc/v3',
                    timeout=15,
                )
                r = wcapi.get('orders', params={'per_page': 50, 'orderby': 'date', 'order': 'desc'})
                if r.status_code == 200:
                    live_total = r.headers.get('X-WP-Total')
                    for o in r.json():
                        billing = o.get('billing', {})
                        shipping = o.get('shipping', {})
                        addr_parts = [shipping.get('address_1') or billing.get('address_1'), shipping.get('city') or billing.get('city')]
                        live_orders.append({
                            'platform_id': str(o.get('id')),
                            'name': f"#{o.get('number')}",
                            'customer': f"{billing.get('first_name','')} {billing.get('last_name','')}".strip(),
                            'phone': billing.get('phone', ''),
                            'address': ', '.join(filter(None, addr_parts)),
                            'country': shipping.get('country') or billing.get('country', ''),
                            'cod': o.get('total'),
                            'currency': o.get('currency'),
                            'status': o.get('status'),
                            'fulfillment': None,
                            'date': o.get('date_created'),
                            'source': 'woocommerce',
                        })
                else:
                    live_error = f"API error {r.status_code}: {r.text[:200]}"

            elif selected_api.api_type == 'google_sheet':
                import re
                import gspread
                from google.oauth2.credentials import Credentials
                from google.auth.transport.requests import Request
                from django.conf import settings as django_settings
                from pathlib import Path

                sheet_url = selected_api.google_sheet_url or selected_api.site_api_url or ''
                token_path = Path(django_settings.BASE_DIR) / getattr(
                    django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
                )
                if not token_path.exists():
                    raise Exception(
                        'Google Sheets not authorized yet. '
                        'Run: python google_sheets_auth.py (one-time setup)'
                    )

                import json as _json
                _td = _json.loads(token_path.read_text())
                creds = Credentials(
                    token=_td.get('access_token'),
                    refresh_token=_td.get('refresh_token'),
                    token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
                    client_id=_td.get('client_id'),
                    client_secret=_td.get('client_secret'),
                    scopes=_td.get('scope', '').split(),
                )
                if creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                    _td['access_token'] = creds.token
                    token_path.write_text(_json.dumps(_td, indent=2))

                gc = gspread.authorize(creds)

                match = re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
                if not match:
                    raise Exception('Invalid Google Sheet URL.')
                sheet_id = match.group(1)
                gid_match = re.search(r'gid=(\d+)', sheet_url)
                gid = int(gid_match.group(1)) if gid_match else 0

                spreadsheet = gc.open_by_key(sheet_id)
                worksheet = None
                for ws in spreadsheet.worksheets():
                    if ws.id == gid:
                        worksheet = ws
                        break
                if worksheet is None:
                    worksheet = spreadsheet.sheet1

                all_values = worksheet.get_all_values()
                if not all_values or len(all_values) <= 1:
                    # Try other worksheets if current one is empty
                    for ws in spreadsheet.worksheets():
                        if ws.id == worksheet.id:
                            continue
                        candidate = ws.get_all_values()
                        if len(candidate) > 1:
                            worksheet = ws
                            all_values = candidate
                            break

                if not all_values or len(all_values) <= 1:
                    raise Exception('Sheet is empty — no data rows found.')

                headers = all_values[0]
                data_rows = all_values[1:]
                live_total = len(data_rows)

                # Use saved column mapping if available, else auto-detect
                saved_mapping = selected_api.column_mapping or {}

                def col_idx_by_header(header_name):
                    """Find column index by exact header name."""
                    for j, h in enumerate(headers):
                        if h.strip() == header_name.strip():
                            return j
                    return None

                def col_idx_auto(*keys):
                    """Auto-detect column index by common name variants."""
                    for k in keys:
                        for j, h in enumerate(headers):
                            if h.strip().lower() == k.lower():
                                return j
                    return None

                def get_idx(field_name, *auto_keys):
                    """Get column index: saved mapping first, then auto-detect."""
                    if field_name in saved_mapping:
                        return col_idx_by_header(saved_mapping[field_name])
                    return col_idx_auto(*auto_keys)

                idx_name = get_idx('customer_name', 'name', 'customer name', 'customer', 'client name', 'client', 'first name')
                idx_phone = get_idx('customer_phone', 'phone', 'mobile', 'phone number', 'contact')
                idx_address = get_idx('customer_address', 'address', 'delivery address', 'location', 'area', 'city')
                idx_cod = get_idx('cod_amount', 'cod', 'amount', 'cod amount', 'price', 'total')
                idx_status = col_idx_auto('status', 'order status', 'state')
                idx_date = get_idx('order_date', 'date', 'order date', 'created', 'created at')
                idx_country = col_idx_auto('country', 'country code')
                idx_order = get_idx('client_order_code', 'order', 'order number', 'order #', 'order id', 'ref', 'reference')
                idx_product = get_idx('product_name', 'product', 'product name', 'item', 'items')

                def cell(row, idx):
                    if idx is not None and idx < len(row):
                        v = row[idx]
                        return str(v).strip() if v is not None else ''
                    return ''

                # Build display headers: non-empty sheet headers
                sheet_display_headers = [h for h in headers if h.strip()]

                for i, row in enumerate(data_rows):
                    sheet_row = i + 2  # +2 for header + 1-based
                    order_num = cell(row, idx_order) or str(sheet_row)
                    # Build all cell values keyed by header name
                    all_cells = {}
                    for j, h in enumerate(headers):
                        if h.strip():
                            all_cells[h.strip()] = cell(row, j)
                    live_orders.append({
                        'platform_id': f'gs_{i}_{order_num}',
                        'name': order_num,
                        'customer': cell(row, idx_name),
                        'phone': cell(row, idx_phone),
                        'address': cell(row, idx_address),
                        'country': cell(row, idx_country),
                        'cod': cell(row, idx_cod),
                        'currency': '',
                        'status': cell(row, idx_status) or 'pending',
                        'fulfillment': None,
                        'date': cell(row, idx_date),
                        'source': 'google_sheet',
                        'row_number': sheet_row,
                        'cells': all_cells,
                    })

        except Exception as e:
            live_error = str(e)

    # Annotate live_orders with import status from DB
    if live_orders and selected_business and selected_api:
        if selected_api.api_type == 'google_sheet':
            # For Google Sheets, check by row_number stored in original_order_data
            imported_qs = orders_models.Order.objects.filter(
                business=selected_business,
            ).filter(
                Q(original_order_data__source='google_sheet') |
                Q(original_order_data__source='onedrive')
            ).values('original_order_data', 'id', 'order_number', 'client_order_code')

            imported_by_row = {}
            imported_by_pid = {}
            for r in imported_qs:
                od = r.get('original_order_data') or {}
                info = {
                    'id': r['id'],
                    'number': r['order_number'] or r['client_order_code'] or f"#{r['id']}",
                }
                rn = od.get('row_number')
                if rn is not None:
                    imported_by_row[int(rn)] = info
                pid = od.get('platform_id')
                if pid:
                    imported_by_pid[str(pid)] = info

            for o in live_orders:
                match = imported_by_row.get(o.get('row_number')) or imported_by_pid.get(str(o['platform_id']))
                o['imported'] = bool(match)
                o['imported_order_id'] = match['id'] if match else None
                o['imported_order_number'] = match['number'] if match else None
        else:
            platform_ids = [o['platform_id'] for o in live_orders]
            imported_qs = orders_models.Order.objects.filter(
                business=selected_business,
                original_order_data__source=selected_api.api_type,
                original_order_data__platform_id__in=platform_ids,
            ).values('original_order_data__platform_id', 'id', 'order_number', 'client_order_code')
            imported_map = {
                str(r['original_order_data__platform_id']): {
                    'id': r['id'],
                    'number': r['order_number'] or r['client_order_code'] or f"#{r['id']}",
                }
                for r in imported_qs
            }
            for o in live_orders:
                match = imported_map.get(str(o['platform_id']))
                o['imported'] = bool(match)
                o['imported_order_id'] = match['id'] if match else None
                o['imported_order_number'] = match['number'] if match else None

    # Apply import status filter
    if status_filter == 'imported':
        live_orders = [o for o in live_orders if o.get('imported')]
    elif status_filter == 'not_imported':
        live_orders = [o for o in live_orders if not o.get('imported')]

    # Find last imported row number for default "Go to row"
    last_imported_row = 0
    for o in live_orders:
        if o.get('imported') and o.get('row_number'):
            last_imported_row = max(last_imported_row, o['row_number'])
    # Default goto row: last imported - 5 (min row 2)
    default_goto_row = max(2, last_imported_row - 5) if last_imported_row else 0

    # Paginate live_orders (list, not queryset)
    from django.core.paginator import Paginator
    try:
        per_page = max(1, min(200, int(request.GET.get('per_page', 25))))
    except (ValueError, TypeError):
        per_page = 25
    paginator = Paginator(live_orders, per_page)

    # If no page param and we have a default_goto_row, auto-navigate to that page
    explicit_page = request.GET.get('page')
    if not explicit_page and default_goto_row and selected_api and selected_api.api_type == 'google_sheet':
        # row_number is sheet_row (1-based, header=1), data index = row_number - 2
        data_idx = default_goto_row - 2
        page_number = max(1, (data_idx // per_page) + 1)
    else:
        try:
            page_number = int(explicit_page or 1)
        except (ValueError, TypeError):
            page_number = 1
    page_obj = paginator.get_page(page_number)

    # Sheet headers for Google Sheet dynamic columns
    gs_headers = []
    if selected_api and selected_api.api_type == 'google_sheet':
        try:
            gs_headers = sheet_display_headers
        except NameError:
            gs_headers = []

    context = {
        'api_businesses': api_businesses,
        'selected_business': selected_business,
        'selected_business_id': selected_business_id,
        'selected_api': selected_api,
        'status_filter': status_filter,
        'live_orders': page_obj,
        'live_error': live_error,
        'live_total': live_total,
        'total_filtered': len(live_orders),
        'page_obj': page_obj,
        'per_page': per_page,
        'gs_headers': gs_headers,
        'default_goto_row': default_goto_row,
    }
    return render(request, 'workforce/orders_api.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def orders_api_guide(request):
    """
    Display API documentation for order creation.
    """
    # Get all active businesses for the example
    businesses = business_models.Business.objects.filter(
        business_status='active'
    ).order_by('business_name')[:5]

    context = {
        'businesses': businesses,
    }
    return render(request, 'workforce/orders_api_guide.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def get_pickup_locations(request, business_id):
    """AJAX endpoint to get pickup locations for a business"""
    try:
        business = business_models.Business.objects.get(business_id=business_id)
        # Show fulfillment stores first when fulfillment service is enabled
        locations = business_models.PickupLocation.objects.filter(
            business=business
        ).order_by('-is_fulfilment_center', 'pickup_location_title')

        location_list = [{
            'id': loc.id,
            'name': loc.pickup_location_title,
            'address': loc.locality or '',
        } for loc in locations]

        return JsonResponse({'success': True, 'locations': location_list})
    except business_models.Business.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Business not found'}, status=404)


@login_required(login_url='/accounts/login/')
@staff_required
def workforce_pickup_location_add(request, business_id):
    """Workforce view for products management - add products and view products list"""
    try:
        from product import models as product_models
        from django.db.models import Sum, Q

        business = business_models.Business.objects.get(business_id=business_id)

        if request.method == 'POST':
            action = request.POST.get('action')

            if action == 'add_product':
                try:
                    # Create product
                    product = product_models.Product.objects.create(
                        business=business,
                        brand_name=request.POST.get('brand_name'),
                        item_name=request.POST.get('item_name'),
                        item_sku=request.POST.get('item_sku'),
                        barcode=request.POST.get('barcode') or None,
                        item_price=int(request.POST.get('item_price', 0)),
                        item_discription=request.POST.get('item_discription') or '',
                    )

                    # Create inventory record if quantity provided
                    quantity = int(request.POST.get('quantity', 0))
                    if quantity > 0:
                        product_models.ProductInventory.objects.create(
                            item_sku=product,
                            item_quantity=quantity
                        )

                    logger.info(f"Workforce user {request.user.id} added product {product.item_sku} for business {business.business_id}")
                    messages.success(request, f"Product '{product.item_name}' added successfully")

                    # Return JSON for AJAX requests
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': True, 'message': 'Product added successfully'})

                    return redirect('workforce:workforce_pickup_location_add', business_id=business.business_id)
                except Exception as e:
                    logger.error(f"Error adding product: {str(e)}")
                    messages.error(request, f"Error adding product: {str(e)}")

                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': str(e)}, status=400)

        # Get all products for this business with inventory
        products = product_models.Product.objects.filter(
            business=business
        ).select_related(
            'product_category', 'color', 'unit'
        ).prefetch_related(
            'product_inventory'
        ).order_by('-created_at')

        # Annotate with total stock quantity
        products_with_stock = []
        for product in products:
            total_stock = product.product_inventory.aggregate(
                total=Sum('item_quantity')
            )['total'] or 0
            product.stock_quantity = total_stock
            products_with_stock.append(product)

        context = {
            'business': business,
            'products': products_with_stock,
        }
        return render(request, 'workforce/workforce_pickup_location_add.html', context)
    except business_models.Business.DoesNotExist:
        logger.error(f"Business {business_id} not found")
        messages.error(request, "Business not found")
        return redirect('workforce:business_licenses_list')


# Delivery Tasks section  ------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
def dl_list_all(request):
    from django.db.models import Prefetch
    failure_history_qs = orders_models.OrderStatusHistory.objects.filter(
        field_name='dl_task_status', new_value='failed'
    ).order_by('-created_at')
    dl_tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_comments',
        'order__order_items',
        'order__order_items__product',
        'task_qrcode',
        Prefetch('order__status_history', queryset=failure_history_qs, to_attr='failure_history'),
    ).all().order_by('-created_at')

    # Get filter parameters
    dl_code = request.GET.get('dlCode', '')
    c_code = request.GET.get('cCode', '')
    mobile = request.GET.get('mobile', '')
    driver_name = request.GET.get('driverName', '')
    c_status = request.GET.get('cStatus', '')
    dl_status = request.GET.get('dlStatus', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    business_id = request.GET.get('business', '')

    # Apply filters
    if dl_code:
        dl_tasks = dl_tasks.filter(dl_task_number__icontains=dl_code)
    if c_code:
        dl_tasks = dl_tasks.filter(order__client_order_code__icontains=c_code)
    if mobile:
        dl_tasks = dl_tasks.filter(order__customer_phone__icontains=mobile)
    if driver_name:
        dl_tasks = dl_tasks.filter(
            Q(driver__user__first_name__icontains=driver_name) |
            Q(driver__user__last_name__icontains=driver_name) |
            Q(driver__user__username__icontains=driver_name)
        )
    if c_status:
        dl_tasks = dl_tasks.filter(dl_task_status_client=c_status)
    if dl_status:
        dl_tasks = dl_tasks.filter(dl_task_status=dl_status)
    if date_from:
        dl_tasks = dl_tasks.filter(dl_task_date__gte=date_from)
    if date_to:
        dl_tasks = dl_tasks.filter(dl_task_date__lte=date_to)
    if business_id:
        dl_tasks = dl_tasks.filter(order__business_id=business_id)

    # Apply sorting
    sort_field = request.GET.get('sort', '')
    sort_order = request.GET.get('order', 'asc')
    SORT_MAP = {
        'task_number': 'dl_task_number',
        'date': 'created_at',
        'business': 'order__business__business_name',
        'customer': 'order__customer_name',
        'cod': 'order__order_cod_amount',
        'status': 'dl_task_status',
        'driver': 'driver__user__first_name',
    }
    if sort_field in SORT_MAP:
        order_field = SORT_MAP[sort_field]
        if sort_order == 'desc':
            order_field = '-' + order_field
        dl_tasks = dl_tasks.order_by(order_field)

    # Get all businesses for the filter dropdown
    businesses = business_models.Business.objects.filter(
        business_status='active'
    ).order_by('business_name')

    dl_tasks = paginate_queryset(request, dl_tasks)

    data = {
        'dl_tasks': dl_tasks,
        'businesses': businesses,
        'today': timezone.localtime().date(),
        'page_title': 'All Delivery Tasks',
        'page_subtitle': 'Manage and track',
        'page_icon': 'fa-tasks',
        'list_type': 'all',
        'show_filters': True,
        'filters': {
            'dlCode': dl_code,
            'cCode': c_code,
            'mobile': mobile,
            'driverName': driver_name,
            'cStatus': c_status,
            'dlStatus': dl_status,
            'dateFrom': date_from,
            'dateTo': date_to,
            'business': business_id,
        }
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def fulfilled_clients_tasks(request):
    """Tasks from businesses with fulfillment service enabled"""
    dl_tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_comments',
        'order__order_items',
        'order__order_items__product',
        'task_qrcode',
    ).filter(
        order__business__fulfillment_service_enabled=True
    ).order_by('-created_at')

    # Get filter parameters
    dl_code = request.GET.get('dlCode', '')
    c_code = request.GET.get('cCode', '')
    mobile = request.GET.get('mobile', '')
    driver_name = request.GET.get('driverName', '')
    c_status = request.GET.get('cStatus', '')
    dms_status = request.GET.get('dmsStatus', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    business_id = request.GET.get('business', '')

    # Apply filters
    if dl_code:
        dl_tasks = dl_tasks.filter(dl_task_number__icontains=dl_code)
    if c_code:
        dl_tasks = dl_tasks.filter(order__client_order_code__icontains=c_code)
    if mobile:
        dl_tasks = dl_tasks.filter(order__customer_phone__icontains=mobile)
    if driver_name:
        dl_tasks = dl_tasks.filter(
            Q(driver__user__first_name__icontains=driver_name) |
            Q(driver__user__last_name__icontains=driver_name) |
            Q(driver__user__username__icontains=driver_name)
        )
    if c_status:
        dl_tasks = dl_tasks.filter(dl_task_status_client=c_status)
    if date_from:
        dl_tasks = dl_tasks.filter(dl_task_date__gte=date_from)
    if date_to:
        dl_tasks = dl_tasks.filter(dl_task_date__lte=date_to)
    if business_id:
        dl_tasks = dl_tasks.filter(order__business_id=business_id)

    # Get fulfillment-enabled businesses
    businesses = business_models.Business.objects.filter(
        business_status='active',
        fulfillment_service_enabled=True
    ).order_by('business_name')

    dl_tasks = paginate_queryset(request, dl_tasks)

    data = {
        'dl_tasks': dl_tasks,
        'businesses': businesses,
        'today': timezone.localtime().date(),
        'page_title': 'Fulfilled Clients Tasks',
        'page_subtitle': 'Tasks from fulfillment-enabled businesses',
        'page_icon': 'fa-truck-ramp-box',
        'list_type': 'fulfilled',
        'show_filters': True,
        'filters': {
            'dlCode': dl_code,
            'cCode': c_code,
            'mobile': mobile,
            'driverName': driver_name,
            'cStatus': c_status,
            'dmsStatus': dms_status,
            'dateFrom': date_from,
            'dateTo': date_to,
            'business': business_id,
        }
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def non_fulfilled_clients_tasks(request):
    """Tasks from businesses without fulfillment service"""
    dl_tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_comments',
        'order__order_items',
        'order__order_items__product',
        'task_qrcode',
    ).exclude(
        order__business__fulfillment_service_status='active'
    ).order_by('-created_at')

    # Get filter parameters
    dl_code = request.GET.get('dlCode', '')
    c_code = request.GET.get('cCode', '')
    mobile = request.GET.get('mobile', '')
    driver_name = request.GET.get('driverName', '')
    c_status = request.GET.get('cStatus', '')
    dms_status = request.GET.get('dmsStatus', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    business_id = request.GET.get('business', '')

    # Apply filters
    if dl_code:
        dl_tasks = dl_tasks.filter(dl_task_number__icontains=dl_code)
    if c_code:
        dl_tasks = dl_tasks.filter(order__client_order_code__icontains=c_code)
    if mobile:
        dl_tasks = dl_tasks.filter(order__customer_phone__icontains=mobile)
    if driver_name:
        dl_tasks = dl_tasks.filter(
            Q(driver__user__first_name__icontains=driver_name) |
            Q(driver__user__last_name__icontains=driver_name) |
            Q(driver__user__username__icontains=driver_name)
        )
    if c_status:
        dl_tasks = dl_tasks.filter(dl_task_status_client=c_status)
    if date_from:
        dl_tasks = dl_tasks.filter(dl_task_date__gte=date_from)
    if date_to:
        dl_tasks = dl_tasks.filter(dl_task_date__lte=date_to)
    if business_id:
        dl_tasks = dl_tasks.filter(order__business_id=business_id)

    # Get non-fulfillment businesses
    businesses = business_models.Business.objects.filter(
        business_status='active',
        fulfillment_service_enabled=False
    ).order_by('business_name')

    dl_tasks = paginate_queryset(request, dl_tasks)

    data = {
        'dl_tasks': dl_tasks,
        'businesses': businesses,
        'today': timezone.localtime().date(),
        'page_title': 'Non-Fulfilled Clients Tasks',
        'page_subtitle': 'Tasks from standard delivery businesses',
        'page_icon': 'fa-truck-fast',
        'list_type': 'non_fulfilled',
        'show_filters': True,
        'filters': {
            'dlCode': dl_code,
            'cCode': c_code,
            'mobile': mobile,
            'driverName': driver_name,
            'cStatus': c_status,
            'dmsStatus': dms_status,
            'dateFrom': date_from,
            'dateTo': date_to,
            'business': business_id,
        }
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def dl_list_incompleted_details(request):
    # Get incomplete delivery tasks (not delivered, not cancelled)
    dl_tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_items',
    ).exclude(
        dl_task_status__in=['delivered', 'cancelled']
    ).order_by('-created_at')

    # Get filter parameters
    dl_code = request.GET.get('dlCode', '')
    c_code = request.GET.get('cCode', '')
    mobile = request.GET.get('mobile', '')
    driver_name = request.GET.get('driverName', '')
    c_status = request.GET.get('cStatus', '')
    dms_status = request.GET.get('dmsStatus', '')
    date_from = request.GET.get('dateFrom', '')
    date_to = request.GET.get('dateTo', '')
    business_id = request.GET.get('business', '')

    if dl_code:
        dl_tasks = dl_tasks.filter(dl_task_number__icontains=dl_code)
    if c_code:
        dl_tasks = dl_tasks.filter(order__client_order_code__icontains=c_code)
    if mobile:
        dl_tasks = dl_tasks.filter(order__customer_phone__icontains=mobile)
    if driver_name:
        dl_tasks = dl_tasks.filter(
            Q(driver__user__first_name__icontains=driver_name) |
            Q(driver__user__last_name__icontains=driver_name) |
            Q(driver__user__username__icontains=driver_name)
        )
    if c_status:
        dl_tasks = dl_tasks.filter(dl_task_status_client=c_status)
    if date_from:
        dl_tasks = dl_tasks.filter(dl_task_date__gte=date_from)
    if date_to:
        dl_tasks = dl_tasks.filter(dl_task_date__lte=date_to)
    if business_id:
        dl_tasks = dl_tasks.filter(order__business_id=business_id)

    businesses = business_models.Business.objects.filter(
        business_status='active'
    ).order_by('business_name')

    dl_tasks = paginate_queryset(request, dl_tasks)

    data = {
        'dl_tasks': dl_tasks,
        'businesses': businesses,
        'today': timezone.localtime().date(),
        'page_title': 'Incompleted Tasks',
        'page_subtitle': 'Pending delivery tasks',
        'page_icon': 'fa-clock-rotate-left',
        'list_type': 'incompleted',
        'show_filters': True,
        'filters': {
            'dlCode': dl_code, 'cCode': c_code, 'mobile': mobile,
            'driverName': driver_name, 'cStatus': c_status, 'dmsStatus': dms_status,
            'dateFrom': date_from, 'dateTo': date_to, 'business': business_id,
        }
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def dl_list_published_to_dms(request):
    dl_tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_items',
    ).filter(
        dl_task_publish=True
    ).order_by('-created_at')
    dl_tasks = paginate_queryset(request, dl_tasks)

    data = {
        'dl_tasks': dl_tasks,
        'today': timezone.localtime().date(),
        'page_title': 'Published to DMS',
        'page_subtitle': 'Tasks published to delivery management system',
        'page_icon': 'fa-cloud-arrow-up',
        'list_type': 'published',
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', data)


@login_required(login_url='/accounts/login/')
@staff_required
def dl_list_ready_to_published_to_dms(request):
    """List orders ready to be published to DMS (task_created=False)"""
    orders = orders_models.Order.objects.select_related(
        'business', 'pickup_location'
    ).prefetch_related(
        'delivery_task'
    ).filter(task_created=False).order_by('-created_at')
    orders = paginate_queryset(request, orders)

    data = {
        'orders': orders,
    }
    return render(request, 'workforce/parts/lists/dl_list_unpublished.html', data)




# DMS section  ------------------------------------------------------------------------------------------------------


# Workflow Guide -----------------------------------------------------

@login_required(login_url='account_login')
@staff_required
def workflow_guide(request):
    """Display comprehensive workflow guide for workforce/staff members"""

    workflow_steps = [
        {
            'number': 1,
            'title': 'Dashboard Overview',
            'description': 'Familiarize yourself with the workforce dashboard',
            'tasks': [
                'View pending orders count',
                'Check orders awaiting verification',
                'Monitor delivery tasks status',
                'Review today\'s activities',
            ],
            'status': 'completed',
            'url': 'workforce:wf_dashboard',
        },
        {
            'number': 2,
            'title': 'Review Pending Orders',
            'description': 'Check new orders that need verification',
            'tasks': [
                'Go to Orders → Pending Verification',
                'Review order details (customer name, address, COD)',
                'Check if all required information is provided',
                'Identify orders with missing information',
            ],
            'status': 'in_progress',
            'url': 'workforce:orders_pending_verification',
        },
        {
            'number': 3,
            'title': 'Verify Customer Address',
            'description': 'Validate and verify delivery addresses',
            'tasks': [
                'Click "Verify Address" on pending order',
                'Validate address format and completeness',
                'Use map integration to confirm coordinates',
                'Set zone, street, and building numbers',
                'Add GPS coordinates (latitude/longitude)',
                'Mark verification result: Valid, Needs Update, or Invalid',
                'Add notes if address needs customer contact',
            ],
            'status': 'primary',
            'url': None,
        },
        {
            'number': 4,
            'title': 'Contact Customer (If Needed)',
            'description': 'Reach out for address clarification',
            'tasks': [
                'Call customer using provided phone number',
                'Request missing address details',
                'Confirm zone, street, building information',
                'Update order with corrected information',
                'Log communication in order notes',
            ],
            'status': 'conditional',
            'url': None,
        },
        {
            'number': 5,
            'title': 'Verify Complete Order',
            'description': 'Final verification before delivery task creation',
            'tasks': [
                'Ensure address is verified',
                'Confirm customer contact details',
                'Validate COD amount if applicable',
                'Check product list completeness',
                'Click "Verify Order" button',
                'Order automatically becomes "Verified"',
            ],
            'status': 'primary',
            'url': None,
        },
        {
            'number': 6,
            'title': 'Automated Task Creation',
            'description': 'System automatically creates delivery tasks',
            'tasks': [
                'Verified order triggers automatic process',
                'Delivery task created with all order details',
                'Address information duplicated to task',
                'Original order preserved as proof',
                'Task pushed to DMS automatically',
                'You receive confirmation notification',
            ],
            'status': 'automated',
            'url': None,
        },
        {
            'number': 7,
            'title': 'Monitor Delivery Tasks',
            'description': 'Track delivery progress and status',
            'tasks': [
                'Go to Tasks → All Delivery Tasks',
                'Check task status: Published, Assigned, In Transit',
                'Monitor driver assignments',
                'View real-time delivery updates from DMS',
                'Check completion proof and signatures',
            ],
            'status': 'pending',
            'url': 'workforce:dl_list_all',
        },
        {
            'number': 8,
            'title': 'Handle Exceptions',
            'description': 'Manage orders that need special attention',
            'tasks': [
                'Identify rejected or failed deliveries',
                'Contact customers for failed delivery reasons',
                'Reschedule deliveries if needed',
                'Update order status accordingly',
                'Document all actions in order logs',
            ],
            'status': 'conditional',
            'url': None,
        },
        {
            'number': 9,
            'title': 'COD Collection Tracking',
            'description': 'Monitor Cash on Delivery payments',
            'tasks': [
                'Track COD amounts to be collected',
                'Verify driver has collected payment',
                'Update COD status: With Driver, With EZZY, Settled',
                'Generate COD collection reports',
                'Coordinate with finance for settlement',
            ],
            'status': 'pending',
            'url': None,
        },
        {
            'number': 10,
            'title': 'Daily Reporting',
            'description': 'Complete end-of-day activities',
            'tasks': [
                'Review all verified orders for the day',
                'Check outstanding pending verifications',
                'Generate daily activity report',
                'Note any issues or concerns',
                'Prepare task list for next day',
            ],
            'status': 'pending',
            'url': None,
        },
    ]

    important_notes = [
        {
            'title': 'Address Verification is Critical',
            'description': 'Accurate address verification ensures successful deliveries and prevents return trips.',
        },
        {
            'title': 'Original Orders are Preserved',
            'description': 'All order data is saved as proof. Delivery tasks are duplicates for tracking.',
        },
        {
            'title': 'Automated Workflow',
            'description': 'Once you verify an order, the system automatically creates delivery tasks and pushes to DMS.',
        },
        {
            'title': 'Audit Trail',
            'description': 'All your verification actions are logged for accountability and tracking.',
        },
    ]

    context = {
        'workflow_steps': workflow_steps,
        'important_notes': important_notes,
        'page_title': 'Workforce Workflow Guide',
    }

    return render(request, 'workforce/workflow_guide.html', context)


# Order Detail View ------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
def order_detail(request, order_id):
    """Display order detail page"""
    order = get_object_or_404(
        orders_models.Order.objects.select_related('business', 'pickup_location', 'verified_by'),
        id=order_id
    )

    # Get related data with select_related to avoid N+1 queries
    order_items = orders_models.OrderItem.objects.filter(order=order)
    order_comments = orders_models.OrderComments.objects.filter(order=order).select_related('user').order_by('-created_at')
    verification_logs = orders_models.OrderVerificationLog.objects.filter(order=order).select_related('verified_by').order_by('-created_at')

    # Get delivery task if exists with related driver
    delivery_task = delivery_models.DeliveryTask.objects.select_related(
        'driver', 'driver__user', 'pickup_location'
    ).filter(order=order).first()

    # Get status change timeline
    status_history = orders_models.OrderStatusHistory.objects.filter(
        order=order
    ).select_related('changed_by').order_by('created_at')

    context = {
        'order': order,
        'order_items': order_items,
        'order_comments': order_comments,
        'verification_logs': verification_logs,
        'delivery_task': delivery_task,
        'status_history': status_history,
        'timeline_count': status_history.count(),
    }

    # Check if this is being loaded in a panel (via HTMX)
    # Use panel template only when targeting the slide panel, not main content
    is_htmx = request.headers.get('HX-Request') == 'true'
    hx_target = request.headers.get('HX-Target', '')

    # If targeting main-content or using hx-select, use full page template
    # Panel template is only for the slide-out panel (orderDetailContent)
    use_panel = is_htmx and hx_target == 'orderDetailContent'
    template = 'workforce/order_detail_panel.html' if use_panel else 'workforce/order_detail.html'

    return render(request, template, context)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def cancel_order(request, order_id):
    """Cancel an order"""
    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business').prefetch_related('order_items', 'delivery_task'),
            id=order_id
        )

        # Cannot cancel already delivered or already cancelled orders
        if order.order_status == 'delivered':
            return JsonResponse({
                'success': False,
                'error': 'Cannot cancel a delivered order'
            }, status=400)

        if order.order_status == 'cancelled':
            return JsonResponse({
                'success': False,
                'error': 'Order is already cancelled'
            }, status=400)

        # Block cancellation if driver has already physically picked up the package
        delivery_task = delivery_models.DeliveryTask.objects.filter(order=order).first()
        active_pickup_statuses = ('picked_up', 'start_ride', 'in_transit', 'out_for_delivery', 'contacted')
        if delivery_task and delivery_task.dl_task_status in active_pickup_statuses:
            return JsonResponse({
                'success': False,
                'error': f'Cannot cancel — driver has already picked up this order (status: {delivery_task.get_dl_task_status_display()}). Contact the driver directly.'
            }, status=400)

        # Collect optional cancellation reason from request body
        cancellation_reason = None
        cancellation_notes = None
        try:
            body = json.loads(request.body)
            cancellation_reason = body.get('cancellation_reason')
            cancellation_notes = body.get('cancellation_notes')
        except (json.JSONDecodeError, AttributeError):
            pass

        # Update order status
        old_status = order.order_status
        order.order_status = 'cancelled'
        order.cancelled_by = request.user
        if cancellation_reason:
            order.cancellation_reason = cancellation_reason
        if cancellation_notes:
            order.cancellation_notes = cancellation_notes
        order.save()

        # Cancel related delivery task (if not already picked up — already guarded above)
        if delivery_task and delivery_task.dl_task_status != 'cancelled':
            delivery_task.dl_task_status = 'cancelled'
            delivery_task.dl_task_status_client = '9'
            delivery_task.save(update_fields=['dl_task_status', 'dl_task_status_client'])

        # Log the cancellation
        notes_text = f'Order cancelled by {request.user.username}'
        if cancellation_reason:
            notes_text += f' | Reason: {cancellation_reason}'
        if cancellation_notes:
            notes_text += f' | Notes: {cancellation_notes}'
        orders_models.OrderVerificationLog.objects.create(
            order=order,
            verified_by=request.user,
            action='order_cancelled',
            old_status=old_status,
            new_status='cancelled',
            notes=notes_text
        )

        # Return updated row HTML for HTMX
        if request.headers.get('HX-Request'):
            return render(request, 'orders/parts/order_row.html', {'order': order})

        return JsonResponse({
            'success': True,
            'message': 'Order cancelled successfully',
            'order_id': order.id
        })
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error cancelling order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while cancelling the order'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def assign_driver_to_order(request, order_id):
    """Assign a driver to an order (from AI suggest result)"""
    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business'),
            id=order_id
        )
        data = json.loads(request.body)
        driver_id = data.get('driver_id')

        if not driver_id:
            return JsonResponse({
                'success': False,
                'error': 'Driver ID is required'
            }, status=400)

        # Verify driver exists and is approved
        try:
            driver = fleet_models.Driver.objects.get(driver_id=driver_id, driver_status='approved')
        except fleet_models.Driver.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Driver {driver_id} not found or not approved'
            }, status=400)

        # Check if order already has a delivery task - use select_for_update to prevent race conditions
        from django.db import transaction

        with transaction.atomic():
            existing_task = delivery_models.DeliveryTask.objects.select_for_update().filter(order=order).first()

            if existing_task:
                # Update existing task
                old_driver = existing_task.driver
                existing_task.driver = driver
                existing_task.save()

                # Log the update
                orders_models.OrderVerificationLog.objects.create(
                    order=order,
                    verified_by=request.user,
                    action='driver_reassigned',
                    old_status=str(old_driver.driver_id) if old_driver else 'None',
                    new_status=str(driver_id),
                    notes=f'Driver reassigned from AI suggestion: {driver.user.get_full_name() if driver.user else driver.driver_code}'
                )
            else:
                # Create new delivery task
                task = delivery_models.DeliveryTask.objects.create(
                    order=order,
                    driver=driver,
                    business=order.business,
                    pickup_location=order.pickup_location,
                    dl_task_status='pending',
                    dl_task_number=f'DL-{order.order_number}'
                )

                # Log the assignment
                orders_models.OrderVerificationLog.objects.create(
                    order=order,
                    verified_by=request.user,
                    action='driver_assigned',
                    old_status='None',
                    new_status=str(driver_id),
                    notes=f'Driver assigned from AI suggestion: {driver.user.get_full_name() if driver.user else driver.driver_code}'
                )

        driver_name = driver.user.get_full_name() if driver.user else driver.driver_code

        return JsonResponse({
            'success': True,
            'message': f'Driver {driver_name} assigned successfully',
            'driver_id': driver_id,
            'driver_name': driver_name
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception("Error assigning driver to order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while assigning driver'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def delete_order(request, order_id):
    """Permanently delete an order and its related records. Superadmin only."""
    profile = get_cached_profile(request)
    if not profile or not profile.is_superadmin:
        return JsonResponse({'success': False, 'error': 'Superadmin access required'}, status=403)
    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business'),
            id=order_id
        )

        # Block deletion of delivered orders with COD collected
        delivery_task = delivery_models.DeliveryTask.objects.filter(order=order).first()
        if delivery_task and delivery_task.cod_collected:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete — COD has been collected for this order. Cancel it instead.'
            }, status=400)

        order_number = order.order_number or order.client_order_code or str(order.id)
        logger.info("Order %s (ID %s) deleted by %s", order_number, order.id, request.user.username)

        # Delete related records (order uses DO_NOTHING FKs, so manual cleanup)
        # Order matters: delivery tasks reference address updates, so tasks first
        orders_models.OrderVerificationLog.objects.filter(order=order).delete()
        orders_models.OrderStatusHistory.objects.filter(order=order).delete()
        delivery_models.DeliveryTask.objects.filter(order=order).delete()
        delivery_models.DlAddressUpdate.objects.filter(order=order).delete()
        order.order_items.all().delete()
        order.delete()

        if request.headers.get('HX-Request'):
            return HttpResponse('')

        return JsonResponse({
            'success': True,
            'message': f'Order {order_number} deleted'
        })
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error deleting order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while deleting the order'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def update_order_zone(request, order_id):
    """Update order delivery zone (from AI parse result)"""
    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business'),
            id=order_id
        )
        data = json.loads(request.body)
        zone_number = data.get('zone_number')
        street_number = data.get('street_number')
        building_number = data.get('building_number')
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        coords_accuracy = data.get('coords_accuracy')  # optional, sent by JS

        if not zone_number:
            return JsonResponse({
                'success': False,
                'error': 'Zone number is required'
            }, status=400)

        # Look up zone name (optional — save even if zone not in ZoneName table)
        zone = delivery_models.ZoneName.objects.filter(zone_number=zone_number, is_active=True).first()
        zone_display = zone.zone_name if zone else f'Zone {zone_number}'

        # Update order zone and address fields
        old_zone = order.dl_zone
        order.dl_zone = zone_number
        if street_number:
            order.dl_street = str(street_number)
        if building_number:
            order.dl_building = str(building_number)

        # Update order coordinates and accuracy if provided
        coords_saved = False
        if latitude and longitude:
            order.latitude = latitude
            order.longitude = longitude
            # Derive accuracy: use value sent by JS, or infer from building presence
            if coords_accuracy:
                order.coords_accuracy = coords_accuracy
            elif building_number:
                order.coords_accuracy = 'exact'
            else:
                order.coords_accuracy = 'street'
            coords_saved = True

        order.save()

        # Also update delivery task address (dl_to_address) for legacy compatibility
        if latitude and longitude:
            delivery_task = delivery_models.DeliveryTask.objects.filter(order=order).first()
            if delivery_task and delivery_task.dl_to_address:
                dl_address = delivery_task.dl_to_address
                dl_address.dl_latitude = latitude
                dl_address.dl_longitude = longitude
                dl_address.dl_zone = zone_number
                if street_number:
                    dl_address.dl_street = str(street_number)
                if building_number:
                    dl_address.dl_building = str(building_number)
                dl_address.save()

        # Log the update
        notes = f'Zone updated from AI parse: {zone_display}'
        if coords_saved:
            notes += f' | Coordinates saved: {latitude}, {longitude}'
        orders_models.OrderVerificationLog.objects.create(
            order=order,
            verified_by=request.user,
            action='zone_updated',
            old_status=str(old_zone) if old_zone else 'None',
            new_status=str(zone_number),
            notes=notes
        )

        return JsonResponse({
            'success': True,
            'message': f'Zone updated to {zone_number} ({zone_display})' + (' with coordinates' if coords_saved else ''),
            'zone_number': zone_number,
            'zone_name': zone_display,
            'coordinates_saved': coords_saved
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.exception("Error updating zone for order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating zone'
        }, status=400)


# AJAX Endpoints for Orders List ------------------------------------------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def publish_order_to_delivery(request, order_id):
    """AJAX endpoint to publish order to delivery"""
    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business').prefetch_related('order_items', 'delivery_task'),
            id=order_id
        )

        # Update order status to publish (triggers auto delivery task creation signal)
        order.order_status = 'publish'
        order.task_status = 'dl_task_listed'
        order.save()

        # Refresh to get updated delivery_task after signal runs
        order = orders_models.Order.objects.select_related('business').prefetch_related('order_items', 'delivery_task').get(id=order_id)

        # Log the publish action
        orders_models.OrderVerificationLog.objects.create(
            order=order,
            verified_by=request.user,
            action='order_published',
            old_status='to_publish',
            new_status='published',
            notes=f'Order published to delivery by {request.user.username}'
        )

        # Return updated row HTML for HTMX
        if request.headers.get('HX-Request'):
            return render(request, 'orders/parts/order_row.html', {'order': order})

        return JsonResponse({
            'success': True,
            'message': 'Order published to delivery successfully',
            'order_id': order.id,
            'order_status': order.order_status
        })
    except Exception as e:
        logger.exception("Error publishing order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while publishing order'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def update_order_status(request, order_id):
    """AJAX endpoint to update order status"""
    # Valid status values for validation
    VALID_ORDER_STATUSES = [
        'to_review', 'to_publish', 'published', 'processing', 'ready_to_pickup',
        'in_transit', 'delivered', 'failed', 'cancelled', 'returned', 'reported'
    ]
    VALID_TASK_STATUSES = [
        'pending', 'dl_task_listed', 'assigned', 'in_progress', 'completed',
        'failed', 'cancelled'
    ]

    try:
        order = get_object_or_404(
            orders_models.Order.objects.select_related('business'),
            id=order_id
        )

        # Parse JSON body
        data = json.loads(request.body)
        status = data.get('status')
        status_type = data.get('status_type', 'order')  # 'order' or 'task'

        if not status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)

        # Validate status_type
        if status_type not in ('order', 'task'):
            return JsonResponse({
                'success': False,
                'error': 'Invalid status_type. Must be "order" or "task"'
            }, status=400)

        # Validate status against allowed values
        if status_type == 'task' and status not in VALID_TASK_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid task status: {status}'
            }, status=400)
        elif status_type == 'order' and status not in VALID_ORDER_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid order status: {status}'
            }, status=400)

        old_status = order.order_status if status_type == 'order' else order.task_status

        # Update appropriate status field
        if status_type == 'task':
            order.task_status = status
        else:
            order.order_status = status
        order.save()

        # Log the status update
        from orders.models import OrderVerificationLog
        OrderVerificationLog.objects.create(
            order=order,
            verified_by=request.user,
            action=f'{status_type}_status_updated',
            notes=f'{status_type.title()} status changed from {old_status} to {status} by {request.user.username}',
        )

        return JsonResponse({
            'success': True,
            'message': f'Status updated to {status} successfully',
            'order_id': order.id,
            'new_status': status
        })
    except Exception as e:
        logger.exception("Error updating status for order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating order status'
        }, status=400)


@login_required(login_url='/accounts/login/')
@staff_required
@require_http_methods(["POST"])
def bulk_update_order_status(request):
    """Bulk update order status for multiple orders"""
    VALID_ORDER_STATUSES = [
        'to_review', 'to_publish', 'published', 'processing', 'ready_to_pickup',
        'in_transit', 'delivered', 'failed', 'cancelled', 'returned', 'reported'
    ]
    try:
        data = json.loads(request.body)
        order_ids = data.get('order_ids', [])
        status = data.get('status')

        if not order_ids:
            return JsonResponse({
                'success': False,
                'error': 'No orders selected'
            }, status=400)

        if not status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)

        if status not in VALID_ORDER_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status: {status}'
            }, status=400)

        orders = orders_models.Order.objects.filter(id__in=order_ids)
        updated = 0
        for order in orders:
            old_status = order.order_status
            if old_status != status:
                order.order_status = status
                order.save(update_fields=['order_status'])
                from orders.models import OrderVerificationLog
                OrderVerificationLog.objects.create(
                    order=order,
                    verified_by=request.user,
                    action='order_status_updated',
                    notes=f'Bulk status change from {old_status} to {status} by {request.user.username}',
                )
                updated += 1

        return JsonResponse({
            'success': True,
            'message': f'{updated} order(s) updated to {status}',
            'updated_count': updated
        })
    except Exception as e:
        logger.exception("Error in bulk order status update: %s", str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating orders'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def add_order_comment(request, order_id):
    """AJAX endpoint to add comment to order"""
    try:
        order = get_object_or_404(orders_models.Order, id=order_id)

        # Parse JSON body
        data = json.loads(request.body)
        comment_text = data.get('comment')

        if not comment_text:
            return JsonResponse({
                'success': False,
                'error': 'Comment text is required'
            }, status=400)

        # Create comment
        from orders.models import OrderComments
        comment = OrderComments.objects.create(
            order=order,
            name=request.user.username,
            body=comment_text
        )

        return JsonResponse({
            'success': True,
            'message': 'Comment added successfully',
            'order_id': order.id,
            'comment_id': comment.id,
            'comment_text': comment.body,
            'created_at': comment.created_at.strftime('%B %d, %Y %H:%M')
        })
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error adding comment to order %s: %s", order_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while adding comment'
        }, status=400)


@login_required(login_url='/accounts/login/')
@staff_required
def ajax_zone_name(request):
    """Return zone name JSON for a given zone number — used by order edit form."""
    zone_num = request.GET.get('zone', '').strip()
    if zone_num.isdigit():
        name = delivery_models.ZoneName.objects.filter(
            zone_number=int(zone_num)
        ).values_list('zone_name', flat=True).first() or ''
    else:
        name = ''
    return JsonResponse({'zone_name': name})


@login_required(login_url='account_login')
@require_POST
def update_order_coords(request, order_id):
    """AJAX endpoint to update order latitude/longitude and/or qnas_status from QNAS verification"""
    try:
        order = get_object_or_404(orders_models.Order, id=order_id)
        data = json.loads(request.body)
        latitude = data.get('latitude')
        longitude = data.get('longitude')
        qnas_status = data.get('qnas_status')

        update_fields = []

        if latitude and longitude:
            order.latitude = latitude
            order.longitude = longitude
            update_fields.extend(['latitude', 'longitude'])

        if qnas_status and qnas_status in dict(orders_models.Order.QNAS_STATUS):
            order.qnas_status = qnas_status
            update_fields.append('qnas_status')

        if not update_fields:
            return JsonResponse({'success': False, 'error': 'No valid fields to update'}, status=400)

        order.save(update_fields=update_fields)

        response_data = {'success': True, 'qnas_status': order.qnas_status}
        if order.latitude:
            response_data['latitude'] = float(order.latitude)
        if order.longitude:
            response_data['longitude'] = float(order.longitude)

        return JsonResponse(response_data)
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error updating coords for order %s: %s", order_id, str(e))
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Delivery Task Detail View ------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
def delivery_task_detail(request, task_id):
    """
    Display detailed information about a delivery task.
    Shows task status, driver updates, timeline, and allows quick actions.
    """
    task = get_object_or_404(
        delivery_models.DeliveryTask.objects.select_related(
            'order', 'order__business', 'driver', 'business', 'pickup_location'
        ),
        id=task_id
    )

    # Status timeline from OrderStatusHistory (exclude DMS — logged by ShipDay, not useful)
    status_history = orders_models.OrderStatusHistory.objects.filter(
        order=task.order
    ).select_related('changed_by').order_by('created_at')

    # Verification logs as fallback
    verification_logs = orders_models.OrderVerificationLog.objects.filter(
        order=task.order
    ).select_related('verified_by').order_by('-created_at')

    # Driver activity: delivery task status changes only (no DMS)
    driver_status_updates = orders_models.OrderStatusHistory.objects.filter(
        order=task.order,
        field_name__in=['dl_task_status']
    ).select_related('changed_by').order_by('created_at')

    # Driver uploads / documents
    try:
        from ezzy_api.models import TaskDocument
        driver_documents = TaskDocument.objects.filter(
            task=task
        ).select_related('uploaded_by').order_by('-created_at')
    except Exception:
        driver_documents = []

    # Seller / order comments
    seller_comments = orders_models.OrderComments.objects.filter(
        order=task.order
    ).order_by('-created_at')

    # Approved drivers for assignment modal
    approved_drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('driver_code')

    # GPS status points — build lookup by (old_status, new_status) for timeline matching
    status_points = delivery_models.TaskStatusPoint.objects.filter(
        task=task
    ).order_by('created_at')
    # Build lookup: key = new_status value (for dl_task_status entries)
    status_point_map = {}
    for sp in status_points:
        key = f"{sp.old_status}__{sp.new_status}"
        status_point_map[key] = sp

    # Pick list for this order
    from warehouse.models import PickList, PickListItem
    pick_list = None
    pick_list_items = []
    if task.order:
        pick_item = PickListItem.objects.filter(
            order=task.order
        ).select_related('pick_list').first()
        if pick_item:
            pick_list = pick_item.pick_list
            pick_list_items = pick_list.items.select_related(
                'product', 'location', 'order_item'
            ).order_by('location__code')

    context = {
        'page_title': f'Delivery Task #{task.dl_task_number}',
        'task': task,
        'status_history': status_history,
        'verification_logs': verification_logs,
        'driver_status_updates': driver_status_updates,
        'driver_documents': driver_documents,
        'seller_comments': seller_comments,
        'approved_drivers': approved_drivers,
        'status_points': status_points,
        'status_point_map': status_point_map,
        'pick_list': pick_list,
        'pick_list_items': pick_list_items,
    }

    return render(request, 'workforce/parts/delivery_task_detail.html', context)


# AJAX Endpoints for Delivery Tasks ------------------------------------------------------------------------------------------------------

@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def publish_task_to_fleets(request, task_id):
    """AJAX endpoint to publish delivery task to Fleet drivers"""
    try:
        task = get_object_or_404(
            delivery_models.DeliveryTask.objects.select_related('order'), id=task_id)

        # Block if order is cancelled
        if task.order and task.order.order_status == 'cancelled':
            return JsonResponse({
                'success': False,
                'error': 'Cannot publish — order is cancelled'
            }, status=400)

        # Mark task as published to fleet (pending = visible to all fleet drivers)
        task.dl_task_status = 'pending'
        task.dl_task_publish = True
        task.save()

        # Log to timeline
        orders_models.OrderStatusHistory.objects.create(
            order=task.order,
            field_name='dl_task_publish',
            old_value='False',
            new_value='True',
            old_display='Not Published',
            new_display='Published to Fleet',
            changed_by=request.user,
        )

        return JsonResponse({
            'success': True,
            'message': 'Task published to Fleets successfully',
            'task_id': task.id,
            'task_number': task.dl_task_number
        })
    except Exception as e:
        logger.exception("Error publishing task %s to fleets: %s", task_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while publishing task to fleets'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def unpublish_task_from_fleets(request, task_id):
    """AJAX endpoint to unpublish delivery task from Fleet drivers"""
    try:
        task = get_object_or_404(delivery_models.DeliveryTask, id=task_id)
        task.dl_task_publish = False
        task.save(update_fields=['dl_task_publish'])
        return JsonResponse({
            'success': True,
            'message': 'Task unpublished from Fleets',
            'task_id': task.id,
        })
    except Exception as e:
        logger.exception("Error unpublishing task %s: %s", task_id, str(e))
        return JsonResponse({'success': False, 'error': 'An error occurred'}, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def assign_driver_to_task(request, task_id):
    """AJAX endpoint to assign driver to delivery task"""
    from django.db import transaction
    try:
        # Parse JSON body first (before locking)
        data = json.loads(request.body)
        driver_id = data.get('driver_id')

        if not driver_id:
            return JsonResponse({
                'success': False,
                'error': 'Driver ID is required'
            }, status=400)

        # Get driver — must exist and be approved
        try:
            driver = fleet_models.Driver.objects.get(
                driver_id=driver_id, driver_status='approved')
        except fleet_models.Driver.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'Driver {driver_id} not found or not approved'
            }, status=400)

        with transaction.atomic():
            # Lock the task row to prevent race conditions
            task = get_object_or_404(
                delivery_models.DeliveryTask.objects.select_related('order').select_for_update(),
                id=task_id)

            # Block if order is cancelled
            if task.order and task.order.order_status == 'cancelled':
                return JsonResponse({
                    'success': False,
                    'error': 'Cannot assign driver — order is cancelled'
                }, status=400)

            # Block if task already assigned to a different driver
            if task.driver_id and task.driver_id != driver.driver_id:
                existing_name = task.driver.user.get_full_name() if task.driver and task.driver.user else str(task.driver_id)
                return JsonResponse({
                    'success': False,
                    'error': f'Task already assigned to {existing_name}. Unassign first.'
                }, status=400)

            task.driver = driver
            task.dl_task_status = 'assigned'
            task._status_actor = 'staff'
            task.save()

        driver_name = driver.user.get_full_name() if driver.user else driver.driver_code

        return JsonResponse({
            'success': True,
            'message': f'Driver {driver_name} assigned successfully',
            'task_id': task.id,
            'driver_id': driver.driver_id,
            'driver_name': driver_name
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid request data'
        }, status=400)
    except Exception as e:
        logger.exception("Error assigning driver to task %s: %s", task_id, str(e))
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required(login_url='/accounts/login/')
@staff_required
def api_drivers_list(request):
    """API endpoint to get active drivers for dropdowns"""
    drivers = fleet_models.Driver.objects.select_related('user').filter(
        driver_status='approved'
    ).order_by('user__first_name')
    driver_list = []
    for d in drivers:
        name = d.user.get_full_name() or d.user.username
        driver_list.append({'id': d.pk, 'name': name})  # Use d.pk instead of d.id (driver_id is PK)
    return JsonResponse({'drivers': driver_list})


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def update_task_status(request, task_id):
    """AJAX endpoint to update delivery task status"""
    VALID_STATUSES = [
        'for_review', 'pending', 'assigned', 'accepted', 'picked_up',
        'start_ride', 'out_for_delivery', 'in_transit', 'contacted',
        'non_reachable', 'delivered', 'failed', 'rejected', 'cancelled',
    ]
    try:
        task = get_object_or_404(
            delivery_models.DeliveryTask.objects.select_related('order'),
            id=task_id
        )

        # Lock check: Prevent status change on settled tasks
        if task.dl_task_status == 'delivered' and task.order and task.order.cod_status_by_staff == 'cod_settled_with_business':
            return JsonResponse({
                'success': False,
                'error': 'Task is locked. Status cannot be changed after delivery is successful and COD is settled.'
            }, status=403)

        # Parse JSON body
        data = json.loads(request.body)
        status = data.get('status')

        if not status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)

        # Handle publish_to_fleets as a special action (fallback if JS intercept missed it)
        if status == 'publish_to_fleets':
            task.dl_task_status = 'pending'
            task.dl_task_publish = True
            task.save(update_fields=['dl_task_status', 'dl_task_publish'])
            return JsonResponse({
                'success': True,
                'message': 'Task published to Fleet drivers',
                'task_id': task.id,
                'new_status': 'pending'
            })

        if status not in VALID_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status: {status}'
            }, status=400)

        # Map internal status to client-facing status
        STATUS_TO_CLIENT = {
            'for_review': 'for_review',
            'pending': 'for_review',
            'assigned': '0',
            'accepted': '0',
            'picked_up': '0',
            'start_ride': '0',
            'out_for_delivery': '0',
            'in_transit': '0',
            'contacted': '0',
            'non_reachable': '0',
            'delivered': '2',
            'failed': 'rejected',
            'rejected': 'rejected',
            'cancelled': '9',
        }

        # Optional fields
        driver_id = data.get('driver_id')
        notes = data.get('notes', '')
        time_str = data.get('time', '')

        # Update both task status fields
        task.dl_task_status = status
        update_fields = ['dl_task_status']

        client_status = STATUS_TO_CLIENT.get(status)
        if client_status:
            task.dl_task_status_client = client_status
            update_fields.append('dl_task_status_client')

        # Assign driver if provided
        if driver_id:
            try:
                driver = fleet_models.Driver.objects.get(id=driver_id)
                task.driver = driver
                update_fields.append('driver')
            except fleet_models.Driver.DoesNotExist:
                pass
        elif driver_id == '':
            # Explicitly unassign driver
            task.driver = None
            update_fields.append('driver')

        # Set completed_at timestamp for delivered status
        if status == 'delivered' and not task.completed_at:
            if time_str:
                from django.utils.dateparse import parse_datetime
                parsed = parse_datetime(time_str)
                if parsed:
                    if timezone.is_naive(parsed):
                        parsed = timezone.make_aware(parsed)
                    task.completed_at = parsed
                    update_fields.append('completed_at')
            else:
                task.completed_at = timezone.now()
                update_fields.append('completed_at')

        # Save notes to task description if provided
        if notes:
            task.dl_task_description = notes
            update_fields.append('dl_task_description')

        task.save(update_fields=update_fields)

        return JsonResponse({
            'success': True,
            'message': f'Task status updated to {status} successfully',
            'task_id': task.id,
            'new_status': status
        })
    except Exception as e:
        logger.exception("Error updating task %s status: %s", task_id, str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating task status'
        }, status=400)


# USER VERIFICATION VIEWS --------------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
@ensure_csrf_cookie
def business_verification_list(request):
    """Staff view to manage business verification — all businesses with verification status & approval buttons"""
    from django.db.models import Count

    # Get filter
    verification_filter = request.GET.get('status', 'all')
    search = request.GET.get('search', '').strip()

    # Base queryset — only businesses that have a profile (inner join via select_related)
    businesses = business_models.Business.objects.select_related(
        'profile__user', 'business_profile'
    ).filter(profile__isnull=False)

    # Count stats from same base queryset (not from Profile directly)
    status_counts = dict(
        businesses.values_list('profile__verification_status')
        .annotate(cnt=Count('business_id'))
        .values_list('profile__verification_status', 'cnt')
    )

    if verification_filter not in ('all', ''):
        businesses = businesses.filter(profile__verification_status=verification_filter)

    if search:
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search) |
            Q(profile__user__email__icontains=search)
        )

    businesses = businesses.order_by('-profile__verification_applied_at', '-business_id')

    # Paginate
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Business Verification',
        'page_obj': page_obj,
        'status_counts': status_counts,
        'current_filter': verification_filter,
        'search': search,
        'total_count': sum(status_counts.values()),
    }
    return render(request, 'workforce/business_verification_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
@ensure_csrf_cookie
def driver_verification_list(request):
    """Staff view to manage driver verification — all drivers with verification status & approval buttons"""
    from django.db.models import Count

    verification_filter = request.GET.get('status', 'all')
    search = request.GET.get('search', '').strip()

    # Base queryset — drivers that have a profile
    drivers = fleet_models.Driver.objects.select_related(
        'user', 'profile', 'profile__verified_by'
    ).filter(profile__isnull=False)

    # Count stats from same base queryset
    status_counts = dict(
        drivers.values_list('profile__verification_status')
        .annotate(cnt=Count('driver_id'))
        .values_list('profile__verification_status', 'cnt')
    )

    if verification_filter not in ('all', ''):
        drivers = drivers.filter(profile__verification_status=verification_filter)

    if search:
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(driver_phone__icontains=search) |
            Q(driver_code__icontains=search)
        )

    drivers = drivers.order_by('-profile__verification_applied_at', '-driver_id')

    page_obj = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'page_title': 'Driver Verification',
        'page_obj': page_obj,
        'status_counts': status_counts,
        'current_filter': verification_filter,
        'search': search,
        'total_count': sum(status_counts.values()),
    }
    return render(request, 'workforce/driver_verification_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
@ensure_csrf_cookie
def user_verification_list(request):
    """Staff view to see all users pending verification"""
    from core import models as core_models
    from business import models as business_models
    from django.db.models import Count

    # Get status counts for filter tabs
    status_counts = dict(
        core_models.Profile.objects.values_list('verification_status')
        .annotate(cnt=Count('id'))
        .values_list('verification_status', 'cnt')
    )
    status_counts['not_applied'] = core_models.Profile.objects.filter(
        verification_applied_at__isnull=True
    ).count()
    total_count = sum(v for k, v in status_counts.items() if k != 'not_applied')

    # Get all profiles based on filter
    verification_filter = request.GET.get('status', 'all')

    profiles = core_models.Profile.objects.select_related('user')
    if verification_filter == 'not_applied':
        profiles = profiles.filter(verification_applied_at__isnull=True)
    elif verification_filter in ('pending', 'under_review', 'verified', 'rejected', 'incomplete'):
        profiles = profiles.filter(verification_status=verification_filter)

    # Order by application date (most recent first)
    profiles = profiles.order_by('-verification_applied_at', '-created_at')

    # Prefetch related Business and Driver data to avoid N+1 queries
    profile_list = list(profiles)
    user_ids = [p.user_id for p in profile_list]

    # Bulk fetch businesses and drivers
    businesses_by_user = {
        b.user_id: b for b in business_models.Business.objects.filter(user_id__in=user_ids)
    }
    drivers_by_user = {
        d.user_id: d for d in fleet_models.Driver.objects.filter(user_id__in=user_ids)
    }

    # Bulk fetch team memberships
    from collections import defaultdict
    team_memberships_by_user = defaultdict(list)
    team_profiles = business_models.BusinessTeamProfile.objects.select_related('business').filter(
        user_id__in=user_ids
    )
    for tp in team_profiles:
        team_memberships_by_user[tp.user_id].append(tp)

    # Build verification data efficiently
    verification_data = []
    for profile in profile_list:
        team_list = team_memberships_by_user.get(profile.user_id, [])
        # Build JSON for pending team memberships (accepted by user, awaiting staff verification)
        pending_teams_json = json.dumps([
            {'id': tm.id, 'biz': tm.business.business_name, 'role': tm.get_team_role_display()}
            for tm in team_list
            if tm.team_status == 'pending' and tm.team_verifed
        ])
        data = {
            'profile': profile,
            'business': businesses_by_user.get(profile.user_id) if profile.is_business else None,
            'driver': drivers_by_user.get(profile.user_id) if profile.is_driver else None,
            'user': profile.user,
            'team_memberships': team_list,
            'pending_teams_json': pending_teams_json,
        }
        verification_data.append(data)

    context = {
        'verification_data': verification_data,
        'current_filter': verification_filter,
        'total_count': total_count,
        'status_counts': status_counts,
    }

    return render(request, 'workforce/user_verification_list.html', context)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def update_verification_status(request, profile_id):
    """AJAX endpoint to update user verification status"""
    from core import models as core_models
    from django.utils import timezone

    try:
        profile = get_object_or_404(core_models.Profile, id=profile_id)

        # Parse JSON body
        data = json.loads(request.body)
        new_status = data.get('status')
        rejection_reason = data.get('rejection_reason', '')

        if not new_status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)

        # Block approval if user has never applied
        if new_status == 'verified' and not profile.verification_applied_at:
            return JsonResponse({
                'success': False,
                'error': 'Cannot verify — user has not yet applied for verification.'
            }, status=400)

        # Update verification status
        profile.verification_status = new_status
        profile.verified_by = request.user

        if new_status == 'verified':
            profile.verified_at = timezone.now()
            profile.rejection_reason = None
            profile.is_profile_completed = True

            # Update business or driver status to active
            if profile.is_business:
                profile.is_business_profile_completed = True
                try:
                    from business import models as business_models
                    business = business_models.Business.objects.get(user=profile.user)
                    business.business_status = 'active'
                    business.save()
                except business_models.Business.DoesNotExist:
                    pass

            if profile.is_driver:
                profile.is_driver_profile_completed = True
                try:
                    driver = fleet_models.Driver.objects.get(user=profile.user)
                    driver.driver_status = 'approved'
                    driver.save()
                except fleet_models.Driver.DoesNotExist:
                    pass

            # Activate team memberships if requested
            activate_teams = data.get('activate_teams', [])
            if activate_teams:
                from business import models as business_models
                business_models.BusinessTeamProfile.objects.filter(
                    id__in=activate_teams,
                    user=profile.user,
                    team_status='pending',
                    team_verifed=True
                ).update(team_status='active')

        elif new_status == 'rejected':
            profile.rejection_reason = rejection_reason
            profile.verified_at = None

        elif new_status == 'under_review':
            profile.verified_at = None

        profile.save()

        return JsonResponse({
            'success': True,
            'message': f'Verification status updated to {new_status}',
            'profile_id': profile.id,
            'new_status': new_status
        })
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error updating verification status for profile %s: %s", profile_id, str(e))
        return JsonResponse({
            'success': False,
            'error': f'Error: {type(e).__name__}: {str(e)}'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def update_team_status(request, team_id):
    """AJAX endpoint to update team membership status"""
    from business import models as business_models
    try:
        team_profile = get_object_or_404(business_models.BusinessTeamProfile, id=team_id)
        data = json.loads(request.body)
        new_status = data.get('status')

        valid_statuses = ['active', 'pending', 'inactive', 'suspended', 'rejected']
        if new_status not in valid_statuses:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
            }, status=400)

        old_status = team_profile.team_status
        team_profile.team_status = new_status

        # When activating, also mark as verified
        if new_status == 'active':
            team_profile.team_verifed = True

        team_profile.save()

        logger.info(
            "Team membership %s status changed from '%s' to '%s' by staff user %s",
            team_id, old_status, new_status, request.user.id
        )

        return JsonResponse({
            'success': True,
            'message': f'Team status updated to {new_status}',
            'team_id': team_id,
            'new_status': new_status
        })
    except Http404:
        raise
    except Exception as e:
        logger.exception("Error updating team status for team %s: %s", team_id, str(e))
        return JsonResponse({
            'success': False,
            'error': f'Error: {type(e).__name__}: {str(e)}'
        }, status=400)


# ADDITIONAL SIDEBAR FUNCTIONS --------------------------------------------------------------------------------------------------------------


@login_required(login_url='/accounts/login/')
@staff_required
def orders_reported(request):
    """View for reported orders list"""
    from orders import models as orders_models

    orders_list = orders_models.Order.objects.select_related(
        'business'
    ).prefetch_related('order_comments', 'delivery_task').filter(
        order_status='reported'
    ).order_by('-created_at')

    orders_with_pagination = paginate_queryset(request, orders_list, items_per_page=20)

    context = {
        'orders': orders_with_pagination,
        'page_title': 'Reported Orders',
    }
    return render(request, 'workforce/wf_orders_reported.html', context)


# Tasks Section Functions
@login_required(login_url='/accounts/login/')
@staff_required
def tasks_followup_list(request):
    """View for follow-up tasks list — latest driver updates sorted by time"""
    # All task statuses including delivered for follow-up tracking
    active_statuses = [
        'pending', 'assigned', 'accepted', 'picked_up',
        'start_ride', 'out_for_delivery', 'in_transit',
        'contacted', 'non_reachable', 'failed', 'rejected',
        'delivered',
    ]

    tasks_list = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_items',
    ).filter(
        dl_task_status__in=active_statuses
    )

    # Optional filters
    status_filter = request.GET.get('status', '')
    driver_filter = request.GET.get('driver', '')

    if status_filter:
        tasks_list = tasks_list.filter(dl_task_status=status_filter)
    if driver_filter:
        tasks_list = tasks_list.filter(driver_id=driver_filter)

    # Sorting
    sort_param = request.GET.get('sort', 'updated')
    SORT_MAP = {
        'task':         ('dl_task_number', 'asc'),
        'task-desc':    ('-dl_task_number', 'desc'),
        'updated':      ('-updated_at', 'desc'),
        'updated-asc':  ('updated_at', 'asc'),
        'driver':       ('driver__user__first_name', 'asc'),
        'driver-desc':  ('-driver__user__first_name', 'desc'),
        'status':       ('dl_task_status', 'asc'),
        'status-desc':  ('-dl_task_status', 'desc'),
        'cod':          ('-order__cod_amount', 'desc'),
        'cod-asc':      ('order__cod_amount', 'asc'),
    }
    order_field, sort_dir = SORT_MAP.get(sort_param, ('-updated_at', 'desc'))
    tasks_list = tasks_list.order_by(order_field)

    # Extract sort name without direction suffix for template
    current_sort = sort_param.replace('-desc', '').replace('-asc', '')

    # Get active drivers for filter dropdown
    active_drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('user__first_name')

    tasks_with_pagination = paginate_queryset(request, tasks_list, items_per_page=30)

    # Build filter params for pagination
    filter_params = ''
    if status_filter:
        filter_params += f'&status={status_filter}'
    if driver_filter:
        filter_params += f'&driver={driver_filter}'
    if sort_param and sort_param != 'updated':
        filter_params += f'&sort={sort_param}'

    context = {
        'dl_tasks': tasks_with_pagination,
        'page_title': 'Follow-Up Tasks',
        'page_subtitle': 'Latest updates by driver & status',
        'page_icon': 'fa-flag',
        'list_type': 'followup',
        'active_drivers': active_drivers,
        'active_statuses': active_statuses,
        'current_status': status_filter,
        'current_driver': driver_filter,
        'current_sort': current_sort,
        'current_dir': sort_dir,
        'filter_params': filter_params,
        'show_followup_filters': True,
    }
    return render(request, 'workforce/parts/lists/dl_list_followup.html', context)



@login_required(login_url='/accounts/login/')
@staff_required
def tasks_reported(request):
    """View for reported tasks list - showing rejected/cancelled tasks"""
    tasks_list = delivery_models.DeliveryTask.objects.select_related(
        'order', 'driver', 'business', 'pickup_location', 'order__business'
    ).prefetch_related(
        'order__order_items',
    ).filter(
        dl_task_status__in=['rejected', 'cancelled']
    ).order_by('-created_at')

    tasks_with_pagination = paginate_queryset(request, tasks_list, items_per_page=20)

    context = {
        'dl_tasks': tasks_with_pagination,
        'page_title': 'Reported Tasks',
        'page_subtitle': 'Rejected and cancelled tasks',
        'page_icon': 'fa-triangle-exclamation',
        'list_type': 'reported',
    }
    return render(request, 'workforce/parts/lists/dl_list_all.html', context)



# Finance Dashboard Section Functions
@login_required(login_url='/accounts/login/')
@staff_required
def workforce_finance_dashboard(request):
    """Comprehensive finance dashboard for workforce staff"""
    from django.db.models import Sum, Count, Q, F
    from decimal import Decimal
    from datetime import timedelta

    try:
        days = int(request.GET.get('days', 30))
        if days < 1 or days > 365:
            days = 30
    except (ValueError, TypeError):
        days = 30
    start_date = timezone.now() - timedelta(days=days)

    # All drivers summary
    drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user')

    driver_totals = drivers.aggregate(
        total_wallet=Sum('wallet_balance'),
        total_cod_in_hand=Sum('cod_in_hand'),
        total_pending_earnings=Sum('pending_earnings'),
        total_credit_limit=Sum('credit_limit'),
    )

    # Transaction summary for the period
    txns = fleet_models.DriverTransaction.objects.filter(created_at__gte=start_date)

    # COD pipeline
    cod_collected = abs(txns.filter(
        transaction_type='cod_collection'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    cod_driver_settled = abs(txns.filter(
        transaction_type__in=['cod_deposit', 'cod_driver_settle']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    cod_client_settled = abs(txns.filter(
        transaction_type='cod_client_settle'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    cod_returned = abs(txns.filter(
        transaction_type='cod_return'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    # Additional COD metrics
    # COD in Bank (deposited by drivers)
    cod_in_bank = abs(txns.filter(
        transaction_type='cod_deposit'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    # COD to Settle = Total collected - Driver settled - Client settled
    cod_to_settle = cod_collected - cod_driver_settled - cod_client_settled
    if cod_to_settle < 0:
        cod_to_settle = Decimal('0')

    # Earnings & settlements
    earnings_total = txns.filter(
        transaction_type='earning'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    settlements_total = abs(txns.filter(
        transaction_type='settlement'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    # Charges summary
    charges_summary = txns.filter(
        transaction_type__in=['delivery_charge', 'fulfillment_charge', 'inventory_handling', 'other_charge']
    ).values('transaction_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    )

    total_charges = sum(abs(c['total']) for c in charges_summary) if charges_summary else Decimal('0')

    # Bills payable / receivable
    bills_payable = abs(txns.filter(
        transaction_type='bills_payable'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    bills_receivable = abs(txns.filter(
        transaction_type='bills_receivable'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0'))

    # Drivers with highest COD in hand (top 10)
    top_cod_drivers = drivers.filter(
        cod_in_hand__gt=0
    ).order_by('-cod_in_hand')[:10]

    # Recent transactions (last 20)
    recent_transactions = txns.select_related(
        'driver__user', 'delivery_task', 'business'
    ).order_by('-created_at')[:20]

    # Transaction counts by type
    type_breakdown = txns.values('transaction_type').annotate(
        count=Count('id'),
        total=Sum('amount')
    ).order_by('-count')

    # Performance metrics - Delivery statistics
    from orders import models as orders_models

    # Total completed deliveries in period
    total_deliveries = orders_models.Order.objects.filter(
        order_status__in=['delivered', 'fulfilled'],
        delivered_at__gte=start_date
    ).count()

    # Failed deliveries in period
    failed_deliveries = orders_models.Order.objects.filter(
        order_status__in=['failed', 'cancelled'],
        updated_at__gte=start_date
    ).count()

    # Total delivery fees collected
    total_delivery_fees = orders_models.Order.objects.filter(
        delivered_at__gte=start_date,
        order_status__in=['delivered', 'fulfilled']
    ).aggregate(total=Sum('dl_amount'))['total'] or Decimal('0')

    # Average delivery value (COD + DL amount)
    avg_delivery_stats = orders_models.Order.objects.filter(
        delivered_at__gte=start_date,
        order_status__in=['delivered', 'fulfilled']
    ).aggregate(
        total_cod=Sum('cod_amount'),
        total_dl=Sum('dl_amount'),
        count=Count('id')
    )

    avg_delivery_value = Decimal('0')
    if avg_delivery_stats['count'] and avg_delivery_stats['count'] > 0:
        total_value = (avg_delivery_stats['total_cod'] or Decimal('0')) + (avg_delivery_stats['total_dl'] or Decimal('0'))
        avg_delivery_value = total_value / avg_delivery_stats['count']

    # Active drivers (drivers with at least one delivery in period)
    active_drivers = orders_models.Order.objects.filter(
        delivered_at__gte=start_date,
        order_status__in=['delivered', 'fulfilled']
    ).values('delivery_task__driver').distinct().count()

    # Success rate calculation
    total_attempts = total_deliveries + failed_deliveries
    success_rate = Decimal('0')
    if total_attempts > 0:
        success_rate = (Decimal(total_deliveries) / Decimal(total_attempts)) * 100

    # COD with Fleet metrics
    active_drivers_count = drivers.count()
    drivers_with_cod_count = drivers.filter(cod_in_hand__gt=0).count()

    context = {
        'selected_days': days,
        'driver_totals': driver_totals,
        'driver_count': drivers.count(),
        'active_drivers_count': active_drivers_count,
        'drivers_with_cod_count': drivers_with_cod_count,
        'cod_collected': cod_collected,
        'cod_driver_settled': cod_driver_settled,
        'cod_client_settled': cod_client_settled,
        'cod_returned': cod_returned,
        'cod_in_bank': cod_in_bank,
        'cod_to_settle': cod_to_settle,
        'earnings_total': earnings_total,
        'settlements_total': settlements_total,
        'charges_summary': charges_summary,
        'total_charges': total_charges,
        'bills_payable': bills_payable,
        'bills_receivable': bills_receivable,
        'top_cod_drivers': top_cod_drivers,
        'recent_transactions': recent_transactions,
        'type_breakdown': type_breakdown,
        # Performance metrics
        'total_deliveries': total_deliveries,
        'failed_deliveries': failed_deliveries,
        'total_delivery_fees': total_delivery_fees,
        'avg_delivery_value': avg_delivery_value,
        'active_drivers': active_drivers,
        'success_rate': success_rate,
    }

    return render(request, 'workforce/workforce_finance_dashboard.html', context)


# Fleet Accounts Section Functions
@login_required(login_url='/accounts/login/')
@staff_required
def fleet_cod_in_hand(request):
    """View for COD in hand with drivers"""
    from django.db.models import Sum, Value, DecimalField
    from django.db.models.functions import Coalesce
    from datetime import timedelta
    from delivery import models as delivery_models

    # Date preset filter
    date_preset = request.GET.get('date_preset', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    today = timezone.now().date()

    # Calculate dates based on preset
    if date_preset == 'today':
        date_from = today.isoformat()
        date_to = today.isoformat()
    elif date_preset == 'yesterday':
        yesterday = today - timedelta(days=1)
        date_from = yesterday.isoformat()
        date_to = yesterday.isoformat()
    elif date_preset == '3days':
        date_from = (today - timedelta(days=3)).isoformat()
        date_to = today.isoformat()
    elif date_preset == '1week':
        date_from = (today - timedelta(days=7)).isoformat()
        date_to = today.isoformat()
    elif date_preset == '1month':
        date_from = (today - timedelta(days=30)).isoformat()
        date_to = today.isoformat()
    # For 'custom', use the date_from and date_to from request

    # If date filter is applied, calculate COD collected in that period per driver
    if date_from or date_to:
        # Get drivers with COD collected in the date range
        from django.db.models import OuterRef, Subquery

        # Build subquery with date filter - use cod_collected_at for filtering
        cod_subquery = delivery_models.DeliveryTask.objects.filter(
            driver_id=OuterRef('driver_id'),
            cod_collected=True,
        )

        # Apply date filters on cod_collected_at date
        if date_from:
            cod_subquery = cod_subquery.filter(cod_collected_at__date__gte=date_from)
        if date_to:
            cod_subquery = cod_subquery.filter(cod_collected_at__date__lte=date_to)

        cod_subquery = cod_subquery.values('driver').annotate(
            total=Sum('cod_collected_amount')
        ).values('total')

        # Show all drivers (not just approved) for COD tracking
        drivers = fleet_models.Driver.objects.all().select_related('user').annotate(
            period_cod=Coalesce(
                Subquery(cod_subquery),
                Value(0),
                output_field=DecimalField()
            )
        )
    else:
        # No date filter - calculate actual COD from DeliveryTask (collected but not settled)
        from django.db.models import OuterRef, Subquery

        cod_subquery = delivery_models.DeliveryTask.objects.filter(
            driver_id=OuterRef('driver_id'),
            cod_collected=True,
            cod_settled=False,
        ).values('driver').annotate(
            total=Sum('cod_collected_amount')
        ).values('total')

        # Show all drivers (not just approved) for COD tracking
        drivers = fleet_models.Driver.objects.all().select_related('user').annotate(
            period_cod=Coalesce(
                Subquery(cod_subquery),
                Value(0),
                output_field=DecimalField()
            )
        )

    # COD filter (yes/no/custom) - filters on period_cod (which is either cod_in_hand or calculated from date range)
    cod_filter = request.GET.get('cod_filter', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')

    if cod_filter == 'yes':
        # Filter drivers who have COD (using period_cod annotation)
        drivers = drivers.filter(period_cod__gt=0)
    elif cod_filter == 'no':
        # Filter drivers with no COD
        from django.db.models import Q
        drivers = drivers.filter(Q(period_cod=0) | Q(period_cod__isnull=True))
    elif cod_filter == 'custom':
        # Filter by custom amount range on period_cod
        if min_amount:
            drivers = drivers.filter(period_cod__gte=min_amount)
        if max_amount:
            drivers = drivers.filter(period_cod__lte=max_amount)

    # Sorting
    sort_by = request.GET.get('sort', 'name_asc')
    sort_options = {
        'name_asc': 'user__first_name',
        'name_desc': '-user__first_name',
        'amount_asc': 'period_cod',
        'amount_desc': '-period_cod',
        'date_asc': 'last_settlement_date',
        'date_desc': '-last_settlement_date',
    }
    drivers = drivers.order_by(sort_options.get(sort_by, 'user__first_name'))

    # Calculate totals and statistics
    from django.db.models import Avg, Max, Count

    total_cod = drivers.aggregate(total=Sum('period_cod'))['total'] or 0
    pending_settlements = drivers.filter(period_cod__gt=0).count()

    # Additional statistics
    avg_cod = drivers.filter(period_cod__gt=0).aggregate(avg=Avg('period_cod'))['avg'] or 0
    max_cod = drivers.aggregate(max=Max('period_cod'))['max'] or 0

    # Count drivers by COD range
    cod_ranges = {
        'under_500': drivers.filter(period_cod__gt=0, period_cod__lt=500).count(),
        'between_500_2000': drivers.filter(period_cod__gte=500, period_cod__lt=2000).count(),
        'above_2000': drivers.filter(period_cod__gte=2000).count(),
    }

    # Total drivers in system (not just those with COD)
    total_drivers = fleet_models.Driver.objects.filter(driver_status='approved').count()

    # View mode (grid or list) - default to list
    view_mode = request.GET.get('view', 'list')

    # Build filter params for pagination (exclude page and per_page as they're handled separately)
    filter_params = request.GET.copy()
    if 'page' in filter_params:
        del filter_params['page']
    if 'per_page' in filter_params:
        del filter_params['per_page']

    drivers_with_pagination = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'drivers': drivers_with_pagination,
        'page_title': 'COD In Hand',
        'total_cod': total_cod,
        'pending_settlements': pending_settlements,
        'avg_cod': avg_cod,
        'max_cod': max_cod,
        'cod_ranges': cod_ranges,
        'total_drivers': total_drivers,
        'filter_params': filter_params.urlencode(),
        'per_page': get_per_page(request, default=20),
        'date_preset': date_preset,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'cod_filter': cod_filter,
        'min_amount': min_amount or '',
        'max_amount': max_amount or '',
        'sort_by': sort_by,
        'view_mode': view_mode,
        'is_filtered_by_date': bool(date_from or date_to),
    }
    return render(request, 'workforce/fleet_cod_in_hand.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def fleet_drivers_earnings(request):
    """View for drivers earnings"""
    from django.db.models import Sum, Count, Max, Value, DecimalField, IntegerField, OuterRef, Subquery
    from django.db.models.functions import Coalesce
    from delivery import models as delivery_models

    # Subquery: count of completed deliveries per driver (after last settlement)
    deliveries_sub = delivery_models.DeliveryTask.objects.filter(
        driver_id=OuterRef('driver_id'),
        dl_task_status='delivered',
    ).values('driver').annotate(cnt=Count('id')).values('cnt')

    # Subquery: pending earnings (COD collected, not settled)
    earnings_sub = delivery_models.DeliveryTask.objects.filter(
        driver_id=OuterRef('driver_id'),
        cod_collected=True,
        cod_settled=False,
    ).values('driver').annotate(total=Sum('cod_collected_amount')).values('total')

    # Subquery: last settlement date
    last_settled_sub = fleet_models.DriverTransaction.objects.filter(
        driver_id=OuterRef('pk'),
        transaction_type='settlement',
    ).order_by('-created_at').values('created_at')[:1]

    drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').annotate(
        deliveries_after_settlement=Coalesce(
            Subquery(deliveries_sub), Value(0), output_field=IntegerField()
        ),
        real_pending_earnings=Coalesce(
            Subquery(earnings_sub), Value(0), output_field=DecimalField()
        ),
        last_settled_date=Subquery(last_settled_sub),
    ).order_by('user__first_name')

    # Totals for summary stats
    agg = drivers.aggregate(
        total_earnings=Sum('real_pending_earnings'),
        total_deliveries=Sum('deliveries_after_settlement'),
    )

    drivers_with_pagination = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'drivers': drivers_with_pagination,
        'page_title': 'Drivers Earnings',
        'total_earnings': agg['total_earnings'] or 0,
        'total_deliveries': agg['total_deliveries'] or 0,
    }
    return render(request, 'workforce/fleet_drivers_earnings.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def earnings_verification(request):
    """
    Staff view to verify and publish driver earnings.
    Shows completed deliveries with pending earnings verification.
    Allows bulk verify/publish with editable earnings.
    """
    from django.db.models import Sum, Case, When, DecimalField, F
    from delivery import models as delivery_models
    from datetime import timedelta

    # Filters
    driver_id = request.GET.get('driver', '')
    status_filter = request.GET.get('status', 'pending')
    try:
        days = int(request.GET.get('days', 7))
        if days < 1 or days > 365:
            days = 7
    except (ValueError, TypeError):
        days = 7

    start_date = timezone.now() - timedelta(days=days)

    # Base queryset - completed deliveries with COD collected
    tasks = delivery_models.DeliveryTask.objects.filter(
        dl_task_status='delivered',
        dl_task_date__gte=start_date.date(),
        cod_collected=True  # Only show deliveries where COD has been collected
    ).select_related(
        'driver', 'driver__user', 'order', 'order__business',
        'pickup_location', 'dl_to_address', 'earnings_verified_by'
    ).order_by('-completed_at', '-id')

    # Apply filters
    if driver_id:
        tasks = tasks.filter(driver_id=driver_id)

    if status_filter and status_filter != 'all':
        tasks = tasks.filter(earnings_verification_status=status_filter)

    # Calculate stats
    stats = {
        'pending_count': tasks.filter(earnings_verification_status='pending').count(),
        'verified_count': tasks.filter(earnings_verification_status='verified').count(),
        'published_count': tasks.filter(earnings_verification_status='published').count(),
        'total_pending_earnings': tasks.filter(
            earnings_verification_status='pending'
        ).aggregate(
            total=Sum(Case(
                When(calculated_earnings__isnull=False, then='calculated_earnings'),
                When(driver_earnings__isnull=False, then='driver_earnings'),
                default='dl_price',
                output_field=DecimalField(max_digits=10, decimal_places=2)
            ))
        )['total'] or 0,
    }

    # Get approved drivers for filter dropdown
    drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('user__first_name')

    # Paginate
    tasks_paginated = paginate_queryset(request, tasks, items_per_page=50)

    context = {
        'tasks': tasks_paginated,
        'drivers': drivers,
        'stats': stats,
        'selected_driver': driver_id,
        'selected_status': status_filter,
        'selected_days': days,
        'page_title': 'Earnings Verification',
    }
    return render(request, 'workforce/earnings_verification.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def earnings_verification_action(request):
    """
    Process bulk earnings verification actions.
    Actions: verify, publish, reject, update_amount
    """
    from django.http import JsonResponse
    from delivery import models as delivery_models
    from decimal import Decimal, InvalidOperation

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    action = request.POST.get('action', '')
    task_ids = request.POST.getlist('task_ids[]')
    earnings_updates = request.POST.get('earnings_updates', '{}')

    if not task_ids:
        return JsonResponse({'error': 'No tasks selected'}, status=400)

    try:
        import json
        earnings_dict = json.loads(earnings_updates) if earnings_updates else {}
    except json.JSONDecodeError:
        earnings_dict = {}

    updated_count = 0
    errors = []

    # Fetch all tasks at once to avoid N+1 queries
    tasks_queryset = delivery_models.DeliveryTask.objects.select_related('driver').filter(id__in=task_ids)
    tasks_dict = {str(task.id): task for task in tasks_queryset}

    for task_id in task_ids:
        try:
            task = tasks_dict.get(str(task_id))
            if not task:
                errors.append(f"Task {task_id} not found")
                continue

            # Update earnings amount if provided
            if str(task_id) in earnings_dict:
                try:
                    new_amount = Decimal(str(earnings_dict[str(task_id)]))
                    task.verified_earnings = new_amount
                except (ValueError, InvalidOperation):
                    errors.append(f"Invalid amount for task {task_id}")
                    continue

            if action == 'verify':
                task.earnings_verification_status = 'verified'
                task.earnings_verified_by = request.user
                task.earnings_verified_at = timezone.now()
                # Set verified_earnings to calculated amount if not already set
                if not task.verified_earnings:
                    task.verified_earnings = task.calculate_driver_earnings()
                task.save()
                updated_count += 1

            elif action == 'publish':
                # Can only publish verified tasks
                if task.earnings_verification_status not in ['verified', 'pending']:
                    errors.append(f"Task {task_id} cannot be published")
                    continue

                task.earnings_verification_status = 'published'
                task.earnings_published_at = timezone.now()
                if not task.earnings_verified_by:
                    task.earnings_verified_by = request.user
                    task.earnings_verified_at = timezone.now()

                # Update driver_earnings with verified amount
                final_earnings = task.verified_earnings or task.calculate_driver_earnings()
                task.driver_earnings = final_earnings
                task.save()

                # Update driver's pending_earnings atomically using F() and Coalesce to prevent race conditions
                if task.driver:
                    from django.db.models import F
                    from django.db.models.functions import Coalesce
                    fleet_models.Driver.objects.filter(driver_id=task.driver.driver_id).update(
                        pending_earnings=Coalesce(F('pending_earnings'), Decimal('0')) + Decimal(str(final_earnings))
                    )

                updated_count += 1

            elif action == 'reject':
                task.earnings_verification_status = 'rejected'
                task.earnings_verified_by = request.user
                task.earnings_verified_at = timezone.now()
                task.save()
                updated_count += 1

            elif action == 'update':
                # Just update the earnings amount
                task.save()
                updated_count += 1

        except Exception as e:
            logger.exception("Error processing task %s: %s", task_id, str(e))
            errors.append(f"Error processing task {task_id}")

    return JsonResponse({
        'success': True,
        'updated': updated_count,
        'errors': errors,
        'message': f'{updated_count} task(s) updated successfully'
    })


@login_required(login_url='/accounts/login/')
@staff_required
def cod_settlement_report(request):
    """COD Settlement Report - Select drivers, settle COD, export PDF"""
    from django.db.models import Sum, F
    from delivery import models as delivery_models
    from datetime import timedelta

    # Date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    today = timezone.now().date()
    if not date_from:
        date_from = (today - timedelta(days=7)).isoformat()
    if not date_to:
        date_to = today.isoformat()

    # Get drivers with unsettled COD (all drivers, not just approved)
    drivers = fleet_models.Driver.objects.filter(
        cod_in_hand__gt=0
    ).select_related('user').order_by('-cod_in_hand')

    # Get unsettled COD deliveries grouped by driver
    unsettled_tasks = delivery_models.DeliveryTask.objects.filter(
        cod_collected=True,
        dl_task_status='delivered',
        order__cod_status_by_staff='cod_with_driver'
    ).select_related('driver', 'driver__user', 'order').order_by('-completed_at')

    # Apply date filter if provided
    if date_from:
        unsettled_tasks = unsettled_tasks.filter(completed_at__date__gte=date_from)
    if date_to:
        unsettled_tasks = unsettled_tasks.filter(completed_at__date__lte=date_to)

    # Calculate totals
    total_unsettled = drivers.aggregate(total=Sum('cod_in_hand'))['total'] or 0
    drivers_with_cod = drivers.count()

    context = {
        'drivers': drivers,
        'unsettled_tasks': unsettled_tasks[:100],
        'page_title': 'COD Settlement Report',
        'total_unsettled': total_unsettled,
        'drivers_with_cod': drivers_with_cod,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'workforce/cod_settlement_report.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def cod_settlement_action(request):
    """Process COD settlement for selected drivers"""
    from django.http import JsonResponse
    from django.db import transaction
    from fleet.wallet_service import WalletService

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    driver_ids = request.POST.getlist('driver_ids[]')
    payment_method = request.POST.get('payment_method', 'cash')
    reference = request.POST.get('reference', '')

    if not driver_ids:
        return JsonResponse({'error': 'No drivers selected'}, status=400)

    # Limit number of drivers to prevent DoS
    if len(driver_ids) > 50:
        return JsonResponse({'error': 'Maximum 50 drivers can be settled at once'}, status=400)

    wallet_service = WalletService()
    settled_drivers = []
    total_settled = 0

    for driver_id in driver_ids:
        try:
            # Use transaction.atomic and select_for_update to prevent race conditions
            with transaction.atomic():
                driver = fleet_models.Driver.objects.select_for_update().get(driver_id=driver_id)
                cod_amount = driver.cod_in_hand

                if cod_amount > 0:
                    # Record COD deposit transaction with payment method
                    # Pass no delivery_ids so submit_cod_to_admin settles oldest tasks automatically
                    wallet_service.submit_cod_to_admin(
                        driver=driver,
                        amount=cod_amount,
                        created_by=request.user,
                        reference_number=reference,
                        payment_method=payment_method,
                        notes=f"COD settlement via {payment_method}"
                    )
                    settled_drivers.append({
                        'name': driver.driver_name,
                        'amount': float(cod_amount),
                        'payment_method': payment_method
                    })
                    total_settled += cod_amount
        except fleet_models.Driver.DoesNotExist:
            continue

    return JsonResponse({
        'success': True,
        'settled_count': len(settled_drivers),
        'total_settled': float(total_settled),
        'settled_drivers': settled_drivers,
        'payment_method': payment_method
    })


@login_required(login_url='/accounts/login/')
@staff_required
def cod_settlement_pdf(request):
    """Generate PDF report for COD settlement"""
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from io import BytesIO
    from django.db.models import Sum

    driver_ids = request.GET.getlist('driver_ids')

    # Get selected drivers or all with COD
    if driver_ids:
        drivers = fleet_models.Driver.objects.filter(
            driver_id__in=driver_ids,
            cod_in_hand__gt=0
        ).select_related('user').order_by('-cod_in_hand')
    else:
        drivers = fleet_models.Driver.objects.filter(
            driver_status='approved',
            cod_in_hand__gt=0
        ).select_related('user').order_by('-cod_in_hand')

    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    elements.append(Paragraph('COD Settlement Report', title_style))
    elements.append(Paragraph(f'Date: {timezone.now().strftime("%d %b %Y, %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary
    total_cod = drivers.aggregate(total=Sum('cod_in_hand'))['total'] or 0
    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=12, spaceAfter=10)
    elements.append(Paragraph(f'<b>Total Drivers:</b> {drivers.count()}', summary_style))
    elements.append(Paragraph(f'<b>Total COD to Collect:</b> {total_cod} QR', summary_style))
    elements.append(Spacer(1, 20))

    # Table data
    data = [['#', 'Driver Name', 'Driver ID', 'Phone', 'COD Amount (QR)']]
    for i, driver in enumerate(drivers, 1):
        data.append([
            str(i),
            driver.driver_name,
            driver.driver_id,
            driver.driver_mobile or '-',
            f'{driver.cod_in_hand:.2f}'
        ])

    # Add total row
    data.append(['', '', '', 'TOTAL:', f'{total_cod:.2f}'])

    # Create table
    table = Table(data, colWidths=[30, 120, 80, 90, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#001f3f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(table)

    # Signature section
    elements.append(Spacer(1, 40))
    sig_data = [
        ['Staff Signature:', '_' * 30, 'Driver Signature:', '_' * 30],
        ['Name:', '_' * 30, 'Name:', '_' * 30],
        ['Date:', '_' * 30, 'Date:', '_' * 30],
    ]
    sig_table = Table(sig_data, colWidths=[80, 130, 80, 130])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(sig_table)

    doc.build(elements)

    # Return PDF response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f'cod_settlement_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required(login_url='/accounts/login/')
@staff_required
def fleet_transactions(request):
    """View for fleet transactions with filtering and sorting"""
    from django.db.models import Sum
    from decimal import Decimal
    from datetime import timedelta

    # Get all approved drivers for filter dropdown
    all_drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('user__first_name')

    # Get selected driver
    driver_id = request.GET.get('driver_id')
    selected_driver = None
    transactions = fleet_models.DriverTransaction.objects.none()

    if driver_id:
        try:
            selected_driver = fleet_models.Driver.objects.select_related('user').get(driver_id=driver_id)
            transactions = fleet_models.DriverTransaction.objects.filter(
                driver=selected_driver
            ).select_related('delivery_task', 'settlement')
        except fleet_models.Driver.DoesNotExist:
            selected_driver = None

    # Date preset filter
    date_preset = request.GET.get('date_preset', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    today = timezone.now().date()

    # Calculate dates based on preset
    if date_preset == 'today':
        date_from = today.isoformat()
        date_to = today.isoformat()
    elif date_preset == 'yesterday':
        yesterday = today - timedelta(days=1)
        date_from = yesterday.isoformat()
        date_to = yesterday.isoformat()
    elif date_preset == '3days':
        date_from = (today - timedelta(days=3)).isoformat()
        date_to = today.isoformat()
    elif date_preset == '1week':
        date_from = (today - timedelta(days=7)).isoformat()
        date_to = today.isoformat()
    elif date_preset == '1month':
        date_from = (today - timedelta(days=30)).isoformat()
        date_to = today.isoformat()
    # For 'custom', use the date_from and date_to from request

    # Default to cod_collection if no type filter specified
    txn_type = request.GET.get('type', 'cod_collection') if 'type' not in request.GET else request.GET.get('type')
    status = request.GET.get('status', '')
    min_amount = request.GET.get('min_amount', '')
    max_amount = request.GET.get('max_amount', '')

    if selected_driver:
        if date_from:
            transactions = transactions.filter(created_at__date__gte=date_from)
        if date_to:
            transactions = transactions.filter(created_at__date__lte=date_to)
        if txn_type:
            transactions = transactions.filter(transaction_type=txn_type)
        if status == 'settled':
            transactions = transactions.filter(settlement__isnull=False)
        elif status == 'pending':
            transactions = transactions.filter(settlement__isnull=True)
        if min_amount:
            transactions = transactions.filter(amount__gte=min_amount)
        if max_amount:
            transactions = transactions.filter(amount__lte=max_amount)

    # Sorting
    sort_by = request.GET.get('sort', 'date_desc')
    sort_options = {
        'date_asc': 'created_at',
        'date_desc': '-created_at',
        'amount_asc': 'amount',
        'amount_desc': '-amount',
        'type_asc': 'transaction_type',
        'type_desc': '-transaction_type',
    }
    if selected_driver:
        transactions = transactions.order_by(sort_options.get(sort_by, '-created_at'))

    # Transaction type choices for filter dropdown
    transaction_types = fleet_models.DriverTransaction.TRANSACTION_TYPES

    # Calculate totals (on unfiltered queryset for selected driver)
    total_cod = Decimal('0.00')
    total_earnings = Decimal('0.00')
    cod_unsettled = Decimal('0.00')
    earnings_unsettled = Decimal('0.00')
    deposited_refs = set()

    if selected_driver:
        from delivery import models as delivery_models

        # Real COD from DeliveryTask (source of truth)
        dl_tasks = delivery_models.DeliveryTask.objects.filter(driver=selected_driver)
        total_cod = dl_tasks.filter(
            cod_collected=True
        ).aggregate(total=Sum('cod_collected_amount'))['total'] or Decimal('0.00')
        cod_unsettled = dl_tasks.filter(
            cod_collected=True, cod_settled=False
        ).aggregate(total=Sum('cod_collected_amount'))['total'] or Decimal('0.00')

        # Get all transactions for earnings totals and deposit refs
        all_txns = fleet_models.DriverTransaction.objects.filter(driver=selected_driver)

        # Fetch all COD deposit reference numbers in one query to avoid N+1
        deposited_refs = set(
            all_txns.filter(transaction_type='cod_deposit').values_list('reference_number', flat=True)
        )

        for txn in all_txns:
            if txn.transaction_type == 'earning':
                total_earnings += txn.amount
                # Earnings not yet settled
                if not txn.settlement_id:
                    earnings_unsettled += txn.amount

    # Build filter params for pagination (exclude page and per_page as they're handled separately)
    filter_params = request.GET.copy()
    if 'page' in filter_params:
        del filter_params['page']
    if 'per_page' in filter_params:
        del filter_params['per_page']

    transactions_paginated = paginate_queryset(request, transactions, items_per_page=20)

    context = {
        'page_title': 'Fleet Transactions',
        'transactions': transactions_paginated,
        'all_drivers': all_drivers,
        'selected_driver': selected_driver,
        'total_cod': total_cod,
        'total_earnings': total_earnings,
        'cod_unsettled': cod_unsettled,
        'earnings_unsettled': earnings_unsettled,
        'deposited_refs': deposited_refs,
        'filter_params': filter_params.urlencode(),
        'per_page': get_per_page(request, default=20),
        # Filter values for form
        'date_preset': date_preset,
        'date_from': date_from or '',
        'date_to': date_to or '',
        'txn_type': txn_type or '',
        'status': status or '',
        'min_amount': min_amount or '',
        'max_amount': max_amount or '',
        'sort_by': sort_by,
        'transaction_types': transaction_types,
    }
    return render(request, 'workforce/fleet_transactions.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def generate_demo_transactions(request):
    """Generate demo transactions for a driver"""
    from django.contrib import messages
    from decimal import Decimal
    from datetime import timedelta
    import random

    if request.method != 'POST':
        return redirect('workforce:fleet_transactions')

    driver_id = request.POST.get('driver_id')
    if not driver_id:
        messages.error(request, 'Driver ID is required')
        return redirect('workforce:fleet_transactions')

    try:
        driver = fleet_models.Driver.objects.get(driver_id=driver_id)
    except fleet_models.Driver.DoesNotExist:
        messages.error(request, 'Driver not found')
        return redirect('workforce:fleet_transactions')

    # Check if driver already has transactions
    existing_count = fleet_models.DriverTransaction.objects.filter(driver=driver).count()
    if existing_count > 0:
        messages.warning(request, f'Driver already has {existing_count} transactions. Demo data not generated.')
        return redirect(f"{reverse('workforce:fleet_transactions')}?driver_id={driver_id}")

    now = timezone.now()
    cod_in_hand = Decimal('0.00')
    pending_earnings = Decimal('0.00')
    wallet_balance = driver.wallet_balance or Decimal('5000.00')

    transactions_to_create = []

    # Generate 10-15 COD collection transactions over past 7 days
    for i in range(random.randint(10, 15)):
        days_ago = random.randint(0, 7)
        hours_ago = random.randint(0, 23)
        txn_date = now - timedelta(days=days_ago, hours=hours_ago)

        cod_amount = Decimal(random.randint(50, 500))
        earning_amount = Decimal(random.randint(15, 50))
        cod_in_hand += cod_amount
        wallet_balance -= cod_amount  # COD collection decreases wallet
        pending_earnings += earning_amount

        # COD Collection transaction
        transactions_to_create.append(fleet_models.DriverTransaction(
            driver=driver,
            transaction_type='cod_collection',
            amount=cod_amount,
            description=f'COD collected for delivery #{random.randint(1000, 9999)}',
            reference_number=f'COD-{random.randint(10000, 99999)}',
            wallet_balance_after=wallet_balance,
            cod_in_hand_after=cod_in_hand,
            pending_earnings_after=pending_earnings,
            created_by=request.user,
            created_at=txn_date,
        ))

        # Earning transaction
        transactions_to_create.append(fleet_models.DriverTransaction(
            driver=driver,
            transaction_type='earning',
            amount=earning_amount,
            description=f'Delivery earning for task #{random.randint(1000, 9999)}',
            reference_number=f'EARN-{random.randint(10000, 99999)}',
            wallet_balance_after=wallet_balance,
            cod_in_hand_after=cod_in_hand,
            pending_earnings_after=pending_earnings,
            created_by=request.user,
            created_at=txn_date + timedelta(minutes=5),
        ))

    # Generate 2-3 COD deposit (cash settled) transactions
    for i in range(random.randint(2, 3)):
        days_ago = random.randint(1, 5)
        txn_date = now - timedelta(days=days_ago, hours=random.randint(10, 18))

        deposit_amount = Decimal(random.randint(200, 800))
        deposit_amount = min(deposit_amount, cod_in_hand)  # Can't deposit more than in hand
        cod_in_hand = max(Decimal('0.00'), cod_in_hand - deposit_amount)
        wallet_balance += deposit_amount  # COD deposit increases wallet

        transactions_to_create.append(fleet_models.DriverTransaction(
            driver=driver,
            transaction_type='cod_deposit',
            amount=-deposit_amount,  # Negative because money leaving driver
            description=f'COD deposited to admin - Cash settlement',
            reference_number=f'DEP-{random.randint(10000, 99999)}',
            wallet_balance_after=wallet_balance,
            cod_in_hand_after=cod_in_hand,
            pending_earnings_after=pending_earnings,
            created_by=request.user,
            created_at=txn_date,
        ))

    # Generate 1 earnings settlement transaction
    if pending_earnings > Decimal('100.00'):
        settlement_amount = pending_earnings * Decimal('0.6')  # Settle 60%
        pending_earnings -= settlement_amount

        transactions_to_create.append(fleet_models.DriverTransaction(
            driver=driver,
            transaction_type='settlement',
            amount=-settlement_amount,  # Negative because paid out
            description='Weekly earnings settlement - Bank transfer',
            reference_number=f'SETTLE-{random.randint(10000, 99999)}',
            wallet_balance_after=wallet_balance,
            cod_in_hand_after=cod_in_hand,
            pending_earnings_after=pending_earnings,
            created_by=request.user,
            created_at=now - timedelta(days=3),
        ))

    # Add a bonus transaction
    bonus_amount = Decimal(random.randint(50, 150))
    pending_earnings += bonus_amount
    transactions_to_create.append(fleet_models.DriverTransaction(
        driver=driver,
        transaction_type='bonus',
        amount=bonus_amount,
        description='Weekly performance bonus',
        reference_number=f'BONUS-{random.randint(10000, 99999)}',
        wallet_balance_after=wallet_balance,
        cod_in_hand_after=cod_in_hand,
        pending_earnings_after=pending_earnings,
        created_by=request.user,
        created_at=now - timedelta(days=1),
    ))

    # Bulk create all transactions
    fleet_models.DriverTransaction.objects.bulk_create(transactions_to_create)

    # Update driver's current balances including wallet_balance
    driver.cod_in_hand = cod_in_hand
    driver.pending_earnings = pending_earnings
    driver.wallet_balance = wallet_balance
    driver.save(update_fields=['cod_in_hand', 'pending_earnings', 'wallet_balance'])

    messages.success(request, f'Successfully generated {len(transactions_to_create)} demo transactions for {driver.user.first_name} {driver.user.last_name}')
    return redirect(f"{reverse('workforce:fleet_transactions')}?driver_id={driver_id}")


@login_required(login_url='/accounts/login/')
@staff_required
def bulk_settle_transactions(request):
    """Bulk settle selected transactions for a driver"""
    from django.contrib import messages
    from decimal import Decimal
    import json

    if request.method != 'POST':
        return redirect('workforce:fleet_transactions')

    driver_id = request.POST.get('driver_id')
    transaction_ids_json = request.POST.get('transaction_ids', '[]')

    try:
        transaction_ids = json.loads(transaction_ids_json)
    except json.JSONDecodeError:
        messages.error(request, 'Invalid transaction data')
        return redirect('workforce:fleet_transactions')

    if not driver_id or not transaction_ids:
        messages.error(request, 'Driver ID and transactions are required')
        return redirect('workforce:fleet_transactions')

    try:
        driver = fleet_models.Driver.objects.get(driver_id=driver_id)
    except fleet_models.Driver.DoesNotExist:
        messages.error(request, 'Driver not found')
        return redirect('workforce:fleet_transactions')

    # Get transactions that can be settled (pending earnings and COD collections)
    transactions = fleet_models.DriverTransaction.objects.filter(
        id__in=transaction_ids,
        driver=driver,
        settlement__isnull=True  # Not already settled
    ).exclude(
        transaction_type__in=['settlement', 'cod_deposit']  # These are already finalized
    )

    if not transactions.exists():
        messages.warning(request, 'No eligible transactions found for settlement')
        return redirect(f"{reverse('workforce:fleet_transactions')}?driver_id={driver_id}")

    # Calculate total settlement amount
    total_amount = sum(txn.amount for txn in transactions)

    # Wrap entire settlement operation in transaction.atomic for data consistency
    from django.db import transaction as db_transaction
    from django.db.models import F
    from django.db.models.functions import Greatest

    # Pre-calculate amounts before entering atomic block (transactions queryset will be re-evaluated)
    transaction_list = list(transactions)
    earnings_settled = sum(
        (txn.amount for txn in transaction_list if txn.transaction_type == 'earning'),
        Decimal('0.00')
    )
    cod_settled = sum(
        (abs(txn.amount) for txn in transaction_list if txn.transaction_type == 'cod_collection'),
        Decimal('0.00')
    )

    with db_transaction.atomic():
        # Create a settlement record
        settlement_code = f"STL-{timezone.now().strftime('%Y%m%d%H%M%S')}-{driver_id}"

        # Check if transactions still exist (might have been settled by another request)
        first_txn = transactions.order_by('created_at').first()
        if not first_txn:
            messages.warning(request, 'Transactions were already settled by another request')
            return redirect(f"{reverse('workforce:fleet_transactions')}?driver_id={driver_id}")

        settlement = fleet_models.DriverSettlement.objects.create(
            driver=driver,
            settlement_code=settlement_code,
            period_start=first_txn.created_at.date(),
            period_end=timezone.now().date(),
            total_deliveries=transactions.filter(transaction_type='earning').count(),
            gross_earnings=total_amount,
            net_amount=total_amount,
            created_by=request.user,
            status='paid',
            paid_at=timezone.now(),
        )

        # Link transactions to settlement
        transactions.update(settlement=settlement)

        # Update driver balances atomically using F() to prevent race conditions
        if earnings_settled > 0:
            # Use Greatest to ensure we don't go below zero
            fleet_models.Driver.objects.filter(driver_id=driver_id).update(
                pending_earnings=Greatest(
                    F('pending_earnings') - earnings_settled,
                    Decimal('0.00')
                )
            )
        if cod_settled > 0:
            fleet_models.Driver.objects.filter(driver_id=driver_id).update(
                cod_in_hand=Greatest(
                    F('cod_in_hand') - cod_settled,
                    Decimal('0.00')
                )
            )

        # Update last_settlement_date
        fleet_models.Driver.objects.filter(driver_id=driver_id).update(
            last_settlement_date=timezone.now()
        )

    messages.success(
        request,
        f'Successfully settled {transactions.count()} transactions for QAR {total_amount:.2f}. '
        f'Settlement code: {settlement_code}'
    )
    return redirect(f"{reverse('workforce:fleet_transactions')}?driver_id={driver_id}")


# ==========================================
# RECEIPT TEMPLATES SECTION
# ==========================================

@login_required(login_url='/accounts/login/')
@staff_required
def receipt_templates_list(request):
    """List all receipt templates"""
    templates = fleet_models.ReceiptTemplate.objects.all()

    # Filter by type if specified
    template_type = request.GET.get('type')
    if template_type:
        templates = templates.filter(template_type=template_type)

    context = {
        'page_title': 'Receipt Templates',
        'templates': templates,
        'template_types': fleet_models.ReceiptTemplate.TEMPLATE_TYPES,
        'selected_type': template_type,
    }
    return render(request, 'workforce/receipt_templates_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def receipt_template_create(request):
    """Create a new receipt template"""
    from django.contrib import messages

    if request.method == 'POST':
        template = fleet_models.ReceiptTemplate(
            name=request.POST.get('name'),
            template_type=request.POST.get('template_type'),
            paper_size=request.POST.get('paper_size', 'thermal_80'),
            is_default=request.POST.get('is_default') == 'on',
            show_logo=request.POST.get('show_logo') == 'on',
            logo_url=request.POST.get('logo_url', ''),
            company_name=request.POST.get('company_name', 'Ezzy Delivery'),
            company_address=request.POST.get('company_address', ''),
            company_phone=request.POST.get('company_phone', ''),
            company_email=request.POST.get('company_email') or None,
            primary_color=request.POST.get('primary_color', '#2196F3'),
            font_family=request.POST.get('font_family', 'Courier New, monospace'),
            font_size=int(request.POST.get('font_size', 12)),
            show_signature_line=request.POST.get('show_signature_line') == 'on',
            show_qr_code=request.POST.get('show_qr_code') == 'on',
            footer_message=request.POST.get('footer_message', ''),
            custom_css=request.POST.get('custom_css', ''),
            created_by=request.user,
        )
        template.save()
        messages.success(request, f'Template "{template.name}" created successfully.')
        return redirect('workforce:receipt_templates_list')

    context = {
        'page_title': 'Create Receipt Template',
        'template_types': fleet_models.ReceiptTemplate.TEMPLATE_TYPES,
        'paper_sizes': fleet_models.ReceiptTemplate.PAPER_SIZES,
    }
    return render(request, 'workforce/receipt_template_edit.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def receipt_template_edit(request, template_id):
    """Edit an existing receipt template"""
    from django.contrib import messages
    from django.shortcuts import get_object_or_404

    template = get_object_or_404(fleet_models.ReceiptTemplate, template_id=template_id)

    if request.method == 'POST':
        template.name = request.POST.get('name')
        template.template_type = request.POST.get('template_type')
        template.paper_size = request.POST.get('paper_size', 'thermal_80')
        template.is_default = request.POST.get('is_default') == 'on'
        template.show_logo = request.POST.get('show_logo') == 'on'
        template.logo_url = request.POST.get('logo_url', '')
        template.company_name = request.POST.get('company_name', 'Ezzy Delivery')
        template.company_address = request.POST.get('company_address', '')
        template.company_phone = request.POST.get('company_phone', '')
        template.company_email = request.POST.get('company_email') or None
        template.primary_color = request.POST.get('primary_color', '#2196F3')
        template.font_family = request.POST.get('font_family', 'Courier New, monospace')
        template.font_size = int(request.POST.get('font_size', 12))
        template.show_signature_line = request.POST.get('show_signature_line') == 'on'
        template.show_qr_code = request.POST.get('show_qr_code') == 'on'
        template.footer_message = request.POST.get('footer_message', '')
        template.custom_css = request.POST.get('custom_css', '')
        template.save()
        messages.success(request, f'Template "{template.name}" updated successfully.')
        return redirect('workforce:receipt_templates_list')

    context = {
        'page_title': f'Edit Template: {template.name}',
        'template': template,
        'template_types': fleet_models.ReceiptTemplate.TEMPLATE_TYPES,
        'paper_sizes': fleet_models.ReceiptTemplate.PAPER_SIZES,
    }
    return render(request, 'workforce/receipt_template_edit.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def receipt_template_preview(request, template_id):
    """Preview a receipt template with sample data"""
    from django.shortcuts import get_object_or_404
    from decimal import Decimal

    template = get_object_or_404(fleet_models.ReceiptTemplate, template_id=template_id)

    # Create sample data for preview
    sample_settlement = {
        'settlement_code': 'STL-PREVIEW-001',
        'created_at': timezone.now(),
        'period_start': timezone.now().date() - timezone.timedelta(days=7),
        'period_end': timezone.now().date(),
        'net_amount': Decimal('1250.00'),
        'gross_earnings': Decimal('1300.00'),
        'deductions': Decimal('50.00'),
        'bonuses': Decimal('0.00'),
        'status': 'paid',
        'paid_at': timezone.now(),
    }

    sample_driver = {
        'driver_id': 45,
        'user': {'first_name': 'Ahmed', 'last_name': 'Khan'},
        'driver_phone': '+974-5555-1234',
    }

    sample_transactions = [
        {'description': 'COD collected for order #1234', 'created_at': timezone.now() - timezone.timedelta(days=1), 'amount': Decimal('350.00')},
        {'description': 'COD collected for order #1235', 'created_at': timezone.now() - timezone.timedelta(days=2), 'amount': Decimal('275.00')},
        {'description': 'COD collected for order #1236', 'created_at': timezone.now() - timezone.timedelta(days=3), 'amount': Decimal('425.00')},
        {'description': 'COD collected for order #1237', 'created_at': timezone.now() - timezone.timedelta(days=4), 'amount': Decimal('200.00')},
    ]

    context = {
        'template': template,
        'settlement': sample_settlement,
        'driver': sample_driver,
        'transactions': sample_transactions,
        'is_preview': True,
    }
    return render(request, 'workforce/receipt_template_preview.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def receipt_template_delete(request, template_id):
    """Delete a receipt template"""
    from django.shortcuts import get_object_or_404
    from django.contrib import messages

    template = get_object_or_404(fleet_models.ReceiptTemplate, template_id=template_id)

    if request.method == 'POST':
        name = template.name
        template.delete()
        messages.success(request, f'Template "{name}" deleted successfully.')
        return redirect('workforce:receipt_templates_list')

    return redirect('workforce:receipt_templates_list')


@login_required(login_url='/accounts/login/')
@staff_required
def settlement_receipt_print(request, settlement_id):
    """Print a settlement receipt using the default or specified template"""
    from django.shortcuts import get_object_or_404

    settlement = get_object_or_404(fleet_models.DriverSettlement, settlement_id=settlement_id)
    transactions = fleet_models.DriverTransaction.objects.filter(settlement=settlement)

    # Get template (use specified or default)
    template_id = request.GET.get('template_id')
    if template_id:
        template = fleet_models.ReceiptTemplate.objects.filter(template_id=template_id).first()
    else:
        template = fleet_models.ReceiptTemplate.objects.filter(
            template_type='settlement',
            is_default=True
        ).first()

    context = {
        'settlement': settlement,
        'transactions': transactions,
        'template': template,
        'is_preview': False,
    }
    return render(request, 'workforce/settlement_receipt.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def fleet_task_sheet(request, driver_id):
    """Generate printable A4 task sheet for a driver's daily deliveries"""
    from django.shortcuts import get_object_or_404
    from django.db.models import Sum
    from datetime import datetime
    from decimal import Decimal

    driver = get_object_or_404(fleet_models.Driver.objects.select_related('user'), driver_id=driver_id)

    # Get date from query param or use today
    date_str = request.GET.get('date')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()

    # Get delivery tasks for this driver on the selected date
    tasks = delivery_models.DeliveryTask.objects.filter(
        driver=driver,
        dl_task_date=selected_date
    ).select_related('order').prefetch_related('order__order_items').order_by('dl_task_date')

    # Calculate total COD amount
    total_cod = Decimal('0.00')
    for task in tasks:
        if task.order and task.order.payment_type == 'cod':
            total_cod += task.order.cod_amount or task.order.order_total or Decimal('0.00')

    # Get unique delivery areas
    areas = set()
    for task in tasks:
        if task.order and task.order.delivery_area:
            areas.add(task.order.delivery_area)
    delivery_area = ', '.join(areas) if areas else 'Multiple'

    # Sheet number (can be customized based on shift or slot)
    sheet_number = request.GET.get('sheet', '01')
    slot_number = request.GET.get('slot', '01')

    context = {
        'driver': driver,
        'tasks': tasks,
        'date': selected_date,
        'total_cod': total_cod,
        'delivery_area': delivery_area,
        'sheet_number': sheet_number,
        'slot_number': slot_number,
    }
    return render(request, 'workforce/fleet_task_sheet.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def fleet_task_sheets_list(request):
    """List view to select driver and date for task sheet printing"""
    from django.db.models import Count, Q
    from datetime import datetime

    # Get all approved drivers with task counts for today
    today = timezone.now().date()
    selected_date = request.GET.get('date')
    if selected_date:
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = today
    else:
        selected_date = today

    drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').annotate(
        task_count=Count(
            'deliverytask',
            filter=Q(deliverytask__dl_task_date=selected_date)
        )
    ).order_by('user__first_name')

    # Get drivers with tasks for quick filter
    drivers_with_tasks = drivers.filter(task_count__gt=0)

    context = {
        'page_title': 'Fleet Task Sheets',
        'drivers': drivers,
        'drivers_with_tasks': drivers_with_tasks,
        'selected_date': selected_date,
        'today': today,
    }
    return render(request, 'workforce/fleet_task_sheets_list.html', context)


# Inventory Section Functions
@login_required(login_url='/accounts/login/')
@staff_required
def inventory_reports(request):
    """View for inventory reports"""
    context = {
        'page_title': 'Inventory Reports',
    }
    return render(request, 'workforce/inventory_reports.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def inventory_restock_list(request):
    """View for restock list"""
    context = {
        'page_title': 'Restock List',
    }
    return render(request, 'workforce/inventory_restock_list.html', context)


# Quick Links Functions
@login_required(login_url='/accounts/login/')
@staff_required
def staff_reports(request):
    """View for staff reports dashboard"""
    from orders import models as orders_models
    from django.db.models import Count, Q

    # Get order statistics in single query
    order_stats = orders_models.Order.objects.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(order_status='pending')),
        verified=Count('id', filter=Q(order_status='verified')),
        completed=Count('id', filter=Q(order_status='completed')),
    )

    # Get task statistics in single query
    task_stats = delivery_models.DeliveryTask.objects.aggregate(
        total=Count('id'),
        active=Count('id', filter=Q(dl_task_status__in=['in_transit', 'pending', 'address_pending'])),
        completed=Count('id', filter=Q(dl_task_status='delivered')),
    )

    # Get driver statistics in single query
    driver_stats = fleet_models.Driver.objects.aggregate(
        total=Count('driver_id'),
        active=Count('driver_id', filter=Q(driver_status='approved')),
    )

    context = {
        'page_title': 'Reports Dashboard',
        'total_orders': order_stats['total'],
        'pending_orders': order_stats['pending'],
        'verified_orders': order_stats['verified'],
        'completed_orders': order_stats['completed'],
        'total_tasks': task_stats['total'],
        'active_tasks': task_stats['active'],
        'completed_tasks': task_stats['completed'],
        'total_drivers': driver_stats['total'],
        'active_drivers': driver_stats['active'],
    }
    return render(request, 'workforce/staff_reports.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def staff_contacts(request):
    """View for staff contacts directory"""
    from core import models as core_models
    from django.db.models import Count, Q

    # Get all staff members
    staff_profiles = core_models.Profile.objects.select_related('user').filter(
        is_staff=True
    ).order_by('first_name')

    # Get business and driver counts in a single query
    profile_counts = core_models.Profile.objects.aggregate(
        business_count=Count('id', filter=Q(is_business=True, verification_status='verified')),
        driver_count=Count('id', filter=Q(is_driver=True, verification_status='verified')),
    )

    staff_with_pagination = paginate_queryset(request, staff_profiles, items_per_page=50)

    context = {
        'page_title': 'Contacts Directory',
        'staff_profiles': staff_with_pagination,
        'business_count': profile_counts['business_count'],
        'driver_count': profile_counts['driver_count'],
    }
    return render(request, 'workforce/staff_contacts.html', context)



# Documents section  ------------------------------------------------------------------------------------------------------


@login_required(login_url='/accounts/login/')
@staff_required
def driver_documents_list(request):
    """View for listing all driver documents with search and card/table toggle"""
    from django.db.models import Q

    # Get search query and view type
    search_query = request.GET.get('search', '').strip()
    view_type = request.GET.get('view', 'card')  # 'card' or 'table'

    # Start with all driver documents
    documents = fleet_models.DriverDocument.objects.select_related('driver', 'driver__user', 'driver__profile').all()

    # Apply search filter
    if search_query:
        documents = documents.filter(
            Q(document_no__icontains=search_query) |
            Q(document_type__icontains=search_query) |
            Q(driver__user__first_name__icontains=search_query) |
            Q(driver__user__last_name__icontains=search_query) |
            Q(driver__driver_code__icontains=search_query)
        )

    # Order by most recent
    documents = documents.order_by('-created_at')

    # Paginate results
    page_obj = paginate_queryset(request, documents, items_per_page=20)

    context = {
        'page_title': 'Driver ID Documents',
        'documents': page_obj,
        'search_query': search_query,
        'view_type': view_type,
    }

    return render(request, 'workforce/driver_documents_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def driver_document_detail(request, document_id):
    """View for viewing and updating a specific driver document"""
    document = get_object_or_404(fleet_models.DriverDocument, id=document_id)

    if request.method == 'POST':
        # Handle document update
        try:
            document.document_type = request.POST.get('document_type', document.document_type)
            document.document_no = request.POST.get('document_no', document.document_no)
            document.document_issued_from = request.POST.get('document_issued_from', document.document_issued_from)

            expiry_date = request.POST.get('document_expiry_date')
            if expiry_date:
                document.document_expiry_date = expiry_date

            # Handle file upload with validation
            if 'document_file' in request.FILES:
                uploaded_file = request.FILES['document_file']
                # Validate file extension
                allowed_extensions = ['.jpg', '.jpeg', '.png', '.pdf', '.gif']
                import os
                ext = os.path.splitext(uploaded_file.name)[1].lower()
                if ext not in allowed_extensions:
                    return JsonResponse({
                        'success': False,
                        'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
                    }, status=400)
                # Validate file size (max 10MB)
                max_size = 10 * 1024 * 1024
                if uploaded_file.size > max_size:
                    return JsonResponse({
                        'success': False,
                        'error': 'File too large. Maximum size is 10MB'
                    }, status=400)
                document.document_file = uploaded_file

            document.save()

            return JsonResponse({
                'success': True,
                'message': 'Document updated successfully'
            })
        except Exception as e:
            logger.exception("Error updating driver document %s: %s", document_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating document'
            }, status=400)

    context = {
        'page_title': 'Driver Document Detail',
        'document': document,
    }

    return render(request, 'workforce/driver_document_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def vehicle_documents_list(request):
    """View for listing all vehicle documents with search and card/table toggle"""
    from django.db.models import Q

    # Get search query and view type
    search_query = request.GET.get('search', '').strip()
    view_type = request.GET.get('view', 'card')  # 'card' or 'table'

    # Start with all driver vehicles
    vehicles = fleet_models.DriverVehicle.objects.select_related('driver', 'driver__user', 'driver__profile').all()

    # Apply search filter
    if search_query:
        vehicles = vehicles.filter(
            Q(vehicle_no__icontains=search_query) |
            Q(vehicle_type__icontains=search_query) |
            Q(vehicle_model__icontains=search_query) |
            Q(driver__user__first_name__icontains=search_query) |
            Q(driver__user__last_name__icontains=search_query) |
            Q(driver__driver_code__icontains=search_query)
        )

    # Order by most recent
    vehicles = vehicles.order_by('-created_at')

    # Paginate results
    page_obj = paginate_queryset(request, vehicles, items_per_page=20)

    context = {
        'page_title': 'Vehicle Documents',
        'vehicles': page_obj,
        'search_query': search_query,
        'view_type': view_type,
    }

    return render(request, 'workforce/vehicle_documents_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def vehicle_document_detail(request, driver_id):
    """View for viewing and updating vehicle documents for a specific driver"""
    driver = get_object_or_404(fleet_models.Driver, driver_id=driver_id)
    vehicles = fleet_models.DriverVehicle.objects.filter(driver=driver).order_by('-created_at')

    if request.method == 'POST':
        # Handle vehicle update
        try:
            vehicle_id = request.POST.get('vehicle_id')
            if vehicle_id:
                vehicle = get_object_or_404(fleet_models.DriverVehicle, id=vehicle_id)
                vehicle.vehicle_type = request.POST.get('vehicle_type', vehicle.vehicle_type)
                vehicle.vehicle_no = request.POST.get('vehicle_no', vehicle.vehicle_no)
                vehicle.vehicle_model = request.POST.get('vehicle_model', vehicle.vehicle_model)
                vehicle.vehicle_color = request.POST.get('vehicle_color', vehicle.vehicle_color)
                vehicle.vehicle_status = request.POST.get('vehicle_status', vehicle.vehicle_status)
                vehicle.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Vehicle updated successfully'
                })
        except Exception as e:
            logger.exception("Error updating vehicle for driver %s: %s", driver_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating vehicle'
            }, status=400)

    context = {
        'page_title': 'Vehicle Documents Detail',
        'driver': driver,
        'vehicles': vehicles,
    }

    return render(request, 'workforce/vehicle_document_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def store_documents_list(request):
    """View for listing all store/business documents with search and card/table toggle"""
    from django.db.models import Q

    # Get search query and view type
    search_query = request.GET.get('search', '').strip()
    view_type = request.GET.get('view', 'card')  # 'card' or 'table'

    # Start with all businesses
    businesses = business_models.Business.objects.select_related('business_profile', 'user').all()

    # Apply search filter
    if search_query:
        businesses = businesses.filter(
            Q(business_name__icontains=search_query) |
            Q(business_code__icontains=search_query) |
            Q(business_phone__icontains=search_query) |
            Q(business_email__icontains=search_query) |
            Q(business_qid__icontains=search_query)
        )

    # Order by most recent
    businesses = businesses.order_by('-created_at')

    # Paginate results
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Store Documents',
        'businesses': page_obj,
        'search_query': search_query,
        'view_type': view_type,
    }

    return render(request, 'workforce/store_documents_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def store_document_detail(request, business_id):
    """View for viewing and updating store documents for a specific business"""
    from django.db.models import Count, Sum, Q

    business = get_object_or_404(
        business_models.Business.objects.select_related('user', 'profile'),
        business_id=business_id
    )

    try:
        business_profile = business.business_profile
    except business_models.Business.business_profile.RelatedObjectDoesNotExist:
        business_profile = None

    # Get user profile
    try:
        user_profile = business.user.profile if business.user else None
    except Exception:
        user_profile = None

    # Pickup locations
    pickup_locations = business.pickup_location.all()

    # Order stats
    order_stats = orders_models.Order.objects.filter(business=business).aggregate(
        total_orders=Count('id'),
        delivered_orders=Count('id', filter=Q(order_status='delivered')),
    )

    # Registration completion
    required_fields = ['business_name', 'business_phone', 'business_whatsapp',
                       'business_email', 'business_product_category', 'business_qid']
    filled = sum(1 for f in required_fields if getattr(business, f, None))
    completion_pct = round(filled / len(required_fields) * 100)

    if request.method == 'POST':
        # Handle business document update via HTMX or fetch
        try:
            business.business_name = request.POST.get('business_name', business.business_name)
            business.business_phone = request.POST.get('business_phone', business.business_phone)
            business.business_whatsapp = request.POST.get('business_whatsapp', business.business_whatsapp)
            business.business_email = request.POST.get('business_email', business.business_email)
            business.business_qid = request.POST.get('business_qid', business.business_qid)
            business.business_product_category = request.POST.get('business_product_category', business.business_product_category)
            business.business_status = request.POST.get('business_status', business.business_status)
            business.save()

            # Update business profile if exists
            if business_profile:
                business_profile.business_city = request.POST.get('business_city', business_profile.business_city) or None
                business_profile.business_country = request.POST.get('business_country', business_profile.business_country) or 'Qatar'
                business_profile.business_founters_name = request.POST.get('business_founters_name', business_profile.business_founters_name) or None
                business_profile.save()

            return JsonResponse({
                'success': True,
                'message': 'Store document updated successfully'
            })
        except Exception as e:
            logger.exception("Error updating store document for business %s: %s", business_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating store document'
            }, status=400)

    context = {
        'page_title': f'Store: {business.business_name}',
        'business': business,
        'business_profile': business_profile,
        'user_profile': user_profile,
        'pickup_locations': pickup_locations,
        'order_stats': order_stats,
        'completion_pct': completion_pct,
        'required_fields': required_fields,
    }

    return render(request, 'workforce/store_document_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def business_licenses_list(request):
    """View for listing all business licenses with search and card/table toggle"""
    from django.db.models import Q

    # Get search query and view type
    search_query = request.GET.get('search', '').strip()
    view_type = request.GET.get('view', 'card')  # 'card' or 'table'

    # Start with all businesses
    businesses = business_models.Business.objects.select_related('business_profile', 'user').all()

    # Apply search filter
    if search_query:
        businesses = businesses.filter(
            Q(business_name__icontains=search_query) |
            Q(business_code__icontains=search_query) |
            Q(business_qid__icontains=search_query)
        )

    # Order by most recent
    businesses = businesses.order_by('-created_at')

    # Paginate results
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Business Licenses',
        'businesses': page_obj,
        'search_query': search_query,
        'view_type': view_type,
    }

    return render(request, 'workforce/business_licenses_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def business_license_detail(request, business_id):
    """View for viewing and updating business license details"""
    business = get_object_or_404(business_models.Business, business_id=business_id)

    try:
        business_profile = business.business_profile
    except business_models.Business.business_profile.RelatedObjectDoesNotExist:
        business_profile = None

    if request.method == 'POST':
        section = request.POST.get('section', 'all')
        try:
            if section == 'license':
                business.business_name = request.POST.get('business_name', business.business_name)
                business.business_qid = request.POST.get('business_qid', business.business_qid)
                business.business_product_category = request.POST.get('business_product_category', business.business_product_category)
                business_since = request.POST.get('business_since')
                if business_since:
                    business.business_since = business_since
                business.save()

            elif section == 'contact':
                business.business_phone = request.POST.get('business_phone', business.business_phone)
                business.business_email = request.POST.get('business_email', business.business_email)
                business.business_whatsapp = request.POST.get('business_whatsapp', business.business_whatsapp)
                business.business_facebook_page = request.POST.get('business_facebook_page', business.business_facebook_page)
                business.business_instagram = request.POST.get('business_instagram', business.business_instagram)
                business.save()

            elif section == 'address':
                if business_profile:
                    business_profile.business_address = request.POST.get('business_address', business_profile.business_address)
                    business_profile.business_city = request.POST.get('business_city', business_profile.business_city)
                    business_profile.business_country = request.POST.get('business_country', business_profile.business_country)
                    business_profile.save()

            elif section == 'status':
                business.business_status = request.POST.get('business_status', business.business_status)
                business.save()

            elif section == 'fulfillment':
                from warehouse import models as warehouse_models

                fulfillment_status = request.POST.get('fulfillment_service_status', business.fulfillment_service_status)
                warehouse_location_id = request.POST.get('warehouse_location_id', '').strip()

                logger.info(f"Fulfillment update for business {business.business_id}: status={fulfillment_status}, warehouse_id={warehouse_location_id}")

                # If activating fulfillment service, warehouse location is required
                if fulfillment_status == 'active' and not warehouse_location_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'Please select a warehouse location when activating fulfillment service'
                    }, status=400)

                # Update fulfillment status
                try:
                    business.fulfillment_service_status = fulfillment_status
                    if fulfillment_status == 'active':
                        business.fulfillment_service_enabled = True
                        from django.utils import timezone
                        if not business.fulfillment_activated_at:
                            business.fulfillment_activated_at = timezone.now()
                    business.save()
                    logger.info(f"Business {business.business_id} fulfillment status updated to {fulfillment_status}")
                except Exception as e:
                    logger.exception(f"Error saving business fulfillment status: {e}")
                    return JsonResponse({
                        'success': False,
                        'error': f'Error updating fulfillment status: {str(e)}'
                    }, status=400)

                # Link warehouse if provided
                if warehouse_location_id and fulfillment_status == 'active':
                    try:
                        warehouse_location = warehouse_models.WarehouseLocation.objects.select_related('warehouse').get(id=warehouse_location_id)
                        logger.info(f"Found warehouse location: {warehouse_location.warehouse.code}/{warehouse_location.code}")

                        # Create or update warehouse link
                        link, created = warehouse_models.SellerWarehouseLink.objects.get_or_create(
                            business=business,
                            warehouse=warehouse_location.warehouse,
                            defaults={
                                'default_location': warehouse_location,  # Correct field name
                                'is_default': True,  # First warehouse becomes default
                                'linked_by': request.user,
                            }
                        )

                        if not created:
                            link.default_location = warehouse_location  # Correct field name
                            link.save()

                        action = 'created' if created else 'updated'
                        logger.info(f"Warehouse link {action} for business {business.business_id} -> {warehouse_location.warehouse.code}")

                        # Create or update a PickupLocation for the warehouse fulfillment center
                        pickup_location, pickup_created = business_models.PickupLocation.objects.update_or_create(
                            business=business,
                            warehouse=warehouse_location.warehouse,
                            defaults={
                                'pickup_location_title': f"{warehouse_location.warehouse.name} - Fulfillment",
                                'locality': warehouse_location.address or warehouse_location.warehouse.city or 'Warehouse Location',
                                'is_fulfilment_center': True,
                                'pickup_status': 'active',
                                'pickup_zone_no': warehouse_location.zone_number,
                                'pickup_lat': warehouse_location.latitude,
                                'pickup_lon': warehouse_location.longitude,
                            }
                        )
                        pickup_action = 'created' if pickup_created else 'updated'
                        logger.info(f"Fulfillment pickup location {pickup_action}: {pickup_location.pickup_location_title}")
                    except warehouse_models.WarehouseLocation.DoesNotExist:
                        logger.error(f"Warehouse location {warehouse_location_id} not found")
                        return JsonResponse({
                            'success': False,
                            'error': 'Selected warehouse location not found'
                        }, status=400)
                    except Exception as e:
                        logger.exception(f"Error linking warehouse: {e}")
                        return JsonResponse({
                            'success': False,
                            'error': f'Error linking warehouse: {str(e)}'
                        }, status=400)

            else:
                # Legacy: save all fields
                business.business_name = request.POST.get('business_name', business.business_name)
                business.business_qid = request.POST.get('business_qid', business.business_qid)
                business.business_status = request.POST.get('business_status', business.business_status)
                business.fulfillment_service_status = request.POST.get('fulfillment_service_status', business.fulfillment_service_status)
                business_since = request.POST.get('business_since')
                if business_since:
                    business.business_since = business_since
                business.save()
                if business_profile:
                    business_profile.business_address = request.POST.get('business_address', business_profile.business_address)
                    business_profile.business_city = request.POST.get('business_city', business_profile.business_city)
                    business_profile.save()

            return JsonResponse({
                'success': True,
                'message': f'{section.title()} updated successfully'
            })
        except Exception as e:
            logger.exception("Error updating business license %s: %s", business_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating business license'
            }, status=400)

    # Order stats
    order_stats = orders_models.Order.objects.filter(business=business).aggregate(
        total=Count('id'),
        delivered=Count('id', filter=Q(order_status='delivered')),
        pending=Count('id', filter=Q(order_status__in=['to_review', 'ready_to_pickup', 'publish'])),
    )

    # Get linked warehouse (if any)
    from warehouse import models as warehouse_models
    from product import models as product_models
    from django.db.models import Sum

    linked_warehouse = None
    if business.fulfillment_service_status == 'active':
        try:
            linked_warehouse = warehouse_models.SellerWarehouseLink.objects.select_related(
                'warehouse', 'default_location'
            ).filter(business=business, is_active=True).first()
        except Exception as e:
            logger.warning(f"Error fetching linked warehouse for business {business.business_id}: {e}")

    # Get product statistics for fulfillment businesses
    product_stats = {
        'total_products': 0,
        'active_skus': 0,
        'total_stock': 0,
        'low_stock_items': 0,
    }
    if business.fulfillment_service_status == 'active':
        try:
            products = product_models.Product.objects.filter(business=business)
            product_stats['total_products'] = products.count()
            product_stats['active_skus'] = products.count()  # All products are considered active SKUs

            # Get total stock from inventory
            total_stock_result = product_models.ProductInventory.objects.filter(
                item_sku__business=business
            ).aggregate(total=Sum('item_quantity'))
            product_stats['total_stock'] = total_stock_result['total'] or 0

            # Count low stock items (less than 10 units)
            for product in products:
                stock = product_models.ProductInventory.objects.filter(
                    item_sku=product
                ).aggregate(total=Sum('item_quantity'))['total'] or 0
                if stock < 10:
                    product_stats['low_stock_items'] += 1
        except Exception as e:
            logger.warning(f"Error fetching product stats for business {business.business_id}: {e}")

    # Get all products for this business (regardless of fulfillment status)
    products_list = product_models.Product.objects.filter(
        business=business
    ).order_by('item_sku')

    # Update product_stats for all businesses (not just fulfillment)
    if not product_stats['total_products'] and products_list.exists():
        product_stats['total_products'] = products_list.count()
        product_stats['active_skus'] = products_list.count()

    context = {
        'page_title': f'{business.business_name} - Business Detail',
        'business': business,
        'business_profile': business_profile,
        'order_stats': order_stats,
        'linked_warehouse': linked_warehouse,
        'product_stats': product_stats,
        'products_list': products_list,
    }

    return render(request, 'workforce/business_license_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def api_warehouse_locations(request):
    """API endpoint to get all active warehouse locations"""
    from warehouse import models as warehouse_models

    try:
        locations = warehouse_models.WarehouseLocation.objects.select_related(
            'warehouse'
        ).filter(
            is_active=True
        ).order_by('warehouse__name', 'name')

        location_list = [{
            'id': loc.id,
            'warehouse_name': loc.warehouse.name,
            'name': loc.name,
            'code': loc.code,
            'warehouse_code': loc.warehouse.code,
            'address': loc.address or '',
            'city': loc.warehouse.city or '',
        } for loc in locations]

        return JsonResponse({'success': True, 'locations': location_list})
    except Exception as e:
        logger.exception("Error fetching warehouse locations: %s", str(e))
        return JsonResponse({'success': False, 'error': 'Failed to load warehouse locations'}, status=500)


# Sellers section  ------------------------------------------------------------------------------------------------------

@login_required(login_url='/accounts/login/')
@staff_required
def wf_seller_api_configs(request):
    """
    Staff view: list all businesses' API configurations.
    Allows staff to test and approve each API config.
    """
    from django.db.models import Count

    # All API configs across all businesses
    api_configs = business_models.BusinessApiSettings.objects.select_related(
        'business'
    ).order_by('-is_verify_api', 'business__business_name', 'api_type')

    # Optional filter by business
    selected_business_id = request.GET.get('business', '')
    selected_business = None
    if selected_business_id:
        try:
            selected_business = business_models.Business.objects.get(business_id=selected_business_id)
            api_configs = api_configs.filter(business=selected_business)
        except business_models.Business.DoesNotExist:
            pass

    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter == 'approved':
        api_configs = api_configs.filter(is_verify_api=True)
    elif status_filter == 'pending':
        api_configs = api_configs.filter(is_verify_api=False)

    businesses_with_api = business_models.Business.objects.filter(
        business_settings_api__isnull=False
    ).distinct().order_by('business_name')

    context = {
        'api_configs': api_configs,
        'businesses_with_api': businesses_with_api,
        'selected_business': selected_business,
        'selected_business_id': selected_business_id,
        'status_filter': status_filter,
        'total_count': api_configs.count(),
        'approved_count': business_models.BusinessApiSettings.objects.filter(is_verify_api=True).count(),
        'pending_count': business_models.BusinessApiSettings.objects.filter(is_verify_api=False).count(),
    }
    return render(request, 'workforce/seller_api_configs.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def wf_approve_api_config(request, api_id):
    """
    Staff action: toggle approval (is_verify_api) for a BusinessApiSettings.
    POST only.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        api = business_models.BusinessApiSettings.objects.select_related('business').get(pk=api_id)
        api.is_verify_api = not api.is_verify_api
        api.save(update_fields=['is_verify_api', 'updated_at'])
        return JsonResponse({
            'success': True,
            'is_approved': api.is_verify_api,
            'label': 'Approved' if api.is_verify_api else 'Pending',
        })
    except business_models.BusinessApiSettings.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Not found'}, status=404)
    except Exception as e:
        logger.exception('wf_approve_api_config error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def wf_save_google_sheet(request):
    """
    POST: Save a Google Sheet URL for a business as a BusinessApiSettings record
    with api_type='google_sheet'.
    Params: business_id, sheet_url, sheet_name (optional label)
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    business_id = request.POST.get('business_id', '').strip()
    sheet_url = request.POST.get('sheet_url', '').strip()
    sheet_name = request.POST.get('sheet_name', '').strip() or 'Google Sheet'

    if not business_id or not sheet_url:
        return JsonResponse({'success': False, 'error': 'business_id and sheet_url required'}, status=400)

    try:
        business = business_models.Business.objects.get(business_id=business_id)
    except business_models.Business.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Business not found'}, status=404)

    try:
        api, created = business_models.BusinessApiSettings.objects.update_or_create(
            business=business,
            api_type='google_sheet',
            defaults={
                'google_sheet_url': sheet_url,
                'order_api_endpoint': sheet_name,
                'is_verify_api': True,
            }
        )
        return JsonResponse({
            'success': True,
            'created': created,
            'api_id': api.id,
            'business_name': business.business_name,
        })
    except Exception as e:
        logger.exception('wf_save_google_sheet error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def google_sheets_auth_start(request):
    """Start Google Sheets OAuth2 flow — redirects to Google consent screen."""
    from google_auth_oauthlib.flow import Flow
    from django.conf import settings as django_settings
    from pathlib import Path

    client_file = Path(django_settings.BASE_DIR) / 'google_sheets_client.json'
    if not client_file.exists():
        return HttpResponse('google_sheets_client.json not found on server.', status=500)

    redirect_uri = request.build_absolute_uri('/workforce/google-sheets/auth/callback/')
    flow = Flow.from_client_secrets_file(
        str(client_file),
        scopes=[
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.readonly',
        ],
        redirect_uri=redirect_uri,
    )
    auth_url, state = flow.authorization_url(access_type='offline', prompt='consent')
    request.session['google_sheets_oauth_state'] = state
    request.session['google_sheets_code_verifier'] = flow.code_verifier
    return redirect(auth_url)


@login_required(login_url='/accounts/login/')
@staff_required
def google_sheets_auth_callback(request):
    """Handle Google OAuth2 callback — save token file."""
    from google_auth_oauthlib.flow import Flow
    from django.conf import settings as django_settings
    from pathlib import Path
    import json as _json

    client_file = Path(django_settings.BASE_DIR) / 'google_sheets_client.json'
    token_path = Path(django_settings.BASE_DIR) / getattr(
        django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
    )

    redirect_uri = request.build_absolute_uri('/workforce/google-sheets/auth/callback/')
    flow = Flow.from_client_secrets_file(
        str(client_file),
        scopes=[
            'https://www.googleapis.com/auth/spreadsheets.readonly',
            'https://www.googleapis.com/auth/drive.readonly',
        ],
        redirect_uri=redirect_uri,
    )
    flow.code_verifier = request.session.pop('google_sheets_code_verifier', None)

    try:
        flow.fetch_token(authorization_response=request.build_absolute_uri())
    except Exception as e:
        return HttpResponse(f'OAuth error: {e}', status=400)

    creds = flow.credentials
    token_data = {
        'access_token': creds.token,
        'refresh_token': creds.refresh_token,
        'token_uri': creds.token_uri,
        'client_id': creds.client_id,
        'client_secret': creds.client_secret,
        'scope': ' '.join(creds.scopes or []),
    }
    token_path.write_text(_json.dumps(token_data, indent=2))
    return HttpResponse(
        '<h2 style="font-family:sans-serif;margin:40px;">Google Sheets authorized successfully!</h2>'
        '<p style="font-family:sans-serif;margin:40px;">Token saved. You can now test Google Sheet connections.</p>'
        '<a href="/workforce/sellers/api-configs/" style="font-family:sans-serif;margin:40px;display:inline-block;">Back to API Configs</a>'
    )


@login_required(login_url='/accounts/login/')
@staff_required
def wf_test_api_config(request, api_id):
    """Show the test console page with config panel and Test Connection button."""
    try:
        api = business_models.BusinessApiSettings.objects.select_related('business').get(pk=api_id)
    except business_models.BusinessApiSettings.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'API config not found'}, status=404)

    context = {
        'business': api.business,
        'api': api,
    }
    return render(request, 'workforce/seller_api_test.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def wf_test_api_config_result(request, api_id):
    """
    Staff proxy for testing a BusinessApiSettings (any business).
    Fetches orders + products and returns an HTMX partial with results.
    """
    try:
        api = business_models.BusinessApiSettings.objects.select_related('business').get(pk=api_id)
    except business_models.BusinessApiSettings.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'API config not found'}, status=404)

    business = api.business
    from django.utils import timezone as dj_timezone
    update_time = dj_timezone.localtime().strftime('%Y-%m-%d  Time : %H:%M:%S')

    order_response = None
    product_response = None
    order_status_code = None
    product_status_code = None
    error_message = None
    # api_stats: platform-specific summary shown in the status line
    api_stats = {}

    try:
        if api.api_type == 'shopify':
            import shopify
            shop_name = (api.site_api_url or '').replace('https://', '').replace('http://', '').replace('.myshopify.com', '').strip()
            session = shopify.Session(shop_name, api.api_version or '2023-10', api.api_access_token)
            shopify.ShopifyResource.activate_session(session)

            # Fetch orders (last 5 for preview)
            orders = shopify.Order.find(limit=5, status='any')
            order_response = [{'id': o.id, 'name': o.name, 'financial_status': o.financial_status, 'fulfillment_status': o.fulfillment_status, 'total_price': o.total_price, 'created_at': o.created_at} for o in orders]
            order_status_code = 200

            # Order count
            try:
                order_count_obj = shopify.Order.count(status='any')
                api_stats['order_count'] = order_count_obj
            except Exception:
                api_stats['order_count'] = None

            # Product count
            try:
                product_count_obj = shopify.Product.count()
                api_stats['product_count'] = product_count_obj
                products = shopify.Product.find(limit=5)
                product_response = [{'id': p.id, 'title': p.title, 'status': p.status, 'variants': len(p.variants)} for p in products]
                product_status_code = 200
            except Exception as pe:
                api_stats['product_count'] = None
                product_status_code = 500

            # Shop info
            try:
                shop = shopify.Shop.current()
                api_stats['shop_name'] = shop.name
                api_stats['shop_email'] = shop.email
                api_stats['plan'] = shop.plan_display_name
                api_stats['currency'] = shop.currency
            except Exception:
                pass

            shopify.ShopifyResource.clear_session()

        elif api.api_type == 'woocommerce':
            from woocommerce import API as WooAPI
            wcapi = WooAPI(
                url=api.site_api_url or '',
                consumer_key=api.api_key or '',
                consumer_secret=api.api_secret or '',
                version='wc/v3',
                timeout=10,
            )

            # Fetch orders (last 5 for preview)
            r = wcapi.get('orders', params={'per_page': 5})
            order_status_code = r.status_code
            if r.status_code == 200:
                orders_data = r.json()
                order_response = [{'id': o.get('id'), 'number': o.get('number'), 'status': o.get('status'), 'total': o.get('total'), 'currency': o.get('currency'), 'date_created': o.get('date_created')} for o in orders_data]
                # Total order count from response headers (WooCommerce sends X-WP-Total)
                api_stats['order_count'] = r.headers.get('X-WP-Total')
            else:
                order_response = {'error': r.text}

            # Product count + preview
            try:
                rp = wcapi.get('products', params={'per_page': 5})
                product_status_code = rp.status_code
                if rp.status_code == 200:
                    products_data = rp.json()
                    product_response = [{'id': p.get('id'), 'name': p.get('name'), 'status': p.get('status'), 'stock_status': p.get('stock_status'), 'price': p.get('price')} for p in products_data]
                    api_stats['product_count'] = rp.headers.get('X-WP-Total')
            except Exception:
                product_status_code = 500

            # WC system status for extra info
            try:
                rs = wcapi.get('system_status')
                if rs.status_code == 200:
                    ss = rs.json()
                    env = ss.get('environment', {})
                    api_stats['wc_version'] = env.get('version')
                    api_stats['wp_version'] = env.get('wp_version')
                    api_stats['currency'] = ss.get('settings', {}).get('currency')
            except Exception:
                pass

        elif api.api_type == 'google_sheet':
            import re as _re
            import gspread
            from google.oauth2.credentials import Credentials as _GCreds
            from google.auth.transport.requests import Request as _GReq
            from django.conf import settings as django_settings
            from pathlib import Path as _Path
            import json as _json

            sheet_url = api.google_sheet_url or api.site_api_url or ''
            if not sheet_url:
                raise Exception('No Google Sheet URL configured for this integration.')

            token_path = _Path(django_settings.BASE_DIR) / getattr(
                django_settings, 'GOOGLE_SHEETS_TOKEN_FILE', 'google_sheets_token.json'
            )
            if not token_path.exists():
                raise Exception(
                    'Google Sheets not authorized yet. '
                    'Run: python google_sheets_auth.py (one-time setup)'
                )

            _td = _json.loads(token_path.read_text())
            creds = _GCreds(
                token=_td.get('access_token'),
                refresh_token=_td.get('refresh_token'),
                token_uri=_td.get('token_uri', 'https://oauth2.googleapis.com/token'),
                client_id=_td.get('client_id'),
                client_secret=_td.get('client_secret'),
                scopes=_td.get('scope', '').split(),
            )
            if creds.expired and creds.refresh_token:
                creds.refresh(_GReq())
                _td['access_token'] = creds.token
                token_path.write_text(_json.dumps(_td, indent=2))

            gc = gspread.authorize(creds)

            match = _re.search(r'/spreadsheets/d/([^/]+)', sheet_url)
            if not match:
                raise Exception('Invalid Google Sheet URL.')
            sheet_id = match.group(1)
            gid_match = _re.search(r'gid=(\d+)', sheet_url)
            gid = int(gid_match.group(1)) if gid_match else 0

            spreadsheet = gc.open_by_key(sheet_id)
            worksheet = None
            for ws in spreadsheet.worksheets():
                if ws.id == gid:
                    worksheet = ws
                    break
            if worksheet is None:
                worksheet = spreadsheet.sheet1

            all_values = worksheet.get_all_values()

            # If selected worksheet is empty, try to find one with data
            if len(all_values) <= 1:
                for ws in spreadsheet.worksheets():
                    if ws.id == worksheet.id:
                        continue
                    candidate = ws.get_all_values()
                    if len(candidate) > 1:
                        worksheet = ws
                        all_values = candidate
                        break

            if not all_values or len(all_values) <= 1:
                raise Exception('Sheet is empty — no data rows found.')

            headers = all_values[0]  # first row = headers
            data_rows = all_values[1:]  # remaining rows = data
            total_rows = len(data_rows)
            order_status_code = 200

            api_stats['order_count'] = total_rows
            api_stats['sheet_title'] = spreadsheet.title
            api_stats['worksheet_name'] = worksheet.title
            api_stats['column_count'] = len(headers)

            # Build preview of last 5 rows
            preview_start = max(0, total_rows - 5)
            preview_rows = data_rows[preview_start:]
            order_response = []
            for i, row in enumerate(preview_rows):
                row_data = {}
                for j, h in enumerate(headers):
                    val = row[j] if j < len(row) else ''
                    row_data[h or f'Col {j+1}'] = str(val)[:80]
                order_response.append({
                    'row_num': preview_start + i + 2,  # +2 for header row + 0-index
                    'data': row_data,
                })

        else:
            error_message = f"Test not implemented for api_type={api.api_type}"
            order_status_code = 0

    except Exception as e:
        error_message = str(e)
        order_status_code = 500

    context = {
        'business': business,
        'api': api,
        'order_response': order_response,
        'product_response': product_response,
        'order_status_code': order_status_code,
        'product_status_code': product_status_code,
        'api_stats': api_stats,
        'update_time': update_time,
        'error_message': error_message,
    }
    return render(request, 'workforce/seller_api_test_result.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def sellers_list(request):
    """
    View to display all sellers/businesses in the system
    """
    # Get all businesses
    businesses = business_models.Business.objects.select_related(
        'profile', 'business_profile'
    ).all()

    # Apply filters
    search = request.GET.get('search', '').strip()
    status = request.GET.get('status', '').strip()
    verification_status = request.GET.get('verification', '').strip()

    if search:
        from django.db.models import Q
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search) |
            Q(profile__user__email__icontains=search)
        )

    if status:
        businesses = businesses.filter(business_status=status)

    if verification_status:
        businesses = businesses.filter(profile__verification_status=verification_status)

    # Order by most recent
    businesses = businesses.order_by('-business_since', '-business_id')

    # Count statistics using single aggregation query
    from django.db.models import Count, Q as DQ
    stats = business_models.Business.objects.aggregate(
        total=Count('business_id'),
        active=Count('business_id', filter=DQ(business_status='active')),
        pending=Count('business_id', filter=DQ(profile__verification_status='pending')),
        inactive=Count('business_id', filter=DQ(business_status='inactive')),
    )
    total_sellers = stats['total']
    active_sellers = stats['active']
    pending_sellers = stats['pending']
    inactive_sellers = stats['inactive']

    # Paginate
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'All Sellers',
        'page_obj': page_obj,
        'total_sellers': total_sellers,
        'active_sellers': active_sellers,
        'pending_sellers': pending_sellers,
        'inactive_sellers': inactive_sellers,
        'search': search,
        'status': status,
        'verification': verification_status,
    }

    return render(request, 'workforce/sellers_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def sellers_pending(request):
    """
    View to display sellers pending approval
    """
    # Get businesses with pending verification
    businesses = business_models.Business.objects.select_related(
        'profile', 'business_profile'
    ).filter(
        profile__verification_status='pending'
    ).order_by('-business_since', '-business_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search)
        )

    # Paginate
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Pending Sellers',
        'page_obj': page_obj,
        'search': search,
    }

    return render(request, 'workforce/sellers_pending.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def sellers_active(request):
    """
    View to display active verified sellers
    """
    from django.db.models import Count, Q

    # Get active and verified businesses with order count annotation
    businesses = business_models.Business.objects.select_related(
        'profile', 'business_profile'
    ).annotate(
        total_orders=Count('order')
    ).filter(
        business_status='active',
        profile__verification_status='verified'
    ).order_by('-business_since', '-business_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search)
        )

    # Paginate
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Active Sellers',
        'page_obj': page_obj,
        'search': search,
    }

    return render(request, 'workforce/sellers_active.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def sellers_inactive(request):
    """
    View to display inactive or suspended sellers
    """
    # Get inactive businesses
    businesses = business_models.Business.objects.select_related(
        'profile', 'business_profile'
    ).filter(
        business_status='inactive'
    ).order_by('-business_since', '-business_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        businesses = businesses.filter(
            Q(business_name__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search)
        )

    # Paginate
    page_obj = paginate_queryset(request, businesses, items_per_page=20)

    context = {
        'page_title': 'Inactive Sellers',
        'page_obj': page_obj,
        'search': search,
    }

    return render(request, 'workforce/sellers_inactive.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def seller_detail(request, business_id):
    """
    View for viewing comprehensive seller/business details including
    teams, API settings, pickup locations, and documents.
    """
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta

    # Fetch business with all related data
    business = get_object_or_404(
        business_models.Business.objects.select_related(
            'user', 'profile'
        ).prefetch_related(
            'team_members__user',
            'business_settings_api',
            'pickup_location',
        ),
        business_id=business_id
    )

    # Get business profile (handle missing OneToOne relation)
    try:
        business_profile = business.business_profile
    except Exception:
        business_profile = None

    # Get team members
    team_members = business.team_members.select_related('user', 'profile').all()

    # Get API settings
    api_settings = business.business_settings_api.all()

    # Get pickup locations
    pickup_locations = business.pickup_location.all()

    # Get comprehensive order statistics
    order_stats = orders_models.Order.objects.filter(business=business).aggregate(
        total_orders=Count('id'),
        to_review_orders=Count('id', filter=Q(order_status='to_review')),
        ready_orders=Count('id', filter=Q(order_status='ready_to_pickup')),
        active_orders=Count('id', filter=Q(order_status='publish')),
        delivered_orders=Count('id', filter=Q(order_status='delivered')),
        cancelled_orders=Count('id', filter=Q(order_status='cancelled')),
    )

    # Get COD statistics
    cod_stats = orders_models.Order.objects.filter(
        business=business,
        cod_amount__gt=0
    ).aggregate(
        total_cod_orders=Count('id'),
        total_cod_amount=Sum('cod_amount'),
        collected_cod=Sum('cod_amount', filter=Q(order_status='delivered')),
        pending_cod=Sum('cod_amount', filter=~Q(order_status__in=['delivered', 'cancelled'])),
    )

    # Calculate delivery success rate
    total_completed = (order_stats.get('delivered_orders') or 0) + (order_stats.get('cancelled_orders') or 0)
    delivery_success_rate = 0
    if total_completed > 0:
        delivery_success_rate = round((order_stats.get('delivered_orders') or 0) / total_completed * 100, 1)

    # Get recent orders count (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_orders_count = orders_models.Order.objects.filter(
        business=business,
        created_at__gte=thirty_days_ago
    ).count()

    # Get last 7 days orders for trend
    seven_days_ago = timezone.now() - timedelta(days=7)
    last_7_days_orders = orders_models.Order.objects.filter(
        business=business,
        created_at__gte=seven_days_ago
    ).count()

    # Get recent orders for activity timeline (last 10)
    # Use .only() to avoid missing columns (delivered_at, fulfilled_at not migrated)
    recent_orders = orders_models.Order.objects.filter(
        business=business
    ).only(
        'id', 'order_number', 'client_order_code', 'order_status',
        'cod_amount', 'customer_name', 'customer_phone', 'created_at'
    ).order_by('-created_at')[:10]

    # Get delivery task statistics
    delivery_stats = delivery_models.DeliveryTask.objects.filter(
        business=business
    ).aggregate(
        total_tasks=Count('id'),
        in_transit=Count('id', filter=Q(dl_task_status='in_transit')),
        completed=Count('id', filter=Q(dl_task_status='delivered')),
    )

    # Calculate average orders per month (if business_since exists)
    avg_orders_per_month = 0
    if business.business_since:
        months_active = max(1, (timezone.now().date() - business.business_since).days / 30)
        avg_orders_per_month = round((order_stats.get('total_orders') or 0) / months_active, 1)

    # Handle POST for status update
    if request.method == 'POST':
        try:
            # Basic Information
            business.business_code = request.POST.get('business_code', business.business_code) or None
            business.business_name = request.POST.get('business_name', business.business_name)
            business.business_qid = request.POST.get('business_qid', business.business_qid) or None
            business.business_product_category = request.POST.get('business_product_category', business.business_product_category) or None
            business.business_bio = request.POST.get('business_bio', business.business_bio) or None
            business.business_status = request.POST.get('business_status', business.business_status)

            # Handle business_since date
            business_since_str = request.POST.get('business_since', '')
            if business_since_str:
                from datetime import datetime
                try:
                    business.business_since = datetime.strptime(business_since_str, '%Y-%m-%d').date()
                except ValueError:
                    pass  # Keep existing value if parsing fails
            else:
                business.business_since = None

            # Contact Information
            business.business_phone = request.POST.get('business_phone', business.business_phone) or None
            business.business_whatsapp = request.POST.get('business_whatsapp', business.business_whatsapp) or None
            business.business_email = request.POST.get('business_email', business.business_email) or None

            # Social Media
            business.business_instagram = request.POST.get('business_instagram', business.business_instagram) or None
            business.business_facebook_page = request.POST.get('business_facebook_page', business.business_facebook_page) or None

            # Handle fulfillment service toggle
            fulfillment_enabled = request.POST.get('fulfillment_service_enabled') == 'on'
            if fulfillment_enabled and not business.fulfillment_service_enabled:
                business.fulfillment_service_enabled = True
                business.fulfillment_activated_at = timezone.now()
            elif not fulfillment_enabled:
                business.fulfillment_service_enabled = False

            business.save()

            # Update business profile if exists, create if not
            if not business_profile:
                business_profile = business_models.BusinessProfile.objects.create(business=business)

            business_profile.business_address = request.POST.get('business_address', business_profile.business_address) or None
            business_profile.business_city = request.POST.get('business_city', business_profile.business_city) or None
            business_profile.business_state = request.POST.get('business_state', business_profile.business_state) or None
            business_profile.business_zip_code = request.POST.get('business_zip_code', business_profile.business_zip_code) or None
            business_profile.business_country = request.POST.get('business_country', business_profile.business_country) or 'Qatar'
            business_profile.business_website = request.POST.get('business_website', business_profile.business_website) or None
            business_profile.save()

            return JsonResponse({
                'success': True,
                'message': 'Seller updated successfully'
            })
        except Exception as e:
            logger.exception("Error updating seller %s: %s", business_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating seller'
            }, status=400)

    # Build documents list from business profile fields (if available)
    documents = []
    # Always show QID row (editable even if empty)
    documents.append({
        'document_type': 'QID / CR',
        'document_no': business.business_qid or '—',
        'document_file': None,
        'document_expiry_date': None,
        'field_name': 'business_qid',
        'field_type': 'text',
    })
    # Add business logo row — logo is on BusinessLogo related model
    logo_obj = business.business_logo.first()
    logo_file = None
    has_real_file = False
    if logo_obj and logo_obj.business_logo:
        try:
            has_real_file = (
                bool(logo_obj.business_logo.name)
                and logo_obj.business_logo.name != 'business/avatar.png'
                and logo_obj.business_logo.size > 0
            )
        except Exception:
            has_real_file = False
        if has_real_file:
            logo_file = logo_obj.business_logo
    documents.append({
        'document_type': 'Business Logo',
        'document_no': 'Custom Logo' if has_real_file else 'Default Avatar',
        'document_file': logo_file,
        'document_expiry_date': None,
        'field_name': 'business_logo',
        'field_type': 'file',
        'current_logo': logo_obj.business_logo if logo_obj else None,
    })

    # Calculate business registration completion percentage
    required_fields = ['business_name', 'business_phone', 'business_whatsapp',
                       'business_email', 'business_product_category', 'business_qid']
    filled = sum(1 for f in required_fields if getattr(business, f, None))
    completion_percentage = round(filled / len(required_fields) * 100)

    # Get user profile
    try:
        user_profile = business.user.profile if business.user else None
    except Exception:
        user_profile = None

    # User profile completion
    user_required_fields = ['first_name', 'last_name', 'email', 'phone']
    user_filled = 0
    if user_profile:
        user_filled = sum(1 for f in user_required_fields if getattr(user_profile, f, None))
    user_completion = round(user_filled / len(user_required_fields) * 100) if user_required_fields else 0

    # Get products for this business
    products_list = product_models.Product.objects.filter(
        business=business
    ).select_related('color', 'unit', 'product_category').order_by('item_sku')

    context = {
        'page_title': f'Seller: {business.business_name}',
        'business': business,
        'business_profile': business_profile,
        'team_members': team_members,
        'api_settings': api_settings,
        'pickup_locations': pickup_locations,
        'order_stats': order_stats,
        'cod_stats': cod_stats,
        'delivery_stats': delivery_stats,
        'delivery_success_rate': delivery_success_rate,
        'recent_orders_count': recent_orders_count,
        'last_7_days_orders': last_7_days_orders,
        'recent_orders': recent_orders,
        'avg_orders_per_month': avg_orders_per_month,
        'documents': documents,
        'products_list': products_list,
        'now': timezone.now(),
        'completion_percentage': completion_percentage,
        'user_profile': user_profile,
        'user_completion': user_completion,
    }

    return render(request, 'workforce/seller_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def seller_doc_field_update(request, business_id):
    """Update or clear a single document field (business_qid or business_logo) for a seller."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)

    business = get_object_or_404(business_models.Business, business_id=business_id)
    field = request.POST.get('field', '')
    action = request.POST.get('action', 'update')

    ALLOWED_FIELDS = {'business_qid', 'business_logo'}
    if field not in ALLOWED_FIELDS:
        return JsonResponse({'success': False, 'error': 'Invalid field'}, status=400)

    try:
        if field == 'business_logo':
            # Logo lives on BusinessLogo related model; upload_path_handler needs instance.path
            # so we save the file manually to avoid that crash.
            import os
            from django.core.files.storage import default_storage
            from django.core.files.base import ContentFile

            logo_obj = business.business_logo.first()

            if action == 'clear':
                if logo_obj and logo_obj.business_logo and logo_obj.business_logo.name != 'business/avatar.png':
                    try:
                        default_storage.delete(logo_obj.business_logo.name)
                    except Exception:
                        pass
                    logo_obj.business_logo = 'business/avatar.png'
                    logo_obj.save(update_fields=['business_logo'])
                return JsonResponse({'success': True, 'message': 'Logo cleared'})

            # update — save file manually to a fixed upload path
            file = request.FILES.get('value')
            if not file:
                return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)

            ext = os.path.splitext(file.name)[1].lower()
            upload_path = f'business/logos/{business_id}/logo{ext}'
            # Delete old file if it's not the default
            if logo_obj and logo_obj.business_logo and logo_obj.business_logo.name != 'business/avatar.png':
                try:
                    default_storage.delete(logo_obj.business_logo.name)
                except Exception:
                    pass
            file.seek(0)
            saved_path = default_storage.save(upload_path, ContentFile(file.read()))

            if logo_obj:
                logo_obj.business_logo = saved_path
                logo_obj.save(update_fields=['business_logo'])
            else:
                business_models.BusinessLogo.objects.create(business=business, business_logo=saved_path)
            return JsonResponse({'success': True, 'message': 'Logo updated successfully'})

        # Text fields on Business model directly
        if action == 'clear':
            setattr(business, field, None)
            business.save(update_fields=[field])
            return JsonResponse({'success': True, 'message': 'Field cleared'})

        value = request.POST.get('value', '').strip() or None
        setattr(business, field, value)
        business.save(update_fields=[field])
        return JsonResponse({'success': True, 'message': 'Updated successfully'})

    except Exception as e:
        logger.exception("Error updating doc field %s for business %s: %s", field, business_id, str(e))
        return JsonResponse({'success': False, 'error': 'Update failed'}, status=500)


# =============================================================================
# DRIVER MANAGEMENT VIEWS
# =============================================================================

@login_required(login_url='/accounts/login/')
@staff_required
def drivers_list(request):
    """
    View to display all drivers with search and filtering
    """
    drivers = fleet_models.Driver.objects.select_related(
        'user', 'profile'
    ).prefetch_related(
        'driver_vehicle', 'driver_document'
    ).order_by('-driver_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(driver_code__icontains=search) |
            Q(driver_phone__icontains=search)
        )

    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        drivers = drivers.filter(driver_status=status_filter)

    # Pagination
    page_obj = paginate_queryset(request, drivers, items_per_page=20)

    # Get counts for stats
    total_count = fleet_models.Driver.objects.count()
    active_count = fleet_models.Driver.objects.filter(driver_status='approved').count()
    pending_count = fleet_models.Driver.objects.filter(driver_status__in=['pending', 'processing']).count()
    inactive_count = fleet_models.Driver.objects.filter(driver_status__in=['rejected', 'blocked', 'suspended']).count()

    context = {
        'page_title': 'All Drivers',
        'page_obj': page_obj,
        'search': search,
        'status_filter': status_filter,
        'total_count': total_count,
        'active_count': active_count,
        'pending_count': pending_count,
        'inactive_count': inactive_count,
    }

    return render(request, 'workforce/drivers_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def drivers_pending(request):
    """
    View to display drivers pending approval
    """
    drivers = fleet_models.Driver.objects.select_related(
        'user', 'profile'
    ).prefetch_related(
        'driver_vehicle', 'driver_document'
    ).filter(
        driver_status__in=['pending', 'processing']
    ).order_by('-driver_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(driver_code__icontains=search) |
            Q(driver_phone__icontains=search)
        )

    # Pagination
    page_obj = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'page_title': 'Pending Drivers',
        'page_obj': page_obj,
        'search': search,
        'status_type': 'pending',
    }

    return render(request, 'workforce/drivers_pending.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def drivers_active(request):
    """
    View to display active (approved) drivers
    """
    drivers = fleet_models.Driver.objects.select_related(
        'user', 'profile'
    ).prefetch_related(
        'driver_vehicle', 'driver_document'
    ).filter(
        driver_status='approved'
    ).order_by('-driver_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(driver_code__icontains=search) |
            Q(driver_phone__icontains=search)
        )

    # Pagination
    page_obj = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'page_title': 'Active Drivers',
        'page_obj': page_obj,
        'search': search,
        'status_type': 'active',
    }

    return render(request, 'workforce/drivers_active.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def drivers_inactive(request):
    """
    View to display inactive, rejected, or blocked drivers
    """
    drivers = fleet_models.Driver.objects.select_related(
        'user', 'profile'
    ).prefetch_related(
        'driver_vehicle', 'driver_document'
    ).filter(
        driver_status__in=['rejected', 'blocked', 'suspended']
    ).order_by('-driver_id')

    # Apply search filter
    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q
        drivers = drivers.filter(
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(driver_code__icontains=search) |
            Q(driver_phone__icontains=search)
        )

    # Pagination
    page_obj = paginate_queryset(request, drivers, items_per_page=20)

    context = {
        'page_title': 'Inactive Drivers',
        'page_obj': page_obj,
        'search': search,
        'status_type': 'inactive',
    }

    return render(request, 'workforce/drivers_inactive.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def driver_detail(request, driver_id):
    """
    View for viewing comprehensive driver details including
    documents, vehicles, transactions, and delivery stats.
    """
    from django.db.models import Count, Sum, Q
    from django.utils import timezone
    from datetime import timedelta

    # Fetch driver with all related data
    driver = get_object_or_404(
        fleet_models.Driver.objects.select_related(
            'user', 'profile'
        ).prefetch_related(
            'driver_vehicle',
            'driver_document',
        ),
        driver_id=driver_id
    )

    # Get vehicles
    vehicles = driver.driver_vehicle.all()

    # Get documents and check if files actually exist
    raw_documents = driver.driver_document.all()
    documents = []
    for doc in raw_documents:
        doc_dict = {
            'id': doc.id,
            'document_type': doc.document_type,
            'document_no': doc.document_no,
            'document_issued_from': getattr(doc, 'document_issued_from', ''),
            'document_expiry_date': getattr(doc, 'document_expiry_date', None),
            'document_file': None,
            'document_file_back': None,
        }
        if doc.document_file and doc.document_file.name:
            try:
                if doc.document_file.storage.exists(doc.document_file.name):
                    doc_dict['document_file'] = doc.document_file
            except Exception:
                pass
        if hasattr(doc, 'document_file_back') and doc.document_file_back and doc.document_file_back.name:
            try:
                if doc.document_file_back.storage.exists(doc.document_file_back.name):
                    doc_dict['document_file_back'] = doc.document_file_back
            except Exception:
                pass
        documents.append(doc_dict)

    # Get delivery task statistics
    delivery_stats = delivery_models.DeliveryTask.objects.filter(
        driver=driver
    ).aggregate(
        total_tasks=Count('id'),
        completed_tasks=Count('id', filter=Q(dl_task_status='delivered')),
        in_transit=Count('id', filter=Q(dl_task_status='in_transit')),
        failed_tasks=Count('id', filter=Q(dl_task_status__in=['failed', 'cancelled'])),
    )

    # Calculate success rate
    total_completed = (delivery_stats.get('completed_tasks') or 0) + (delivery_stats.get('failed_tasks') or 0)
    success_rate = 0
    if total_completed > 0:
        success_rate = round((delivery_stats.get('completed_tasks') or 0) / total_completed * 100, 1)

    # Get recent tasks count (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_tasks_count = delivery_models.DeliveryTask.objects.filter(
        driver=driver,
        created_at__gte=thirty_days_ago
    ).count()

    # Get last 7 days tasks for trend
    seven_days_ago = timezone.now() - timedelta(days=7)
    last_7_days_tasks = delivery_models.DeliveryTask.objects.filter(
        driver=driver,
        created_at__gte=seven_days_ago
    ).count()

    # Get recent tasks for activity timeline (last 10)
    recent_tasks = delivery_models.DeliveryTask.objects.filter(
        driver=driver
    ).select_related('business').order_by('-created_at')[:10]

    # Get COD statistics
    cod_stats = delivery_models.DeliveryTask.objects.filter(
        driver=driver,
        order__cod_amount__gt=0
    ).aggregate(
        total_cod_tasks=Count('id'),
        total_cod_amount=Sum('order__cod_amount'),
        collected_cod=Sum('order__cod_amount', filter=Q(dl_task_status='delivered')),
    )

    # Get recent transactions (last 10)
    recent_transactions = fleet_models.DriverTransaction.objects.filter(
        driver=driver
    ).order_by('-created_at')[:10]

    from core.decorators import is_superadmin as check_superadmin
    user_is_superadmin = check_superadmin(request.user)

    # Handle POST request for updates
    if request.method == 'POST':
        try:
            driver.driver_phone = request.POST.get('driver_phone', driver.driver_phone)
            driver.driver_whatsapp = request.POST.get('driver_whatsapp', driver.driver_whatsapp)
            driver.driver_bio = request.POST.get('driver_bio', driver.driver_bio)
            # Update wallet limit if provided
            credit_limit = request.POST.get('credit_limit')
            if credit_limit:
                from decimal import Decimal
                driver.credit_limit = Decimal(credit_limit)

            # Status change — superadmin only
            new_status = request.POST.get('driver_status')
            if new_status and new_status != driver.driver_status:
                if user_is_superadmin:
                    driver.driver_status = new_status
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Only super admins can change driver status.'
                    }, status=403)

            driver.save()
            return JsonResponse({
                'success': True,
                'message': 'Driver updated successfully'
            })
        except Exception as e:
            logger.exception("Error updating driver %s: %s", driver_id, str(e))
            return JsonResponse({
                'success': False,
                'error': 'An error occurred while updating driver'
            }, status=400)

    context = {
        'page_title': f'Driver: {driver.user.first_name} {driver.user.last_name}',
        'driver': driver,
        'vehicles': vehicles,
        'documents': documents,
        'delivery_stats': delivery_stats,
        'success_rate': success_rate,
        'recent_tasks_count': recent_tasks_count,
        'last_7_days_tasks': last_7_days_tasks,
        'recent_tasks': recent_tasks,
        'cod_stats': cod_stats,
        'recent_transactions': recent_transactions,
        'now': timezone.now(),
        'user_is_superadmin': user_is_superadmin,
        'vehicle_type_choices': fleet_models.VEHICLE_CHOICES,
        'document_type_choices': fleet_models.DriverDocument.document_choices,
    }

    return render(request, 'workforce/driver_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def driver_vehicle_save(request, driver_id, vehicle_id=None):
    """Add or edit a driver vehicle (AJAX POST)."""
    driver = get_object_or_404(fleet_models.Driver, driver_id=driver_id)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        if vehicle_id:
            vehicle = get_object_or_404(fleet_models.DriverVehicle, id=vehicle_id, driver=driver)
        else:
            vehicle = fleet_models.DriverVehicle(driver=driver)
        vehicle.vehicle_type = request.POST.get('vehicle_type', 'none')
        vehicle.vehicle_no = request.POST.get('vehicle_no', '') or ''
        vehicle.vehicle_model = request.POST.get('vehicle_model', '') or ''
        vehicle.vehicle_color = request.POST.get('vehicle_color', '') or ''
        vehicle.vehicle_status = request.POST.get('vehicle_status', 'active')
        if request.FILES.get('vehicle_photo'):
            vehicle.vehicle_photo = request.FILES['vehicle_photo']
        vehicle.save()
        return JsonResponse({'success': True, 'message': 'Vehicle saved successfully'})
    except Exception as e:
        logger.exception('driver_vehicle_save error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='/accounts/login/')
@staff_required
def driver_vehicle_delete(request, driver_id, vehicle_id):
    """Delete a driver vehicle (AJAX POST)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    vehicle = get_object_or_404(fleet_models.DriverVehicle, id=vehicle_id, driver__driver_id=driver_id)
    vehicle.delete()
    return JsonResponse({'success': True, 'message': 'Vehicle deleted'})


@login_required(login_url='/accounts/login/')
@staff_required
def driver_document_save(request, driver_id, document_id=None):
    """Add or edit a driver document (AJAX POST)."""
    driver = get_object_or_404(fleet_models.Driver, driver_id=driver_id)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        if document_id:
            doc = get_object_or_404(fleet_models.DriverDocument, id=document_id, driver=driver)
        else:
            doc = fleet_models.DriverDocument(driver=driver)
        doc.document_type = request.POST.get('document_type', '')
        doc.document_no = request.POST.get('document_no', '')
        doc.document_issued_from = request.POST.get('document_issued_from', '') or ''
        expiry = request.POST.get('document_expiry_date', '')
        doc.document_expiry_date = expiry if expiry else None
        if request.FILES.get('document_file'):
            doc.document_file = request.FILES['document_file']
        if request.FILES.get('document_file_back'):
            doc.document_file_back = request.FILES['document_file_back']
        doc.save()
        return JsonResponse({'success': True, 'message': 'Document saved successfully'})
    except Exception as e:
        logger.exception('driver_document_save error: %s', e)
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='/accounts/login/')
@staff_required
def driver_document_delete(request, driver_id, document_id):
    """Delete a driver document (AJAX POST)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    doc = get_object_or_404(fleet_models.DriverDocument, id=document_id, driver__driver_id=driver_id)
    doc.delete()
    return JsonResponse({'success': True, 'message': 'Document deleted'})


# =============================================================================
# FULFILLMENT SERVICE & PURCHASE ORDERS
# =============================================================================


@login_required
@staff_required
def suppliers_list(request):
    """
    List all businesses (sellers) with fulfillment service status sections.
    Section 1: Active fulfillment service
    Section 2: Requested and Non-active
    """
    from django.db.models import Count, Q as DjangoQ

    # Base queryset - all businesses
    base_qs = business_models.Business.objects.all()

    # Search filter
    search = request.GET.get('search', '').strip()
    if search:
        base_qs = base_qs.filter(
            Q(business_name__icontains=search) |
            Q(business_code__icontains=search) |
            Q(business_email__icontains=search) |
            Q(business_phone__icontains=search)
        )

    # Annotate with order stats
    base_qs = base_qs.annotate(
        total_orders=Count('order'),
        fulfilled_orders=Count('order', filter=DjangoQ(order__order_status__in=['delivered', 'fulfilled'])),
        pending_orders=Count('order', filter=DjangoQ(order__order_status__in=['to_review', 'ready_to_pickup', 'publish']))
    )

    # Section 1: Active fulfillment sellers
    active_sellers = list(base_qs.filter(
        fulfillment_service_status='active'
    ).order_by('-fulfillment_activated_at', 'business_name'))

    # Section 2: Requested and Non-active sellers
    requested_sellers = list(base_qs.filter(
        fulfillment_service_status='requested'
    ).order_by('business_name'))

    nonactive_sellers = list(base_qs.filter(
        fulfillment_service_status='none'
    ).order_by('business_name'))

    # Summary stats
    active_count = len(active_sellers)
    requested_count = len(requested_sellers)
    nonactive_count = len(nonactive_sellers)
    total_sellers = active_count + requested_count + nonactive_count

    context = {
        'page_title': 'Businesses - Fulfillment Service',
        'search': search,
        'active_sellers': active_sellers,
        'requested_sellers': requested_sellers,
        'nonactive_sellers': nonactive_sellers,
        'total_sellers': total_sellers,
        'active_count': active_count,
        'requested_count': requested_count,
        'nonactive_count': nonactive_count,
    }

    return render(request, 'workforce/suppliers_list.html', context)


@login_required
@staff_required
def fulfilled_orders_list(request):
    """
    List all fulfilled/delivered orders (Purchase Orders).
    Shows orders that have been successfully delivered.
    """
    from django.db.models import Q

    orders = orders_models.Order.objects.filter(
        order_status__in=['delivered', 'fulfilled']
    ).select_related('business').order_by('-delivered_at', '-updated_at')

    # Filters
    business_id = request.GET.get('business', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '')
    order_status = request.GET.get('status', '')

    if business_id:
        orders = orders.filter(business__business_id=business_id)

    if date_from:
        orders = orders.filter(delivered_at__date__gte=date_from)

    if date_to:
        orders = orders.filter(delivered_at__date__lte=date_to)

    if search:
        orders = orders.filter(
            Q(order_number__icontains=search) |
            Q(client_order_code__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(customer_phone__icontains=search)
        )

    if order_status:
        orders = orders.filter(order_status=order_status)

    # Get suppliers for filter dropdown (businesses with fulfillment enabled)
    suppliers = business_models.Business.objects.filter(
        fulfillment_service_enabled=True,
        business_status='active'
    ).order_by('business_name')

    # Summary stats
    total_count = orders.count()
    fulfilled_count = orders.filter(order_status='fulfilled').count()
    delivered_count = orders.filter(order_status='delivered').count()

    # Calculate total COD collected
    from django.db.models import Sum
    total_cod = orders.filter(
        cod_amount__gt=0
    ).aggregate(total=Sum('cod_amount'))['total'] or 0

    # Paginate
    page_obj = paginate_queryset(request, orders, items_per_page=25)

    context = {
        'page_title': 'Purchase Orders (Fulfilled)',
        'page_obj': page_obj,
        'suppliers': suppliers,
        'filters': {
            'business': business_id,
            'date_from': date_from,
            'date_to': date_to,
            'search': search,
            'status': order_status,
        },
        'total_count': total_count,
        'fulfilled_count': fulfilled_count,
        'delivered_count': delivered_count,
        'total_cod': total_cod,
    }

    return render(request, 'workforce/fulfilled_orders_list.html', context)


# ==========================================
# DELIVERY TASK BULK ACTIONS
# ==========================================

@login_required(login_url='/accounts/login/')
@staff_required
def delivery_task_edit(request, task_id):
    """Edit delivery task details and associated order"""
    task = get_object_or_404(
        delivery_models.DeliveryTask.objects.select_related(
            'order', 'order__business', 'driver', 'business', 'pickup_location'
        ).prefetch_related('order__order_items__product'),
        id=task_id
    )

    if request.method == 'POST':
        try:
            # Lock check: Prevent editing settled tasks
            if task.dl_task_status == 'delivered' and task.order and task.order.cod_status_by_staff == 'cod_settled_with_business':
                messages.error(request, 'Task is locked. Cannot edit after delivery is successful and COD is settled.')
                return redirect(request.path)

            # Handle task fields
            driver_id = request.POST.get('driver')
            status = request.POST.get('status')
            task_description = request.POST.get('task_description', '').strip()

            if driver_id:
                task.driver_id = driver_id
            if status:
                task.dl_task_status = status
            if task_description:
                task.dl_task_description = task_description[:100]  # max_length=100
            pickup_location_id = request.POST.get('pickup_location', '').strip()
            if pickup_location_id:
                task.pickup_location_id = int(pickup_location_id)
            elif pickup_location_id == '':
                task.pickup_location_id = None
            valid_time_slots = ('9am-1pm', '2pm-6pm', '6pm-10pm')
            preferred_time = request.POST.get('preferred_time', '').strip()
            task.preferred_time = preferred_time if preferred_time in valid_time_slots else ''
            task.save()

            # Handle order fields if order exists
            if task.order:
                order = task.order
                from decimal import Decimal, InvalidOperation

                # Snapshot old values
                old_vals = {
                    'customer_name': order.customer_name or '', 'customer_phone': order.customer_phone or '',
                    'customer_whatsapp': order.customer_whatsapp or '', 'customer_address': order.customer_address or '',
                    'dl_zone': order.dl_zone, 'dl_street': order.dl_street, 'dl_building': order.dl_building,
                    'package_description': order.package_description or '', 'total_quantity': order.total_quantity or 0,
                    'cod_amount': order.cod_amount or 0, 'dl_amount': order.dl_amount or 0,
                    'order_notes': order.order_notes or '',
                    'latitude': str(order.latitude or ''), 'longitude': str(order.longitude or ''),
                }

                # Customer details
                order.customer_name = request.POST.get('customer_name', order.customer_name)
                order.customer_phone = request.POST.get('customer_phone', order.customer_phone)
                order.customer_whatsapp = request.POST.get('customer_whatsapp', order.customer_whatsapp)
                order.customer_address = request.POST.get('customer_address', order.customer_address)

                # Address components
                dl_zone = request.POST.get('dl_zone', '')
                dl_street = request.POST.get('dl_street', '')
                dl_building = request.POST.get('dl_building', '')

                order.dl_zone = int(dl_zone) if dl_zone and dl_zone.isdigit() else None
                order.dl_street = int(dl_street) if dl_street and dl_street.isdigit() else None
                order.dl_building = int(dl_building) if dl_building and dl_building.isdigit() else None

                # Order notes
                order.order_notes = request.POST.get('order_notes', order.order_notes)

                # Package details
                product_desc = request.POST.get('package_description', '').strip()
                if product_desc is not None:
                    order.package_description = product_desc[:255]
                total_qty = request.POST.get('total_quantity', '').strip()
                if total_qty and total_qty.isdigit():
                    order.total_quantity = int(total_qty)

                # COD details
                cod_amount = request.POST.get('cod_amount', '0').strip()
                try:
                    order.cod_amount = Decimal(cod_amount) if cod_amount else 0
                except (InvalidOperation, ValueError):
                    order.cod_amount = 0

                dl_amount = request.POST.get('dl_amount', '0')
                order.dl_amount = int(dl_amount) if dl_amount and dl_amount.isdigit() else 0

                # Coordinates & accuracy
                lat_raw = request.POST.get('latitude', '').strip()
                lng_raw = request.POST.get('longitude', '').strip()
                try:
                    order.latitude = Decimal(lat_raw) if lat_raw else None
                except InvalidOperation:
                    pass
                try:
                    order.longitude = Decimal(lng_raw) if lng_raw else None
                except InvalidOperation:
                    pass
                coords_accuracy = request.POST.get('coords_accuracy', '').strip()
                if coords_accuracy:
                    order.coords_accuracy = coords_accuracy

                order.save()

                # Update order item quantities
                from orders.models import OrderItem
                for key, val in request.POST.items():
                    if key.startswith('item_qty_'):
                        try:
                            item_id = int(key.split('item_qty_')[1])
                            qty = int(val)
                            item = OrderItem.objects.get(id=item_id, order=order)
                            item.quantity = max(0, qty)
                            if item.unit_price:
                                item.total_price = Decimal(str(item.unit_price)) * item.quantity
                            item.save()
                        except (ValueError, OrderItem.DoesNotExist):
                            pass

                # Build change summary
                new_vals = {
                    'customer_name': order.customer_name or '', 'customer_phone': order.customer_phone or '',
                    'customer_whatsapp': order.customer_whatsapp or '', 'customer_address': order.customer_address or '',
                    'dl_zone': order.dl_zone, 'dl_street': order.dl_street, 'dl_building': order.dl_building,
                    'package_description': order.package_description or '', 'total_quantity': order.total_quantity or 0,
                    'cod_amount': order.cod_amount or 0, 'dl_amount': order.dl_amount or 0,
                    'order_notes': order.order_notes or '',
                    'latitude': str(order.latitude or ''), 'longitude': str(order.longitude or ''),
                }
                field_labels = {
                    'customer_name': 'Name', 'customer_phone': 'Phone', 'customer_whatsapp': 'WhatsApp',
                    'customer_address': 'Address', 'dl_zone': 'Zone', 'dl_street': 'Street', 'dl_building': 'Building',
                    'package_description': 'Package Desc', 'total_quantity': 'Qty', 'cod_amount': 'COD',
                    'dl_amount': 'DL Amount', 'order_notes': 'Notes',
                    'latitude': 'Lat', 'longitude': 'Lng',
                }
                changes = []
                for k, old_v in old_vals.items():
                    new_v = new_vals[k]
                    if str(old_v) != str(new_v):
                        label = field_labels.get(k, k)
                        changes.append(f"{label}: {old_v or '—'} → {new_v or '—'}")

                change_summary = ', '.join(changes) if changes else 'No changes'

                # Log to timeline
                orders_models.OrderStatusHistory.objects.create(
                    order=order,
                    field_name='task_edited',
                    old_value='',
                    new_value='edited',
                    old_display='',
                    new_display=f'{len(changes)} field{"s" if len(changes) != 1 else ""} changed',
                    changed_by=request.user,
                    notes=change_summary[:255],
                )

            messages.success(request, f'Task #{task.dl_task_number} updated successfully.')

            # Check if HTMX request
            if request.headers.get('HX-Request'):
                return redirect('workforce:delivery_task_detail', task_id=task_id)
            return redirect('workforce:delivery_task_detail', task_id=task_id)

        except Exception as e:
            logger.exception("Error updating task %s: %s", task_id, str(e))
            messages.error(request, 'An error occurred while updating the task')

    # Get available drivers for dropdown (approved drivers)
    drivers = fleet_models.Driver.objects.select_related('user').filter(
        driver_status='approved'
    ).order_by('user__first_name')

    # Pickup locations scoped to the task's business, including fulfillment center
    from business.models import PickupLocation
    from warehouse.models import SellerWarehouseLink
    from django.db.models import Q
    task_business = task.business or (task.order.business if task.order else None)

    if task_business:
        # Get the linked warehouse (if any) for this business
        wh_link = SellerWarehouseLink.objects.filter(
            business=task_business
        ).select_related('warehouse').first()
        linked_wh = wh_link.warehouse if wh_link else None

        # Include: business's own pickup locations + any fulfillment center linked to their warehouse
        q = Q(business=task_business)
        if linked_wh:
            q |= Q(warehouse=linked_wh, is_fulfilment_center=True)
        pickup_locations = list(
            PickupLocation.objects.filter(q)
            .select_related('warehouse')
            .order_by('-is_fulfilment_center', 'pickup_location_title')
        )
    else:
        pickup_locations = []

    context = {
        'page_title': f'Edit Task #{task.dl_task_number}',
        'task': task,
        'drivers': drivers,
        'pickup_locations': pickup_locations,
    }

    return render(request, 'workforce/parts/delivery_task_edit.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def bulk_print_tasks(request):
    """Generate printable view for selected tasks"""
    task_ids = request.GET.get('ids', '').split(',')
    task_ids = [int(id) for id in task_ids if id.isdigit() and len(id) <= 10][:100]  # Limit to 100 tasks

    if not task_ids:
        return render(request, 'workforce/parts/bulk_print_tasks.html', {
            'page_title': 'Print Tasks',
            'tasks': [],
            'print_mode': True,
        })

    tasks = list(
        delivery_models.DeliveryTask.objects.filter(
            id__in=task_ids
        ).select_related('order', 'driver', 'driver__user', 'business', 'pickup_location')
        .prefetch_related('task_qrcode', 'order__order_items__product')
    )

    # Generate QR code for any task missing one
    for task in tasks:
        qr_qs = task.task_qrcode.all()
        if not qr_qs.exists() and task.dl_task_number:
            qr_obj = delivery_models.DeliveryTaskQRCode(delivery_task=task, task_number=task.dl_task_number)
            qr_obj.generate_qrcode()
            qr_obj.save()

    # Build zone_number → zone_name lookup for all zones used in these tasks
    zone_numbers = {t.order.dl_zone for t in tasks if t.order and t.order.dl_zone}
    zone_names = {}
    if zone_numbers:
        zone_names = dict(
            delivery_models.ZoneName.objects.filter(zone_number__in=zone_numbers)
            .values_list('zone_number', 'zone_name')
        )

    context = {
        'page_title': 'Print Tasks',
        'tasks': tasks,
        'zone_names': zone_names,
        'print_mode': True,
    }

    return render(request, 'workforce/parts/bulk_print_tasks.html', context)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def bulk_publish_fleets(request):
    """Bulk publish tasks to Fleet drivers"""
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])

        if not task_ids:
            return JsonResponse({
                'success': False,
                'error': 'No tasks selected'
            }, status=400)

        # Update all selected tasks (non-cancelled orders)
        updated = delivery_models.DeliveryTask.objects.filter(
            id__in=task_ids,
        ).exclude(
            order__order_status='cancelled'
        ).update(
            dl_task_status='pending',
            dl_task_publish=True
        )

        return JsonResponse({
            'success': True,
            'message': f'{updated} task(s) published to Fleets',
            'updated_count': updated
        })
    except Exception as e:
        logger.exception("Error bulk publishing to fleets: %s", str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while publishing tasks to Fleets'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def bulk_publish_app(request):
    """Bulk publish tasks to Driver App"""
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])

        if not task_ids:
            return JsonResponse({
                'success': False,
                'error': 'No tasks selected'
            }, status=400)

        # Update all selected tasks to be available in driver app
        updated = delivery_models.DeliveryTask.objects.filter(
            id__in=task_ids
        ).update()  # Unassigned - available for drivers

        return JsonResponse({
            'success': True,
            'message': f'{updated} task(s) published to Driver App',
            'updated_count': updated
        })
    except Exception as e:
        logger.exception("Error bulk publishing to driver app: %s", str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while publishing tasks to driver app'
        }, status=400)


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def bulk_update_status(request):
    """Bulk update task status"""
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])
        status = data.get('status')

        if not task_ids:
            return JsonResponse({
                'success': False,
                'error': 'No tasks selected'
            }, status=400)

        if not status:
            return JsonResponse({
                'success': False,
                'error': 'Status is required'
            }, status=400)

        # Validate status against allowed choices
        VALID_STATUSES = [
            'for_review', 'pending', 'assigned', 'accepted', 'picked_up',
            'start_ride', 'out_for_delivery', 'in_transit', 'contacted',
            'non_reachable', 'delivered', 'failed', 'rejected', 'cancelled',
        ]
        if status not in VALID_STATUSES:
            return JsonResponse({
                'success': False,
                'error': f'Invalid status: {status}'
            }, status=400)

        # DMS status mapping for sync

        # Update all selected tasks (excluding locked tasks)
        tasks = delivery_models.DeliveryTask.objects.select_related('order').filter(
            id__in=task_ids
        ).exclude(
            order__cod_status_by_staff='cod_settled_with_business'
        )
        updated = 0
        for task in tasks:
            task.dl_task_status = status
            task.save(update_fields=['dl_task_status'])
            updated += 1

        return JsonResponse({
            'success': True,
            'message': f'{updated} task(s) updated to {status}',
            'updated_count': updated
        })
    except Exception as e:
        logger.exception("Error bulk updating task status: %s", str(e))
        return JsonResponse({
            'success': False,
            'error': 'An error occurred while updating task status'
        }, status=400)


def _sanitize_csv_value(value):
    """Sanitize value to prevent CSV injection attacks"""
    if value is None:
        return ''
    value = str(value)
    # Prevent formula injection by prefixing dangerous characters
    if value and value[0] in ('=', '+', '-', '@', '\t', '\r', '\n'):
        return "'" + value
    return value


@login_required(login_url='/accounts/login/')
@staff_required
def bulk_export_tasks(request):
    """Export selected tasks to CSV"""
    task_ids = request.GET.get('ids', '').split(',')
    task_ids = [int(id) for id in task_ids if id.isdigit() and len(id) <= 10][:500]  # Limit to 500 tasks

    if not task_ids:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="delivery_tasks_export.csv"'
        return response

    tasks = delivery_models.DeliveryTask.objects.filter(
        id__in=task_ids
    ).select_related('order', 'driver', 'driver__user', 'business', 'pickup_location')

    # Log export for audit
    logger.info(f"CSV export by user {request.user.id}: {len(task_ids)} tasks")

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="delivery_tasks_export.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Task Number', 'Date', 'Customer Name', 'Customer Phone',
        'Delivery Address', 'Driver', 'Client Status', 'DMS Status',
        'Pickup Location', 'COD Amount', 'Notes'
    ])

    for task in tasks:
        writer.writerow([
            _sanitize_csv_value(task.dl_task_number),
            _sanitize_csv_value(task.dl_task_date),
            _sanitize_csv_value(task.order.customer_name if task.order else ''),
            _sanitize_csv_value(task.order.customer_phone if task.order else ''),
            _sanitize_csv_value(task.order.customer_address if task.order else ''),
            _sanitize_csv_value(str(task.driver) if task.driver else ''),
            _sanitize_csv_value(task.get_dl_task_status_client_display() if hasattr(task, 'get_dl_task_status_client_display') else task.dl_task_status_client),
            _sanitize_csv_value(task.pickup_location.pickup_location_title if task.pickup_location else ''),
            _sanitize_csv_value(task.order.cod_amount if task.order else ''),
            _sanitize_csv_value(task.notes if hasattr(task, 'notes') else ''),
        ])

    return response


@login_required(login_url='/accounts/login/')
@staff_required
def get_active_drivers(request):
    """API endpoint to get active drivers with details for bulk assignment"""
    drivers = fleet_models.Driver.objects.select_related('user').filter(
        driver_status='approved'
    ).order_by('user__first_name')

    driver_list = []
    for driver in drivers:
        name = driver.user.get_full_name() or driver.user.username
        driver_list.append({
            'driver_id': driver.driver_id,  # Use driver_id (PK)
            'name': name,
            'driver_code': driver.driver_code or '',
            'phone': driver.driver_phone or '',
            'status': 'available',  # Could be enhanced with real-time status
        })

    return JsonResponse({'success': True, 'drivers': driver_list})


@require_http_methods(["POST"])
@login_required(login_url='/accounts/login/')
@staff_required
def bulk_assign_driver(request):
    """Bulk assign driver to selected tasks"""
    try:
        data = json.loads(request.body)
        task_ids = data.get('task_ids', [])
        driver_id = data.get('driver_id')

        if not task_ids or not driver_id:
            return JsonResponse({
                'success': False,
                'message': 'Missing task IDs or driver ID'
            }, status=400)

        # Validate driver
        try:
            driver = fleet_models.Driver.objects.get(driver_id=driver_id, driver_status='approved')
        except fleet_models.Driver.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Driver not found or not approved'
            }, status=400)

        # Get tasks
        tasks = delivery_models.DeliveryTask.objects.filter(id__in=task_ids)

        if not tasks.exists():
            return JsonResponse({
                'success': False,
                'message': 'No tasks found with provided IDs'
            }, status=400)

        # Assign driver to all tasks
        assigned_count = 0
        for task in tasks:
            task.driver = driver
            # Update status to assigned if currently pending/for_review
            if task.dl_task_status in ['for_review', 'pending', None]:
                task.dl_task_status = 'assigned'
                task._status_actor = 'staff'
            task.save()
            assigned_count += 1

        # Log assignment
        driver_name = driver.user.get_full_name() or driver.user.username
        logger.info(f"Bulk assign by user {request.user.id}: {assigned_count} tasks to driver {driver_id} ({driver_name})")

        return JsonResponse({
            'success': True,
            'assigned': assigned_count,
            'message': f'Successfully assigned {assigned_count} task(s) to {driver_name}'
        })

    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        logger.error(f"Error in bulk assign driver: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Error assigning tasks: {str(e)}'
        }, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
def order_edit(request, order_id):
    """Edit order details - Staff view for correcting order information"""
    order = get_object_or_404(
        orders_models.Order.objects.select_related(
            'business', 'pickup_location'
        ),
        id=order_id
    )

    # Get pickup locations for the business
    # Show fulfillment stores first when fulfillment service is enabled
    pickup_locations = business_models.PickupLocation.objects.filter(
        business=order.business
    ).order_by('-is_fulfilment_center', 'pickup_location_title')

    if request.method == 'POST':
        # Handle form submission
        try:
            from decimal import Decimal, InvalidOperation

            # Snapshot old values for change tracking
            old_vals = {
                'customer_name': order.customer_name or '',
                'customer_phone': order.customer_phone or '',
                'customer_whatsapp': order.customer_whatsapp or '',
                'customer_address': order.customer_address or '',
                'dl_zone': order.dl_zone,
                'dl_street': order.dl_street,
                'dl_building': order.dl_building,
                'package_description': order.package_description or '',
                'total_quantity': order.total_quantity or 0,
                'cod_amount': order.cod_amount or 0,
                'dl_amount': order.dl_amount or 0,
                'order_notes': order.order_notes or '',
                'order_status': order.order_status or '',
                'verification_status': order.verification_status or '',
                'latitude': str(order.latitude or ''),
                'longitude': str(order.longitude or ''),
                'pickup_location_id': order.pickup_location_id,
            }

            # Customer details
            order.customer_name = request.POST.get('customer_name', order.customer_name)
            order.customer_phone = request.POST.get('customer_phone', order.customer_phone)
            order.customer_whatsapp = request.POST.get('customer_whatsapp', order.customer_whatsapp)
            order.customer_address = request.POST.get('customer_address', order.customer_address)

            # Address components
            dl_zone = request.POST.get('dl_zone', '')
            dl_street = request.POST.get('dl_street', '')
            dl_building = request.POST.get('dl_building', '')

            order.dl_zone = int(dl_zone) if dl_zone and dl_zone.isdigit() else None
            order.dl_street = int(dl_street) if dl_street and dl_street.isdigit() else None
            order.dl_building = int(dl_building) if dl_building and dl_building.isdigit() else None

            # Package details
            product_desc = request.POST.get('package_description', '').strip()
            order.package_description = product_desc[:255]
            total_qty = request.POST.get('total_quantity', '').strip()
            if total_qty and total_qty.isdigit():
                order.total_quantity = int(total_qty)

            # COD details
            cod_amount = request.POST.get('cod_amount', '0').strip()
            try:
                order.cod_amount = Decimal(cod_amount) if cod_amount else 0
            except (InvalidOperation, ValueError):
                order.cod_amount = 0
            cod_status = request.POST.get('cod_status_by_client')
            if cod_status:
                order.cod_status_by_client = cod_status
            elif order.cod_amount > 0 and order.cod_status_by_client == 'no_cod':
                order.cod_status_by_client = 'pending'

            # Delivery amount
            dl_amount = request.POST.get('dl_amount', '0')
            order.dl_amount = int(dl_amount) if dl_amount and dl_amount.isdigit() else 0

            # Pickup location
            pickup_id = request.POST.get('pickup_location')
            if pickup_id:
                order.pickup_location_id = pickup_id

            # Coordinates
            lat_raw = request.POST.get('latitude', '').strip()
            lng_raw = request.POST.get('longitude', '').strip()
            try:
                order.latitude = Decimal(lat_raw) if lat_raw else None
            except (InvalidOperation, ValueError):
                pass
            try:
                order.longitude = Decimal(lng_raw) if lng_raw else None
            except (InvalidOperation, ValueError):
                pass

            # Order notes
            order.order_notes = request.POST.get('order_notes', order.order_notes)

            # Status
            order.order_status = request.POST.get('order_status', order.order_status)
            order.verification_status = request.POST.get('verification_status', order.verification_status)

            order.save()

            # Build change summary
            new_vals = {
                'customer_name': order.customer_name or '',
                'customer_phone': order.customer_phone or '',
                'customer_whatsapp': order.customer_whatsapp or '',
                'customer_address': order.customer_address or '',
                'dl_zone': order.dl_zone,
                'dl_street': order.dl_street,
                'dl_building': order.dl_building,
                'package_description': order.package_description or '',
                'total_quantity': order.total_quantity or 0,
                'cod_amount': order.cod_amount or 0,
                'dl_amount': order.dl_amount or 0,
                'order_notes': order.order_notes or '',
                'order_status': order.order_status or '',
                'verification_status': order.verification_status or '',
                'latitude': str(order.latitude or ''),
                'longitude': str(order.longitude or ''),
                'pickup_location_id': order.pickup_location_id,
            }
            field_labels = {
                'customer_name': 'Name', 'customer_phone': 'Phone', 'customer_whatsapp': 'WhatsApp',
                'customer_address': 'Address', 'dl_zone': 'Zone', 'dl_street': 'Street', 'dl_building': 'Building',
                'package_description': 'Package Desc', 'total_quantity': 'Qty', 'cod_amount': 'COD',
                'dl_amount': 'DL Amount', 'order_notes': 'Notes', 'order_status': 'Status',
                'verification_status': 'Verification', 'latitude': 'Lat', 'longitude': 'Lng',
                'pickup_location_id': 'Pickup Location',
            }
            changes = []
            for k, old_v in old_vals.items():
                new_v = new_vals[k]
                if str(old_v) != str(new_v):
                    label = field_labels.get(k, k)
                    changes.append(f"{label}: {old_v or '—'} → {new_v or '—'}")

            change_summary = ', '.join(changes) if changes else 'No changes'

            # Log to timeline
            orders_models.OrderStatusHistory.objects.create(
                order=order,
                field_name='order_edited',
                old_value='',
                new_value='edited',
                old_display='',
                new_display=f'{len(changes)} field{"s" if len(changes) != 1 else ""} changed',
                changed_by=request.user,
                notes=change_summary[:255],
            )

            messages.success(request, f'Order {order.order_number} updated successfully.')

            # Check if HTMX request
            if request.headers.get('HX-Request'):
                return redirect('workforce:order_detail', order_id=order_id)
            return redirect('workforce:order_detail', order_id=order_id)

        except Exception as e:
            logger.exception("Error updating order %s: %s", order_id, str(e))
            messages.error(request, 'An error occurred while updating the order')

    # Zone name lookup for pre-fill
    zone_name = ''
    if order.dl_zone:
        zn = delivery_models.ZoneName.objects.filter(zone_number=order.dl_zone).values_list('zone_name', flat=True).first()
        zone_name = zn or ''

    # Order items
    order_items = orders_models.OrderItem.objects.filter(
        order=order
    ).select_related('product').order_by('id')

    context = {
        'page_title': f'Edit Order - {order.order_number}',
        'order': order,
        'order_items': order_items,
        'pickup_locations': pickup_locations,
        'order_statuses': orders_models.ORDER_STATUS_BY_CLIENT,
        'verification_statuses': orders_models.Order.VERIFICATION_STATUS,
        'cod_statuses': orders_models.COD_STATUS_BY_CLIENT,
        'zone_name': zone_name,
    }

    return render(request, 'workforce/order_edit.html', context)


# --- Order Item AJAX endpoints -----------------------------------------------

@login_required
@staff_required
@require_http_methods(["POST"])
def order_item_add(request, order_id):
    """Add a new item to an order via AJAX."""
    order = get_object_or_404(orders_models.Order, id=order_id)
    product_id = request.POST.get('product_id')
    quantity = int(request.POST.get('quantity', 1) or 1)
    unit_price = request.POST.get('unit_price', '')

    product = None
    if product_id:
        from product.models import Product
        product = Product.objects.filter(id=product_id).first()

    price = None
    if unit_price:
        from decimal import Decimal
        price = Decimal(unit_price)
    elif product and product.item_price:
        price = product.item_price

    item = orders_models.OrderItem.objects.create(
        order=order,
        product=product,
        quantity=quantity,
        unit_price=price,
    )

    return JsonResponse({
        'success': True,
        'item': {
            'id': item.id,
            'product_id': product.id if product else None,
            'product_name': product.item_name if product else '',
            'sku': product.item_sku if product else '',
            'quantity': item.quantity,
            'unit_price': str(item.unit_price or 0),
            'total_price': str(item.total_price or 0),
        }
    })


@login_required
@staff_required
@require_http_methods(["POST"])
def order_item_update(request, order_id, item_id):
    """Update an existing order item via AJAX."""
    item = get_object_or_404(orders_models.OrderItem, id=item_id, order_id=order_id)
    quantity = request.POST.get('quantity')
    unit_price = request.POST.get('unit_price')
    product_id = request.POST.get('product_id')

    if product_id:
        from product.models import Product
        product = Product.objects.filter(id=product_id).first()
        if product:
            item.product = product

    if quantity:
        item.quantity = int(quantity)
    if unit_price:
        from decimal import Decimal
        item.unit_price = Decimal(unit_price)

    item.save()

    return JsonResponse({
        'success': True,
        'item': {
            'id': item.id,
            'product_id': item.product_id,
            'product_name': item.product.item_name if item.product else '',
            'sku': item.product.item_sku if item.product else '',
            'quantity': item.quantity,
            'unit_price': str(item.unit_price or 0),
            'total_price': str(item.total_price or 0),
        }
    })


@login_required
@staff_required
@require_http_methods(["POST"])
def order_item_delete(request, order_id, item_id):
    """Delete an order item via AJAX."""
    item = get_object_or_404(orders_models.OrderItem, id=item_id, order_id=order_id)
    item_name = item.product.item_name if item.product else (item.notes or 'Item')
    item_qty = item.quantity
    item.delete()

    # Log to timeline
    orders_models.OrderStatusHistory.objects.create(
        order_id=order_id,
        field_name='item_deleted',
        old_value=f'{item_name} x{item_qty}',
        new_value='deleted',
        old_display=f'{item_name} x{item_qty}',
        new_display='Deleted',
        changed_by=request.user,
        notes=f'Item removed: {item_name} x{item_qty}',
    )
    return JsonResponse({'success': True})


# =============================================================================
# WAREHOUSE-BUSINESS LINK MANAGEMENT
# =============================================================================


@login_required
@staff_required
def warehouses_list(request):
    """
    List all warehouses with linked businesses and link/unlink controls.
    """
    from warehouse.models import Warehouse, SellerWarehouseLink

    search = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')

    warehouses = Warehouse.objects.prefetch_related(
        'seller_links__business'
    ).order_by('-is_default', 'name')

    if search:
        warehouses = warehouses.filter(
            Q(name__icontains=search) | Q(code__icontains=search) |
            Q(city__icontains=search)
        )
    if status_filter == 'active':
        warehouses = warehouses.filter(is_active=True)
    elif status_filter == 'inactive':
        warehouses = warehouses.filter(is_active=False)

    # Get fulfillment-enabled businesses for the link dropdown
    linkable_businesses = business_models.Business.objects.filter(
        fulfillment_service_enabled=True,
        business_status='active',
    ).order_by('business_name')

    context = {
        'page_title': 'Warehouse - Business Links',
        'warehouses': warehouses,
        'linkable_businesses': linkable_businesses,
        'search': search,
        'status_filter': status_filter,
        'total_warehouses': Warehouse.objects.count(),
        'active_warehouses': Warehouse.objects.filter(is_active=True).count(),
        'total_links': SellerWarehouseLink.objects.filter(is_active=True).count(),
    }

    return render(request, 'workforce/warehouses_list.html', context)


@login_required
@staff_required
@require_http_methods(["POST"])
def warehouse_link_business(request):
    """
    Link a business to a warehouse. Creates SellerWarehouseLink which
    triggers signal to auto-create PickupLocation.
    """
    from warehouse.models import Warehouse, SellerWarehouseLink

    warehouse_id = request.POST.get('warehouse_id')
    business_id = request.POST.get('business_id')

    if not warehouse_id or not business_id:
        return JsonResponse({'success': False, 'error': 'Missing warehouse or business ID'}, status=400)

    warehouse = get_object_or_404(Warehouse, pk=warehouse_id)
    business = get_object_or_404(business_models.Business, pk=business_id)

    # Check if link already exists
    if SellerWarehouseLink.objects.filter(business=business, warehouse=warehouse).exists():
        return JsonResponse({'success': False, 'error': 'This business is already linked to this warehouse'}, status=400)

    SellerWarehouseLink.objects.create(
        business=business,
        warehouse=warehouse,
        is_active=True,
        linked_by=request.user,
    )

    messages.success(request, f'{business.business_name} linked to {warehouse.name}')
    return JsonResponse({'success': True, 'message': f'{business.business_name} linked to {warehouse.name}'})


@login_required
@staff_required
@require_http_methods(["POST"])
def warehouse_unlink_business(request):
    """
    Unlink a business from a warehouse. Deletes SellerWarehouseLink which
    triggers signal to deactivate PickupLocation.
    """
    from warehouse.models import SellerWarehouseLink

    link_id = request.POST.get('link_id')

    if not link_id:
        return JsonResponse({'success': False, 'error': 'Missing link ID'}, status=400)

    link = get_object_or_404(SellerWarehouseLink, pk=link_id)
    biz_name = link.business.business_name
    wh_name = link.warehouse.name
    link.delete()

    messages.success(request, f'{biz_name} unlinked from {wh_name}')
    return JsonResponse({'success': True, 'message': f'{biz_name} unlinked from {wh_name}'})


# =============================================================================
# PRODUCT REQUEST MANAGEMENT VIEWS
# =============================================================================


@login_required(login_url='/accounts/login/')
@staff_required
def product_requests_list(request):
    """
    Staff view to manage all inbound and outbound product requests.

    Shows combined list of requests from all businesses with filtering and stats.
    Staff can view, approve, and complete requests from this interface.
    """
    from warehouse.models import InboundProductRequest, OutboundProductRequest
    from django.core.paginator import Paginator

    # Get both types of requests
    inbound_qs = InboundProductRequest.objects.select_related(
        'business', 'warehouse', 'created_by', 'approved_by', 'completed_by'
    ).prefetch_related('items__product')

    outbound_qs = OutboundProductRequest.objects.select_related(
        'business', 'warehouse', 'created_by', 'approved_by', 'completed_by'
    ).prefetch_related('items__product')

    # Apply filters
    request_type = request.GET.get('type', 'all')  # all, inbound, outbound
    status_filter = request.GET.get('status', '')
    business_id = request.GET.get('business', '')

    if status_filter:
        inbound_qs = inbound_qs.filter(status=status_filter)
        outbound_qs = outbound_qs.filter(status=status_filter)

    if business_id:
        try:
            business_id = int(business_id)
            inbound_qs = inbound_qs.filter(business_id=business_id)
            outbound_qs = outbound_qs.filter(business_id=business_id)
        except (ValueError, TypeError):
            pass

    # Combine or filter by type
    if request_type == 'inbound':
        requests_list = list(inbound_qs)
    elif request_type == 'outbound':
        requests_list = list(outbound_qs)
    else:
        # Combine both types and sort by created_at
        requests_list = sorted(
            list(inbound_qs) + list(outbound_qs),
            key=lambda x: x.created_at,
            reverse=True
        )

    # Stats
    stats = {
        'total': len(requests_list),
        'pending_inbound': InboundProductRequest.objects.filter(status='pending').count(),
        'pending_outbound': OutboundProductRequest.objects.filter(status='pending').count(),
        'approved': len([r for r in requests_list if r.status == 'approved']),
        'completed': len([r for r in requests_list if r.status == 'completed']),
    }

    # Manual pagination since we combined querysets
    paginator = Paginator(requests_list, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Get all businesses for filter dropdown
    businesses = business_models.Business.objects.filter(
        fulfillment_service_enabled=True
    ).order_by('business_name')

    context = {
        'page_title': 'Product Requests',
        'page_obj': page_obj,
        'stats': stats,
        'request_type': request_type,
        'status_filter': status_filter,
        'business_id': business_id,
        'businesses': businesses,
    }
    return render(request, 'workforce/product_requests_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
@require_http_methods(["POST"])
def approve_product_request(request, request_id, request_type):
    """
    Approve an inbound or outbound product request.

    Changes status from pending to approved and records who approved it.
    """
    from django.utils import timezone
    from warehouse.models import InboundProductRequest, OutboundProductRequest

    if request_type == 'inbound':
        req = get_object_or_404(InboundProductRequest, id=request_id)
    elif request_type == 'outbound':
        req = get_object_or_404(OutboundProductRequest, id=request_id)
    else:
        messages.error(request, "Invalid request type")
        return redirect('workforce:product_requests_list')

    if req.status != 'pending':
        messages.warning(request, "Only pending requests can be approved.")
        return redirect('workforce:product_requests_list')

    req.status = 'approved'
    req.approved_by = request.user
    req.approved_at = timezone.now()
    req.save()

    messages.success(request, f"Request {req.request_number} approved successfully.")
    return redirect('workforce:product_requests_list')


@login_required(login_url='/accounts/login/')
@staff_required
@require_http_methods(["POST"])
def complete_product_request(request, request_id, request_type):
    """
    Mark a product request as completed.

    Changes status from approved to completed and records who completed it.
    Only approved requests can be marked as completed.
    """
    from django.utils import timezone
    from warehouse.models import InboundProductRequest, OutboundProductRequest

    if request_type == 'inbound':
        req = get_object_or_404(InboundProductRequest, id=request_id)
    elif request_type == 'outbound':
        req = get_object_or_404(OutboundProductRequest, id=request_id)
    else:
        messages.error(request, "Invalid request type")
        return redirect('workforce:product_requests_list')

    if req.status != 'approved':
        messages.warning(request, "Only approved requests can be marked as completed.")
        return redirect('workforce:product_requests_list')

    req.status = 'completed'
    req.completed_by = request.user
    req.completed_at = timezone.now()
    req.save()

    messages.success(request, f"Request {req.request_number} marked as completed.")
    return redirect('workforce:product_requests_list')




# QNAS Coordinate Lookup Tool
@login_required(login_url="/accounts/login/")
@staff_required
def qnas_lookup_tool(request):
    """
    QNAS coordinate lookup tool - enter zone/street/building to get coordinates
    """
    return render(request, "workforce/qnas_lookup_tool.html")



@login_required(login_url="/accounts/login/")
@staff_required
def qnas_test(request):
    """
    QNAS API connection test page - diagnostics and troubleshooting
    """
    return render(request, "workforce/qnas_test.html")


# =============================================================================
# FORMS / INQUIRIES
# =============================================================================

@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiries_list(request):
    """List all 3PL pricing inquiry form submissions."""
    inquiries = webpages_models.PricingEnquiry.objects.all().order_by('-date_created')

    search = request.GET.get('search', '').strip()
    if search:
        inquiries = inquiries.filter(
            Q(full_name__icontains=search) |
            Q(business_name__icontains=search) |
            Q(business_contact_number__icontains=search) |
            Q(product_category__icontains=search)
        )

    cod_filter = request.GET.get('cod', '').strip()
    if cod_filter == '1':
        inquiries = inquiries.filter(is_required_COD_service=True)
    elif cod_filter == '0':
        inquiries = inquiries.filter(is_required_COD_service=False)

    fulfillment_filter = request.GET.get('fulfillment', '').strip()
    if fulfillment_filter == '1':
        inquiries = inquiries.filter(
            Q(is_required_fulfillment_service_for_operate_from_outside_qatar=True) |
            Q(is_required_fulfillment_service_for_make_hub_in_doha=True)
        )

    status_filter = request.GET.get('status', '').strip()
    if status_filter:
        inquiries = inquiries.filter(crm_status=status_filter)

    total_count = webpages_models.PricingEnquiry.objects.count()
    cod_count = webpages_models.PricingEnquiry.objects.filter(is_required_COD_service=True).count()
    fulfillment_count = webpages_models.PricingEnquiry.objects.filter(
        Q(is_required_fulfillment_service_for_operate_from_outside_qatar=True) |
        Q(is_required_fulfillment_service_for_make_hub_in_doha=True)
    ).count()

    page_obj = paginate_queryset(request, inquiries, items_per_page=20)

    context = {
        'page_title': 'Pricing Inquiries',
        'page_obj': page_obj,
        'search': search,
        'cod_filter': cod_filter,
        'fulfillment_filter': fulfillment_filter,
        'status_filter': status_filter,
        'total_count': total_count,
        'cod_count': cod_count,
        'fulfillment_count': fulfillment_count,
    }
    return render(request, 'workforce/forms/pricing_inquiries_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def whatsapp_inquiries_list(request):
    """List all WhatsApp quick inquiry submissions."""
    inquiries = webpages_models.WhatsAppInquiry.objects.all().order_by('-created_at')

    search = request.GET.get('search', '').strip()
    if search:
        inquiries = inquiries.filter(
            Q(company_name__icontains=search) |
            Q(contact_person__icontains=search) |
            Q(contact_number__icontains=search) |
            Q(product_category__icontains=search)
        )

    total_count = webpages_models.WhatsAppInquiry.objects.count()
    page_obj = paginate_queryset(request, inquiries, items_per_page=20)

    context = {
        'page_title': 'WhatsApp Inquiries',
        'page_obj': page_obj,
        'search': search,
        'total_count': total_count,
    }
    return render(request, 'workforce/forms/whatsapp_inquiries_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiry_detail(request, inquiry_id):
    """Full detail view for a single PricingEnquiry submission."""
    inquiry = get_object_or_404(
        webpages_models.PricingEnquiry.objects.prefetch_related('activities__created_by'),
        pk=inquiry_id
    )
    from django.contrib.auth.models import User
    staff_users = User.objects.filter(is_staff=True).order_by('first_name', 'username')
    context = {
        'page_title': f'Pricing Inquiry – {inquiry.business_name}',
        'inquiry': inquiry,
        'activities': inquiry.activities.select_related('created_by').order_by('-created_at'),
        'staff_users': staff_users,
        'status_choices': webpages_models.PricingEnquiry.STATUS_CHOICES,
    }
    return render(request, 'workforce/forms/pricing_inquiry_detail.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiry_update_status(request, inquiry_id):
    """AJAX: update crm_status, assigned_to, or staff_notes and log activity."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    inquiry = get_object_or_404(webpages_models.PricingEnquiry, pk=inquiry_id)
    from django.contrib.auth.models import User

    new_status = request.POST.get('crm_status', '').strip()
    assigned_to_id = request.POST.get('assigned_to', '').strip()
    staff_notes = request.POST.get('staff_notes', '').strip()

    changes = []
    if new_status and new_status != inquiry.crm_status:
        old_label = inquiry.get_crm_status_display()
        inquiry.crm_status = new_status
        new_label = inquiry.get_crm_status_display()
        changes.append(f'Status changed: {old_label} → {new_label}')

    if assigned_to_id == '':
        if inquiry.assigned_to_id is not None:
            changes.append('Assignment cleared')
            inquiry.assigned_to = None
    else:
        try:
            user = User.objects.get(pk=int(assigned_to_id))
            if inquiry.assigned_to_id != user.pk:
                changes.append(f'Assigned to {user.get_full_name() or user.username}')
                inquiry.assigned_to = user
        except (User.DoesNotExist, ValueError):
            pass

    if staff_notes != (inquiry.staff_notes or ''):
        inquiry.staff_notes = staff_notes or None
        changes.append('Notes updated')

    inquiry.save()

    if changes:
        webpages_models.PricingEnquiryActivity.objects.create(
            inquiry=inquiry,
            activity_type=webpages_models.PricingEnquiryActivity.TYPE_STATUS_CHANGE,
            body='; '.join(changes),
            created_by=request.user,
        )

    return JsonResponse({
        'success': True,
        'crm_status': inquiry.crm_status,
        'crm_status_display': inquiry.get_crm_status_display(),
        'assigned_to': inquiry.assigned_to.get_full_name() or inquiry.assigned_to.username if inquiry.assigned_to else '',
    })


@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiry_edit(request, inquiry_id):
    """AJAX: edit PricingEnquiry fields by staff."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    inquiry = get_object_or_404(webpages_models.PricingEnquiry, pk=inquiry_id)

    # Editable fields (all model fields except CRM/meta)
    EDITABLE_FIELDS = {
        'full_name', 'business_name', 'business_contact_number', 'operation_team_contact_number',
        'website_url', 'social_profile', 'product_category', 'average_order_value_qar',
        'business_operating_age', 'business_location_country',
        'avarage_number_of_order_last_week', 'avarage_number_of_order_done_last_month',
        'avarage_number_of_order_expect_next_month', 'orders_expected_in_next_3_months_milestone',
        'speed_delivery_offer_to_customers', 'preferred_delivery_time_window',
        'typical_package_size', 'delivery_coverage', 'current_courier_provider',
        'order_management_system', 'preferred_communication_channel',
        'is_delivery_free_to_customers', 'preferred_start_date',
        'type_of_pickup_location', 'pickup_Location_area_name',
        'pickup_location_time_slab', 'number_of_pickup_times_in_day',
    }
    BOOL_FIELDS = {
        'is_personalized_product', 'is_registered_company_in_qatar',
        'is_located_in_qatar', 'is_team_available_in_qatar',
        'is_required_COD_service', 'is_required_fulfillment_service_for_operate_from_outside_qatar',
        'is_required_fulfillment_service_for_make_hub_in_doha',
        'is_frequent_same_day_pick_and_delivery_required',
        'is_special_handling_required', 'is_return_logistics_required',
    }

    changes = []
    for field in EDITABLE_FIELDS:
        if field in request.POST:
            new_val = request.POST[field].strip()
            old_val = getattr(inquiry, field) or ''
            if str(old_val) != new_val:
                setattr(inquiry, field, new_val or None)
                changes.append(f'{field}: {old_val} → {new_val}')

    for field in BOOL_FIELDS:
        if field in request.POST:
            new_val = request.POST[field] in ('true', '1', 'on', 'True')
            old_val = getattr(inquiry, field)
            if old_val != new_val:
                setattr(inquiry, field, new_val)
                changes.append(f'{field}: {old_val} → {new_val}')

    if changes:
        inquiry.save()
        webpages_models.PricingEnquiryActivity.objects.create(
            inquiry=inquiry,
            activity_type=webpages_models.PricingEnquiryActivity.TYPE_STATUS_CHANGE,
            body='Fields edited: ' + '; '.join(changes[:10]),
            created_by=request.user,
        )

    return JsonResponse({'success': True, 'changes': len(changes)})


@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiry_add_activity(request, inquiry_id):
    """AJAX: add a followup note or activity to a PricingEnquiry."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    inquiry = get_object_or_404(webpages_models.PricingEnquiry, pk=inquiry_id)

    activity_type = request.POST.get('activity_type', webpages_models.PricingEnquiryActivity.TYPE_NOTE)
    body = request.POST.get('body', '').strip()
    if not body:
        return JsonResponse({'error': 'Body is required'}, status=400)

    valid_types = [t[0] for t in webpages_models.PricingEnquiryActivity.TYPE_CHOICES]
    if activity_type not in valid_types:
        activity_type = webpages_models.PricingEnquiryActivity.TYPE_NOTE

    activity = webpages_models.PricingEnquiryActivity.objects.create(
        inquiry=inquiry,
        activity_type=activity_type,
        body=body,
        created_by=request.user,
    )

    return JsonResponse({
        'success': True,
        'activity': {
            'id': activity.pk,
            'activity_type': activity.activity_type,
            'activity_type_display': activity.get_activity_type_display(),
            'body': activity.body,
            'created_by': activity.created_by.get_full_name() or activity.created_by.username,
            'created_at': activity.created_at.strftime('%d %b %Y, %H:%M'),
        }
    })


@login_required(login_url='/accounts/login/')
@staff_required
def pricing_inquiry_delete_activity(request, inquiry_id, activity_id):
    """AJAX: delete an activity entry (staff only, own entries)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    activity = get_object_or_404(
        webpages_models.PricingEnquiryActivity,
        pk=activity_id, inquiry_id=inquiry_id
    )
    # Only allow deletion by the creator or superuser
    if activity.created_by_id != request.user.pk and not request.user.is_superuser:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    activity.delete()
    return JsonResponse({'success': True})


@login_required(login_url='/accounts/login/')
@staff_required
def whatsapp_inquiry_detail(request, inquiry_id):
    """Full detail view for a single WhatsAppInquiry submission."""
    inquiry = get_object_or_404(webpages_models.WhatsAppInquiry, pk=inquiry_id)
    context = {
        'page_title': f'WhatsApp Inquiry – {inquiry.company_name}',
        'inquiry': inquiry,
    }
    return render(request, 'workforce/forms/whatsapp_inquiry_detail.html', context)


# =============================================================================
# HUB OPERATIONS — Hub Pickup Batch Management
# =============================================================================

@login_required
@staff_required
def hub_batch_list(request):
    """List all hub pickup batches with status filters."""
    from warehouse.models import WarehouseLocation

    status_filter = request.GET.get('status', '')
    batches = (
        delivery_models.HubPickupBatch.objects
        .select_related('pickup_location', 'hub_warehouse', 'hub_warehouse__warehouse', 'driver', 'driver__user', 'created_by')
        .prefetch_related('orders')
        .order_by('-created_at')
    )
    if status_filter:
        batches = batches.filter(status=status_filter)

    context = {
        'page_title': 'Hub Pickup Batches',
        'batches': batches,
        'status_filter': status_filter,
        'status_choices': delivery_models.HubPickupBatch.BATCH_STATUS_CHOICES,
    }
    return render(request, 'workforce/hub_batch_list.html', context)


@login_required
@staff_required
def hub_batch_create(request):
    """Create a new hub pickup batch from selected orders."""
    from warehouse.models import WarehouseLocation
    from django.utils.timezone import now as tz_now

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, ValueError):
            data = request.POST

        pickup_location_id = data.get('pickup_location_id')
        hub_warehouse_id = data.get('hub_warehouse_id')
        order_ids = data.get('order_ids', [])
        driver_id = data.get('driver_id') or None
        driver_earnings = data.get('driver_earnings') or 0
        notes = data.get('notes', '')

        errors = []
        if not pickup_location_id:
            errors.append("Pickup location is required.")
        if not hub_warehouse_id:
            errors.append("Hub warehouse is required.")
        if not order_ids:
            errors.append("At least one order must be selected.")

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        try:
            pickup_location = business_models.PickupLocation.objects.get(id=pickup_location_id)
            hub_warehouse = WarehouseLocation.objects.get(id=hub_warehouse_id, is_active=True)
        except (business_models.PickupLocation.DoesNotExist, WarehouseLocation.DoesNotExist) as e:
            return JsonResponse({'success': False, 'errors': [str(e)]}, status=400)

        driver = None
        if driver_id:
            try:
                driver = fleet_models.Driver.objects.get(driver_id=driver_id, driver_status='approved')
            except fleet_models.Driver.DoesNotExist:
                return JsonResponse({'success': False, 'errors': ['Driver not found or not approved.']}, status=400)

        # Generate batch number
        today = tz_now().strftime('%Y%m%d')
        existing_count = delivery_models.HubPickupBatch.objects.filter(
            batch_number__startswith=f'BATCH-{today}-'
        ).count()
        batch_number = f'BATCH-{today}-{existing_count + 1:03d}'

        from decimal import Decimal as _Decimal
        try:
            earnings = _Decimal(str(driver_earnings))
        except Exception:
            earnings = _Decimal('0')

        batch = delivery_models.HubPickupBatch.objects.create(
            batch_number=batch_number,
            pickup_location=pickup_location,
            hub_warehouse=hub_warehouse,
            driver=driver,
            driver_earnings=earnings,
            notes=notes,
            created_by=request.user,
            status='assigned' if driver else 'pending',
        )

        # Assign orders to batch
        orders = orders_models.Order.objects.filter(
            id__in=order_ids,
            is_hub_delivery=True,
            hub_pickup_batch__isnull=True,
        )
        orders.update(hub_pickup_batch=batch)

        return JsonResponse({
            'success': True,
            'batch_id': batch.id,
            'batch_number': batch_number,
            'order_count': orders.count(),
            'redirect_url': f'/workforce/hub/batches/{batch.id}/',
        })

    # GET — render create form
    from warehouse.models import WarehouseLocation
    hub_warehouses = WarehouseLocation.objects.filter(is_active=True).select_related('warehouse').order_by('warehouse__name', 'name')
    pickup_locations = business_models.PickupLocation.objects.filter(
        pickup_status='active'
    ).select_related('business').order_by('business__business_name', 'pickup_location_title')
    eligible_orders = orders_models.Order.objects.filter(
        is_hub_delivery=True,
        hub_pickup_batch__isnull=True,
        verification_status='verified',
    ).select_related('business', 'pickup_location').order_by('-created_at')
    approved_drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('user__first_name')

    context = {
        'page_title': 'Create Hub Pickup Batch',
        'hub_warehouses': hub_warehouses,
        'pickup_locations': pickup_locations,
        'eligible_orders': eligible_orders,
        'approved_drivers': approved_drivers,
    }
    return render(request, 'workforce/hub_batch_create.html', context)


@login_required
@staff_required
def hub_batch_detail(request, batch_id):
    """Show hub pickup batch detail — orders, status, driver assignment."""
    from warehouse.models import WarehouseLocation

    batch = get_object_or_404(
        delivery_models.HubPickupBatch.objects.select_related(
            'pickup_location', 'hub_warehouse', 'hub_warehouse__warehouse',
            'driver', 'driver__user', 'created_by',
        ).prefetch_related('orders', 'orders__delivery_task', 'delivery_tasks'),
        id=batch_id,
    )
    approved_drivers = fleet_models.Driver.objects.filter(
        driver_status='approved'
    ).select_related('user').order_by('user__first_name')

    context = {
        'page_title': f'Hub Batch — {batch.batch_number}',
        'batch': batch,
        'approved_drivers': approved_drivers,
    }
    return render(request, 'workforce/hub_batch_detail.html', context)


@login_required
@staff_required
@require_POST
def hub_batch_assign_driver(request, batch_id):
    """Assign or reassign a driver to a hub pickup batch."""
    batch = get_object_or_404(delivery_models.HubPickupBatch, id=batch_id)

    if batch.status in ('at_hub', 'cancelled'):
        return JsonResponse({'success': False, 'error': f'Cannot assign driver to batch in status: {batch.status}'}, status=400)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        data = request.POST

    driver_id = data.get('driver_id')
    driver_earnings = data.get('driver_earnings')

    try:
        driver = fleet_models.Driver.objects.get(driver_id=driver_id, driver_status='approved')
    except fleet_models.Driver.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Driver not found or not approved.'}, status=400)

    update_fields = ['driver', 'status']
    batch.driver = driver
    if batch.status == 'pending':
        batch.status = 'assigned'
    if driver_earnings is not None:
        from decimal import Decimal as _Decimal
        try:
            batch.driver_earnings = _Decimal(str(driver_earnings))
            update_fields.append('driver_earnings')
        except Exception:
            pass
    batch.save(update_fields=update_fields)

    return JsonResponse({
        'success': True,
        'driver_name': driver.user.get_full_name() or driver.user.username,
        'batch_status': batch.status,
    })


@login_required
@staff_required
@require_POST
def hub_batch_update_status(request, batch_id):
    """Staff updates hub batch status (e.g. cancel, mark at_hub manually)."""
    batch = get_object_or_404(delivery_models.HubPickupBatch, id=batch_id)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        data = request.POST

    new_status = data.get('status')
    valid_statuses = [s[0] for s in delivery_models.HubPickupBatch.BATCH_STATUS_CHOICES]

    if new_status not in valid_statuses:
        return JsonResponse({'success': False, 'error': f'Invalid status: {new_status}'}, status=400)

    if batch.status in ('at_hub', 'cancelled'):
        return JsonResponse({'success': False, 'error': f'Batch already in terminal status: {batch.status}'}, status=400)

    batch._old_batch_status = batch.status
    batch.status = new_status
    batch.save(update_fields=['status'])

    return JsonResponse({'success': True, 'status': new_status})


# =============================================================================
# TEAM VERIFICATION
# =============================================================================

@login_required
@staff_required
def team_verification_list(request):
    """Staff view: list BusinessTeamProfile entries pending verification."""
    status_filter = request.GET.get('status', 'pending')

    qs = business_models.BusinessTeamProfile.objects.select_related(
        'user', 'user__profile', 'business', 'invited_by'
    )

    valid_filters = ('pending', 'active', 'rejected', 'suspended', 'inactive', 'all')
    if status_filter not in valid_filters:
        status_filter = 'pending'

    if status_filter != 'all':
        qs = qs.filter(team_status=status_filter)

    qs = qs.order_by('-created_at')

    pending_count = business_models.BusinessTeamProfile.objects.filter(team_status='pending').count()
    active_count = business_models.BusinessTeamProfile.objects.filter(team_status='active').count()

    context = {
        'team_profiles': qs,
        'pending_count': pending_count,
        'active_count': active_count,
        'current_filter': status_filter,
    }
    return render(request, 'workforce/team_verification_list.html', context)


@login_required(login_url='/accounts/login/')
@staff_required
def tasks_live_map(request):
    """Live map showing all active delivery task pins and driver locations."""
    from django.db.models import Subquery, OuterRef

    active_statuses = [
        'for_review', 'pending', 'assigned', 'accepted', 'picked_up',
        'start_ride', 'out_for_delivery', 'in_transit', 'contacted',
        'non_reachable', 'delivered', 'failed', 'rejected', 'cancelled',
    ]

    # Get all active tasks with coordinates
    tasks = delivery_models.DeliveryTask.objects.select_related(
        'order', 'order__business', 'driver', 'driver__user', 'dl_to_address',
    ).filter(
        dl_task_status__in=active_statuses,
    )

    pins = []
    for t in tasks:
        lat = lng = None
        if t.order.latitude and t.order.longitude:
            lat = float(t.order.latitude)
            lng = float(t.order.longitude)
        elif t.dl_to_address and t.dl_to_address.dl_latitude and t.dl_to_address.dl_longitude:
            lat = float(t.dl_to_address.dl_latitude)
            lng = float(t.dl_to_address.dl_longitude)

        pins.append({
            'id': t.id,
            'task_number': t.dl_task_number or str(t.id),
            'status': t.dl_task_status,
            'status_display': t.get_dl_task_status_display(),
            'customer_name': t.order.customer_name or '',
            'customer_phone': t.order.customer_phone or '',
            'zone': t.order.dl_zone or '',
            'street': t.order.dl_street or '',
            'building': t.order.dl_building or '',
            'address': t.order.customer_address or '',
            'driver_name': str(t.driver) if t.driver else 'Unassigned',
            'driver_id': t.driver_id,
            'business_name': t.order.business.business_name if t.order.business else '',
            'lat': lat,
            'lng': lng,
        })

    # Get latest GPS location for drivers
    # Include drivers with active tasks AND any driver with recent GPS pings (online)
    active_driver_ids = set(t.driver_id for t in tasks if t.driver_id)
    driver_locations = []
    drivers_with_gps = set()

    # Get ALL recent GPS pings (last 2 hours) — catches online drivers with no tasks
    recent_cutoff = timezone.now() - timezone.timedelta(hours=2)
    latest_locs = fleet_models.DriverLocation.objects.filter(
        created_at__gte=recent_cutoff,
    ).order_by('driver_id', '-created_at').distinct('driver_id')

    # Build a driver lookup for names
    gps_driver_ids = set(loc.driver_id for loc in latest_locs)
    all_relevant_driver_ids = active_driver_ids | gps_driver_ids
    driver_map = {}
    if all_relevant_driver_ids:
        for d in fleet_models.Driver.objects.select_related('user').filter(
            driver_id__in=all_relevant_driver_ids
        ):
            driver_map[d.driver_id] = d

    for loc in latest_locs:
        drivers_with_gps.add(loc.driver_id)
        driver = driver_map.get(loc.driver_id)
        task_count = sum(1 for t in tasks if t.driver_id == loc.driver_id)
        driver_locations.append({
            'driver_id': loc.driver_id,
            'driver_name': str(driver) if driver else f'Driver #{loc.driver_id}',
            'lat': float(loc.latitude),
            'lng': float(loc.longitude),
            'accuracy': loc.accuracy,
            'speed': loc.speed,
            'updated': loc.created_at.strftime('%H:%M'),
            'minutes_ago': int((timezone.now() - loc.created_at).total_seconds() / 60),
            'has_gps': True,
            'task_count': task_count,
        })

    # For drivers with active tasks but no GPS pings, show at task location
    for driver_id in active_driver_ids - drivers_with_gps:
        driver = driver_map.get(driver_id)
        fallback_lat = fallback_lng = None
        task_count = 0
        for t in tasks:
            if t.driver_id == driver_id:
                task_count += 1
                if fallback_lat is None:
                    if t.order.latitude and t.order.longitude:
                        fallback_lat = float(t.order.latitude)
                        fallback_lng = float(t.order.longitude)
                    elif t.dl_to_address and t.dl_to_address.dl_latitude and t.dl_to_address.dl_longitude:
                        fallback_lat = float(t.dl_to_address.dl_latitude)
                        fallback_lng = float(t.dl_to_address.dl_longitude)
        driver_locations.append({
            'driver_id': driver_id,
            'driver_name': str(driver) if driver else f'Driver #{driver_id}',
            'lat': fallback_lat,
            'lng': fallback_lng,
            'accuracy': None,
            'speed': None,
            'updated': None,
            'minutes_ago': -1,
            'has_gps': False,
            'task_count': task_count,
        })

    # Static mode: show each driver at their latest task's delivery location
    # Drivers with GPS but no tasks still show at GPS position
    mode = request.GET.get('mode', 'live')
    if mode == 'static':
        static_locations = []
        # Drivers with active tasks — show at latest task location
        for driver_id in active_driver_ids:
            driver = driver_map.get(driver_id)
            latest_task = None
            task_count = 0
            for t in tasks:
                if t.driver_id == driver_id:
                    task_count += 1
                    if latest_task is None or t.updated_at > latest_task.updated_at:
                        latest_task = t
            if latest_task:
                lat = lng = None
                if latest_task.order.latitude and latest_task.order.longitude:
                    lat = float(latest_task.order.latitude)
                    lng = float(latest_task.order.longitude)
                elif latest_task.dl_to_address and latest_task.dl_to_address.dl_latitude and latest_task.dl_to_address.dl_longitude:
                    lat = float(latest_task.dl_to_address.dl_latitude)
                    lng = float(latest_task.dl_to_address.dl_longitude)
                minutes_ago = int((timezone.now() - latest_task.updated_at).total_seconds() / 60)
                static_locations.append({
                    'driver_id': driver_id,
                    'driver_name': str(driver) if driver else f'Driver #{driver_id}',
                    'lat': lat,
                    'lng': lng,
                    'accuracy': None,
                    'speed': None,
                    'updated': latest_task.updated_at.strftime('%H:%M'),
                    'minutes_ago': minutes_ago,
                    'has_gps': False,
                    'task_count': task_count,
                    'last_status': latest_task.get_dl_task_status_display(),
                    'last_task_number': latest_task.dl_task_number or str(latest_task.id),
                })
        # GPS-only drivers (no tasks) — keep their GPS position in static mode too
        for loc_entry in driver_locations:
            if loc_entry['driver_id'] not in active_driver_ids and loc_entry.get('has_gps'):
                loc_entry['last_status'] = 'Online (no tasks)'
                static_locations.append(loc_entry)
        driver_locations = static_locations

    # Return JSON for AJAX refresh requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'pins': pins,
            'drivers': driver_locations,
            'pin_count': len([p for p in pins if p['lat']]),
            'driver_count': len(driver_locations),
            'mode': mode,
        })

    context = {
        'pins_json': json.dumps(pins),
        'drivers_json': json.dumps(driver_locations),
        'pin_count': len([p for p in pins if p['lat']]),
        'total_count': len(pins),
        'driver_count': len(driver_locations),
    }
    return render(request, 'workforce/tasks_live_map.html', context)


# =============================================================================
# ONEDRIVE IMPORT SOURCES
# =============================================================================

@login_required(login_url='/accounts/login/')
@staff_required
def onedrive_sources(request):
    """List and manage OneDrive import sources."""
    sources = orders_models.OneDriveSource.objects.select_related(
        'business', 'last_import_by'
    ).order_by('-created_at')

    businesses = business_models.Business.objects.filter(
        business_status='active'
    ).order_by('business_name')

    # Handle add source
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            business_id = request.POST.get('business_id')
            label = request.POST.get('label', '').strip()
            share_link = request.POST.get('share_link', '').strip()

            if business_id and label and share_link:
                try:
                    business = business_models.Business.objects.get(business_id=business_id)
                    orders_models.OneDriveSource.objects.create(
                        business=business, label=label, share_link=share_link
                    )
                    return JsonResponse({'success': True, 'message': 'Source added'})
                except business_models.Business.DoesNotExist:
                    return JsonResponse({'success': False, 'error': 'Business not found'}, status=400)
            return JsonResponse({'success': False, 'error': 'All fields required'}, status=400)

        elif action == 'edit':
            source_id = request.POST.get('source_id')
            label = request.POST.get('label', '').strip()
            share_link = request.POST.get('share_link', '').strip()
            if source_id and label and share_link:
                orders_models.OneDriveSource.objects.filter(id=source_id).update(
                    label=label, share_link=share_link
                )
                return JsonResponse({'success': True, 'message': 'Source updated'})
            return JsonResponse({'success': False, 'error': 'Label and link required'}, status=400)

        elif action == 'delete':
            source_id = request.POST.get('source_id')
            orders_models.OneDriveSource.objects.filter(id=source_id).delete()
            return JsonResponse({'success': True, 'message': 'Source deleted'})

        elif action == 'toggle':
            source_id = request.POST.get('source_id')
            source = orders_models.OneDriveSource.objects.get(id=source_id)
            source.is_active = not source.is_active
            source.save()
            return JsonResponse({'success': True, 'is_active': source.is_active})

    import json as json_lib
    # Serialize saved mappings and imported rows per source for JS
    source_mappings = {}
    source_imported_rows = {}
    for s in sources:
        source_mappings[s.id] = s.last_column_mapping or {}
        source_imported_rows[s.id] = s.last_imported_rows or []

    context = {
        'page_title': 'OneDrive Import Sources',
        'sources': sources,
        'businesses': businesses,
        'source_mappings_json': json_lib.dumps(source_mappings),
        'source_imported_rows_json': json_lib.dumps(source_imported_rows),
    }
    return render(request, 'workforce/onedrive_sources.html', context)


def _onedrive_download_file(source):
    """Download Excel file from OneDrive source. Returns (bytes, error_msg)."""
    import requests as http_requests

    link = source.share_link.strip()

    try:
        # Strategy: Follow the share link without redirects to get the intermediate URL,
        # then append &download=1 to force file download instead of web view.
        resp1 = http_requests.get(link, timeout=30, allow_redirects=False)

        if resp1.status_code in (301, 302, 307, 308):
            redirect_url = resp1.headers.get('Location', '')
            if redirect_url:
                # Append download=1 to the redirect URL
                sep = '&' if '?' in redirect_url else '?'
                download_url = redirect_url + sep + 'download=1'
                resp = http_requests.get(download_url, timeout=60, allow_redirects=True)
                content_type = resp.headers.get('Content-Type', '')

                if resp.status_code == 200 and 'text/html' not in content_type:
                    return resp.content, None

        # Fallback: try the model's get_download_url() method
        download_url = source.get_download_url()
        resp = http_requests.get(download_url, timeout=30, allow_redirects=True)
        content_type = resp.headers.get('Content-Type', '')

        if resp.status_code == 200 and 'text/html' not in content_type:
            return resp.content, None

        return None, 'Could not download the file. Make sure it is shared with "Anyone with the link" in OneDrive.'

    except http_requests.exceptions.Timeout:
        return None, 'Download timed out. The file may be too large or OneDrive is slow.'
    except Exception as e:
        return None, f'Download error: {str(e)}'


@login_required(login_url='/accounts/login/')
@staff_required
@require_POST
def onedrive_fetch_sheets(request, source_id):
    """Fetch Excel from OneDrive and return list of sheet names with row counts."""
    import openpyxl
    from io import BytesIO

    source = get_object_or_404(orders_models.OneDriveSource, id=source_id)

    try:
        file_bytes, err = _onedrive_download_file(source)
        if err:
            return JsonResponse({'success': False, 'error': err}, status=400)

        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as parse_err:
            return JsonResponse({
                'success': False,
                'error': f'Could not read Excel file: {str(parse_err)}. Make sure it is an .xlsx file.'
            }, status=400)

        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            row_count = 0
            for _ in ws.iter_rows(values_only=True):
                row_count += 1
            sheets.append({'name': name, 'rows': max(0, row_count - 1)})  # exclude header
        wb.close()

        return JsonResponse({'success': True, 'sheets': sheets})

    except Exception as e:
        logger.exception("OneDrive fetch sheets error for source %s: %s", source_id, str(e))
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
@require_POST
def onedrive_sheet_preview(request, source_id):
    """Fetch Excel and return preview rows from a specific sheet. Paginates from the top."""
    import openpyxl
    from io import BytesIO

    source = get_object_or_404(orders_models.OneDriveSource, id=source_id)
    sheet_name = request.POST.get('sheet_name', '')
    offset = int(request.POST.get('offset', 0))
    limit = int(request.POST.get('limit', 10))

    try:
        file_bytes, err = _onedrive_download_file(source)
        if err:
            return JsonResponse({'success': False, 'error': err}, status=400)

        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as parse_err:
            return JsonResponse({'success': False, 'error': str(parse_err)}, status=400)

        if sheet_name not in wb.sheetnames:
            wb.close()
            return JsonResponse({'success': False, 'error': f'Sheet "{sheet_name}" not found'}, status=400)

        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            return JsonResponse({'success': True, 'headers': [], 'rows': [], 'total': 0, 'has_more': False})

        # First row = headers
        headers = [str(h).strip() if h else '' for h in all_rows[0]]
        data_rows = all_rows[1:]
        total = len(data_rows)

        if total == 0:
            return JsonResponse({'success': True, 'headers': headers, 'rows': [], 'total': 0, 'has_more': False})

        # Paginate from top: offset=0 means rows 1-10, offset=10 means rows 11-20, etc.
        start_idx = offset
        end_idx = min(total, offset + limit)

        if start_idx >= total:
            return JsonResponse({'success': True, 'headers': headers, 'rows': [], 'total': total, 'has_more': False})

        sliced = data_rows[start_idx:end_idx]
        row_list = []
        for i, row in enumerate(sliced):
            row_num = start_idx + i + 2  # +2 for 1-indexed + header row
            cells = [str(c).strip() if c is not None else '' for c in row]
            row_list.append({'row_num': row_num, 'cells': cells})

        has_more = end_idx < total

        return JsonResponse({
            'success': True,
            'headers': headers,
            'rows': row_list,
            'total': total,
            'has_more': has_more,
            'next_offset': offset + limit,
        })

    except Exception as e:
        logger.exception("OneDrive sheet preview error for source %s: %s", source_id, str(e))
        return JsonResponse({'success': False, 'error': f'Error: {str(e)}'}, status=500)


@login_required(login_url='/accounts/login/')
@staff_required
@require_POST
def onedrive_import_trigger(request, source_id):
    """Import selected rows from OneDrive Excel using user-defined column mapping."""
    import openpyxl
    import uuid
    import json as json_lib
    from io import BytesIO
    from core.utils import (
        contains_arabic, translate_to_english,
        convert_arabic_numerals, format_whatsapp_number
    )

    source = get_object_or_404(orders_models.OneDriveSource, id=source_id)
    business = source.business

    try:
        body = json_lib.loads(request.body)
    except (json_lib.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

    sheet_name = body.get('sheet_name', '')
    row_numbers = set(body.get('row_numbers', []))  # Excel row numbers to import
    column_mapping = body.get('column_mapping', {})  # {col_idx_str: db_field}

    if not row_numbers:
        return JsonResponse({'success': False, 'error': 'No rows selected'}, status=400)
    if not column_mapping:
        return JsonResponse({'success': False, 'error': 'No column mapping provided'}, status=400)

    # Build field_indices: {db_field: col_idx}
    field_indices = {}
    for col_idx_str, db_field in column_mapping.items():
        if db_field:
            field_indices[db_field] = int(col_idx_str)

    # Check required
    required = ['customer_name', 'customer_phone', 'customer_address']
    missing = [f for f in required if f not in field_indices]
    if missing:
        return JsonResponse({
            'success': False,
            'error': f'Required fields not mapped: {", ".join(missing)}'
        }, status=400)

    try:
        file_bytes, err = _onedrive_download_file(source)
        if err:
            return JsonResponse({'success': False, 'error': err}, status=400)

        try:
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as parse_err:
            return JsonResponse({'success': False, 'error': f'Could not read file: {str(parse_err)}'}, status=400)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        def get_val(row_data, field):
            idx = field_indices.get(field)
            if idx is not None and idx < len(row_data):
                val = row_data[idx]
                return str(val).strip() if val is not None else ''
            return ''

        def safe_int(val):
            try:
                return int(float(val)) if val else 0
            except (ValueError, TypeError):
                return 0

        def safe_int_or_none(val):
            try:
                if not val or (isinstance(val, str) and not val.strip()):
                    return None
                return int(float(val))
            except (ValueError, TypeError):
                return None

        saved = 0
        skipped = 0
        errors = []
        imported_rows_data = []
        start_row = min(row_numbers) if row_numbers else 2

        data_rows = rows[1:]  # skip header
        for row_num, row_data in enumerate(data_rows, start=2):
            if row_num not in row_numbers:
                continue
            if not any(row_data):
                skipped += 1
                continue

            customer_name = get_val(row_data, 'customer_name')
            customer_phone = get_val(row_data, 'customer_phone')
            customer_address = get_val(row_data, 'customer_address')

            if not customer_name or not customer_phone:
                skipped += 1
                continue

            client_order_code = get_val(row_data, 'client_order_code')
            if not client_order_code:
                client_order_code = f"OD-{uuid.uuid4().hex[:8].upper()}"

            if orders_models.Order.objects.filter(client_order_code=client_order_code).exists():
                skipped += 1
                errors.append(f"Row {row_num}: Duplicate '{client_order_code}'")
                continue

            try:
                customer_phone = convert_arabic_numerals(customer_phone)
                raw_whatsapp = get_val(row_data, 'customer_whatsapp')
                customer_whatsapp = format_whatsapp_number(raw_whatsapp) if raw_whatsapp else format_whatsapp_number(customer_phone)

                if contains_arabic(customer_name):
                    customer_name = translate_to_english(customer_name)
                if contains_arabic(customer_address):
                    customer_address = translate_to_english(customer_address)

                seller_notes = get_val(row_data, 'seller_notes')
                internal_notes = get_val(row_data, 'internal_notes')
                notes_parts = []
                if seller_notes:
                    notes_parts.append(f"Seller: {seller_notes}")
                if internal_notes:
                    notes_parts.append(f"Ezzy: {internal_notes}")
                combined_notes = ' | '.join(notes_parts) if notes_parts else ''

                order = orders_models.Order(
                    business=business,
                    client_order_code=client_order_code,
                    customer_name=customer_name,
                    customer_phone=customer_phone,
                    customer_whatsapp=customer_whatsapp,
                    customer_address=customer_address or '',
                    dl_zone=safe_int_or_none(get_val(row_data, 'dl_zone')),
                    dl_street=safe_int_or_none(get_val(row_data, 'dl_street')),
                    dl_building=safe_int_or_none(get_val(row_data, 'dl_building')),
                    cod_amount=safe_int(get_val(row_data, 'cod_amount')),
                    dl_amount=safe_int(get_val(row_data, 'dl_amount') if 'dl_amount' in field_indices else '0'),
                    order_notes=combined_notes[:100] if combined_notes else '',
                    deadline_date=get_val(row_data, 'deadline_date'),
                    order_status='to_review',
                    verification_status='pending',
                    original_order_data={
                        'source': 'onedrive_import',
                        'onedrive_source_id': source.id,
                        'sheet_name': sheet_name,
                        'row_number': row_num,
                    },
                )
                # Save product description + total qty on order (not as items)
                desc_parts = []
                total_qty = 0
                product_name = get_val(row_data, 'product_name')
                if product_name:
                    main_qty = safe_int(get_val(row_data, 'quantity')) or 1
                    desc_parts.append(f"{product_name} x{main_qty}")
                    total_qty += main_qty

                for i in range(1, 6):
                    pn = get_val(row_data, f'product_{i}')
                    if pn:
                        pc = safe_int(get_val(row_data, f'count_{i}')) or 1
                        desc_parts.append(f"{pn} x{pc}")
                        total_qty += pc

                if desc_parts:
                    order.package_description = ', '.join(desc_parts)[:255]
                    order.total_quantity = total_qty

                order.save()

                imported_rows_data.append({'row': row_num, 'order': order.order_number})
                saved += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        from django.utils import timezone
        # Merge with previously imported rows (keep old + add new)
        existing_imported = source.last_imported_rows or []
        existing_imported.extend(imported_rows_data)
        source.last_import_at = timezone.now()
        source.last_import_count = saved
        source.last_import_by = request.user
        source.last_sheet_name = sheet_name
        source.last_start_row = start_row
        source.last_import_max_row = max(row_numbers) if row_numbers else 0
        source.last_imported_rows = existing_imported
        source.last_column_mapping = column_mapping
        source.save()

        return JsonResponse({
            'success': True,
            'saved': saved,
            'skipped': skipped,
            'errors': errors[:20],
            'imported': imported_rows_data[:20],
            'message': f'{saved} orders imported, {skipped} skipped'
        })

    except Exception as e:
        logger.exception("OneDrive import error for source %s: %s", source_id, str(e))
        return JsonResponse({'success': False, 'error': f'Import failed: {str(e)}'}, status=500)
